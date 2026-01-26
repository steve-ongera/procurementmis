from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from pms.models import *


def login_view(request):
    """
    Universal login view that authenticates users and redirects to role-based dashboards
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active_user:
                login(request, user)
                
                # Set session expiry based on remember me
                if not remember_me:
                    request.session.set_expiry(0)  # Session expires when browser closes
                else:
                    request.session.set_expiry(1209600)  # 2 weeks
                
                # Log the login
                AuditLog.objects.create(
                    user=user,
                    action='LOGIN',
                    model_name='User',
                    object_id=str(user.id),
                    object_repr=str(user),
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect to next parameter or dashboard
                next_url = request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('dashboard')
            else:
                messages.error(request, 'Your account has been deactivated. Please contact the administrator.')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'auth/login.html')


def logout_view(request):
    """
    Logout view
    """
    if request.user.is_authenticated:
        # Log the logout
        AuditLog.objects.create(
            user=request.user,
            action='LOGOUT',
            model_name='User',
            object_id=str(request.user.id),
            object_repr=str(request.user),
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
    
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


@login_required
def dashboard_view(request):
    """
    Role-based dashboard router
    """
    user = request.user
    role = user.role
    
    # Route to appropriate dashboard based on role
    if role == 'ADMIN':
        return admin_dashboard(request)
    elif role == 'STAFF':
        return staff_dashboard(request)
    elif role == 'HOD':
        return hod_dashboard(request)
    elif role == 'PROCUREMENT':
        return procurement_dashboard_view(request)
    elif role == 'FINANCE':
        return finance_dashboard_view(request)
    elif role == 'STORES':
        return stores_dashboard(request)
    elif role == 'SUPPLIER':
        return supplier_dashboard(request)
    elif role == 'AUDITOR':
        return auditor_dashboard(request)
    else:
        messages.error(request, 'Invalid user role.')
        return redirect('login')



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings
from django.utils.crypto import get_random_string
from .models import User, Supplier, ItemCategory, AuditLog
import logging

logger = logging.getLogger(__name__)

def register_supplier(request):
    """
    Supplier self-registration view.
    Only suppliers can register themselves; other user types are created by admin.
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Extract form data
                company_name = request.POST.get('company_name')
                registration_number = request.POST.get('registration_number')
                tax_id = request.POST.get('tax_id', '')
                email = request.POST.get('email')
                phone_number = request.POST.get('phone_number')
                physical_address = request.POST.get('physical_address')
                postal_address = request.POST.get('postal_address', '')
                website = request.POST.get('website', '')
                
                # Contact person details
                contact_person = request.POST.get('contact_person')
                contact_person_phone = request.POST.get('contact_person_phone')
                contact_person_email = request.POST.get('contact_person_email')
                
                # Banking details
                bank_name = request.POST.get('bank_name')
                bank_branch = request.POST.get('bank_branch')
                account_number = request.POST.get('account_number')
                account_name = request.POST.get('account_name')
                swift_code = request.POST.get('swift_code', '')
                
                # Categories
                category_ids = request.POST.getlist('categories')
                
                # Validation
                if User.objects.filter(email=email).exists():
                    messages.error(request, 'A user with this email already exists.')
                    return redirect('register_supplier')
                
                if Supplier.objects.filter(registration_number=registration_number).exists():
                    messages.error(request, 'A supplier with this registration number already exists.')
                    return redirect('register_supplier')
                
                # Generate username and password
                username = email.split('@')[0] + '_' + get_random_string(4)
                temp_password = get_random_string(12)
                
                # Create User account
                user = User.objects.create(
                    username=username,
                    email=email,
                    first_name=contact_person.split()[0] if contact_person else '',
                    last_name=' '.join(contact_person.split()[1:]) if contact_person and len(contact_person.split()) > 1 else '',
                    role='SUPPLIER',
                    phone_number=contact_person_phone,
                    is_active=False,  # Inactive until admin approval
                    password=make_password(temp_password)
                )
                
                # Generate supplier number
                year = timezone.now().year
                last_supplier = Supplier.objects.filter(
                    supplier_number__startswith=f'SUP-{year}'
                ).order_by('-supplier_number').first()
                
                if last_supplier:
                    last_number = int(last_supplier.supplier_number.split('-')[-1])
                    new_number = last_number + 1
                else:
                    new_number = 1
                
                supplier_number = f'SUP-{year}-{new_number:06d}'
                
                # Create Supplier profile
                supplier = Supplier.objects.create(
                    user=user,
                    supplier_number=supplier_number,
                    name=company_name,
                    registration_number=registration_number,
                    tax_id=tax_id,
                    email=email,
                    phone_number=phone_number,
                    physical_address=physical_address,
                    postal_address=postal_address,
                    website=website,
                    contact_person=contact_person,
                    contact_person_phone=contact_person_phone,
                    contact_person_email=contact_person_email,
                    bank_name=bank_name,
                    bank_branch=bank_branch,
                    account_number=account_number,
                    account_name=account_name,
                    swift_code=swift_code,
                    status='PENDING',
                    created_by=None
                )
                
                # Add categories
                if category_ids:
                    valid_categories = [cid for cid in category_ids if cid and cid.strip()]
                    if valid_categories:
                        supplier.categories.set(valid_categories)
                
                # Handle document uploads
                documents = {
                    'REGISTRATION': request.FILES.get('registration_cert'),
                    'TAX': request.FILES.get('tax_cert'),
                    'BANK': request.FILES.get('bank_statement'),
                }
                
                from .models import SupplierDocument
                for doc_type, file in documents.items():
                    if file:
                        SupplierDocument.objects.create(
                            supplier=supplier,
                            document_type=doc_type,
                            document_name=file.name,
                            file=file,
                            is_verified=False
                        )
                
                # Create audit log
                AuditLog.objects.create(
                    user=user,
                    action='CREATE',
                    model_name='Supplier',
                    object_id=str(supplier.id),
                    object_repr=supplier.supplier_number,
                    changes={'status': 'Self-registered, pending approval'}
                )
                
                # Send welcome email with credentials
                try:
                    email_subject = 'Welcome to University Procurement System'
                    email_body = f"""
Dear {contact_person},

Thank you for registering with the University Procurement System!

Your supplier account has been created successfully. Here are your login credentials:

Supplier Number: {supplier_number}
Username: {username}
Temporary Password: {temp_password}
Login URL: {request.build_absolute_uri('/login/')}

IMPORTANT SECURITY NOTICE:
- Please change your password immediately after your first login
- Keep your credentials confidential
- Your account is currently pending approval by our procurement team

NEXT STEPS:
1. Your account will be reviewed by our procurement team
2. You may be contacted for additional verification documents
3. Once approved, you will receive a confirmation email
4. You can then participate in tenders and quotations

If you have any questions, please contact us at:
Email: procurement@university.edu
Phone: [Your Contact Number]

Best regards,
University Procurement Team

---
This is an automated message. Please do not reply to this email.
                    """
                    
                    send_mail(
                        email_subject,
                        email_body,
                        settings.DEFAULT_FROM_EMAIL,
                        [email, contact_person_email],
                        fail_silently=False,
                    )
                    
                    logger.info(f'Registration email sent to {email}')
                    
                except Exception as e:
                    logger.error(f'Failed to send registration email: {str(e)}')
                    # Don't fail registration if email fails
                
                messages.success(
                    request, 
                    f'Registration successful! Your supplier number is {supplier_number}. '
                    f'Login credentials have been sent to {email}. '
                    f'Your account is pending approval by the procurement team.'
                )
                return redirect('login')
                
        except Exception as e:
            messages.error(request, f'Registration failed: {str(e)}')
            logger.error(f'Supplier registration error: {str(e)}')
            import traceback
            logger.error(traceback.format_exc())
    
    # GET request - show registration form
    context = {
        'categories': ItemCategory.objects.filter(is_active=True).order_by('name'),
    }
    
    return render(request, 'accounts/register_supplier.html', context)


from django.shortcuts import render
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncMonth, TruncDate
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
import json

from .models import (
    User, Requisition, PurchaseOrder, Supplier, Contract, Invoice,
    Department, AuditLog, Budget, Tender, Bid, GoodsReceivedNote,
    Payment, StockItem, Asset, StockMovement
)


def admin_dashboard(request):
    """Admin Dashboard with Comprehensive Analytics"""
    
    # Time ranges
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    ninety_days_ago = today - timedelta(days=90)
    year_start = today.replace(month=1, day=1)
    
    # ==================== Basic Statistics ====================
    total_users = User.objects.filter(is_active_user=True).count()
    total_requisitions = Requisition.objects.count()
    total_pos = PurchaseOrder.objects.count()
    total_suppliers = Supplier.objects.filter(status='APPROVED').count()
    pending_approvals = Requisition.objects.filter(
        status__in=['SUBMITTED', 'HOD_APPROVED', 'BUDGET_APPROVED']
    ).count()
    
    # ==================== Financial Analytics ====================
    # Total spend by status
    total_spend = Invoice.objects.filter(status='PAID').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    
    # Monthly spend trend (last 12 months)
    twelve_months_ago = today - timedelta(days=365)
    monthly_spend_data = Invoice.objects.filter(
        status='PAID',
        payment_date__gte=twelve_months_ago,
        payment_date__isnull=False
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('total_amount')
    ).order_by('month')
    
    # Ensure we have data for all 12 months
    from dateutil.relativedelta import relativedelta
    spend_labels = []
    spend_values = []
    
    current_date = twelve_months_ago.replace(day=1)
    monthly_data_dict = {item['month'].strftime('%Y-%m'): float(item['total']) for item in monthly_spend_data}
    
    for i in range(12):
        month_key = current_date.strftime('%Y-%m')
        spend_labels.append(current_date.strftime('%b %Y'))
        spend_values.append(monthly_data_dict.get(month_key, 0))
        current_date = current_date + relativedelta(months=1)
    
    # Budget utilization
    budget_stats = Budget.objects.filter(
        is_active=True
    ).aggregate(
        total_allocated=Sum('allocated_amount'),
        total_committed=Sum('committed_amount'),
        total_spent=Sum('actual_spent')
    )
    
    budget_allocated = float(budget_stats['total_allocated'] or 0)
    budget_committed = float(budget_stats['total_committed'] or 0)
    budget_spent = float(budget_stats['total_spent'] or 0)
    budget_available = budget_allocated - budget_committed - budget_spent
    
    # ==================== Requisition Analytics ====================
    # Requisitions by status
    req_by_status = Requisition.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    req_status_labels = [item['status'].replace('_', ' ').title() for item in req_by_status]
    req_status_values = [item['count'] for item in req_by_status]
    
    # Requisition trends (last 90 days)
    req_trend_data = Requisition.objects.filter(
        created_at__date__gte=ninety_days_ago
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id')
    ).order_by('date')
    
    req_trend_labels = [item['date'].strftime('%b %d') for item in req_trend_data]
    req_trend_values = [item['count'] for item in req_trend_data]
    
    # Average approval time (in days)
    avg_approval_time = 0
    approved_reqs = Requisition.objects.filter(
        status='APPROVED',
        submitted_at__isnull=False
    )
    
    if approved_reqs.exists():
        total_days = 0
        count = 0
        for req in approved_reqs:
            if req.submitted_at:
                days = (req.updated_at - req.submitted_at).days
                total_days += days
                count += 1
        avg_approval_time = round(total_days / count, 1) if count > 0 else 0
    
    # ==================== Procurement Analytics ====================
    # PO by status
    po_by_status = PurchaseOrder.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    po_status_labels = [item['status'].replace('_', ' ').title() for item in po_by_status]
    po_status_values = [item['count'] for item in po_by_status]
    
    # Monthly PO trend (last 6 months)
    six_months_ago = today - timedelta(days=180)
    po_monthly_data = PurchaseOrder.objects.filter(
        created_at__date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id'),
        total_value=Sum('total_amount')
    ).order_by('month')
    
    po_trend_labels = [item['month'].strftime('%b %Y') for item in po_monthly_data]
    po_trend_count = [item['count'] for item in po_monthly_data]
    po_trend_value = [float(item['total_value'] or 0) for item in po_monthly_data]
    
    # ==================== Supplier Analytics ====================
    # Top 10 suppliers by transaction value
    top_suppliers = Supplier.objects.filter(
        purchase_orders__status__in=['DELIVERED', 'CLOSED']
    ).annotate(
        total_value=Sum('purchase_orders__total_amount'),
        po_count=Count('purchase_orders')
    ).order_by('-total_value')[:10]
    
    top_supplier_names = [s.name[:20] + '...' if len(s.name) > 20 else s.name for s in top_suppliers]
    top_supplier_values = [float(s.total_value) if s.total_value else 0 for s in top_suppliers]
    
    # Supplier by status
    supplier_status = Supplier.objects.values('status').annotate(
        count=Count('id')
    )
    
    supplier_status_labels = [item['status'].title() for item in supplier_status]
    supplier_status_values = [item['count'] for item in supplier_status]
    
    # Average supplier rating
    avg_supplier_rating = Supplier.objects.filter(
        status='APPROVED'
    ).aggregate(avg_rating=Avg('rating'))['avg_rating'] or 0
    
    # ==================== Department Analytics ====================
    # Top 5 departments by spend
    top_depts = Department.objects.filter(
        is_active=True,
        requisitions__purchase_orders__isnull=False
    ).annotate(
        total_spend=Sum('requisitions__purchase_orders__total_amount')
    ).order_by('-total_spend')[:5]
    
    dept_names = [dept.name for dept in top_depts]
    dept_values = [float(dept.total_spend) if dept.total_spend else 0 for dept in top_depts]
    
    # Requisitions by department
    req_by_dept = Department.objects.filter(
        is_active=True
    ).annotate(
        req_count=Count('requisitions')
    ).order_by('-req_count')[:8]
    
    req_dept_labels = [dept.code for dept in req_by_dept]
    req_dept_values = [dept.req_count for dept in req_by_dept]
    
    # ==================== Tender & Bid Analytics ====================
    # Active tenders
    active_tenders = Tender.objects.filter(status='PUBLISHED').count()
    
    # Tender by status
    tender_status = Tender.objects.values('status').annotate(
        count=Count('id')
    )
    
    tender_status_labels = [item['status'].title() for item in tender_status]
    tender_status_values = [item['count'] for item in tender_status]
    
    # Average bids per tender
    avg_bids = Tender.objects.filter(
        status__in=['CLOSED', 'EVALUATING', 'AWARDED']
    ).annotate(
        bid_count=Count('bids')
    ).aggregate(avg=Avg('bid_count'))['avg'] or 0
    
    # ==================== Contract Analytics ====================
    active_contracts = Contract.objects.filter(status='ACTIVE').count()
    expiring_soon = Contract.objects.filter(
        status='ACTIVE',
        end_date__lte=today + timedelta(days=30),
        end_date__gte=today
    ).count()
    
    # Contract value by type
    contract_by_type = Contract.objects.filter(
        status='ACTIVE'
    ).values('contract_type').annotate(
        total=Sum('contract_value')
    )
    
    contract_type_labels = [item['contract_type'].replace('_', ' ').title() for item in contract_by_type]
    contract_type_values = [float(item['total']) for item in contract_by_type]
    
    # ==================== Inventory Analytics ====================
    # Low stock items
    low_stock_items = StockItem.objects.filter(
        quantity_on_hand__lte=F('reorder_level')
    ).count()
    
    # Total inventory value
    total_inventory_value = StockItem.objects.aggregate(
        total=Sum('total_value')
    )['total'] or Decimal('0')
    
    # Stock movements (last 30 days)
    stock_movements = StockMovement.objects.filter(
        movement_date__date__gte=thirty_days_ago
    ).values('movement_type').annotate(
        count=Count('id')
    )
    
    stock_movement_labels = [item['movement_type'].replace('_', ' ').title() for item in stock_movements]
    stock_movement_values = [item['count'] for item in stock_movements]
    
    # ==================== Recent Activities ====================
    recent_activities = AuditLog.objects.select_related('user').all()[:15]
    
    # ==================== System Health Indicators ====================
    # User activity (last 7 days)
    active_users_week = AuditLog.objects.filter(
        action='LOGIN',
        timestamp__gte=today - timedelta(days=7)
    ).values('user').distinct().count()
    
    # Pending items summary
    pending_summary = {
        'requisitions': Requisition.objects.filter(status='SUBMITTED').count(),
        'approvals': pending_approvals,
        'invoices': Invoice.objects.filter(status='SUBMITTED').count(),
        'grns': GoodsReceivedNote.objects.filter(status='DRAFT').count(),
    }
    
    # ==================== Performance Metrics ====================
    # On-time delivery compliance
    on_time_deliveries = GoodsReceivedNote.objects.filter(
        delivery_date__lte=F('purchase_order__delivery_date')
    ).count()
    total_deliveries = GoodsReceivedNote.objects.count()
    delivery_compliance = round((on_time_deliveries / total_deliveries * 100), 1) if total_deliveries > 0 else 0
    
    # ==================== Context Assembly ====================
    context = {
        # Basic stats
        'total_users': total_users,
        'total_requisitions': total_requisitions,
        'total_pos': total_pos,
        'total_suppliers': total_suppliers,
        'pending_approvals': pending_approvals,
        'recent_activities': recent_activities,
        
        # System stats
        'system_stats': {
            'departments': Department.objects.filter(is_active=True).count(),
            'active_contracts': active_contracts,
            'total_spend': total_spend,
            'active_tenders': active_tenders,
            'low_stock_items': low_stock_items,
            'expiring_contracts': expiring_soon,
            'active_users_week': active_users_week,
            'avg_supplier_rating': round(float(avg_supplier_rating), 2),
            'inventory_value': total_inventory_value,
        },
        
        # Financial charts
        'spend_chart': {
            'labels': json.dumps(spend_labels),
            'values': json.dumps(spend_values),
        },
        'budget_chart': {
            'labels': json.dumps(['Spent', 'Committed', 'Available']),
            'values': json.dumps([budget_spent, budget_committed, budget_available]),
        },
        
        # Requisition charts
        'req_status_chart': {
            'labels': json.dumps(req_status_labels),
            'values': json.dumps(req_status_values),
        },
        'req_trend_chart': {
            'labels': json.dumps(req_trend_labels),
            'values': json.dumps(req_trend_values),
        },
        
        # Purchase Order charts
        'po_status_chart': {
            'labels': json.dumps(po_status_labels),
            'values': json.dumps(po_status_values),
        },
        'po_trend_chart': {
            'labels': json.dumps(po_trend_labels),
            'count': json.dumps(po_trend_count),
            'value': json.dumps(po_trend_value),
        },
        
        # Supplier charts
        'top_suppliers_chart': {
            'labels': json.dumps(top_supplier_names),
            'values': json.dumps(top_supplier_values),
        },
        'supplier_status_chart': {
            'labels': json.dumps(supplier_status_labels),
            'values': json.dumps(supplier_status_values),
        },
        
        # Department charts
        'dept_spend_chart': {
            'labels': json.dumps(dept_names),
            'values': json.dumps(dept_values),
        },
        'req_by_dept_chart': {
            'labels': json.dumps(req_dept_labels),
            'values': json.dumps(req_dept_values),
        },
        
        # Contract charts
        'contract_type_chart': {
            'labels': json.dumps(contract_type_labels),
            'values': json.dumps(contract_type_values),
        },
        
        # Tender charts
        'tender_status_chart': {
            'labels': json.dumps(tender_status_labels),
            'values': json.dumps(tender_status_values),
        },
        
        # Inventory charts
        'stock_movement_chart': {
            'labels': json.dumps(stock_movement_labels),
            'values': json.dumps(stock_movement_values),
        },
        
        # Performance metrics
        'performance': {
            'avg_approval_time': avg_approval_time,
            'avg_bids_per_tender': round(float(avg_bids), 1),
            'avg_po_creation': 3.5,  # You can calculate this based on your data
            'avg_delivery_time': 15.2,  # You can calculate this based on your data
            'delivery_compliance': delivery_compliance,
        },
        
        # Pending summary
        'pending_summary': pending_summary,
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)






def procurement_dashboard(request):
    """Procurement Officer Dashboard"""
    context = {
        'pending_requisitions': Requisition.objects.filter(
            status='HOD_APPROVED'
        ).count(),
        'active_tenders': Tender.objects.filter(
            status__in=['PUBLISHED', 'EVALUATING']
        ).order_by('-created_at')[:10],
        'pending_pos': PurchaseOrder.objects.filter(
            status='PENDING_APPROVAL'
        ).count(),
        'recent_bids': Bid.objects.filter(
            status='SUBMITTED'
        ).order_by('-submitted_at')[:10],
        'supplier_stats': {
            'total': Supplier.objects.count(),
            'approved': Supplier.objects.filter(status='APPROVED').count(),
            'pending': Supplier.objects.filter(status='PENDING').count(),
        },
        'procurement_stats': {
            'total_pos': PurchaseOrder.objects.count(),
            'total_value': PurchaseOrder.objects.aggregate(
                total=Sum('total_amount')
            )['total'] or Decimal('0'),
            'active_contracts': Contract.objects.filter(status='ACTIVE').count(),
        }
    }
    return render(request, 'dashboards/procurement_dashboard.html', context)


def finance_dashboard(request):
    """Finance Officer Dashboard"""
    context = {
        'pending_invoices': Invoice.objects.filter(
            status='SUBMITTED'
        ).order_by('-created_at')[:10],
        'pending_payments': Invoice.objects.filter(
            status='APPROVED'
        ).count(),
        'recent_payments': Payment.objects.filter(
            status='COMPLETED'
        ).order_by('-created_at')[:10],
        'financial_stats': {
            'total_invoices': Invoice.objects.count(),
            'paid_amount': Payment.objects.filter(
                status='COMPLETED'
            ).aggregate(total=Sum('payment_amount'))['total'] or Decimal('0'),
            'pending_amount': Invoice.objects.filter(
                status__in=['SUBMITTED', 'APPROVED']
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
        },
        'budget_utilization': get_budget_utilization(),
    }
    return render(request, 'dashboards/finance_dashboard.html', context)


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q, F, DecimalField
from django.db.models.functions import TruncMonth, Coalesce
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
import json

from .models import (
    GoodsReceivedNote, StockItem, StockIssue, Store,
    StockMovement, Asset, PurchaseOrder, Department
)


@login_required
def stores_dashboard(request):
    """
    Stores Officer Dashboard with comprehensive analytics and charts
    """
    # Date ranges
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_6_months = today - timedelta(days=180)
    current_month_start = timezone.datetime(today.year, today.month, 1).date()
    
    # ============================================================================
    # KEY METRICS
    # ============================================================================
    
    # Pending GRNs
    pending_grns_count = GoodsReceivedNote.objects.filter(
        status__in=['DRAFT', 'INSPECTING']
    ).count()
    
    # Low stock items
    low_stock_count = StockItem.objects.filter(
        quantity_on_hand__lte=F('reorder_level')
    ).count()
    
    # Pending stock issues
    pending_issues_count = StockIssue.objects.filter(
        status='PENDING'
    ).count()
    
    # Total inventory value
    total_inventory_value = StockItem.objects.aggregate(
        total=Sum('total_value')
    )['total'] or Decimal('0')
    
    # Active stores
    active_stores_count = Store.objects.filter(is_active=True).count()
    
    # Total stock items
    total_stock_items = StockItem.objects.count()
    
    # Items received this month
    items_received_month = GoodsReceivedNote.objects.filter(
        received_date__gte=current_month_start,
        status__in=['ACCEPTED', 'PARTIAL']
    ).count()
    
    # Items issued this month
    items_issued_month = StockIssue.objects.filter(
        issue_date__gte=current_month_start,
        status='ISSUED'
    ).count()
    
    # ============================================================================
    # CHART DATA 1: Monthly Stock Movements (Line Chart)
    # ============================================================================
    
    monthly_movements = StockMovement.objects.filter(
        movement_date__gte=last_6_months
    ).annotate(
        month=TruncMonth('movement_date')
    ).values('month', 'movement_type').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity')
    ).order_by('month')
    
    # Organize data by movement type
    receipts = {}
    issues = {}
    
    for movement in monthly_movements:
        month_str = movement['month'].strftime('%b %Y')
        if movement['movement_type'] == 'RECEIPT':
            receipts[month_str] = int(movement['count'])
        elif movement['movement_type'] == 'ISSUE':
            issues[month_str] = int(movement['count'])
    
    # Get all months
    all_months = sorted(set(list(receipts.keys()) + list(issues.keys())))
    
    movements_chart_data = {
        'labels': all_months,
        'receipts': [receipts.get(m, 0) for m in all_months],
        'issues': [issues.get(m, 0) for m in all_months]
    }
    
    # ============================================================================
    # CHART DATA 2: Stock Status Distribution (Donut Chart)
    # ============================================================================
    
    # Categorize stock items
    stock_categories = {
        'Normal Stock': StockItem.objects.filter(
            quantity_on_hand__gt=F('reorder_level')
        ).count(),
        'Low Stock': StockItem.objects.filter(
            quantity_on_hand__lte=F('reorder_level'),
            quantity_on_hand__gt=0
        ).count(),
        'Out of Stock': StockItem.objects.filter(
            quantity_on_hand=0
        ).count(),
        'Overstock': StockItem.objects.filter(
            quantity_on_hand__gt=F('max_stock_level'),
            max_stock_level__isnull=False
        ).count()
    }
    
    stock_status_chart = {
        'labels': list(stock_categories.keys()),
        'values': list(stock_categories.values()),
        'colors': ['#10B981', '#F59E0B', '#EF4444', '#6366F1']
    }
    
    # ============================================================================
    # CHART DATA 3: Inventory Value by Store (Bar Chart)
    # ============================================================================
    
    store_inventory = StockItem.objects.values(
        'store__name', 'store__code'
    ).annotate(
        total_value=Sum('total_value'),
        item_count=Count('id')
    ).order_by('-total_value')[:10]
    
    store_inventory_chart = {
        'labels': [f"{item['store__code']}" for item in store_inventory],
        'values': [float(item['total_value']) for item in store_inventory],
        'counts': [item['item_count'] for item in store_inventory]
    }
    
    # ============================================================================
    # CHART DATA 4: Stock Movement Types (Radar Chart)
    # ============================================================================
    
    movement_type_stats = {}
    movement_types = ['RECEIPT', 'ISSUE', 'TRANSFER', 'ADJUSTMENT', 'RETURN']
    
    for m_type in movement_types:
        count = StockMovement.objects.filter(
            movement_type=m_type,
            movement_date__gte=last_30_days
        ).count()
        movement_type_stats[m_type] = count
    
    # Normalize to 0-100 scale
    max_count = max(movement_type_stats.values()) if movement_type_stats.values() else 1
    
    movement_radar = {
        'labels': [m.replace('_', ' ').title() for m in movement_types],
        'values': [
            (movement_type_stats.get(m, 0) / max_count * 100) if max_count > 0 else 0 
            for m in movement_types
        ]
    }
    
    # ============================================================================
    # CHART DATA 5: GRN Processing Status (Bar Chart)
    # ============================================================================
    
    grn_status_distribution = GoodsReceivedNote.objects.values(
        'status'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    grn_status_chart = {
        'labels': [item['status'].replace('_', ' ').title() for item in grn_status_distribution],
        'values': [item['count'] for item in grn_status_distribution]
    }
    
    # ============================================================================
    # CHART DATA 6: Top Requested Items (Bar Chart)
    # ============================================================================
    
    from django.db.models import Subquery, OuterRef
    
    top_items = StockMovement.objects.filter(
        movement_type='ISSUE',
        movement_date__gte=last_30_days
    ).values(
        'stock_item__item__name',
        'stock_item__item__code'
    ).annotate(
        total_quantity=Sum('quantity'),
        issue_count=Count('id')
    ).order_by('-total_quantity')[:10]
    
    top_items_chart = {
        'labels': [item['stock_item__item__code'] for item in top_items],
        'values': [float(item['total_quantity']) for item in top_items],
        'counts': [item['issue_count'] for item in top_items]
    }
    
    # ============================================================================
    # RECENT ACTIVITIES - Lists for display
    # ============================================================================
    
    pending_grns = GoodsReceivedNote.objects.filter(
        status__in=['DRAFT', 'INSPECTING']
    ).select_related('purchase_order', 'store').order_by('-created_at')[:10]
    
    low_stock_items = StockItem.objects.filter(
        quantity_on_hand__lte=F('reorder_level')
    ).select_related('item', 'store').order_by('quantity_on_hand')[:10]
    
    pending_issues = StockIssue.objects.filter(
        status='PENDING'
    ).select_related('department', 'store').order_by('-created_at')[:10]
    
    recent_movements = StockMovement.objects.select_related(
        'stock_item__item', 'performed_by'
    ).order_by('-movement_date')[:10]
    
    # Critical alerts
    out_of_stock_items = StockItem.objects.filter(
        quantity_on_hand=0
    ).select_related('item', 'store')[:5]
    
    overdue_grns = GoodsReceivedNote.objects.filter(
        status__in=['DRAFT', 'INSPECTING'],
        created_at__lt=timezone.now() - timedelta(days=3)
    ).count()
    
    # ============================================================================
    # CONTEXT
    # ============================================================================
    
    context = {
        # Key Metrics
        'pending_grns_count': pending_grns_count,
        'low_stock_count': low_stock_count,
        'pending_issues_count': pending_issues_count,
        'total_inventory_value': total_inventory_value,
        'active_stores_count': active_stores_count,
        'total_stock_items': total_stock_items,
        'items_received_month': items_received_month,
        'items_issued_month': items_issued_month,
        'overdue_grns': overdue_grns,
        
        # Chart Data (JSON encoded for JavaScript)
        'movements_chart_data': json.dumps(movements_chart_data),
        'stock_status_chart': json.dumps(stock_status_chart),
        'store_inventory_chart': json.dumps(store_inventory_chart),
        'movement_radar': json.dumps(movement_radar),
        'grn_status_chart': json.dumps(grn_status_chart),
        'top_items_chart': json.dumps(top_items_chart),
        
        # Activity Lists
        'pending_grns': pending_grns,
        'low_stock_items': low_stock_items,
        'pending_issues': pending_issues,
        'recent_movements': recent_movements,
        'out_of_stock_items': out_of_stock_items,
        
        # Additional Stats
        'inventory_stats': {
            'total_items': total_stock_items,
            'total_value': total_inventory_value,
            'stores': active_stores_count,
        },
    }
    
    return render(request, 'dashboards/stores_dashboard.html', context)




# Helper Functions
def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def get_monthly_spend(department):
    """Get monthly spending for a department"""
    start_date = timezone.now().date() - timedelta(days=30)
    return Invoice.objects.filter(
        purchase_order__requisition__department=department,
        status='PAID',
        invoice_date__gte=start_date
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')


def get_budget_utilization():
    """Get overall budget utilization"""
    budgets = Budget.objects.filter(budget_year__is_active=True)
    total_allocated = budgets.aggregate(total=Sum('allocated_amount'))['total'] or Decimal('0')
    total_spent = budgets.aggregate(total=Sum('actual_spent'))['total'] or Decimal('0')
    
    utilization = (total_spent / total_allocated * 100) if total_allocated > 0 else 0
    
    return {
        'total_allocated': total_allocated,
        'total_spent': total_spent,
        'utilization_percentage': round(utilization, 2)
    }

def custom_404(request, exception):
    return render(request, 'errors/404.html', status=404)

def custom_500(request):
    return render(request, 'errors/500.html', status=500)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, F, Q, DecimalField, ExpressionWrapper
from django.db.models.functions import TruncMonth, TruncYear, Coalesce
from django.utils import timezone
from django.http import HttpResponse
from decimal import Decimal
from datetime import timedelta, datetime
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# ADMIN ANALYTICS DASHBOARD
# ============================================================================
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import (
    Sum, Count, Avg, F
)
from django.db.models.functions import TruncMonth
from django.shortcuts import render, redirect
from django.utils import timezone


@login_required
def admin_analytics_dashboard(request):
    """
    Comprehensive admin analytics dashboard
    """

    if request.user.role != 'ADMIN':
        messages.error(request, 'Access denied. Administrators only.')
        return redirect('dashboard')

    # =========================================================
    # DATE RANGE FILTERS
    # =========================================================
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=365)

    filter_start = request.GET.get('start_date')
    filter_end = request.GET.get('end_date')

    if filter_start:
        start_date = datetime.strptime(filter_start, '%Y-%m-%d').date()
    if filter_end:
        end_date = datetime.strptime(filter_end, '%Y-%m-%d').date()

    # =========================================================
    # SYSTEM STATISTICS
    # =========================================================
    total_requisitions = Requisition.objects.count()
    total_pos = PurchaseOrder.objects.count()
    total_suppliers = Supplier.objects.filter(status='APPROVED').count()
    total_contracts = Contract.objects.count()

    total_spend = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED']
    ).aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    # =========================================================
    # MONTHLY SPEND TREND
    # =========================================================
    monthly_spend = PurchaseOrder.objects.filter(
        po_date__gte=start_date,
        po_date__lte=end_date,
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED']
    ).annotate(
        month=TruncMonth('po_date')
    ).values('month').annotate(
        total=Sum('total_amount')
    ).order_by('month')

    spend_chart = {
        'labels': [m['month'].strftime('%b %Y') for m in monthly_spend],
        'values': [float(m['total']) for m in monthly_spend]
    }

    # =========================================================
    # BUDGET UTILIZATION
    # =========================================================
    budget_data = Budget.objects.filter(
        budget_year__is_active=True
    ).aggregate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent')
    )

    total_allocated = budget_data['allocated'] or Decimal('0.00')
    total_committed = budget_data['committed'] or Decimal('0.00')
    total_spent = budget_data['spent'] or Decimal('0.00')
    available = total_allocated - total_committed - total_spent

    budget_chart = {
        'labels': ['Spent', 'Committed', 'Available'],
        'values': [
            float(total_spent),
            float(total_committed),
            float(available)
        ]
    }

    # =========================================================
    # REQUISITION STATUS
    # =========================================================
    req_status = Requisition.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')

    req_status_chart = {
        'labels': [
            dict(Requisition.STATUS_CHOICES).get(r['status'], r['status'])
            for r in req_status
        ],
        'values': [r['count'] for r in req_status]
    }

    # =========================================================
    # PURCHASE ORDER TRENDS
    # =========================================================
    po_trend = PurchaseOrder.objects.filter(
        po_date__gte=start_date,
        po_date__lte=end_date
    ).annotate(
        month=TruncMonth('po_date')
    ).values('month').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('month')

    po_trend_chart = {
        'labels': [p['month'].strftime('%b %Y') for p in po_trend],
        'count': [p['count'] for p in po_trend],
        'values': [float(p['total']) for p in po_trend]
    }

    # =========================================================
    # TOP SUPPLIERS
    # =========================================================
    top_suppliers = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED']
    ).values('supplier__name').annotate(
        total=Sum('total_amount')
    ).order_by('-total')[:10]

    top_suppliers_chart = {
        'labels': [s['supplier__name'] for s in top_suppliers],
        'values': [float(s['total']) for s in top_suppliers]
    }

    # =========================================================
    # DEPARTMENT SPEND
    # =========================================================
    dept_spend = Requisition.objects.filter(
        status='APPROVED',
        purchase_orders__status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED']
    ).values('department__name').annotate(
        total=Sum('purchase_orders__total_amount')
    ).order_by('-total')[:10]

    dept_spend_chart = {
        'labels': [d['department__name'] for d in dept_spend],
        'values': [float(d['total'] or 0) for d in dept_spend]
    }

    # =========================================================
    # TENDER STATUS
    # =========================================================
    tender_status = Tender.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')

    tender_status_chart = {
        'labels': [
            dict(Tender.STATUS_CHOICES).get(t['status'], t['status'])
            for t in tender_status
        ],
        'values': [t['count'] for t in tender_status]
    }

    # =========================================================
    # CONTRACT TYPES
    # =========================================================
    contract_type = Contract.objects.values('contract_type').annotate(
        count=Count('id'),
        total_value=Sum('contract_value')
    ).order_by('-count')

    contract_type_chart = {
        'labels': [
            dict(Contract.CONTRACT_TYPES).get(c['contract_type'], c['contract_type'])
            for c in contract_type
        ],
        'values': [c['count'] for c in contract_type]
    }

    # =========================================================
    # PERFORMANCE METRICS (FIXED)
    # =========================================================
    approved_reqs = Requisition.objects.filter(
        status='APPROVED',
        submitted_at__isnull=False,
        updated_at__isnull=False
    )

    avg_approval_days = 0
    if approved_reqs.exists():
        total_seconds = sum(
            (r.updated_at - r.submitted_at).total_seconds()
            for r in approved_reqs
        )
        avg_approval_days = int(
            total_seconds / approved_reqs.count() / 86400
        )

    avg_bids = Bid.objects.values('tender').annotate(
        bid_count=Count('id')
    ).aggregate(
        avg=Avg('bid_count')
    )['avg'] or 0

    delivered_pos = PurchaseOrder.objects.filter(
        status='DELIVERED',
        sent_at__isnull=False,
        updated_at__isnull=False
    )

    avg_delivery_days = 0
    if delivered_pos.exists():
        total_delivery_seconds = sum(
            (po.updated_at - po.sent_at).total_seconds()
            for po in delivered_pos
        )
        avg_delivery_days = int(
            total_delivery_seconds / delivered_pos.count() / 86400
        )

    total_delivered = delivered_pos.count()
    on_time = delivered_pos.filter(
        updated_at__lte=F('delivery_date')
    ).count()

    delivery_compliance = int(
        (on_time / total_delivered) * 100
    ) if total_delivered > 0 else 0

    performance = {
        'avg_approval_time': avg_approval_days,
        'avg_bids_per_tender': round(avg_bids, 1),
        'avg_delivery_time': avg_delivery_days,
        'delivery_compliance': delivery_compliance
    }

    # =========================================================
    # CATEGORY SPEND
    # =========================================================
    category_spend = RequisitionItem.objects.filter(
        requisition__status='APPROVED',
        requisition__purchase_orders__status__in=[
            'APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED'
        ]
    ).values(
        'item__category__name'
    ).annotate(
        total=Sum('estimated_total')
    ).order_by('-total')[:8]

    category_spend_chart = {
        'labels': [
            c['item__category__name'] or 'Uncategorized'
            for c in category_spend
        ],
        'values': [float(c['total']) for c in category_spend]
    }

    # =========================================================
    # SUPPLIER RATINGS
    # =========================================================
    supplier_ratings = SupplierPerformance.objects.values(
        'supplier__name'
    ).annotate(
        avg_rating=Avg('overall_rating'),
        count=Count('id')
    ).order_by('-avg_rating')[:10]

    supplier_rating_chart = {
        'labels': [s['supplier__name'] for s in supplier_ratings],
        'values': [float(s['avg_rating']) for s in supplier_ratings]
    }

    # =========================================================
    # PAYMENT STATUS
    # =========================================================
    payment_status = Invoice.objects.values('status').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('-count')

    payment_status_chart = {
        'labels': [
            dict(Invoice.STATUS_CHOICES).get(p['status'], p['status'])
            for p in payment_status
        ],
        'values': [float(p['total'] or 0) for p in payment_status]
    }

    # =========================================================
    # STOCK VALUE
    # =========================================================
    stock_value = StockItem.objects.values(
        'store__name'
    ).annotate(
        total_value=Sum('total_value')
    ).order_by('-total_value')

    stock_value_chart = {
        'labels': [s['store__name'] for s in stock_value],
        'values': [float(s['total_value']) for s in stock_value]
    }

    # =========================================================
    # CONTEXT
    # =========================================================
    context = {
        'title': 'Admin Analytics Dashboard',
        'total_requisitions': total_requisitions,
        'total_pos': total_pos,
        'total_suppliers': total_suppliers,
        'total_contracts': total_contracts,
        'system_stats': {
            'total_spend': total_spend
        },
        'spend_chart': spend_chart,
        'budget_chart': budget_chart,
        'req_status_chart': req_status_chart,
        'po_trend_chart': po_trend_chart,
        'top_suppliers_chart': top_suppliers_chart,
        'dept_spend_chart': dept_spend_chart,
        'tender_status_chart': tender_status_chart,
        'contract_type_chart': contract_type_chart,
        'category_spend_chart': category_spend_chart,
        'supplier_rating_chart': supplier_rating_chart,
        'payment_status_chart': payment_status_chart,
        'stock_value_chart': stock_value_chart,
        'performance': performance,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'admin/analytics_dashboard.html', context)


# ============================================================================
# REPORTS VIEW WITH FILTERS
# ============================================================================

@login_required
def admin_reports(request):
    """
    Reports page with advanced filtering and Excel export
    """
    if request.user.role != 'ADMIN':
        messages.error(request, 'Access denied. Administrators only.')
        return redirect('dashboard')
    
    # Get filter parameters
    report_type = request.GET.get('report_type', 'requisitions')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department_id = request.GET.get('department')
    supplier_id = request.GET.get('supplier')
    status_filter = request.GET.get('status')
    category_id = request.GET.get('category')
    
    # Set default date range (last 90 days)
    if not start_date:
        start_date = (timezone.now() - timedelta(days=90)).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.now().strftime('%Y-%m-%d')
    
    # Initialize data
    report_data = []
    chart_data = {}
    summary_stats = {}
    
    # ========== REQUISITIONS REPORT ==========
    if report_type == 'requisitions':
        queryset = Requisition.objects.select_related(
            'department', 'requested_by', 'budget'
        ).filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        report_data = queryset.order_by('-created_at')
        
        # Summary stats
        summary_stats = {
            'total_count': queryset.count(),
            'total_value': queryset.aggregate(Sum('estimated_amount'))['estimated_amount__sum'] or Decimal('0.00'),
            'avg_value': queryset.aggregate(Avg('estimated_amount'))['estimated_amount__avg'] or Decimal('0.00'),
            'pending': queryset.filter(status='SUBMITTED').count(),
            'approved': queryset.filter(status='APPROVED').count(),
            'rejected': queryset.filter(status='REJECTED').count(),
        }
        
        # Chart: Status breakdown
        status_data = queryset.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        chart_data['status_chart'] = {
            'labels': [dict(Requisition.STATUS_CHOICES).get(item['status'], item['status']) for item in status_data],
            'values': [item['count'] for item in status_data]
        }
        
        # Chart: Monthly trend
        monthly_data = queryset.annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            count=Count('id'),
            total=Sum('estimated_amount')
        ).order_by('month')
        
        chart_data['monthly_chart'] = {
            'labels': [item['month'].strftime('%b %Y') for item in monthly_data],
            'count': [item['count'] for item in monthly_data],
            'values': [float(item['total']) for item in monthly_data]
        }
    
    # ========== PURCHASE ORDERS REPORT ==========
    elif report_type == 'purchase_orders':
        queryset = PurchaseOrder.objects.select_related(
            'supplier', 'requisition', 'created_by'
        ).filter(
            po_date__gte=start_date,
            po_date__lte=end_date
        )
        
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        report_data = queryset.order_by('-po_date')
        
        # Summary stats
        summary_stats = {
            'total_count': queryset.count(),
            'total_value': queryset.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00'),
            'avg_value': queryset.aggregate(Avg('total_amount'))['total_amount__avg'] or Decimal('0.00'),
            'pending': queryset.filter(status='PENDING_APPROVAL').count(),
            'approved': queryset.filter(status='APPROVED').count(),
            'delivered': queryset.filter(status='DELIVERED').count(),
        }
        
        # Chart: Status breakdown
        status_data = queryset.values('status').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        ).order_by('-count')
        
        chart_data['status_chart'] = {
            'labels': [dict(PurchaseOrder.STATUS_CHOICES).get(item['status'], item['status']) for item in status_data],
            'values': [float(item['total']) for item in status_data]
        }
        
        # Chart: Top suppliers
        supplier_data = queryset.values('supplier__name').annotate(
            total=Sum('total_amount')
        ).order_by('-total')[:10]
        
        chart_data['supplier_chart'] = {
            'labels': [item['supplier__name'] for item in supplier_data],
            'values': [float(item['total']) for item in supplier_data]
        }
    
    # ========== SUPPLIER REPORT ==========
    elif report_type == 'suppliers':
        queryset = Supplier.objects.annotate(
            po_count=Count('purchase_orders'),
            total_value=Sum('purchase_orders__total_amount')
        ).filter(
            purchase_orders__po_date__gte=start_date,
            purchase_orders__po_date__lte=end_date
        ).distinct()
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        report_data = queryset.order_by('-total_value')
        
        # Summary stats
        summary_stats = {
            'total_count': queryset.count(),
            'total_spend': queryset.aggregate(Sum('total_value'))['total_value__sum'] or Decimal('0.00'),
            'active': queryset.filter(status='APPROVED').count(),
            'suspended': queryset.filter(status='SUSPENDED').count(),
        }
        
        # Chart: Supplier ratings
        rating_data = SupplierPerformance.objects.filter(
            reviewed_at__date__gte=start_date,
            reviewed_at__date__lte=end_date
        ).values('supplier__name').annotate(
            avg_rating=Avg('overall_rating')
        ).order_by('-avg_rating')[:10]
        
        chart_data['rating_chart'] = {
            'labels': [item['supplier__name'] for item in rating_data],
            'values': [float(item['avg_rating']) for item in rating_data]
        }
    
    # ========== BUDGET REPORT ==========
    elif report_type == 'budget':
        queryset = Budget.objects.select_related(
            'department', 'budget_year', 'category'
        ).filter(
            budget_year__is_active=True
        )
        
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        report_data = queryset.order_by('department__name')
        
        # Summary stats
        summary_stats = {
            'total_allocated': queryset.aggregate(Sum('allocated_amount'))['allocated_amount__sum'] or Decimal('0.00'),
            'total_committed': queryset.aggregate(Sum('committed_amount'))['committed_amount__sum'] or Decimal('0.00'),
            'total_spent': queryset.aggregate(Sum('actual_spent'))['actual_spent__sum'] or Decimal('0.00'),
        }
        summary_stats['available'] = summary_stats['total_allocated'] - summary_stats['total_committed'] - summary_stats['total_spent']
        summary_stats['utilization_rate'] = (summary_stats['total_spent'] / summary_stats['total_allocated'] * 100) if summary_stats['total_allocated'] > 0 else 0
        
        # Chart: Department utilization
        dept_data = queryset.values('department__name').annotate(
            allocated=Sum('allocated_amount'),
            spent=Sum('actual_spent')
        ).order_by('-spent')[:10]
        
        chart_data['dept_chart'] = {
            'labels': [item['department__name'] for item in dept_data],
            'allocated': [float(item['allocated']) for item in dept_data],
            'spent': [float(item['spent']) for item in dept_data]
        }
    
    # ========== INVENTORY REPORT ==========
    elif report_type == 'inventory':
        queryset = StockItem.objects.select_related(
            'store', 'item'
        ).all()
        
        report_data = queryset.order_by('store__name', 'item__name')
        
        # Summary stats
        summary_stats = {
            'total_items': queryset.count(),
            'total_value': queryset.aggregate(Sum('total_value'))['total_value__sum'] or Decimal('0.00'),
            'low_stock': queryset.filter(quantity_on_hand__lte=F('reorder_level')).count(),
        }
        
        # Chart: Stock value by store
        store_data = queryset.values('store__name').annotate(
            total=Sum('total_value')
        ).order_by('-total')
        
        chart_data['store_chart'] = {
            'labels': [item['store__name'] for item in store_data],
            'values': [float(item['total']) for item in store_data]
        }
    
    # Get filter options
    departments = Department.objects.filter(is_active=True).order_by('name')
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    categories = ItemCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'title': 'Reports & Analytics',
        'report_type': report_type,
        'report_data': report_data,
        'chart_data': chart_data,
        'summary_stats': summary_stats,
        'start_date': start_date,
        'end_date': end_date,
        'department_id': department_id,
        'supplier_id': supplier_id,
        'status_filter': status_filter,
        'category_id': category_id,
        'departments': departments,
        'suppliers': suppliers,
        'categories': categories,
        'requisition_statuses': Requisition.STATUS_CHOICES,
        'po_statuses': PurchaseOrder.STATUS_CHOICES,
        'supplier_statuses': Supplier.STATUS_CHOICES,
    }
    
    return render(request, 'admin/reports.html', context)


# ============================================================================
# EXCEL EXPORT
# ============================================================================

@login_required
def export_report_excel(request):
    """
    Export filtered report data to Excel
    """
    if request.user.role != 'ADMIN':
        return HttpResponse('Access denied', status=403)
    
    report_type = request.GET.get('report_type', 'requisitions')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department_id = request.GET.get('department')
    supplier_id = request.GET.get('supplier')
    status_filter = request.GET.get('status')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styling
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== REQUISITIONS EXPORT ==========
    if report_type == 'requisitions':
        ws.title = 'Requisitions Report'
        
        # Headers
        headers = ['Req Number', 'Date', 'Department', 'Title', 'Requested By', 
                   'Status', 'Priority', 'Estimated Amount', 'Required Date']
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Get data
        queryset = Requisition.objects.select_related(
            'department', 'requested_by'
        ).filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Add data rows
        for req in queryset.order_by('-created_at'):
            ws.append([
                req.requisition_number,
                req.created_at.strftime('%Y-%m-%d'),
                req.department.name,
                req.title,
                req.requested_by.get_full_name(),
                dict(Requisition.STATUS_CHOICES).get(req.status, req.status),
                dict(Requisition.PRIORITY_CHOICES).get(req.priority, req.priority),
                float(req.estimated_amount),
                req.required_date.strftime('%Y-%m-%d')
            ])
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # ========== PURCHASE ORDERS EXPORT ==========
    elif report_type == 'purchase_orders':
        ws.title = 'Purchase Orders Report'
        
        headers = ['PO Number', 'Date', 'Supplier', 'Requisition', 'Status', 
                   'Subtotal', 'Tax', 'Total', 'Delivery Date']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        queryset = PurchaseOrder.objects.select_related(
            'supplier', 'requisition'
        ).filter(
            po_date__gte=start_date,
            po_date__lte=end_date
        )
        
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        for po in queryset.order_by('-po_date'):
            ws.append([
                po.po_number,
                po.po_date.strftime('%Y-%m-%d'),
                po.supplier.name,
                po.requisition.requisition_number,
                dict(PurchaseOrder.STATUS_CHOICES).get(po.status, po.status),
                float(po.subtotal),
                float(po.tax_amount),
                float(po.total_amount),
                po.delivery_date.strftime('%Y-%m-%d')
            ])
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # ========== SUPPLIER EXPORT ==========
    elif report_type == 'suppliers':
        ws.title = 'Suppliers Report'
        
        headers = ['Supplier Number', 'Name', 'Email', 'Phone', 'Status', 
                   'PO Count', 'Total Value', 'Rating']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        queryset = Supplier.objects.annotate(
            po_count=Count('purchase_orders'),
            total_value=Sum('purchase_orders__total_amount')
        )
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        for supplier in queryset.order_by('name'):
            ws.append([
                supplier.supplier_number,
                supplier.name,
                supplier.email,
                supplier.phone_number,
                dict(Supplier.STATUS_CHOICES).get(supplier.status, supplier.status),
                supplier.po_count or 0,
                float(supplier.total_value or 0),
                float(supplier.rating)
            ])
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # ========== BUDGET EXPORT ==========
    elif report_type == 'budget':
        ws.title = 'Budget Report'
        
        headers = ['Department', 'Category', 'Budget Year', 'Allocated', 
                   'Committed', 'Spent', 'Available', 'Utilization %']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            
            

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from decimal import Decimal
import json

from .models import (
    Requisition, RequisitionItem, RequisitionAttachment, RequisitionApproval,
    Department, Budget, Item, User, AuditLog, Notification
)


@login_required
def requisition_list(request):
    """List all requisitions with filters"""
    requisitions = Requisition.objects.select_related(
        'department', 'budget', 'requested_by'
    ).prefetch_related('items')
    
    # Filters
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    department_filter = request.GET.get('department', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status_filter:
        requisitions = requisitions.filter(status=status_filter)
    
    if priority_filter:
        requisitions = requisitions.filter(priority=priority_filter)
    
    if department_filter:
        requisitions = requisitions.filter(department_id=department_filter)
    
    if search_query:
        requisitions = requisitions.filter(
            Q(requisition_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(justification__icontains=search_query)
        )
    
    if date_from:
        requisitions = requisitions.filter(created_at__gte=date_from)
    
    if date_to:
        requisitions = requisitions.filter(created_at__lte=date_to)
    
    # Role-based filtering
    if request.user.role == 'STAFF':
        requisitions = requisitions.filter(requested_by=request.user)
    elif request.user.role == 'HOD':
        requisitions = requisitions.filter(department=request.user.department)
    
    # Statistics
    total_count = requisitions.count()
    pending_count = requisitions.filter(status='SUBMITTED').count()
    approved_count = requisitions.filter(status='APPROVED').count()
    rejected_count = requisitions.filter(status='REJECTED').count()
    total_amount = requisitions.aggregate(Sum('estimated_amount'))['estimated_amount__sum'] or 0
    
    # Pagination
    paginator = Paginator(requisitions.order_by('-created_at'), 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get departments for filter
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'requisitions': page_obj,
        'page_obj': page_obj,
        'departments': departments,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'department_filter': department_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': total_count,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'total_amount': total_amount,
        'status_choices': Requisition.STATUS_CHOICES,
        'priority_choices': Requisition.PRIORITY_CHOICES,
    }
    
    return render(request, 'requisitions/requisition_list.html', context)


@login_required
def requisition_create(request):
    """Create new requisition"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Create requisition
                requisition = Requisition.objects.create(
                    title=request.POST.get('title'),
                    department_id=request.POST.get('department'),
                    budget_id=request.POST.get('budget') if request.POST.get('budget') else None,
                    requested_by=request.user,
                    justification=request.POST.get('justification'),
                    estimated_amount=Decimal(request.POST.get('estimated_amount', 0)),
                    required_date=request.POST.get('required_date'),
                    priority=request.POST.get('priority', 'MEDIUM'),
                    is_emergency=request.POST.get('is_emergency') == 'on',
                    emergency_justification=request.POST.get('emergency_justification', ''),
                    notes=request.POST.get('notes', ''),
                    status='DRAFT'
                )
                
                # Create requisition items
                items_data = json.loads(request.POST.get('items_json', '[]'))
                total_estimated = Decimal('0')
                
                for item_data in items_data:
                    item = RequisitionItem.objects.create(
                        requisition=requisition,
                        item_id=item_data.get('item_id') if item_data.get('item_id') else None,
                        item_description=item_data.get('description'),
                        specifications=item_data.get('specifications', ''),
                        quantity=Decimal(item_data.get('quantity')),
                        unit_of_measure=item_data.get('unit'),
                        estimated_unit_price=Decimal(item_data.get('unit_price')),
                        notes=item_data.get('notes', '')
                    )
                    total_estimated += item.estimated_total
                
                # Update estimated amount
                requisition.estimated_amount = total_estimated
                requisition.save()
                
                # Handle file attachments
                if request.FILES:
                    for file_key in request.FILES:
                        file = request.FILES[file_key]
                        RequisitionAttachment.objects.create(
                            requisition=requisition,
                            attachment_type=request.POST.get(f'{file_key}_type', 'OTHER'),
                            file_name=file.name,
                            file=file,
                            description=request.POST.get(f'{file_key}_description', ''),
                            uploaded_by=request.user
                        )
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='CREATE',
                    model_name='Requisition',
                    object_id=str(requisition.id),
                    object_repr=requisition.requisition_number,
                    changes={'status': 'DRAFT', 'created': True}
                )
                
                messages.success(request, f'Requisition {requisition.requisition_number} created successfully!')
                
                # Check if submit action
                if request.POST.get('action') == 'submit':
                    return redirect('requisition_submit', pk=requisition.id)
                
                return redirect('requisition_detail', pk=requisition.id)
                
        except Exception as e:
            messages.error(request, f'Error creating requisition: {str(e)}')
            return redirect('requisition_create')
    
    # GET request - show form
    departments = Department.objects.filter(is_active=True).order_by('name')
    budgets = Budget.objects.filter(is_active=True).select_related('category', 'department')
    items = Item.objects.filter(is_active=True).select_related('category').order_by('name')
    
    # Filter budgets by user's department if not admin
    if request.user.role not in ['ADMIN', 'PROCUREMENT', 'FINANCE']:
        if request.user.department:
            budgets = budgets.filter(department=request.user.department)
    
    context = {
        'departments': departments,
        'budgets': budgets,
        'items': items,
        'priority_choices': Requisition.PRIORITY_CHOICES,
    }
    
    return render(request, 'requisitions/requisition_create.html', context)


@login_required
def requisition_detail(request, pk):
    """View requisition details"""
    requisition = get_object_or_404(
        Requisition.objects.select_related(
            'department', 'budget', 'requested_by'
        ).prefetch_related(
            'items__item',
            'attachments',
            'approvals__approver',
            'purchase_orders'
        ),
        pk=pk
    )
    
    # Check permissions
    can_edit = False
    can_approve = False
    can_submit = False
    
    # Debug info
    debug_info = {
        'user_role': request.user.role,
        'user_dept': request.user.department.name if request.user.department else 'None',
        'req_status': requisition.status,
        'req_dept': requisition.department.name,
        'is_requester': request.user == requisition.requested_by,
    }
    
    # ADMIN can do everything
    if request.user.role == 'ADMIN':
        can_edit = True
        can_approve = requisition.status == 'PROCUREMENT_APPROVED'
        debug_info['approve_reason'] = f'User is ADMIN, status is {requisition.status}, can_approve={can_approve}'
    
    # Requester can edit and submit drafts
    elif request.user == requisition.requested_by and requisition.status == 'DRAFT':
        can_edit = True
        can_submit = True
        debug_info['approve_reason'] = 'User is requester, status is DRAFT'
    
    # HOD can approve submitted requisitions from their department
    elif request.user.role == 'HOD' and request.user.department == requisition.department:
        can_approve = requisition.status == 'SUBMITTED'
        debug_info['approve_reason'] = f'User is HOD, status is {requisition.status}, can_approve={can_approve}'
    
    # PROCUREMENT can approve any HOD_APPROVED requisition (no department restriction)
    elif request.user.role == 'PROCUREMENT':
        can_approve = requisition.status == 'HOD_APPROVED'
        debug_info['approve_reason'] = f'User is PROCUREMENT, status is {requisition.status}, can_approve={can_approve}'
    
    else:
        debug_info['approve_reason'] = 'No matching approval condition'
    
    # Get approval history
    approvals = requisition.approvals.all().order_by('sequence')
    
    # Calculate totals
    items_total = requisition.items.aggregate(Sum('estimated_total'))['estimated_total__sum'] or 0
    
    context = {
        'requisition': requisition,
        'can_edit': can_edit,
        'can_approve': can_approve,
        'can_submit': can_submit,
        'approvals': approvals,
        'items_total': items_total,
        'debug_info': debug_info,
    }
    
    return render(request, 'requisitions/requisition_detail.html', context)

from django.contrib import messages
from django.utils import timezone

@login_required
def requisition_approve(request, pk):
    """Approve a requisition"""
    requisition = get_object_or_404(Requisition, pk=pk)
    
    # Determine next status based on current user role
    if request.user.role == 'HOD':
        if requisition.status != 'SUBMITTED':
            messages.error(request, 'This requisition cannot be approved at this stage.')
            return redirect('requisition_detail', pk=pk)
        
        # Update status
        requisition.status = 'HOD_APPROVED'
        
        # Create or update approval record
        approval, created = RequisitionApproval.objects.get_or_create(
            requisition=requisition,
            approval_stage='HOD',
            defaults={
                'approver': request.user,
                'sequence': 1
            }
        )
        approval.status = 'APPROVED'
        approval.approval_date = timezone.now()
        approval.approver = request.user
        approval.save()
        
        messages.success(request, 'Requisition approved as HOD.')
        
    elif request.user.role == 'PROCUREMENT':
        if requisition.status != 'HOD_APPROVED':
            messages.error(request, 'This requisition cannot be approved at this stage.')
            return redirect('requisition_detail', pk=pk)
        
        # Update status
        requisition.status = 'PROCUREMENT_APPROVED'
        
        # Create or update approval record
        approval, created = RequisitionApproval.objects.get_or_create(
            requisition=requisition,
            approval_stage='PROCUREMENT',
            defaults={
                'approver': request.user,
                'sequence': 2
            }
        )
        approval.status = 'APPROVED'
        approval.approval_date = timezone.now()
        approval.approver = request.user
        approval.save()
        
        messages.success(request, 'Requisition approved as Procurement Officer.')
        
    elif request.user.role == 'ADMIN':
        if requisition.status != 'PROCUREMENT_APPROVED':
            messages.error(request, 'This requisition cannot be approved at this stage.')
            return redirect('requisition_detail', pk=pk)
        
        # Final approval
        requisition.status = 'APPROVED'
        
        # Create or update approval record
        approval, created = RequisitionApproval.objects.get_or_create(
            requisition=requisition,
            approval_stage='FINAL',
            defaults={
                'approver': request.user,
                'sequence': 3
            }
        )
        approval.status = 'APPROVED'
        approval.approval_date = timezone.now()
        approval.approver = request.user
        approval.save()
        
        messages.success(request, 'Requisition fully approved! Ready for procurement.')
        
    else:
        messages.error(request, 'You do not have permission to approve this requisition.')
        return redirect('requisition_detail', pk=pk)
    
    # Save the requisition
    requisition.save()
    
    # Create notification for next approver or requester
    if requisition.status == 'APPROVED':
        Notification.objects.create(
            user=requisition.requested_by,
            notification_type='APPROVAL',
            priority='HIGH',
            title='Requisition Fully Approved',
            message=f'Your requisition {requisition.requisition_number} has been fully approved.',
            link_url=f'/requisitions/{requisition.id}/'
        )
    
    return redirect('requisition_detail', pk=pk)

@login_required
def requisition_update(request, pk):
    """Update requisition"""
    requisition = get_object_or_404(Requisition, pk=pk)
    
    # Check permissions
    if requisition.status != 'DRAFT' and request.user.role != 'ADMIN':
        messages.error(request, 'You can only edit draft requisitions.')
        return redirect('requisition_detail', pk=pk)
    
    if requisition.requested_by != request.user and request.user.role != 'ADMIN':
        messages.error(request, 'You do not have permission to edit this requisition.')
        return redirect('requisition_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Store old values for audit
                old_values = {
                    'title': requisition.title,
                    'estimated_amount': str(requisition.estimated_amount),
                    'status': requisition.status
                }
                
                # Update requisition
                requisition.title = request.POST.get('title')
                requisition.department_id = request.POST.get('department')
                requisition.budget_id = request.POST.get('budget') if request.POST.get('budget') else None
                requisition.justification = request.POST.get('justification')
                requisition.required_date = request.POST.get('required_date')
                requisition.priority = request.POST.get('priority', 'MEDIUM')
                requisition.is_emergency = request.POST.get('is_emergency') == 'on'
                requisition.emergency_justification = request.POST.get('emergency_justification', '')
                requisition.notes = request.POST.get('notes', '')
                
                # Delete existing items
                requisition.items.all().delete()
                
                # Create new items
                items_data = json.loads(request.POST.get('items_json', '[]'))
                total_estimated = Decimal('0')
                
                for item_data in items_data:
                    item = RequisitionItem.objects.create(
                        requisition=requisition,
                        item_id=item_data.get('item_id') if item_data.get('item_id') else None,
                        item_description=item_data.get('description'),
                        specifications=item_data.get('specifications', ''),
                        quantity=Decimal(item_data.get('quantity')),
                        unit_of_measure=item_data.get('unit'),
                        estimated_unit_price=Decimal(item_data.get('unit_price')),
                        notes=item_data.get('notes', '')
                    )
                    total_estimated += item.estimated_total
                
                requisition.estimated_amount = total_estimated
                requisition.save()
                
                # Handle new attachments
                if request.FILES:
                    for file_key in request.FILES:
                        file = request.FILES[file_key]
                        RequisitionAttachment.objects.create(
                            requisition=requisition,
                            attachment_type=request.POST.get(f'{file_key}_type', 'OTHER'),
                            file_name=file.name,
                            file=file,
                            description=request.POST.get(f'{file_key}_description', ''),
                            uploaded_by=request.user
                        )
                
                # Create audit log
                new_values = {
                    'title': requisition.title,
                    'estimated_amount': str(requisition.estimated_amount),
                    'status': requisition.status
                }
                
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Requisition',
                    object_id=str(requisition.id),
                    object_repr=requisition.requisition_number,
                    changes={'old': old_values, 'new': new_values}
                )
                
                messages.success(request, f'Requisition {requisition.requisition_number} updated successfully!')
                return redirect('requisition_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error updating requisition: {str(e)}')
    
    # GET request
    departments = Department.objects.filter(is_active=True).order_by('name')
    budgets = Budget.objects.filter(is_active=True).select_related('category', 'department')
    items = Item.objects.filter(is_active=True).select_related('category').order_by('name')
    
    context = {
        'requisition': requisition,
        'departments': departments,
        'budgets': budgets,
        'items': items,
        'priority_choices': Requisition.PRIORITY_CHOICES,
    }
    
    return render(request, 'requisitions/requisition_update.html', context)


@login_required
def requisition_delete(request, pk):
    """Delete requisition"""
    requisition = get_object_or_404(Requisition, pk=pk)
    
    # Check permissions
    if requisition.status != 'DRAFT' and request.user.role != 'ADMIN':
        messages.error(request, 'You can only delete draft requisitions.')
        return redirect('requisition_detail', pk=pk)
    
    if requisition.requested_by != request.user and request.user.role != 'ADMIN':
        messages.error(request, 'You do not have permission to delete this requisition.')
        return redirect('requisition_detail', pk=pk)
    
    if request.method == 'POST':
        req_number = requisition.requisition_number
        
        # Create audit log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            model_name='Requisition',
            object_id=str(requisition.id),
            object_repr=req_number,
            changes={'deleted': True}
        )
        
        requisition.delete()
        messages.success(request, f'Requisition {req_number} deleted successfully!')
        return redirect('requisition_list')
    
    return render(request, 'requisitions/requisition_delete.html', {'requisition': requisition})


@login_required
def requisition_submit(request, pk):
    """Submit requisition for approval"""
    requisition = get_object_or_404(Requisition, pk=pk)
    
    # Check permissions
    if requisition.requested_by != request.user and request.user.role != 'ADMIN':
        messages.error(request, 'You do not have permission to submit this requisition.')
        return redirect('requisition_detail', pk=pk)
    
    if requisition.status != 'DRAFT':
        messages.error(request, 'Only draft requisitions can be submitted.')
        return redirect('requisition_detail', pk=pk)
    
    if not requisition.items.exists():
        messages.error(request, 'Cannot submit requisition without items.')
        return redirect('requisition_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update status
                requisition.status = 'SUBMITTED'
                requisition.submitted_at = timezone.now()
                requisition.save()
                
                # Create approval workflow
                sequence = 1
                
                # HOD Approval
                if requisition.department.hod:
                    RequisitionApproval.objects.create(
                        requisition=requisition,
                        approval_stage='HOD',
                        approver=requisition.department.hod,
                        sequence=sequence
                    )
                    sequence += 1
                    
                    # Create notification
                    Notification.objects.create(
                        user=requisition.department.hod,
                        notification_type='APPROVAL',
                        priority='HIGH',
                        title='Requisition Pending Approval',
                        message=f'Requisition {requisition.requisition_number} requires your approval.',
                        link_url=f'/requisitions/{requisition.id}/'
                    )
                
                # Budget approval
                finance_users = User.objects.filter(role='FINANCE', is_active=True).first()
                if finance_users:
                    RequisitionApproval.objects.create(
                        requisition=requisition,
                        approval_stage='BUDGET',
                        approver=finance_users,
                        sequence=sequence
                    )
                    sequence += 1
                
                # Procurement approval
                procurement_users = User.objects.filter(role='PROCUREMENT', is_active=True).first()
                if procurement_users:
                    RequisitionApproval.objects.create(
                        requisition=requisition,
                        approval_stage='PROCUREMENT',
                        approver=procurement_users,
                        sequence=sequence
                    )
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='SUBMIT',
                    model_name='Requisition',
                    object_id=str(requisition.id),
                    object_repr=requisition.requisition_number,
                    changes={'status': 'SUBMITTED'}
                )
                
                messages.success(request, f'Requisition {requisition.requisition_number} submitted for approval!')
                return redirect('requisition_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error submitting requisition: {str(e)}')
    
    return render(request, 'requisitions/requisition_submit.html', {'requisition': requisition})


@login_required
def pending_requisitions(request):
    """View pending requisitions for approval"""
    requisitions = Requisition.objects.none()
    
    # Get requisitions based on user role
    if request.user.role == 'HOD':
        requisitions = Requisition.objects.filter(
            department=request.user.department,
            status='SUBMITTED'
        ).select_related('department', 'requested_by')
    
    elif request.user.role == 'FINANCE':
        requisitions = Requisition.objects.filter(
            status__in=['HOD_APPROVED', 'FACULTY_APPROVED']
        ).select_related('department', 'requested_by')
    
    elif request.user.role == 'PROCUREMENT':
        requisitions = Requisition.objects.filter(
            status__in=['HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED']
        ).select_related('department', 'requested_by')
    
    elif request.user.role == 'ADMIN':
        requisitions = Requisition.objects.filter(
            status__in=['SUBMITTED', 'HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED']
        ).select_related('department', 'requested_by')
    
    # Pagination
    paginator = Paginator(requisitions.order_by('-created_at'), 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    total_count = requisitions.count()
    total_amount = requisitions.aggregate(Sum('estimated_amount'))['estimated_amount__sum'] or 0
    
    context = {
        'requisitions': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'total_amount': total_amount,
    }
    
    return render(request, 'requisitions/pending_requisitions.html', context)


# API Views
@login_required
def get_budget_info(request, budget_id):
    """API: Get budget information"""
    try:
        budget = Budget.objects.select_related('category', 'department').get(id=budget_id)
        return JsonResponse({
            'success': True,
            'data': {
                'id': str(budget.id),
                'category': budget.category.name,
                'allocated_amount': float(budget.allocated_amount),
                'available_balance': float(budget.available_balance),
                'department': budget.department.name
            }
        })
    except Budget.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Budget not found'}, status=404)


@login_required
def get_item_info(request, item_id):
    """API: Get item information"""
    try:
        item = Item.objects.select_related('category').get(id=item_id)
        return JsonResponse({
            'success': True,
            'data': {
                'id': str(item.id),
                'name': item.name,
                'code': item.code,
                'description': item.description,
                'unit_of_measure': item.unit_of_measure,
                'standard_price': float(item.standard_price) if item.standard_price else None,
                'specifications': item.specifications,
                'category': item.category.name
            }
        })
    except Item.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)


@login_required
def delete_attachment(request, attachment_id):
    """API: Delete requisition attachment"""
    try:
        attachment = get_object_or_404(RequisitionAttachment, id=attachment_id)
        requisition = attachment.requisition
        
        # Check permissions
        if requisition.requested_by != request.user and request.user.role != 'ADMIN':
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        if requisition.status != 'DRAFT':
            return JsonResponse({'success': False, 'error': 'Can only delete attachments from draft requisitions'}, status=400)
        
        attachment.delete()
        return JsonResponse({'success': True, 'message': 'Attachment deleted successfully'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg, Max, Min
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from decimal import Decimal
import json

from .models import (
    Tender, TenderDocument, Bid, BidItem, BidDocument, BidEvaluation,
    EvaluationCriteria, Requisition, Supplier, User, AuditLog, Notification,
    RequisitionItem
)


@login_required
def tender_list(request):
    """List all tenders with filters"""
    # Check permissions
    if request.user.role not in ['PROCUREMENT', 'ADMIN', 'HOD', 'FINANCE']:
        messages.error(request, 'You do not have permission to access tenders.')
        return redirect('dashboard')
    
    tenders = Tender.objects.select_related(
        'requisition__department', 'created_by'
    ).prefetch_related('invited_suppliers', 'bids')
    
    # Filters
    status_filter = request.GET.get('status', '')
    tender_type_filter = request.GET.get('tender_type', '')
    method_filter = request.GET.get('method', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status_filter:
        tenders = tenders.filter(status=status_filter)
    
    if tender_type_filter:
        tenders = tenders.filter(tender_type=tender_type_filter)
    
    if method_filter:
        tenders = tenders.filter(procurement_method=method_filter)
    
    if search_query:
        tenders = tenders.filter(
            Q(tender_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if date_from:
        tenders = tenders.filter(created_at__gte=date_from)
    
    if date_to:
        tenders = tenders.filter(created_at__lte=date_to)
    
    # Statistics
    total_count = tenders.count()
    published_count = tenders.filter(status='PUBLISHED').count()
    closed_count = tenders.filter(status='CLOSED').count()
    evaluating_count = tenders.filter(status='EVALUATING').count()
    awarded_count = tenders.filter(status='AWARDED').count()
    total_value = tenders.aggregate(Sum('estimated_budget'))['estimated_budget__sum'] or 0
    
    # Pagination
    paginator = Paginator(tenders.order_by('-created_at'), 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'tenders': page_obj,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'tender_type_filter': tender_type_filter,
        'method_filter': method_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': total_count,
        'published_count': published_count,
        'closed_count': closed_count,
        'evaluating_count': evaluating_count,
        'awarded_count': awarded_count,
        'total_value': total_value,
        'status_choices': Tender.STATUS_CHOICES,
        'tender_type_choices': Tender.TENDER_TYPES,
        'method_choices': Tender.METHOD_CHOICES,
    }
    
    return render(request, 'tenders/tender_list.html', context)


@login_required
def tender_create(request):
    """Create new tender"""
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to create tenders.')
        return redirect('tender_list')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get requisition
                requisition_id = request.POST.get('requisition')
                requisition = get_object_or_404(Requisition, id=requisition_id)
                
                # Create tender
                tender = Tender.objects.create(
                    requisition=requisition,
                    title=request.POST.get('title'),
                    tender_type=request.POST.get('tender_type'),
                    procurement_method=request.POST.get('procurement_method'),
                    description=request.POST.get('description'),
                    closing_date=request.POST.get('closing_date'),
                    bid_opening_date=request.POST.get('bid_opening_date'),
                    estimated_budget=Decimal(request.POST.get('estimated_budget')),
                    created_by=request.user,
                    status='DRAFT'
                )
                
                # Add invited suppliers if restricted tender
                if tender.procurement_method == 'RESTRICTED':
                    supplier_ids = request.POST.getlist('invited_suppliers')
                    if supplier_ids:
                        tender.invited_suppliers.set(supplier_ids)
                
                # Create evaluation criteria
                criteria_json = request.POST.get('criteria_json', '[]')
                if criteria_json and criteria_json != '[]':
                    criteria_data = json.loads(criteria_json)
                    for idx, criterion in enumerate(criteria_data, 1):
                        EvaluationCriteria.objects.create(
                            tender=tender,
                            criterion_name=criterion.get('criterion_name', ''),
                            criterion_type=criterion.get('criterion_type', 'GENERAL'),
                            description=criterion.get('description', criterion.get('criterion_name', '')),
                            weight=Decimal(str(criterion.get('weight', 0))),
                            max_score=Decimal(str(criterion.get('max_score', 100))),
                            is_mandatory=criterion.get('is_mandatory', False),
                            sequence=idx
                        )
                
                # Handle tender documents
                document_count = 0
                for key in request.POST.keys():
                    if key.startswith('document_') and key.endswith('_file'):
                        document_count += 1
                
                for i in range(1, document_count + 1):
                    file_key = f'document_{i}_file'
                    if file_key in request.FILES:
                        file = request.FILES[file_key]
                        description_key = f'document_{i}_description'
                        mandatory_key = f'document_{i}_mandatory'
                        
                        TenderDocument.objects.create(
                            tender=tender,
                            document_name=file.name,
                            file=file,
                            description=request.POST.get(description_key, ''),
                            is_mandatory=mandatory_key in request.POST
                        )
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='CREATE',
                    model_name='Tender',
                    object_id=str(tender.id),
                    object_repr=tender.tender_number,
                    changes={'status': 'DRAFT', 'created': True}
                )
                
                messages.success(request, f'Tender {tender.tender_number} created successfully!')
                
                # Check if publish action
                if request.POST.get('action') == 'publish':
                    return redirect('tender_publish', pk=tender.id)
                
                return redirect('tender_detail', pk=tender.id)
                
        except Exception as e:
            messages.error(request, f'Error creating tender: {str(e)}')
            import traceback
            print(traceback.format_exc())  # For debugging
            return redirect('tender_create')
    
    # GET request - show form
    # Get approved requisitions without tenders
    requisitions = Requisition.objects.filter(
        status='APPROVED',
        tenders__isnull=True
    ).select_related('department', 'requested_by')
    
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    
    context = {
        'requisitions': requisitions,
        'suppliers': suppliers,
        'tender_type_choices': Tender.TENDER_TYPES,
        'method_choices': Tender.METHOD_CHOICES,
    }
    
    return render(request, 'tenders/tender_create.html', context)




@login_required
def edit_tender(request, tender_id):
    tender = get_object_or_404(Tender, id=tender_id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update basic fields
                tender.title = request.POST.get('title')
                tender.tender_type = request.POST.get('tender_type')
                tender.procurement_method = request.POST.get('procurement_method')
                tender.description = request.POST.get('description')
                tender.closing_date = request.POST.get('closing_date')
                tender.bid_opening_date = request.POST.get('bid_opening_date')
                tender.estimated_budget = request.POST.get('estimated_budget')
                tender.status = request.POST.get('status')

                # Optional publish date logic
                if tender.status == 'PUBLISHED' and not tender.publish_date:
                    tender.publish_date = timezone.now()

                tender.save()

                # Update invited suppliers - Filter out empty strings
                supplier_ids = request.POST.getlist('invited_suppliers')
                # Remove empty strings and invalid UUIDs
                valid_supplier_ids = [
                    sid for sid in supplier_ids 
                    if sid and sid.strip()  # Only non-empty values
                ]
                
                if valid_supplier_ids:
                    tender.invited_suppliers.set(valid_supplier_ids)
                else:
                    # Clear all suppliers if none selected
                    tender.invited_suppliers.clear()

                # Handle document deletion
                delete_doc_ids = request.POST.getlist('delete_documents')
                # Filter out empty strings from delete_doc_ids too
                valid_delete_ids = [
                    did for did in delete_doc_ids 
                    if did and did.strip()
                ]
                if valid_delete_ids:
                    TenderDocument.objects.filter(
                        id__in=valid_delete_ids,
                        tender=tender
                    ).delete()

                # Handle new document uploads
                document_count = 0
                for key in request.POST.keys():
                    if key.startswith('document_') and key.endswith('_file'):
                        try:
                            # Extract index from key like 'document_1_file'
                            index = key.replace('document_', '').replace('_file', '')
                            document_count = max(document_count, int(index))
                        except ValueError:
                            continue
                
                for i in range(1, document_count + 1):
                    file_key = f'document_{i}_file'
                    if file_key in request.FILES:
                        file = request.FILES[file_key]
                        description_key = f'document_{i}_description'
                        mandatory_key = f'document_{i}_mandatory'
                        
                        TenderDocument.objects.create(
                            tender=tender,
                            document_name=file.name,
                            file=file,
                            description=request.POST.get(description_key, ''),
                            is_mandatory=mandatory_key in request.POST
                        )

                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Tender',
                    object_id=str(tender.id),
                    object_repr=tender.tender_number,
                    changes={
                        'title': tender.title,
                        'status': tender.status,
                        'suppliers_count': tender.invited_suppliers.count()
                    }
                )

                messages.success(request, 'Tender updated successfully.')
                return redirect('tender_detail', pk=tender.id)

        except Exception as e:
            messages.error(request, f'Error updating tender: {str(e)}')
            import traceback
            print(traceback.format_exc())  # For debugging

    # GET request context
    context = {
        'tender': tender,
        'existing_documents': tender.documents.all(),
        'suppliers': Supplier.objects.filter(status='APPROVED').order_by('name'),
        'tender_types': Tender.TENDER_TYPES,
        'procurement_methods': Tender.METHOD_CHOICES,
        'status_choices': Tender.STATUS_CHOICES,
    }

    return render(request, 'tenders/tender_edit.html', context)


@login_required
def tender_detail(request, pk):
    """View tender details"""
    tender = get_object_or_404(
        Tender.objects.select_related(
            'requisition__department', 'created_by'
        ).prefetch_related(
            'documents',
            'bids__supplier',
            'bids__evaluations',
            'evaluation_criteria',
            'invited_suppliers'
        ),
        pk=pk
    )
    
    # Check permissions
    can_edit = False
    can_publish = False
    can_evaluate = False
    can_award = False
    
    if request.user.role == 'ADMIN':
        can_edit = True
        can_publish = True
        can_evaluate = True
        can_award = True
    elif request.user.role == 'PROCUREMENT':
        can_edit = tender.status == 'DRAFT'
        can_publish = tender.status == 'DRAFT'
        can_evaluate = tender.status in ['CLOSED', 'EVALUATING']
        can_award = tender.status == 'EVALUATING'
    
    # Get bids statistics
    bids_stats = {
        'total': tender.bids.count(),
        'submitted': tender.bids.filter(status='SUBMITTED').count(),
        'qualified': tender.bids.filter(status='QUALIFIED').count(),
        'disqualified': tender.bids.filter(status='DISQUALIFIED').count(),
        'lowest_bid': tender.bids.aggregate(Min('bid_amount'))['bid_amount__min'],
        'highest_bid': tender.bids.aggregate(Max('bid_amount'))['bid_amount__max'],
        'average_bid': tender.bids.aggregate(Avg('bid_amount'))['bid_amount__avg'],
    }
    
    # Get evaluation summary
    evaluation_complete = False
    if tender.status == 'EVALUATING':
        total_bids = tender.bids.exclude(status='DISQUALIFIED').count()
        evaluated_bids = tender.bids.exclude(status='DISQUALIFIED').filter(
            evaluations__isnull=False
        ).distinct().count()
        evaluation_complete = total_bids > 0 and total_bids == evaluated_bids
    
    context = {
        'tender': tender,
        'can_edit': can_edit,
        'can_publish': can_publish,
        'can_evaluate': can_evaluate,
        'can_award': can_award,
        'bids_stats': bids_stats,
        'evaluation_complete': evaluation_complete,
    }
    
    return render(request, 'tenders/tender_detail.html', context)


@login_required
def tender_publish(request, pk):
    """Publish tender"""
    tender = get_object_or_404(Tender, pk=pk)
    
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to publish tenders.')
        return redirect('tender_detail', pk=pk)
    
    if tender.status != 'DRAFT':
        messages.error(request, 'Only draft tenders can be published.')
        return redirect('tender_detail', pk=pk)
    
    
    
    if not tender.evaluation_criteria.exists():
        messages.error(request, 'Please add evaluation criteria before publishing.')
        return redirect('tender_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                tender.status = 'PUBLISHED'
                tender.publish_date = timezone.now()
                tender.save()
                
                # Create notifications for invited suppliers (if restricted)
                if tender.procurement_method == 'RESTRICTED':
                    for supplier in tender.invited_suppliers.all():
                        if supplier.user_set.exists():
                            for user in supplier.user_set.filter(role='SUPPLIER'):
                                Notification.objects.create(
                                    user=user,
                                    notification_type='TENDER',
                                    priority='HIGH',
                                    title='New Tender Invitation',
                                    message=f'You are invited to bid on tender {tender.tender_number}: {tender.title}',
                                    link_url=f'/tenders/{tender.id}/'
                                )
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Tender',
                    object_id=str(tender.id),
                    object_repr=tender.tender_number,
                    changes={'status': 'PUBLISHED', 'publish_date': str(timezone.now())}
                )
                
                messages.success(request, f'Tender {tender.tender_number} published successfully!')
                return redirect('tender_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error publishing tender: {str(e)}')
    
    return render(request, 'tenders/tender_publish.html', {'tender': tender})


@login_required
def tender_close(request, pk):
    """Close tender for bidding"""
    tender = get_object_or_404(Tender, pk=pk)
    
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to close tenders.')
        return redirect('tender_detail', pk=pk)
    
    if tender.status != 'PUBLISHED':
        messages.error(request, 'Only published tenders can be closed.')
        return redirect('tender_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                tender.status = 'CLOSED'
                tender.save()
                
                # Update all submitted bids
                tender.bids.filter(status='SUBMITTED').update(status='OPENED')
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Tender',
                    object_id=str(tender.id),
                    object_repr=tender.tender_number,
                    changes={'status': 'CLOSED'}
                )
                
                messages.success(request, f'Tender {tender.tender_number} closed successfully! You can now proceed to evaluation.')
                return redirect('tender_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error closing tender: {str(e)}')
    
    return render(request, 'tenders/tender_close.html', {'tender': tender})


@login_required
def tender_evaluate(request, pk):
    """Start tender evaluation"""
    tender = get_object_or_404(Tender, pk=pk)
    
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to evaluate tenders.')
        return redirect('tender_detail', pk=pk)
    
    if tender.status not in ['CLOSED', 'EVALUATING']:
        messages.error(request, 'Only closed tenders can be evaluated.')
        return redirect('tender_detail', pk=pk)
    
    if tender.status == 'CLOSED':
        tender.status = 'EVALUATING'
        tender.save()
    
    # Get bids for evaluation
    bids = tender.bids.exclude(status='DISQUALIFIED').select_related('supplier')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                bid_id = request.POST.get('bid_id')
                bid = get_object_or_404(Bid, id=bid_id, tender=tender)
                
                # Create evaluation
                evaluation = BidEvaluation.objects.create(
                    bid=bid,
                    evaluator=request.user,
                    technical_compliance=request.POST.get('technical_compliance') == 'on',
                    financial_compliance=request.POST.get('financial_compliance') == 'on',
                    technical_score=Decimal(request.POST.get('technical_score')),
                    financial_score=Decimal(request.POST.get('financial_score')),
                    strengths=request.POST.get('strengths', ''),
                    weaknesses=request.POST.get('weaknesses', ''),
                    recommendation=request.POST.get('recommendation')
                )
                
                # Update bid scores
                bid.technical_score = evaluation.technical_score
                bid.financial_score = evaluation.financial_score
                bid.evaluation_score = evaluation.total_score
                
                # Determine if qualified
                if evaluation.technical_compliance and evaluation.financial_compliance:
                    bid.status = 'QUALIFIED'
                else:
                    bid.status = 'DISQUALIFIED'
                    bid.disqualification_reason = request.POST.get('disqualification_reason', '')
                
                bid.save()
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Bid',
                    object_id=str(bid.id),
                    object_repr=bid.bid_number,
                    changes={'evaluated': True, 'score': float(evaluation.total_score)}
                )
                
                messages.success(request, f'Bid {bid.bid_number} evaluated successfully!')
                return redirect('tender_evaluate', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error evaluating bid: {str(e)}')
    
    context = {
        'tender': tender,
        'bids': bids,
        'criteria': tender.evaluation_criteria.all().order_by('sequence'),
    }
    
    return render(request, 'tenders/tender_evaluate.html', context)


@login_required
def tender_award(request, pk):
    """Award tender to winning bidder"""
    tender = get_object_or_404(Tender, pk=pk)
    
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to award tenders.')
        return redirect('tender_detail', pk=pk)
    
    if tender.status != 'EVALUATING':
        messages.error(request, 'Only evaluated tenders can be awarded.')
        return redirect('tender_detail', pk=pk)
    
    # Get qualified bids ranked by score
    qualified_bids = tender.bids.filter(
        status='QUALIFIED'
    ).order_by('-evaluation_score', 'bid_amount')
    
    if not qualified_bids.exists():
        messages.error(request, 'No qualified bids available for award.')
        return redirect('tender_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                bid_id = request.POST.get('winning_bid')
                winning_bid = get_object_or_404(Bid, id=bid_id, tender=tender)
                
                # Update tender status
                tender.status = 'AWARDED'
                tender.save()
                
                # Update winning bid
                winning_bid.status = 'AWARDED'
                winning_bid.save()
                
                # Rank all qualified bids
                for idx, bid in enumerate(qualified_bids, 1):
                    bid.rank = idx
                    if bid.id != winning_bid.id:
                        bid.status = 'REJECTED'
                    bid.save()
                
                # Create notification for winning supplier
                if winning_bid.supplier.user_set.exists():
                    for user in winning_bid.supplier.user_set.filter(role='SUPPLIER'):
                        Notification.objects.create(
                            user=user,
                            notification_type='TENDER',
                            priority='URGENT',
                            title='Tender Award',
                            message=f'Congratulations! Your bid {winning_bid.bid_number} has been awarded for tender {tender.tender_number}',
                            link_url=f'/bids/{winning_bid.id}/'
                        )
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Tender',
                    object_id=str(tender.id),
                    object_repr=tender.tender_number,
                    changes={'status': 'AWARDED', 'winning_bid': winning_bid.bid_number}
                )
                
                messages.success(request, f'Tender {tender.tender_number} awarded to {winning_bid.supplier.name}!')
                return redirect('tender_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error awarding tender: {str(e)}')
    
    context = {
        'tender': tender,
        'qualified_bids': qualified_bids,
    }
    
    return render(request, 'tenders/tender_award.html', context)


@login_required
def tender_cancel(request, pk):
    """Cancel tender"""
    tender = get_object_or_404(Tender, pk=pk)
    
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to cancel tenders.')
        return redirect('tender_detail', pk=pk)
    
    if tender.status == 'AWARDED':
        messages.error(request, 'Cannot cancel awarded tenders.')
        return redirect('tender_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                cancellation_reason = request.POST.get('cancellation_reason')
                
                # Update tender status
                old_status = tender.status
                tender.status = 'CANCELLED'
                tender.save()
                
                # Create audit log
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    model_name='Tender',
                    object_id=str(tender.id),
                    object_repr=tender.tender_number,
                    changes={
                        'status': 'CANCELLED',
                        'old_status': old_status,
                        'reason': cancellation_reason
                    }
                )
                
                messages.success(request, f'Tender {tender.tender_number} cancelled successfully!')
                return redirect('tender_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error cancelling tender: {str(e)}')
    
    return render(request, 'tenders/tender_cancel.html', {'tender': tender})


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Avg, Q, Count, Min, Max
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import JsonResponse
from decimal import Decimal
from .models import (
    Bid, Tender, Supplier, BidItem, BidDocument, 
    BidEvaluation, PurchaseOrder, AuditLog, Notification,
    RequisitionItem
)


# ============================================================================
# BID MANAGEMENT VIEWS
# ============================================================================

@login_required
def bid_list(request):
    """List all bids in the system with filtering and statistics"""
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN', 'SUPPLIER']:
        messages.error(request, 'You do not have permission to view bids.')
        return redirect('dashboard')
    
    # Base queryset
    bids = Bid.objects.select_related(
        'tender',
        'supplier',
        'opened_by'
    ).prefetch_related(
        'items__requisition_item',
        'documents',
        'evaluations'
    ).annotate(
        evaluation_count=Count('evaluations'),
        items_count=Count('items')
    )
    
    # Filter for suppliers - only show their own bids
    if request.user.role == 'SUPPLIER':
        try:
            supplier = request.user.supplier_profile
            bids = bids.filter(supplier=supplier)
        except:
            messages.error(request, 'Supplier profile not found.')
            return redirect('dashboard')
    
    # Filters
    status_filter = request.GET.get('status', '')
    tender_filter = request.GET.get('tender', '')
    supplier_filter = request.GET.get('supplier', '')
    search_query = request.GET.get('search', '')
    min_amount = request.GET.get('min_amount', '')
    max_amount = request.GET.get('max_amount', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if status_filter:
        bids = bids.filter(status=status_filter)
    
    if tender_filter:
        bids = bids.filter(tender_id=tender_filter)
    
    if supplier_filter and request.user.role in ['PROCUREMENT', 'ADMIN']:
        bids = bids.filter(supplier_id=supplier_filter)
    
    if search_query:
        bids = bids.filter(
            Q(bid_number__icontains=search_query) |
            Q(supplier__name__icontains=search_query) |
            Q(supplier__supplier_number__icontains=search_query) |
            Q(tender__tender_number__icontains=search_query) |
            Q(tender__title__icontains=search_query)
        )
    
    if min_amount:
        try:
            bids = bids.filter(bid_amount__gte=Decimal(min_amount))
        except:
            pass
    
    if max_amount:
        try:
            bids = bids.filter(bid_amount__lte=Decimal(max_amount))
        except:
            pass
    
    if date_from:
        bids = bids.filter(submitted_at__date__gte=date_from)
    
    if date_to:
        bids = bids.filter(submitted_at__date__lte=date_to)
    
    # Sorting
    sort_by = request.GET.get('sort', '-submitted_at')
    valid_sort_fields = [
        'bid_amount', '-bid_amount', 
        'submitted_at', '-submitted_at', 
        'rank', '-rank',
        'bid_number', '-bid_number'
    ]
    if sort_by in valid_sort_fields:
        bids = bids.order_by(sort_by)
    else:
        bids = bids.order_by('-submitted_at')
    
    # Statistics (before pagination)
    all_bids = Bid.objects.all()
    if request.user.role == 'SUPPLIER':
        all_bids = all_bids.filter(supplier=supplier)
    
    stats = {
        'total': all_bids.count(),
        'submitted': all_bids.filter(status='SUBMITTED').count(),
        'opened': all_bids.filter(status='OPENED').count(),
        'evaluating': all_bids.filter(status='EVALUATING').count(),
        'qualified': all_bids.filter(status='QUALIFIED').count(),
        'disqualified': all_bids.filter(status='DISQUALIFIED').count(),
        'awarded': all_bids.filter(status='AWARDED').count(),
        'rejected': all_bids.filter(status='REJECTED').count(),
        'average_amount': all_bids.aggregate(Avg('bid_amount'))['bid_amount__avg'] or 0,
        'total_value': all_bids.aggregate(Sum('bid_amount'))['bid_amount__sum'] or 0,
    }
    
    # Get filter choices for dropdowns (only for procurement/admin)
    tenders_list = []
    suppliers_list = []
    
    if request.user.role in ['PROCUREMENT', 'ADMIN']:
        tenders_list = Tender.objects.filter(
            status__in=['PUBLISHED', 'CLOSED', 'EVALUATING']
        ).order_by('-created_at')[:50]
        
        suppliers_list = Supplier.objects.filter(
            status='APPROVED'
        ).order_by('name')
    
    # Pagination
    paginator = Paginator(bids, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'bids': page_obj,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'tender_filter': tender_filter,
        'supplier_filter': supplier_filter,
        'search_query': search_query,
        'min_amount': min_amount,
        'max_amount': max_amount,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'stats': stats,
        'status_choices': Bid.STATUS_CHOICES,
        'tenders_list': tenders_list,
        'suppliers_list': suppliers_list,
        'total_count': paginator.count,
    }
    
    return render(request, 'bids/bid_list.html', context)


@login_required
def bid_detail(request, pk):
    """View detailed bid information with committee evaluation support"""
    bid = get_object_or_404(
        Bid.objects.select_related(
            'tender',
            'supplier',
            'opened_by'
        ).prefetch_related(
            'items__requisition_item__item',
            'documents',
            'evaluations__evaluator',
            'tender__evaluation_committees__members__user'
        ),
        pk=pk
    )
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN', 'SUPPLIER']:
        messages.error(request, 'You do not have permission to view this bid.')
        return redirect('dashboard')
    
    # Suppliers can only view their own bids
    if request.user.role == 'SUPPLIER':
        try:
            supplier = request.user.supplier_profile
            if bid.supplier != supplier:
                messages.error(request, 'You can only view your own bids.')
                return redirect('dashboard')
        except:
            messages.error(request, 'Supplier profile not found.')
            return redirect('dashboard')
    
    # Calculate totals and statistics
    items_total = bid.items.aggregate(Sum('quoted_total'))['quoted_total__sum'] or 0
    items_count = bid.items.count()
    
    # Get evaluation summary
    evaluations = bid.evaluations.all()
    evaluation_summary = {
        'count': evaluations.count(),
        'avg_technical': evaluations.aggregate(Avg('technical_score'))['technical_score__avg'] or 0,
        'avg_financial': evaluations.aggregate(Avg('financial_score'))['financial_score__avg'] or 0,
        'avg_total': evaluations.aggregate(Avg('total_score'))['total_score__avg'] or 0,
    }
    
    # Compare with tender estimated budget
    variance = bid.bid_amount - bid.tender.estimated_budget
    variance_percentage = (variance / bid.tender.estimated_budget * 100) if bid.tender.estimated_budget > 0 else 0
    
    # Check if user is a committee member for this tender
    is_committee_member = False
    committee_role = None
    active_committee = None
    
    if request.user.role in ['PROCUREMENT', 'ADMIN']:
        # Get active evaluation committee for this tender
        active_committee = bid.tender.evaluation_committees.filter(
            status='ACTIVE'
        ).first()
        
        if active_committee:
            # Check if user is a member
            membership = active_committee.members.filter(
                user=request.user,
                is_active=True
            ).first()
            
            if membership:
                is_committee_member = True
                committee_role = membership.role
    
    # Check if bid can be awarded
    can_award = (
        request.user.role in ['PROCUREMENT', 'ADMIN'] and
        bid.status in ['QUALIFIED'] and
        bid.tender.status in ['EVALUATING', 'CLOSED'] and
        not PurchaseOrder.objects.filter(bid=bid).exists() and
        # Ensure evaluation is complete
        bid.tender.technical_evaluation_complete and
        bid.tender.financial_evaluation_complete
    )
    
    # Check if bid can be opened
    can_open = (
        request.user.role in ['PROCUREMENT', 'ADMIN'] and
        bid.status == 'SUBMITTED' and
        bid.tender.closing_date <= timezone.now()
    )
    
    # Check if bid can be evaluated (must be committee member)
    can_evaluate = (
        is_committee_member and
        bid.status in ['OPENED', 'EVALUATING'] and
        active_committee and
        active_committee.status == 'ACTIVE'
    )
    
    # Check if bid can be disqualified (committee chairperson or admin)
    can_disqualify = (
        (committee_role == 'CHAIRPERSON' or request.user.role == 'ADMIN') and
        bid.status not in ['DISQUALIFIED', 'AWARDED', 'REJECTED']
    )
    
    # Check if user has already evaluated this bid
    user_has_evaluated = False
    user_evaluation = None
    
    if is_committee_member:
        from .models import TechnicalEvaluationScore, FinancialEvaluationScore
        
        # Check for technical evaluation
        technical_scores = TechnicalEvaluationScore.objects.filter(
            bid=bid,
            evaluator=request.user
        ).exists()
        
        # Check for financial evaluation
        financial_scores = FinancialEvaluationScore.objects.filter(
            bid=bid,
            evaluator=request.user
        ).exists()
        
        user_has_evaluated = technical_scores or financial_scores
        
        if user_has_evaluated:
            user_evaluation = {
                'technical_complete': technical_scores,
                'financial_complete': financial_scores,
            }
    
    # Get committee evaluation progress
    committee_progress = None
    if active_committee:
        total_members = active_committee.members.filter(
            is_active=True,
            role__in=['CHAIRPERSON', 'MEMBER']  # Exclude observers
        ).count()
        
        # Count members who have submitted evaluations
        from .models import TechnicalEvaluationScore, FinancialEvaluationScore
        
        technical_evaluators = TechnicalEvaluationScore.objects.filter(
            bid=bid
        ).values('evaluator').distinct().count()
        
        financial_evaluators = FinancialEvaluationScore.objects.filter(
            bid=bid
        ).values('evaluator').distinct().count()
        
        committee_progress = {
            'total_members': total_members,
            'technical_completed': technical_evaluators,
            'financial_completed': financial_evaluators,
            'technical_percentage': (technical_evaluators / total_members * 100) if total_members > 0 else 0,
            'financial_percentage': (financial_evaluators / total_members * 100) if total_members > 0 else 0,
        }
    
    context = {
        'bid': bid,
        'items_total': items_total,
        'items_count': items_count,
        'evaluation_summary': evaluation_summary,
        'variance': variance,
        'variance_percentage': variance_percentage,
        'can_award': can_award,
        'can_open': can_open,
        'can_evaluate': can_evaluate,
        'can_disqualify': can_disqualify,
        'is_committee_member': is_committee_member,
        'committee_role': committee_role,
        'active_committee': active_committee,
        'user_has_evaluated': user_has_evaluated,
        'user_evaluation': user_evaluation,
        'committee_progress': committee_progress,
    }
    
    return render(request, 'bids/bid_detail.html', context)



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Avg, Sum, Count, Q
from django.utils import timezone
from decimal import Decimal
from pms.models import (
    Bid, Tender, TechnicalEvaluationCriteria, FinancialEvaluationCriteria,
    TechnicalEvaluationScore, FinancialEvaluationScore, BidTechnicalResponse,
    EvaluationCommittee, CommitteeMember, CombinedEvaluationResult,
    PurchaseOrder, EvaluationReport
)

# ============================================================================
# TECHNICAL EVALUATION
# ============================================================================

@login_required
def bid_evaluate_technical(request, pk):
    """Technical evaluation of a bid by committee member"""
    bid = get_object_or_404(
        Bid.objects.select_related('tender', 'supplier'),
        pk=pk
    )
    
    # Check if user is committee member
    active_committee = bid.tender.evaluation_committees.filter(
        status='ACTIVE'
    ).first()
    
    if not active_committee:
        messages.error(request, 'No active evaluation committee for this tender.')
        return redirect('bid_detail', pk=bid.id)
    
    membership = active_committee.members.filter(
        user=request.user,
        is_active=True
    ).first()
    
    if not membership:
        messages.error(request, 'You are not a member of the evaluation committee.')
        return redirect('bid_detail', pk=bid.id)
    
    # Check if already evaluated
    existing_scores = TechnicalEvaluationScore.objects.filter(
        bid=bid,
        evaluator=request.user
    )
    
    # Get technical criteria for this tender
    criteria = bid.tender.technical_criteria.all().order_by('sequence')
    
    if not criteria.exists():
        messages.error(request, 'No technical evaluation criteria defined for this tender.')
        return redirect('bid_detail', pk=bid.id)
    
    # Get bid's technical responses
    responses = BidTechnicalResponse.objects.filter(bid=bid)
    
    if request.method == 'POST':
        # Process technical evaluation
        try:
            # Delete existing scores to allow re-evaluation
            existing_scores.delete()
            
            total_weighted_score = Decimal('0')
            all_compliant = True
            
            for criterion in criteria:
                score_value = request.POST.get(f'score_{criterion.id}')
                is_compliant = request.POST.get(f'compliant_{criterion.id}') == 'on'
                comments = request.POST.get(f'comments_{criterion.id}', '')
                justification = request.POST.get(f'justification_{criterion.id}', '')
                
                if not score_value:
                    messages.error(request, f'Please provide score for: {criterion.criterion_name}')
                    return redirect('bid_evaluate_technical', pk=bid.id)
                
                score = Decimal(score_value)
                
                # Validate score
                if score < 0 or score > criterion.max_score:
                    messages.error(request, f'Score for {criterion.criterion_name} must be between 0 and {criterion.max_score}')
                    return redirect('bid_evaluate_technical', pk=bid.id)
                
                # Check pass/fail criteria
                if criterion.is_pass_fail and not is_compliant:
                    all_compliant = False
                
                # Create score record
                eval_score = TechnicalEvaluationScore.objects.create(
                    bid=bid,
                    criterion=criterion,
                    evaluator=request.user,
                    score=score,
                    is_compliant=is_compliant,
                    comments=comments,
                    justification=justification
                )
                
                total_weighted_score += eval_score.weighted_score
            
            # Update bid status
            if bid.status == 'OPENED':
                bid.status = 'EVALUATING'
                bid.save()
            
            messages.success(request, 'Technical evaluation submitted successfully.')
            return redirect('bid_detail', pk=bid.id)
            
        except Exception as e:
            messages.error(request, f'Error submitting evaluation: {str(e)}')
            return redirect('bid_evaluate_technical', pk=bid.id)
    
    # Prepare criteria with existing scores
    criteria_data = []
    for criterion in criteria:
        existing_score = existing_scores.filter(criterion=criterion).first()
        response = responses.filter(criterion=criterion).first()
        
        criteria_data.append({
            'criterion': criterion,
            'response': response,
            'existing_score': existing_score,
        })
    
    context = {
        'bid': bid,
        'committee': active_committee,
        'membership': membership,
        'criteria_data': criteria_data,
        'is_reevaluation': existing_scores.exists(),
    }
    
    return render(request, 'bids/evaluate_technical.html', context)


# ============================================================================
# FINANCIAL EVALUATION
# ============================================================================

@login_required
def bid_evaluate_financial(request, pk):
    """Financial evaluation of a bid by committee member"""
    bid = get_object_or_404(
        Bid.objects.select_related('tender', 'supplier'),
        pk=pk
    )
    
    # Check if user is committee member
    active_committee = bid.tender.evaluation_committees.filter(
        status='ACTIVE'
    ).first()
    
    if not active_committee:
        messages.error(request, 'No active evaluation committee for this tender.')
        return redirect('bid_detail', pk=bid.id)
    
    membership = active_committee.members.filter(
        user=request.user,
        is_active=True
    ).first()
    
    if not membership:
        messages.error(request, 'You are not a member of the evaluation committee.')
        return redirect('bid_detail', pk=bid.id)
    
    # Get financial criteria
    try:
        financial_criteria = bid.tender.financial_criteria
    except:
        messages.error(request, 'No financial evaluation criteria defined for this tender.')
        return redirect('bid_detail', pk=bid.id)
    
    # Check if already evaluated
    existing_score = FinancialEvaluationScore.objects.filter(
        bid=bid,
        evaluator=request.user
    ).first()
    
    # Get all bids for this tender to calculate lowest price
    all_bids = Bid.objects.filter(
        tender=bid.tender,
        status__in=['OPENED', 'EVALUATING', 'QUALIFIED']
    ).order_by('bid_amount')
    
    lowest_bid = all_bids.first()
    
    if request.method == 'POST':
        try:
            # Get form data
            is_arithmetic_correct = request.POST.get('arithmetic_correct') == 'on'
            arithmetic_notes = request.POST.get('arithmetic_notes', '')
            is_within_budget = request.POST.get('within_budget') == 'on'
            is_price_reasonable = request.POST.get('price_reasonable') == 'on'
            comments = request.POST.get('comments', '')
            
            # Calculate variance from estimate
            variance = ((bid.bid_amount - bid.tender.estimated_budget) / bid.tender.estimated_budget * 100)
            
            # Calculate financial score based on evaluation method
            if financial_criteria.evaluation_method == 'LOWEST_PRICE':
                # Formula: (Lowest Price / Bid Price) × 100
                if lowest_bid and bid.bid_amount > 0:
                    financial_score = (lowest_bid.bid_amount / bid.bid_amount * 100)
                else:
                    financial_score = Decimal('100')
            elif financial_criteria.evaluation_method == 'FIXED_BUDGET':
                # Score based on how close to budget
                if is_within_budget:
                    financial_score = Decimal('100')
                else:
                    financial_score = Decimal('0')
            else:  # QUALITY_COST
                # Will be calculated later based on technical score
                financial_score = Decimal('100')
            
            # Cap at 100
            if financial_score > 100:
                financial_score = Decimal('100')
            
            # Delete existing score if re-evaluating
            if existing_score:
                existing_score.delete()
            
            # Create financial evaluation score
            eval_score = FinancialEvaluationScore.objects.create(
                bid=bid,
                evaluator=request.user,
                quoted_amount=bid.bid_amount,
                is_arithmetic_correct=is_arithmetic_correct,
                arithmetic_notes=arithmetic_notes,
                is_within_budget=is_within_budget,
                is_price_reasonable=is_price_reasonable,
                variance_from_estimate=variance,
                financial_score=financial_score,
                weighted_financial_score=financial_score * financial_criteria.financial_weight / 100,
                comments=comments
            )
            
            messages.success(request, 'Financial evaluation submitted successfully.')
            return redirect('bid_detail', pk=bid.id)
            
        except Exception as e:
            messages.error(request, f'Error submitting evaluation: {str(e)}')
            return redirect('bid_evaluate_financial', pk=bid.id)
    
    # Calculate variance
    variance = ((bid.bid_amount - bid.tender.estimated_budget) / bid.tender.estimated_budget * 100) if bid.tender.estimated_budget > 0 else 0
    
    context = {
        'bid': bid,
        'committee': active_committee,
        'membership': membership,
        'financial_criteria': financial_criteria,
        'existing_score': existing_score,
        'lowest_bid': lowest_bid,
        'variance': variance,
        'all_bids': all_bids,
    }
    
    return render(request, 'bids/evaluate_financial.html', context)


# ============================================================================
# VIEW EVALUATION
# ============================================================================

@login_required
def bid_view_evaluation(request, pk):
    """View user's submitted evaluation"""
    bid = get_object_or_404(Bid, pk=pk)
    
    # Get user's technical scores
    technical_scores = TechnicalEvaluationScore.objects.filter(
        bid=bid,
        evaluator=request.user
    ).select_related('criterion')
    
    # Get user's financial score
    financial_score = FinancialEvaluationScore.objects.filter(
        bid=bid,
        evaluator=request.user
    ).first()
    
    # Calculate totals
    total_technical = technical_scores.aggregate(
        total=Sum('weighted_score')
    )['total'] or 0
    
    context = {
        'bid': bid,
        'technical_scores': technical_scores,
        'financial_score': financial_score,
        'total_technical': total_technical,
    }
    
    return render(request, 'bids/view_evaluation.html', context)


# ============================================================================
# COMMITTEE EVALUATION SUMMARY
# ============================================================================

@login_required
def committee_evaluation_summary(request, tender_id):
    """Committee chairperson view of all evaluations"""
    tender = get_object_or_404(Tender, pk=tender_id)
    
    # Check if user is chairperson
    active_committee = tender.evaluation_committees.filter(
        status='ACTIVE'
    ).first()
    
    if not active_committee:
        messages.error(request, 'No active evaluation committee for this tender.')
        return redirect('tender_detail', pk=tender.id)
    
    if active_committee.chairperson != request.user and request.user.role != 'ADMIN':
        messages.error(request, 'Only the committee chairperson can view this summary.')
        return redirect('tender_detail', pk=tender.id)
    
    # Get all bids for this tender
    bids = Bid.objects.filter(tender=tender).select_related('supplier')
    
    # Get evaluation statistics for each bid
    bid_evaluations = []
    for bid in bids:
        # Technical evaluation stats
        tech_scores = TechnicalEvaluationScore.objects.filter(bid=bid)
        tech_avg = tech_scores.aggregate(
            avg_score=Avg('weighted_score')
        )['avg_score'] or 0
        
        # Financial evaluation stats
        fin_scores = FinancialEvaluationScore.objects.filter(bid=bid)
        fin_avg = fin_scores.aggregate(
            avg_score=Avg('financial_score')
        )['avg_score'] or 0
        
        # Count evaluators
        tech_evaluators = tech_scores.values('evaluator').distinct().count()
        fin_evaluators = fin_scores.values('evaluator').distinct().count()
        
        # Get or create combined result
        combined, created = CombinedEvaluationResult.objects.get_or_create(
            bid=bid,
            defaults={
                'average_technical_score': tech_avg,
                'average_financial_score': fin_avg,
                'technical_pass_mark': tender.technical_pass_mark,
                'evaluated_by_committee': active_committee,
                'evaluation_date': timezone.now().date()
            }
        )
        
        # Update if not created
        if not created:
            combined.average_technical_score = tech_avg
            combined.average_financial_score = fin_avg
            combined.save()
            combined.calculate_combined_score()
        
        bid_evaluations.append({
            'bid': bid,
            'tech_avg': tech_avg,
            'fin_avg': fin_avg,
            'combined_score': combined.combined_score,
            'tech_evaluators': tech_evaluators,
            'fin_evaluators': fin_evaluators,
            'is_qualified': combined.is_technically_qualified,
            'combined_result': combined,
        })
    
    # Sort by combined score
    bid_evaluations.sort(key=lambda x: x['combined_score'], reverse=True)
    
    # Assign ranks
    for rank, item in enumerate(bid_evaluations, start=1):
        item['bid'].rank = rank
        item['bid'].save()
        item['rank'] = rank
    
    context = {
        'tender': tender,
        'committee': active_committee,
        'bid_evaluations': bid_evaluations,
        'total_members': active_committee.members.filter(is_active=True).count(),
    }
    
    return render(request, 'bids/committee_summary.html', context)


# ============================================================================
# BID DISQUALIFICATION
# ============================================================================

@login_required
def bid_disqualify(request, pk):
    """Disqualify a bid"""
    bid = get_object_or_404(Bid, pk=pk)
    
    # Permission check
    active_committee = bid.tender.evaluation_committees.filter(
        status='ACTIVE'
    ).first()
    
    is_chairperson = active_committee and active_committee.chairperson == request.user
    is_admin = request.user.role == 'ADMIN'
    
    if not (is_chairperson or is_admin):
        messages.error(request, 'Only committee chairperson or admin can disqualify bids.')
        return redirect('bid_detail', pk=bid.id)
    
    if bid.status in ['AWARDED', 'DISQUALIFIED']:
        messages.error(request, f'Cannot disqualify a bid that is {bid.get_status_display()}.')
        return redirect('bid_detail', pk=bid.id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason')
        
        if not reason:
            messages.error(request, 'Please provide a reason for disqualification.')
            return redirect('bid_disqualify', pk=bid.id)
        
        bid.status = 'DISQUALIFIED'
        bid.disqualification_reason = reason
        bid.save()
        
        # Update combined evaluation result
        try:
            combined = CombinedEvaluationResult.objects.get(bid=bid)
            combined.is_disqualified = True
            combined.disqualification_reason = reason
            combined.save()
        except:
            pass
        
        messages.success(request, f'Bid {bid.bid_number} has been disqualified.')
        return redirect('bid_detail', pk=bid.id)
    
    context = {
        'bid': bid,
        'committee': active_committee,
    }
    
    return render(request, 'bids/disqualify.html', context)


# ============================================================================
# BID OPENING
# ============================================================================

@login_required
def bid_open(request, pk):
    """Open a submitted bid"""
    bid = get_object_or_404(Bid, pk=pk)
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to open bids.')
        return redirect('bid_detail', pk=bid.id)
    
    if bid.status != 'SUBMITTED':
        messages.error(request, 'Only submitted bids can be opened.')
        return redirect('bid_detail', pk=bid.id)
    
    if bid.tender.closing_date > timezone.now():
        messages.error(request, 'Cannot open bids before tender closing date.')
        return redirect('bid_detail', pk=bid.id)
    
    if request.method == 'POST':
        bid.status = 'OPENED'
        bid.opened_at = timezone.now()
        bid.opened_by = request.user
        bid.save()
        
        messages.success(request, f'Bid {bid.bid_number} has been opened.')
        return redirect('bid_detail', pk=bid.id)
    
    context = {
        'bid': bid,
    }
    
    return render(request, 'bids/open_bid.html', context)


# ============================================================================
# BID AWARD
# ============================================================================

@login_required
def bid_award(request, pk):
    """Award tender to a bid"""
    bid = get_object_or_404(Bid, pk=pk)
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to award tenders.')
        return redirect('bid_detail', pk=bid.id)
    
    if bid.status != 'QUALIFIED':
        messages.error(request, 'Only qualified bids can be awarded.')
        return redirect('bid_detail', pk=bid.id)
    
    # Check if evaluations are complete
    if not bid.tender.technical_evaluation_complete or not bid.tender.financial_evaluation_complete:
        messages.error(request, 'Evaluations must be complete before awarding.')
        return redirect('bid_detail', pk=bid.id)
    
    # Check if already awarded to another bid
    existing_award = Bid.objects.filter(
        tender=bid.tender,
        status='AWARDED'
    ).first()
    
    if existing_award and existing_award.id != bid.id:
        messages.error(request, f'Tender already awarded to {existing_award.supplier.name}.')
        return redirect('bid_detail', pk=bid.id)
    
    if request.method == 'POST':
        # Award the bid
        bid.status = 'AWARDED'
        bid.save()
        
        # Update tender status
        bid.tender.status = 'AWARDED'
        bid.tender.save()
        
        # Mark other bids as rejected
        Bid.objects.filter(tender=bid.tender).exclude(id=bid.id).update(
            status='REJECTED'
        )
        
        # Update combined evaluation
        try:
            combined = CombinedEvaluationResult.objects.get(bid=bid)
            combined.is_recommended = True
            combined.save()
        except:
            pass
        
        messages.success(request, f'Tender awarded to {bid.supplier.name}.')
        return redirect('bid_detail', pk=bid.id)
    
    # Get combined evaluation result
    try:
        combined = CombinedEvaluationResult.objects.get(bid=bid)
    except:
        combined = None
    
    context = {
        'bid': bid,
        'combined': combined,
    }
    
    return render(request, 'bids/award.html', context)


# ============================================================================
# BID COMPARISON
# ============================================================================

@login_required
def bid_comparison(request, tender_id):
    """Compare all bids for a tender"""
    tender = get_object_or_404(Tender, pk=tender_id)
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        # Check if committee member
        is_member = tender.evaluation_committees.filter(
            status='ACTIVE',
            members__user=request.user,
            members__is_active=True
        ).exists()
        
        if not is_member:
            messages.error(request, 'You do not have permission to view this comparison.')
            return redirect('tender_detail', pk=tender.id)
    
    # Get all bids with evaluations
    bids = Bid.objects.filter(tender=tender).select_related('supplier')
    
    comparison_data = []
    for bid in bids:
        # Get combined evaluation
        try:
            combined = CombinedEvaluationResult.objects.get(bid=bid)
        except:
            combined = None
        
        comparison_data.append({
            'bid': bid,
            'combined': combined,
            'variance': bid.bid_amount - tender.estimated_budget,
            'variance_pct': ((bid.bid_amount - tender.estimated_budget) / tender.estimated_budget * 100) if tender.estimated_budget > 0 else 0,
        })
    
    # Sort by rank or combined score
    comparison_data.sort(key=lambda x: x['combined'].combined_score if x['combined'] else 0, reverse=True)
    
    context = {
        'tender': tender,
        'comparison_data': comparison_data,
    }
    
    return render(request, 'bids/comparison.html', context)



from decimal import Decimal, InvalidOperation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages

def to_decimal(value, default='0'):
    """Safely convert value to Decimal"""
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


@login_required
def bid_evaluate(request, pk):
    """Evaluate a bid"""
    bid = get_object_or_404(Bid, pk=pk)
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to evaluate bids.')
        return redirect('bid_detail', pk=pk)
    
    # Check if bid can be evaluated
    if bid.status not in ['OPENED', 'EVALUATING']:
        messages.error(request, 'This bid cannot be evaluated.')
        return redirect('bid_detail', pk=pk)
    
    # Check if user already evaluated this bid
    existing_evaluation = BidEvaluation.objects.filter(
        bid=bid, 
        evaluator=request.user
    ).first()
    
    if request.method == 'POST':
        technical_compliance = request.POST.get('technical_compliance') == 'on'
        financial_compliance = request.POST.get('financial_compliance') == 'on'
        
        # Use to_decimal helper for safe conversion
        technical_score = to_decimal(request.POST.get('technical_score', '0'))
        financial_score = to_decimal(request.POST.get('financial_score', '0'))
        
        strengths = request.POST.get('strengths', '')
        weaknesses = request.POST.get('weaknesses', '')
        recommendation = request.POST.get('recommendation', '')
        
        # Validate scores
        if technical_score < 0 or technical_score > 100:
            messages.error(request, 'Technical score must be between 0 and 100.')
            return redirect('bid_evaluate', pk=pk)
        
        if financial_score < 0 or financial_score > 100:
            messages.error(request, 'Financial score must be between 0 and 100.')
            return redirect('bid_evaluate', pk=pk)
        
        if existing_evaluation:
            # Update existing evaluation
            existing_evaluation.technical_compliance = technical_compliance
            existing_evaluation.financial_compliance = financial_compliance
            existing_evaluation.technical_score = technical_score
            existing_evaluation.financial_score = financial_score
            existing_evaluation.strengths = strengths
            existing_evaluation.weaknesses = weaknesses
            existing_evaluation.recommendation = recommendation
            existing_evaluation.save()
            
            messages.success(request, 'Your evaluation has been updated successfully.')
        else:
            # Create new evaluation
            BidEvaluation.objects.create(
                bid=bid,
                evaluator=request.user,
                technical_compliance=technical_compliance,
                financial_compliance=financial_compliance,
                technical_score=technical_score,
                financial_score=financial_score,
                strengths=strengths,
                weaknesses=weaknesses,
                recommendation=recommendation
            )
            
            messages.success(request, 'Bid evaluation submitted successfully.')
        
        # Update bid status to EVALUATING if it's OPENED
        if bid.status == 'OPENED':
            bid.status = 'EVALUATING'
            bid.save()
        
        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='Bid',
            object_id=str(bid.id),
            object_repr=str(bid),
            changes={'evaluation_added': True},
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        return redirect('bid_detail', pk=pk)
    
    # Get evaluation criteria for the tender
    criteria = bid.tender.evaluation_criteria.all().order_by('sequence')
    
    context = {
        'bid': bid,
        'existing_evaluation': existing_evaluation,
        'criteria': criteria,
    }
    
    return render(request, 'bids/bid_evaluate.html', context)


@login_required
def bid_qualify(request, pk):
    """Mark bid as qualified"""
    bid = get_object_or_404(Bid, pk=pk)
    
    # Permission check
    if request.user.role not in ['PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to qualify bids.')
        return redirect('bid_detail', pk=pk)
    
    # Check if bid can be qualified
    if bid.status not in ['EVALUATING']:
        messages.error(request, 'Only bids under evaluation can be qualified.')
        return redirect('bid_detail', pk=pk)
    
    if request.method == 'POST':
        # Calculate average scores from evaluations
        evaluations = bid.evaluations.all()
        if evaluations.count() == 0:
            messages.error(request, 'Bid must be evaluated before qualification.')
            return redirect('bid_detail', pk=pk)
        
        avg_technical = evaluations.aggregate(Avg('technical_score'))['technical_score__avg']
        avg_financial = evaluations.aggregate(Avg('financial_score'))['financial_score__avg']
        
        bid.technical_score = avg_technical
        bid.financial_score = avg_financial
        bid.evaluation_score = (avg_technical * Decimal('0.7')) + (avg_financial * Decimal('0.3'))
        bid.status = 'QUALIFIED'
        bid.save()
        
        # Update rankings for all qualified bids in this tender
        qualified_bids = Bid.objects.filter(
            tender=bid.tender,
            status='QUALIFIED'
        ).order_by('-evaluation_score')
        
        for idx, qualified_bid in enumerate(qualified_bids, start=1):
            qualified_bid.rank = idx
            qualified_bid.save()
        
        # Create audit log
        AuditLog.objects.create(
            user=request.user,
            action='APPROVE',
            model_name='Bid',
            object_id=str(bid.id),
            object_repr=str(bid),
            changes={'status': 'QUALIFIED', 'rank': bid.rank},
            ip_address=request.META.get('REMOTE_ADDR'),
        )
        
        # Notify supplier
        Notification.objects.create(
            user=bid.supplier.user if hasattr(bid.supplier, 'user') else None,
            notification_type='TENDER',
            priority='MEDIUM',
            title=f'Bid Qualified - {bid.tender.tender_number}',
            message=f'Your bid {bid.bid_number} has been qualified and ranked #{bid.rank}.',
            link_url=f'/bids/{bid.id}/'
        )
        
        messages.success(request, f'Bid {bid.bid_number} has been qualified and ranked #{bid.rank}.')
        return redirect('bid_detail', pk=pk)
    
    return render(request, 'bids/bid_qualify_confirm.html', {'bid': bid})





# API ENDPOINTS

@login_required
def get_requisition_items(request, requisition_id):
    """API: Get requisition items for tender creation"""
    try:
        requisition = Requisition.objects.get(id=requisition_id)
        items = requisition.items.all().select_related('item')
        
        items_data = [{
            'id': str(item.id),
            'description': item.item_description,
            'specifications': item.specifications,
            'quantity': float(item.quantity),
            'unit': item.unit_of_measure,
            'estimated_price': float(item.estimated_unit_price),
        } for item in items]
        
        return JsonResponse({
            'success': True,
            'items': items_data,
            'total': float(requisition.estimated_amount)
        })
    except Requisition.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Requisition not found'}, status=404)


@login_required
def get_tender_statistics(request):
    """API: Get tender statistics for dashboard"""
    try:
        stats = {
            'total_tenders': Tender.objects.count(),
            'active_tenders': Tender.objects.filter(status='PUBLISHED').count(),
            'closed_tenders': Tender.objects.filter(status='CLOSED').count(),
            'awarded_tenders': Tender.objects.filter(status='AWARDED').count(),
            'total_value': float(Tender.objects.aggregate(Sum('estimated_budget'))['estimated_budget__sum'] or 0),
            'average_bids': float(Tender.objects.annotate(
                bid_count=Count('bids')
            ).aggregate(Avg('bid_count'))['bid_count__avg'] or 0),
        }
        
        return JsonResponse({'success': True, 'data': stats})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg, F
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json

from .models import (
    PurchaseOrder, PurchaseOrderItem, Requisition, Supplier, 
    Bid, POAmendment, User, Department, Budget, GoodsReceivedNote,
    Invoice, StockMovement
)


@login_required
def po_dashboard(request):
    """Purchase Order Dashboard with statistics and charts"""
    
    # Date filters
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    current_year = today.year
    
    # Basic statistics
    total_pos = PurchaseOrder.objects.count()
    pending_pos = PurchaseOrder.objects.filter(
        status__in=['DRAFT', 'PENDING_APPROVAL']
    ).count()
    active_pos = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY']
    ).count()
    
    # Financial statistics
    total_po_value = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY', 'DELIVERED', 'CLOSED']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    monthly_value = PurchaseOrder.objects.filter(
        po_date__gte=thirty_days_ago
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Status distribution
    status_data = PurchaseOrder.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    status_chart = {
        'labels': [dict(PurchaseOrder.STATUS_CHOICES).get(item['status'], item['status']) 
                   for item in status_data],
        'values': [item['count'] for item in status_data]
    }
    
    # Monthly PO trend (last 6 months)
    months = []
    po_counts = []
    po_values = []
    
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if i > 0:
            next_month = month_date.replace(day=28) + timedelta(days=4)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        else:
            month_end = today
            
        month_pos = PurchaseOrder.objects.filter(
            po_date__gte=month_start,
            po_date__lte=month_end
        )
        
        months.append(month_start.strftime('%b %Y'))
        po_counts.append(month_pos.count())
        po_values.append(float(month_pos.aggregate(
            total=Sum('total_amount'))['total'] or 0))
    
    po_trend_chart = {
        'labels': months,
        'counts': po_counts,
        'values': po_values
    }
    
    # Top suppliers by PO value
    top_suppliers = PurchaseOrder.objects.values(
        'supplier__name'
    ).annotate(
        total_value=Sum('total_amount'),
        po_count=Count('id')
    ).order_by('-total_value')[:5]
    
    supplier_chart = {
        'labels': [item['supplier__name'] for item in top_suppliers],
        'values': [float(item['total_value']) for item in top_suppliers],
        'counts': [item['po_count'] for item in top_suppliers]
    }
    
    # Department spend
    dept_spend = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED', 'CLOSED']
    ).values(
        'requisition__department__name'
    ).annotate(
        total_spend=Sum('total_amount')
    ).order_by('-total_spend')[:8]
    
    dept_chart = {
        'labels': [item['requisition__department__name'] or 'N/A' 
                   for item in dept_spend],
        'values': [float(item['total_spend']) for item in dept_spend]
    }
    
    # Recent POs
    recent_pos = PurchaseOrder.objects.select_related(
        'supplier', 'requisition', 'created_by'
    ).order_by('-created_at')[:10]
    
    # Pending approvals
    pending_approvals = PurchaseOrder.objects.filter(
        status='PENDING_APPROVAL'
    ).select_related('supplier', 'requisition').order_by('-created_at')[:5]
    
    # Delivery tracking
    pending_delivery = PurchaseOrder.objects.filter(
        status__in=['SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY'],
        delivery_date__lte=today + timedelta(days=7)
    ).select_related('supplier').order_by('delivery_date')[:10]
    
    # Performance metrics
    avg_po_value = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED', 'CLOSED']
    ).aggregate(avg=Avg('total_amount'))['avg'] or 0
    
    # Delivery performance
    delivered_pos = PurchaseOrder.objects.filter(status='DELIVERED')
    on_time_delivery = delivered_pos.filter(
        delivery_date__gte=F('po_date')
    ).count()
    total_delivered = delivered_pos.count()
    delivery_rate = (on_time_delivery / total_delivered * 100) if total_delivered > 0 else 0
    
    context = {
        'total_pos': total_pos,
        'pending_pos': pending_pos,
        'active_pos': active_pos,
        'total_po_value': total_po_value,
        'monthly_value': monthly_value,
        'avg_po_value': avg_po_value,
        'delivery_rate': round(delivery_rate, 1),
        'status_chart': status_chart,
        'po_trend_chart': po_trend_chart,
        'supplier_chart': supplier_chart,
        'dept_chart': dept_chart,
        'recent_pos': recent_pos,
        'pending_approvals': pending_approvals,
        'pending_delivery': pending_delivery,
    }
    
    return render(request, 'procurement/po_dashboard.html', context)


@login_required
def po_list(request):
    """List all purchase orders with filtering and search"""
    
    pos = PurchaseOrder.objects.select_related(
        'supplier', 'requisition', 'created_by', 'approved_by'
    ).all()
    
    # Filters
    status_filter = request.GET.get('status')
    supplier_filter = request.GET.get('supplier')
    department_filter = request.GET.get('department')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        pos = pos.filter(status=status_filter)
    
    if supplier_filter:
        pos = pos.filter(supplier_id=supplier_filter)
    
    if department_filter:
        pos = pos.filter(requisition__department_id=department_filter)
    
    if date_from:
        pos = pos.filter(po_date__gte=date_from)
    
    if date_to:
        pos = pos.filter(po_date__lte=date_to)
    
    if search:
        pos = pos.filter(
            Q(po_number__icontains=search) |
            Q(supplier__name__icontains=search) |
            Q(requisition__requisition_number__icontains=search)
        )
    
    pos = pos.order_by('-created_at')
    
    # Get filter options
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Statistics for current filter
    total_value = pos.aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'pos': pos,
        'suppliers': suppliers,
        'departments': departments,
        'status_choices': PurchaseOrder.STATUS_CHOICES,
        'total_value': total_value,
        'filters': {
            'status': status_filter,
            'supplier': supplier_filter,
            'department': department_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    
    return render(request, 'procurement/po_list.html', context)


@login_required
def po_detail(request, po_id):
    """Purchase Order detail view"""
    
    po = get_object_or_404(
        PurchaseOrder.objects.select_related(
            'supplier', 'requisition', 'bid', 'created_by', 'approved_by'
        ),
        id=po_id
    )
    
    # Get PO items
    po_items = po.items.select_related('requisition_item').all()
    
    # Get amendments
    amendments = po.amendments.select_related(
        'requested_by', 'approved_by'
    ).order_by('-created_at')
    
    # Get GRNs
    grns = po.grns.select_related('store', 'received_by').order_by('-created_at')
    
    # Get invoices
    invoices = po.invoices.select_related('supplier').order_by('-created_at')
    
    # Calculate delivery progress
    total_items = po_items.count()
    items_delivered = po_items.filter(
        quantity_delivered__gte=F('quantity')
    ).count()
    delivery_progress = (items_delivered / total_items * 100) if total_items > 0 else 0
    
    # Calculate invoice status
    total_invoiced = invoices.aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    invoice_progress = (total_invoiced / po.total_amount * 100) if po.total_amount > 0 else 0
    
    context = {
        'po': po,
        'po_items': po_items,
        'amendments': amendments,
        'grns': grns,
        'invoices': invoices,
        'delivery_progress': round(delivery_progress, 1),
        'invoice_progress': round(invoice_progress, 1),
    }
    
    return render(request, 'procurement/po_detail.html', context)


@login_required
def po_create(request):
    """Create new purchase order from approved requisition"""
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get form data
                requisition_id = request.POST.get('requisition')
                supplier_id = request.POST.get('supplier')
                bid_id = request.POST.get('bid')
                delivery_date = request.POST.get('delivery_date')
                delivery_address = request.POST.get('delivery_address')
                payment_terms = request.POST.get('payment_terms')
                warranty_terms = request.POST.get('warranty_terms', '')
                special_instructions = request.POST.get('special_instructions', '')
                
                # Validate
                requisition = get_object_or_404(Requisition, id=requisition_id)
                supplier = get_object_or_404(Supplier, id=supplier_id)
                
                # Check if requisition is approved
                if requisition.status != 'APPROVED':
                    messages.error(request, 'Requisition must be fully approved')
                    return redirect('po_create')
                
                # Check if PO already exists for this requisition
                if PurchaseOrder.objects.filter(requisition=requisition).exists():
                    messages.warning(request, 'Purchase order already exists for this requisition')
                    return redirect('po_list')
                
                # Get bid if provided
                bid = None
                if bid_id:
                    bid = get_object_or_404(Bid, id=bid_id)
                
                # Create PO
                po = PurchaseOrder.objects.create(
                    requisition=requisition,
                    supplier=supplier,
                    bid=bid,
                    delivery_date=delivery_date,
                    delivery_address=delivery_address,
                    payment_terms=payment_terms,
                    warranty_terms=warranty_terms,
                    special_instructions=special_instructions,
                    subtotal=0,
                    tax_amount=0,
                    total_amount=0,
                    status='DRAFT',
                    created_by=request.user
                )
                
                # Create PO items from requisition items
                req_items = requisition.items.all()
                subtotal = Decimal('0.00')
                
                for req_item in req_items:
                    # Get price from bid if available
                    unit_price = req_item.estimated_unit_price
                    
                    if bid:
                        bid_item = bid.items.filter(
                            requisition_item=req_item
                        ).first()
                        if bid_item:
                            unit_price = bid_item.quoted_unit_price
                    
                    PurchaseOrderItem.objects.create(
                        purchase_order=po,
                        requisition_item=req_item,
                        item_description=req_item.item_description,
                        specifications=req_item.specifications,
                        quantity=req_item.quantity,
                        unit_of_measure=req_item.unit_of_measure,
                        unit_price=unit_price,
                        total_price=req_item.quantity * unit_price
                    )
                    
                    subtotal += req_item.quantity * unit_price
                
                # Calculate tax (assuming 16% VAT)
                tax_rate = Decimal('0.16')
                tax_amount = subtotal * tax_rate
                total_amount = subtotal + tax_amount
                
                # Update PO totals
                po.subtotal = subtotal
                po.tax_amount = tax_amount
                po.total_amount = total_amount
                po.save()
                
                # Update budget commitment
                if requisition.budget:
                    budget = requisition.budget
                    budget.committed_amount += total_amount
                    budget.save()
                
                messages.success(request, f'Purchase Order {po.po_number} created successfully')
                return redirect('po_detail', po_id=po.id)
                
        except Exception as e:
            messages.error(request, f'Error creating purchase order: {str(e)}')
            return redirect('po_create')
    
    # GET request
    # Get approved requisitions without PO
    approved_requisitions = Requisition.objects.filter(
        status='APPROVED'
    ).exclude(
        id__in=PurchaseOrder.objects.values_list('requisition_id', flat=True)
    ).select_related('department', 'requested_by')
    
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    
    context = {
        'approved_requisitions': approved_requisitions,
        'suppliers': suppliers,
    }
    
    return render(request, 'procurement/po_create.html', context)


@login_required
def po_update(request, po_id):
    """Update purchase order (only in DRAFT status)"""
    
    po = get_object_or_404(PurchaseOrder, id=po_id)
    
    # Check if PO can be edited
    if po.status not in ['DRAFT', 'PENDING_APPROVAL']:
        messages.error(request, 'This purchase order cannot be edited')
        return redirect('po_detail', po_id=po.id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update PO fields
                po.delivery_date = request.POST.get('delivery_date')
                po.delivery_address = request.POST.get('delivery_address')
                po.payment_terms = request.POST.get('payment_terms')
                po.warranty_terms = request.POST.get('warranty_terms', '')
                po.special_instructions = request.POST.get('special_instructions', '')
                po.save()
                
                messages.success(request, 'Purchase order updated successfully')
                return redirect('po_detail', po_id=po.id)
                
        except Exception as e:
            messages.error(request, f'Error updating purchase order: {str(e)}')
    
    context = {
        'po': po,
        'po_items': po.items.all()
    }
    
    return render(request, 'procurement/po_update.html', context)


@login_required
def po_approve(request, po_id):
    """Approve purchase order"""
    
    po = get_object_or_404(PurchaseOrder, id=po_id)
    
    # Check permission (should be procurement/finance officer)
    if request.user.role not in ['PROCUREMENT', 'FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to approve purchase orders')
        return redirect('po_detail', po_id=po.id)
    
  
    if request.method == 'POST':
        po.status = 'APPROVED'
        po.approved_by = request.user
        po.approved_at = timezone.now()
        po.save()
        
        messages.success(request, f'Purchase Order {po.po_number} approved successfully')
        return redirect('po_detail', po_id=po.id)
    
    return render(request, 'procurement/po_approve.html', {'po': po})


@login_required
def po_send(request, po_id):
    """Send PO to supplier"""
    
    po = get_object_or_404(PurchaseOrder, id=po_id)
    
    if po.status != 'APPROVED':
        messages.error(request, 'Only approved POs can be sent to suppliers')
        return redirect('po_detail', po_id=po.id)
    
    if request.method == 'POST':
        po.status = 'SENT'
        po.sent_at = timezone.now()
        po.save()
        
        # TODO: Send email to supplier
        
        messages.success(request, f'Purchase Order {po.po_number} sent to supplier')
        return redirect('po_detail', po_id=po.id)
    
    return render(request, 'procurement/po_send.html', {'po': po})


@login_required
def po_cancel(request, po_id):
    """Cancel purchase order"""
    
    po = get_object_or_404(PurchaseOrder, id=po_id)
    
    # Check if PO can be cancelled
    if po.status in ['DELIVERED', 'CLOSED']:
        messages.error(request, 'This purchase order cannot be cancelled')
        return redirect('po_detail', po_id=po.id)
    
    if request.method == 'POST':
        cancellation_reason = request.POST.get('cancellation_reason')
        
        with transaction.atomic():
            po.status = 'CANCELLED'
            po.save()
            
            # Release budget commitment
            if po.requisition.budget:
                budget = po.requisition.budget
                budget.committed_amount -= po.total_amount
                budget.save()
            
            messages.success(request, f'Purchase Order {po.po_number} cancelled')
            return redirect('po_detail', po_id=po.id)
    
    context = {'po': po}
    return render(request, 'procurement/po_cancel.html', context)


@login_required
def po_amendment_create(request, po_id):
    """Create PO amendment"""
    
    po = get_object_or_404(PurchaseOrder, id=po_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                amendment_type = request.POST.get('amendment_type')
                description = request.POST.get('description')
                justification = request.POST.get('justification')
                old_value = request.POST.get('old_value')
                new_value = request.POST.get('new_value')
                amount_change = Decimal(request.POST.get('amount_change', '0'))
                
                # Generate amendment number
                last_amendment = po.amendments.order_by('-created_at').first()
                if last_amendment:
                    last_num = int(last_amendment.amendment_number.split('-')[-1])
                    amendment_number = f"{po.po_number}-AMD-{last_num + 1:03d}"
                else:
                    amendment_number = f"{po.po_number}-AMD-001"
                
                POAmendment.objects.create(
                    purchase_order=po,
                    amendment_number=amendment_number,
                    amendment_type=amendment_type,
                    description=description,
                    justification=justification,
                    old_value=old_value,
                    new_value=new_value,
                    amount_change=amount_change,
                    requested_by=request.user
                )
                
                messages.success(request, 'Amendment request created successfully')
                return redirect('po_detail', po_id=po.id)
                
        except Exception as e:
            messages.error(request, f'Error creating amendment: {str(e)}')
    
    context = {
        'po': po,
        'amendment_types': POAmendment.AMENDMENT_TYPES
    }
    
    return render(request, 'procurement/po_amendment_create.html', context)


# AJAX endpoints

@login_required
def get_requisition_details(request, req_id):
    """Get requisition details for PO creation"""
    
    requisition = get_object_or_404(Requisition, id=req_id)
    
    items = []
    for item in requisition.items.all():
        items.append({
            'id': str(item.id),
            'description': item.item_description,
            'quantity': float(item.quantity),
            'unit': item.unit_of_measure,
            'price': float(item.estimated_unit_price),
            'total': float(item.estimated_total)
        })
    
    data = {
        'number': requisition.requisition_number,
        'department': requisition.department.name,
        'estimated_amount': float(requisition.estimated_amount),
        'items': items
    }
    
    return JsonResponse(data)


@login_required
def get_supplier_bids(request, req_id, supplier_id):
    """Get supplier bids for a requisition"""
    
    bids = Bid.objects.filter(
        tender__requisition_id=req_id,
        supplier_id=supplier_id,
        status='AWARDED'
    ).select_related('tender')
    
    bid_list = []
    for bid in bids:
        bid_list.append({
            'id': str(bid.id),
            'bid_number': bid.bid_number,
            'amount': float(bid.bid_amount),
            'tender': bid.tender.tender_number
        })
    
    return JsonResponse({'bids': bid_list})



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Avg
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json

from .models import (
    Budget, BudgetYear, BudgetCategory, BudgetReallocation,
    Invoice, InvoiceItem, Payment, PurchaseOrder, Requisition,
    Department, Supplier, User, GoodsReceivedNote
)


# ============================================================================
# FINANCE DASHBOARD
# ============================================================================

@login_required
def finance_dashboard(request):
    """Main finance dashboard with key metrics and charts"""
    
    today = timezone.now().date()
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    # Budget Overview
    total_allocated = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).aggregate(total=Sum('allocated_amount'))['total'] or 0
    
    total_committed = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).aggregate(total=Sum('committed_amount'))['total'] or 0
    
    total_spent = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).aggregate(total=Sum('actual_spent'))['total'] or 0
    
    available_budget = total_allocated - total_committed - total_spent
    utilization_rate = (total_spent / total_allocated * 100) if total_allocated > 0 else 0
    
    # Invoice Statistics
    pending_invoices = Invoice.objects.filter(
        status__in=['SUBMITTED', 'VERIFYING']
    ).count()
    
    approved_invoices_value = Invoice.objects.filter(
        status='APPROVED'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Payment Statistics
    pending_payments = Payment.objects.filter(
        status='PENDING'
    ).count()
    
    payments_this_month = Payment.objects.filter(
        payment_date__month=today.month,
        payment_date__year=today.year,
        status='COMPLETED'
    ).aggregate(total=Sum('payment_amount'))['total'] or 0
    
    # Budget Utilization by Department (Top 10)
    dept_utilization = Budget.objects.filter(
        budget_year=current_year
    ).values(
        'department__name'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        utilization=Sum('actual_spent') / Sum('allocated_amount') * 100
    ).order_by('-spent')[:10]
    
    dept_chart = {
        'labels': [item['department__name'] for item in dept_utilization],
        'allocated': [float(item['allocated']) for item in dept_utilization],
        'spent': [float(item['spent']) for item in dept_utilization]
    }
    
    # Budget by Category
    category_budget = Budget.objects.filter(
        budget_year=current_year
    ).values(
        'category__name'
    ).annotate(
        total=Sum('allocated_amount')
    ).order_by('-total')[:8]
    
    category_chart = {
        'labels': [item['category__name'] for item in category_budget],
        'values': [float(item['total']) for item in category_budget]
    }
    
    # Monthly Expenditure Trend (Last 6 months)
    months = []
    expenditure = []
    
    for i in range(5, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if i > 0:
            next_month = month_date.replace(day=28) + timedelta(days=4)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        else:
            month_end = today
            
        month_spend = Payment.objects.filter(
            payment_date__gte=month_start,
            payment_date__lte=month_end,
            status='COMPLETED'
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        months.append(month_start.strftime('%b %Y'))
        expenditure.append(float(month_spend))
    
    expenditure_chart = {
        'labels': months,
        'values': expenditure
    }
    
    # Invoice Status Distribution
    invoice_status = Invoice.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    invoice_chart = {
        'labels': [dict(Invoice.STATUS_CHOICES).get(item['status'], item['status']) 
                   for item in invoice_status],
        'values': [item['count'] for item in invoice_status]
    }
    
    # Recent Activities
    recent_invoices = Invoice.objects.select_related(
        'supplier', 'purchase_order'
    ).order_by('-created_at')[:5]
    
    recent_payments = Payment.objects.select_related(
        'invoice__supplier'
    ).order_by('-created_at')[:5]
    
    # Pending Approvals
    pending_invoice_approvals = Invoice.objects.filter(
        status='VERIFYING'
    ).select_related('supplier')[:5]
    
    context = {
        'current_year': current_year,
        'total_allocated': total_allocated,
        'total_committed': total_committed,
        'total_spent': total_spent,
        'available_budget': available_budget,
        'utilization_rate': round(utilization_rate, 1),
        'pending_invoices': pending_invoices,
        'approved_invoices_value': approved_invoices_value,
        'pending_payments': pending_payments,
        'payments_this_month': payments_this_month,
        'dept_chart': dept_chart,
        'category_chart': category_chart,
        'expenditure_chart': expenditure_chart,
        'invoice_chart': invoice_chart,
        'recent_invoices': recent_invoices,
        'recent_payments': recent_payments,
        'pending_invoice_approvals': pending_invoice_approvals,
    }
    
    return render(request, 'finance/dashboard.html', context)


# ============================================================================
# BUDGET MANAGEMENT
# ============================================================================

@login_required
def budget_list(request):
    """List all budgets with filtering"""
    
    budgets = Budget.objects.select_related(
        'budget_year', 'department', 'category', 'created_by'
    ).all()
    
    # Filters
    year_filter = request.GET.get('year')
    department_filter = request.GET.get('department')
    category_filter = request.GET.get('category')
    budget_type_filter = request.GET.get('budget_type')
    search = request.GET.get('search')
    
    if year_filter:
        budgets = budgets.filter(budget_year_id=year_filter)
    
    if department_filter:
        budgets = budgets.filter(department_id=department_filter)
    
    if category_filter:
        budgets = budgets.filter(category_id=category_filter)
    
    if budget_type_filter:
        budgets = budgets.filter(budget_type=budget_type_filter)
    
    if search:
        budgets = budgets.filter(
            Q(department__name__icontains=search) |
            Q(category__name__icontains=search)
        )
    
    budgets = budgets.order_by('-created_at')
    
    # Calculate totals
    total_allocated = budgets.aggregate(total=Sum('allocated_amount'))['total'] or 0
    total_spent = budgets.aggregate(total=Sum('actual_spent'))['total'] or 0
    total_available = sum(b.available_balance for b in budgets)
    
    # Get filter options
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    departments = Department.objects.filter(is_active=True).order_by('name')
    categories = BudgetCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'budgets': budgets,
        'budget_years': budget_years,
        'departments': departments,
        'categories': categories,
        'budget_types': Budget.BUDGET_TYPE,
        'total_allocated': total_allocated,
        'total_spent': total_spent,
        'total_available': total_available,
        'filters': {
            'year': year_filter,
            'department': department_filter,
            'category': category_filter,
            'budget_type': budget_type_filter,
            'search': search,
        }
    }
    
    return render(request, 'finance/budget_list.html', context)


@login_required
def budget_detail(request, budget_id):
    """Budget detail view with utilization tracking"""
    
    budget = get_object_or_404(
        Budget.objects.select_related(
            'budget_year', 'department', 'category', 'created_by'
        ),
        id=budget_id
    )
    
    # Get requisitions using this budget
    requisitions = budget.requisitions.select_related(
        'requested_by', 'department'
    ).order_by('-created_at')
    
    # Get reallocations
    reallocations_from = budget.reallocations_from.select_related(
        'to_budget__department', 'requested_by', 'approved_by'
    ).order_by('-created_at')
    
    reallocations_to = budget.reallocations_to.select_related(
        'from_budget__department', 'requested_by', 'approved_by'
    ).order_by('-created_at')
    
    # Calculate utilization percentage
    utilization = (budget.actual_spent / budget.allocated_amount * 100) if budget.allocated_amount > 0 else 0
    commitment = (budget.committed_amount / budget.allocated_amount * 100) if budget.allocated_amount > 0 else 0
    
    context = {
        'budget': budget,
        'requisitions': requisitions,
        'reallocations_from': reallocations_from,
        'reallocations_to': reallocations_to,
        'utilization': round(utilization, 1),
        'commitment': round(commitment, 1),
    }
    
    return render(request, 'finance/budget_detail.html', context)


@login_required
def budget_create(request):
    """Create new budget allocation"""
    
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to create budgets')
        return redirect('budget_list')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                budget_year_id = request.POST.get('budget_year')
                department_id = request.POST.get('department')
                category_id = request.POST.get('category')
                budget_type = request.POST.get('budget_type')
                allocated_amount = Decimal(request.POST.get('allocated_amount'))
                reference_number = request.POST.get('reference_number', '')
                description = request.POST.get('description', '')
                
                # Check for duplicate
                if Budget.objects.filter(
                    budget_year_id=budget_year_id,
                    department_id=department_id,
                    category_id=category_id,
                    reference_number=reference_number
                ).exists():
                    messages.error(request, 'Budget already exists for this combination')
                    return redirect('budget_create')
                
                Budget.objects.create(
                    budget_year_id=budget_year_id,
                    department_id=department_id,
                    category_id=category_id,
                    budget_type=budget_type,
                    allocated_amount=allocated_amount,
                    reference_number=reference_number,
                    description=description,
                    created_by=request.user
                )
                
                messages.success(request, 'Budget created successfully')
                return redirect('budget_list')
                
        except Exception as e:
            messages.error(request, f'Error creating budget: {str(e)}')
    
    # GET request
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    departments = Department.objects.filter(is_active=True).order_by('name')
    categories = BudgetCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'budget_years': budget_years,
        'departments': departments,
        'categories': categories,
        'budget_types': Budget.BUDGET_TYPE,
    }
    
    return render(request, 'finance/budget_create.html', context)


@login_required
def budget_reallocation_create(request, budget_id):
    """Create budget reallocation request"""
    
    from_budget = get_object_or_404(Budget, id=budget_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                to_budget_id = request.POST.get('to_budget')
                amount = Decimal(request.POST.get('amount'))
                justification = request.POST.get('justification')
                
                to_budget = get_object_or_404(Budget, id=to_budget_id)
                
                # Validate amount
                if amount > from_budget.available_balance:
                    messages.error(request, 'Insufficient available budget')
                    return redirect('budget_reallocation_create', budget_id=budget_id)
                
                BudgetReallocation.objects.create(
                    from_budget=from_budget,
                    to_budget=to_budget,
                    amount=amount,
                    justification=justification,
                    requested_by=request.user
                )
                
                messages.success(request, 'Reallocation request submitted successfully')
                return redirect('budget_detail', budget_id=budget_id)
                
        except Exception as e:
            messages.error(request, f'Error creating reallocation: {str(e)}')
    
    # GET request - get available budgets in same year
    available_budgets = Budget.objects.filter(
        budget_year=from_budget.budget_year,
        is_active=True
    ).exclude(id=budget_id).select_related('department', 'category')
    
    context = {
        'from_budget': from_budget,
        'available_budgets': available_budgets,
    }
    
    return render(request, 'finance/budget_reallocation_create.html', context)


# ============================================================================
# INVOICE MANAGEMENT
# ============================================================================

@login_required
def invoice_list(request):
    """List all invoices with filtering"""
    
    invoices = Invoice.objects.select_related(
        'supplier', 'purchase_order', 'verified_by', 'approved_by'
    ).all()
    
    # Filters
    status_filter = request.GET.get('status')
    supplier_filter = request.GET.get('supplier')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        invoices = invoices.filter(status=status_filter)
    
    if supplier_filter:
        invoices = invoices.filter(supplier_id=supplier_filter)
    
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)
    
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(supplier_invoice_number__icontains=search) |
            Q(supplier__name__icontains=search)
        )
    
    invoices = invoices.order_by('-created_at')
    
    # Statistics
    total_value = invoices.aggregate(total=Sum('total_amount'))['total'] or 0
    pending_value = invoices.filter(
        status__in=['SUBMITTED', 'VERIFYING']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Get filter options
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    
    context = {
        'invoices': invoices,
        'suppliers': suppliers,
        'status_choices': Invoice.STATUS_CHOICES,
        'total_value': total_value,
        'pending_value': pending_value,
        'filters': {
            'status': status_filter,
            'supplier': supplier_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    
    return render(request, 'finance/invoice_list.html', context)


@login_required
def invoice_detail(request, invoice_id):
    """Invoice detail view with 3-way matching"""
    
    invoice = get_object_or_404(
        Invoice.objects.select_related(
            'supplier', 'purchase_order', 'grn',
            'verified_by', 'approved_by', 'submitted_by'
        ),
        id=invoice_id
    )
    
    # Get invoice items
    invoice_items = invoice.items.select_related('po_item').all()
    
    # Get payments
    payments = invoice.payments.select_related(
        'processed_by', 'approved_by'
    ).order_by('-created_at')
    
    # Calculate matching status
    if invoice.purchase_order and invoice.grn:
        po_match = abs(invoice.total_amount - invoice.purchase_order.total_amount) < Decimal('0.01')
        grn_match = invoice.grn.status == 'ACCEPTED'
        three_way_match = po_match and grn_match
    else:
        po_match = False
        grn_match = False
        three_way_match = False
    
    # Payment status
    total_paid = payments.filter(
        status='COMPLETED'
    ).aggregate(total=Sum('payment_amount'))['total'] or 0
    
    payment_progress = (total_paid / invoice.total_amount * 100) if invoice.total_amount > 0 else 0
    
    context = {
        'invoice': invoice,
        'invoice_items': invoice_items,
        'payments': payments,
        'po_match': po_match,
        'grn_match': grn_match,
        'three_way_match': three_way_match,
        'total_paid': total_paid,
        'payment_progress': round(payment_progress, 1),
    }
    
    return render(request, 'finance/invoice_detail.html', context)


@login_required
def invoice_verify(request, invoice_id):
    """Verify invoice details"""
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.user.role not in ['FINANCE', 'PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to verify invoices')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if invoice.status != 'SUBMITTED':
        messages.error(request, 'This invoice cannot be verified')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        if action == 'approve':
            invoice.status = 'MATCHED'
            invoice.verified_by = request.user
            invoice.verified_at = timezone.now()
            invoice.matching_notes = notes
            invoice.save()
            
            messages.success(request, f'Invoice {invoice.invoice_number} verified successfully')
        elif action == 'dispute':
            invoice.status = 'DISPUTED'
            invoice.dispute_reason = notes
            invoice.save()
            
            messages.warning(request, f'Invoice {invoice.invoice_number} marked as disputed')
        
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    context = {'invoice': invoice}
    return render(request, 'finance/invoice_verify.html', context)


@login_required
def invoice_approve(request, invoice_id):
    """Approve invoice for payment"""
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to approve invoices')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if invoice.status != 'MATCHED':
        messages.error(request, 'Invoice must be verified and matched before approval')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        invoice.status = 'APPROVED'
        invoice.approved_by = request.user
        invoice.approved_at = timezone.now()
        invoice.save()
        
        messages.success(request, f'Invoice {invoice.invoice_number} approved for payment')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    context = {'invoice': invoice}
    return render(request, 'finance/invoice_approve.html', context)


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q
from decimal import Decimal
from django.utils import timezone

# ============================================================================
# INVOICE CREATION
# ============================================================================

from decimal import Decimal
from django.db import transaction

@login_required
def invoice_create(request, grn_id):
    """Create invoice from GRN"""
    grn = get_object_or_404(
        GoodsReceivedNote.objects.select_related(
            'purchase_order__supplier',
            'store'
        ).prefetch_related('items__po_item'),
        id=grn_id
    )
    
    # Check permissions
    if request.user.role not in ['FINANCE', 'PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to create invoices')
        return redirect('grn_detail', grn_id=grn_id)
    
    # Check if GRN is accepted
    if grn.status not in ['ACCEPTED', 'PARTIAL']:
        messages.error(request, 'Can only create invoices for accepted GRNs')
        return redirect('grn_detail', grn_id=grn_id)
    
    # Check if invoice already exists
    if grn.invoices.exists():
        messages.warning(request, 'Invoice already exists for this GRN')
        return redirect('invoice_detail', invoice_id=grn.invoices.first().id)
    
    # Calculate subtotal and count accepted items
    subtotal = Decimal('0')
    accepted_items = grn.items.filter(item_status='ACCEPTED')
    
    for grn_item in accepted_items:
        item_total = grn_item.quantity_accepted * grn_item.po_item.unit_price
        subtotal += item_total
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                supplier_invoice_number = request.POST.get('supplier_invoice_number', '').strip()
                
                if not supplier_invoice_number:
                    messages.error(request, 'Supplier invoice number is required')
                    return redirect('invoice_create', grn_id=grn_id)
                
                invoice_date = request.POST.get('invoice_date')
                due_date = request.POST.get('due_date')
                tax_rate = Decimal(request.POST.get('tax_rate', '0'))
                other_charges = Decimal(request.POST.get('other_charges', '0'))
                notes = request.POST.get('notes', '')
                
                # Calculate tax and total
                tax_amount = (subtotal * tax_rate) / Decimal('100')
                total_amount = subtotal + tax_amount + other_charges
                
                # Create invoice (invoice_number will be auto-generated)
                invoice = Invoice.objects.create(
                    supplier_invoice_number=supplier_invoice_number,
                    purchase_order=grn.purchase_order,
                    grn=grn,
                    supplier=grn.purchase_order.supplier,
                    invoice_date=invoice_date,
                    due_date=due_date,
                    subtotal=subtotal,
                    tax_amount=tax_amount,
                    other_charges=other_charges,
                    total_amount=total_amount,
                    balance_due=total_amount,
                    status='DRAFT',
                    notes=notes,
                    submitted_by=request.user
                )
                
                # Create invoice items from accepted GRN items
                for grn_item in accepted_items:
                    item_subtotal = grn_item.quantity_accepted * grn_item.po_item.unit_price
                    item_tax = (item_subtotal * tax_rate) / Decimal('100')
                    
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        po_item=grn_item.po_item,
                        description=grn_item.po_item.item_description,
                        quantity=grn_item.quantity_accepted,
                        unit_price=grn_item.po_item.unit_price,
                        total_price=item_subtotal,
                        tax_rate=tax_rate,
                        tax_amount=item_tax
                    )
                
                messages.success(request, f'Invoice {invoice.invoice_number} created successfully!')
                return redirect('invoice_detail', invoice_id=invoice.id)
                
        except Exception as e:
            messages.error(request, f'Error creating invoice: {str(e)}')
            import traceback
            print(traceback.format_exc())  # Log full error for debugging
    
    context = {
        'grn': grn,
        'subtotal': subtotal,
        'po': grn.purchase_order,
        'today': timezone.now().date(),
        'accepted_items_count': accepted_items.count(),  # FIXED: Added this
    }
    
    return render(request, 'finance/invoice_create.html', context)


@login_required
def invoice_submit(request, invoice_id):
    """Submit invoice for verification"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.user.role not in ['FINANCE', 'PROCUREMENT', 'ADMIN']:
        messages.error(request, 'You do not have permission to submit invoices')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if invoice.status != 'DRAFT':
        messages.error(request, 'Invoice has already been submitted')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        invoice.status = 'SUBMITTED'
        invoice.submitted_by = request.user
        invoice.save()
        
        messages.success(request, f'Invoice {invoice.invoice_number} submitted for verification')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    context = {'invoice': invoice}
    return render(request, 'finance/invoice_submit.html', context)


# ============================================================================
# PAYMENT PROCESSING
# ============================================================================

@login_required
def invoice_pay(request, invoice_id):
    """Create payment for invoice"""
    invoice = get_object_or_404(
        Invoice.objects.select_related('supplier', 'purchase_order'),
        id=invoice_id
    )
    
    # Check permissions
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to process payments')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    # Check invoice status
    if invoice.status != 'APPROVED':
        messages.error(request, 'Invoice must be approved before payment')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    # Calculate remaining balance
    total_paid = invoice.payments.filter(
        status__in=['COMPLETED', 'PROCESSING']
    ).aggregate(total=Sum('payment_amount'))['total'] or Decimal('0')
    
    remaining_balance = invoice.total_amount - total_paid
    
    if remaining_balance <= 0:
        messages.warning(request, 'This invoice has been fully paid')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                payment_date = request.POST.get('payment_date')
                payment_amount = Decimal(request.POST.get('payment_amount'))
                payment_method = request.POST.get('payment_method')
                payment_reference = request.POST.get('payment_reference')
                bank_name = request.POST.get('bank_name', '')
                cheque_number = request.POST.get('cheque_number', '')
                notes = request.POST.get('notes', '')
                
                # Validate amount
                if payment_amount <= 0:
                    messages.error(request, 'Payment amount must be greater than zero')
                    return redirect('invoice_pay', invoice_id=invoice_id)
                
                if payment_amount > remaining_balance:
                    messages.error(request, f'Payment amount (KES {payment_amount:,.2f}) exceeds remaining balance (KES {remaining_balance:,.2f})')
                    return redirect('invoice_pay', invoice_id=invoice_id)
                
                # Create payment
                payment = Payment.objects.create(
                    invoice=invoice,
                    payment_date=payment_date,
                    payment_amount=payment_amount,
                    payment_method=payment_method,
                    payment_reference=payment_reference,
                    bank_name=bank_name,
                    cheque_number=cheque_number,
                    notes=notes,
                    processed_by=request.user,
                    status='PENDING'
                )
                
                messages.success(
                    request, 
                    f'Payment {payment.payment_number} created successfully. Please complete the payment process.'
                )
                return redirect('payment_detail', payment_id=payment.id)
                
        except Exception as e:
            messages.error(request, f'Error creating payment: {str(e)}')
    
    context = {
        'invoice': invoice,
        'remaining_balance': remaining_balance,
        'total_paid': total_paid,
        'payment_methods': Payment.PAYMENT_METHODS,
        'today': timezone.now().date(),
    }
    
    return render(request, 'finance/invoice_pay.html', context)


@login_required
def payment_detail(request, payment_id):
    """View payment details"""
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice__supplier',
            'invoice__purchase_order',
            'processed_by',
            'approved_by'
        ),
        id=payment_id
    )
    
    context = {
        'payment': payment,
        'can_complete': request.user.role in ['FINANCE', 'ADMIN'] and payment.status == 'PENDING',
        'can_cancel': request.user.role in ['FINANCE', 'ADMIN'] and payment.status in ['PENDING', 'PROCESSING'],
    }
    
    return render(request, 'finance/payment_detail.html', context)


@login_required
def payment_complete(request, payment_id):
    """Complete/approve payment"""
    payment = get_object_or_404(Payment, id=payment_id)
    
    # Check permissions
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to complete payments')
        return redirect('payment_detail', payment_id=payment_id)
    
    # Check payment status
    if payment.status not in ['PENDING', 'PROCESSING']:
        messages.error(request, 'Payment has already been processed')
        return redirect('payment_detail', payment_id=payment_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update payment status
                payment.status = 'COMPLETED'
                payment.approved_by = request.user
                payment.save()
                
                # Update invoice payment status
                payment.invoice.update_payment_status()
                
                messages.success(
                    request, 
                    f'Payment {payment.payment_number} completed successfully'
                )
                
                # Check if invoice is fully paid
                if payment.invoice.status == 'PAID':
                    messages.success(
                        request,
                        f'Invoice {payment.invoice.invoice_number} has been fully paid!'
                    )
                
                return redirect('payment_detail', payment_id=payment_id)
                
        except Exception as e:
            messages.error(request, f'Error completing payment: {str(e)}')
    
    context = {'payment': payment}
    return render(request, 'finance/payment_complete.html', context)


@login_required
def payment_cancel(request, payment_id):
    """Cancel payment"""
    payment = get_object_or_404(Payment, id=payment_id)
    
    # Check permissions
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to cancel payments')
        return redirect('payment_detail', payment_id=payment_id)
    
    # Check payment status
    if payment.status not in ['PENDING', 'PROCESSING']:
        messages.error(request, 'Cannot cancel completed or failed payments')
        return redirect('payment_detail', payment_id=payment_id)
    
    if request.method == 'POST':
        payment.status = 'CANCELLED'
        payment.notes += f"\n\nCancelled by {request.user.get_full_name()} on {timezone.now()}"
        payment.save()
        
        messages.success(request, f'Payment {payment.payment_number} cancelled')
        return redirect('invoice_detail', invoice_id=payment.invoice.id)
    
    context = {'payment': payment}
    return render(request, 'finance/payment_cancel.html', context)

# ============================================================================
# PAYMENT MANAGEMENT
# ============================================================================

@login_required
def payment_list(request):
    """List all payments with filtering"""
    
    payments = Payment.objects.select_related(
        'invoice__supplier', 'processed_by', 'approved_by'
    ).all()
    
    # Filters
    status_filter = request.GET.get('status')
    method_filter = request.GET.get('method')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    if method_filter:
        payments = payments.filter(payment_method=method_filter)
    
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
    
    if search:
        payments = payments.filter(
            Q(payment_number__icontains=search) |
            Q(payment_reference__icontains=search) |
            Q(invoice__supplier__name__icontains=search)
        )
    
    payments = payments.order_by('-created_at')
    
    # Statistics
    total_paid = payments.filter(
        status='COMPLETED'
    ).aggregate(total=Sum('payment_amount'))['total'] or 0
    
    pending_amount = payments.filter(
        status='PENDING'
    ).aggregate(total=Sum('payment_amount'))['total'] or 0
    
    context = {
        'payments': payments,
        'status_choices': Payment.PAYMENT_STATUS,
        'method_choices': Payment.PAYMENT_METHODS,
        'total_paid': total_paid,
        'pending_amount': pending_amount,
        'filters': {
            'status': status_filter,
            'method': method_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
        }
    }
    
    return render(request, 'finance/payment_list.html', context)


@login_required
def payment_create(request, invoice_id):
    """Create payment for invoice"""
    
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to create payments')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if invoice.status != 'APPROVED':
        messages.error(request, 'Invoice must be approved before payment')
        return redirect('invoice_detail', invoice_id=invoice_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                payment_date = request.POST.get('payment_date')
                payment_amount = Decimal(request.POST.get('payment_amount'))
                payment_method = request.POST.get('payment_method')
                payment_reference = request.POST.get('payment_reference')
                bank_name = request.POST.get('bank_name', '')
                cheque_number = request.POST.get('cheque_number', '')
                notes = request.POST.get('notes', '')
                
                # Validate amount
                total_paid = invoice.payments.filter(
                    status__in=['COMPLETED', 'PROCESSING']
                ).aggregate(total=Sum('payment_amount'))['total'] or 0
                
                if total_paid + payment_amount > invoice.total_amount:
                    messages.error(request, 'Payment amount exceeds invoice balance')
                    return redirect('payment_create', invoice_id=invoice_id)
                
                payment = Payment.objects.create(
                    invoice=invoice,
                    payment_date=payment_date,
                    payment_amount=payment_amount,
                    payment_method=payment_method,
                    payment_reference=payment_reference,
                    bank_name=bank_name,
                    cheque_number=cheque_number,
                    notes=notes,
                    processed_by=request.user,
                    status='PENDING'
                )
                
                messages.success(request, f'Payment {payment.payment_number} created successfully')
                return redirect('invoice_detail', invoice_id=invoice_id)
                
        except Exception as e:
            messages.error(request, f'Error creating payment: {str(e)}')
    
    # Calculate remaining balance
    total_paid = invoice.payments.filter(
        status='COMPLETED'
    ).aggregate(total=Sum('payment_amount'))['total'] or 0
    
    remaining = invoice.total_amount - total_paid
    
    context = {
        'invoice': invoice,
        'remaining': remaining,
        'payment_methods': Payment.PAYMENT_METHODS,
    }
    
    return render(request, 'finance/payment_create.html', context)


@login_required
def payment_process(request, payment_id):
    """Process/complete payment"""
    
    payment = get_object_or_404(Payment, id=payment_id)
    
    if request.user.role not in ['FINANCE', 'ADMIN']:
        messages.error(request, 'You do not have permission to process payments')
        return redirect('payment_list')
    
    if payment.status != 'PENDING':
        messages.error(request, 'Payment has already been processed')
        return redirect('payment_list')
    
    if request.method == 'POST':
        payment.status = 'COMPLETED'
        payment.approved_by = request.user
        payment.save()
        
        # Update invoice status if fully paid
        total_paid = payment.invoice.payments.filter(
            status='COMPLETED'
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        if total_paid >= payment.invoice.total_amount:
            payment.invoice.status = 'PAID'
            payment.invoice.payment_date = timezone.now().date()
            payment.invoice.save()
        
        messages.success(request, f'Payment {payment.payment_number} processed successfully')
        return redirect('payment_list')
    
    context = {'payment': payment}
    return render(request, 'finance/payment_process.html', context)


from django.db.models import Sum, Count, Avg, Q, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth, TruncYear
from django.utils import timezone
from datetime import timedelta, datetime
import json
from decimal import Decimal

@login_required
def financial_reports(request):
    """Financial reports dashboard with dynamic data and audit analytics"""
    
    # Get filters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    budget_year_id = request.GET.get('budget_year')
    
    # Default to current year
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    if budget_year_id:
        selected_year = BudgetYear.objects.filter(id=budget_year_id).first()
    else:
        selected_year = current_year
    
    # Set date range
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = selected_year.start_date if selected_year else timezone.now().date().replace(month=1, day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = selected_year.end_date if selected_year else timezone.now().date()
    
    # ============================================================================
    # 1. BUDGET UTILIZATION DATA
    # ============================================================================
    
    budget_summary = Budget.objects.filter(
        budget_year=selected_year
    ).aggregate(
        total_allocated=Sum('allocated_amount'),
        total_committed=Sum('committed_amount'),
        total_spent=Sum('actual_spent'),
        count=Count('id')
    )
    
    budget_utilization = []
    if budget_summary['total_allocated']:
        utilization_rate = (budget_summary['total_spent'] / budget_summary['total_allocated']) * 100
    else:
        utilization_rate = 0
    
    # Budget by department
    dept_budgets = Budget.objects.filter(
        budget_year=selected_year
    ).values(
        'department__name',
        'department__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent'),
        available=Sum('allocated_amount') - Sum('committed_amount') - Sum('actual_spent')
    ).order_by('-allocated')[:10]
    
    # Budget by category
    category_budgets = Budget.objects.filter(
        budget_year=selected_year
    ).values(
        'category__name',
        'category__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        utilization=ExpressionWrapper(
            Sum('actual_spent') * 100 / Sum('allocated_amount'),
            output_field=DecimalField()
        )
    ).order_by('-allocated')[:8]
    
    # ============================================================================
    # 2. EXPENDITURE ANALYSIS DATA
    # ============================================================================
    
    # Monthly expenditure trend
    monthly_expenditure = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        amount=Sum('payment_amount'),
        count=Count('id')
    ).order_by('month')
    
    # Expenditure by department
    dept_expenditure = Invoice.objects.filter(
        created_at__gte=start_date,
        created_at__lte=end_date,
        status='PAID'
    ).values(
        'purchase_order__requisition__department__name'
    ).annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    # Expenditure by category
    category_expenditure = InvoiceItem.objects.filter(
        invoice__created_at__gte=start_date,
        invoice__created_at__lte=end_date,
        invoice__status='PAID'
    ).values(
        'po_item__requisition_item__item__category__name'
    ).annotate(
        total=Sum('total_price'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    # ============================================================================
    # 3. SUPPLIER PAYMENTS DATA
    # ============================================================================
    
    # Supplier payment summary
    supplier_payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).values(
        'invoice__supplier__name',
        'invoice__supplier__supplier_number'
    ).annotate(
        total=Sum('payment_amount'),
        count=Count('id'),
        avg_payment=Avg('payment_amount')
    ).order_by('-total')[:15]
    
    # Payment method distribution
    payment_methods = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).values('payment_method').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Supplier performance
    supplier_performance = SupplierPerformance.objects.filter(
        reviewed_at__gte=start_date,
        reviewed_at__lte=end_date
    ).values(
        'supplier__name'
    ).annotate(
        avg_rating=Avg('overall_rating'),
        count=Count('id')
    ).order_by('-avg_rating')[:10]
    
    # ============================================================================
    # 4. INVOICE AGING ANALYSIS
    # ============================================================================
    
    # Invoice aging buckets
    today = timezone.now().date()
    
    aging_data = []
    aging_buckets = [
        ('Current', 0, 30),
        ('31-60 days', 31, 60),
        ('61-90 days', 61, 90),
        ('Over 90 days', 91, 365),
    ]
    
    for bucket_name, min_days, max_days in aging_buckets:
        min_date = today - timedelta(days=max_days)
        max_date = today - timedelta(days=min_days) if min_days > 0 else today
        
        invoices = Invoice.objects.filter(
            created_at__range=[min_date, max_date],
            status__in=['SUBMITTED', 'VERIFYING', 'MATCHED', 'APPROVED']
        ).aggregate(
            total=Sum('total_amount'),
            count=Count('id'),
            avg_age=Avg(today - F('created_at'))
        )
        
        aging_data.append({
            'bucket': bucket_name,
            'total': invoices['total'] or 0,
            'count': invoices['count'] or 0,
            'avg_days': invoices['avg_age'].days if invoices['avg_age'] else 0
        })
    
    # Overdue invoices
    overdue_invoices = Invoice.objects.filter(
        due_date__lt=today,
        status__in=['SUBMITTED', 'VERIFYING', 'MATCHED', 'APPROVED']
    ).aggregate(
        total=Sum('total_amount'),
        count=Count('id')
    )
    
    # ============================================================================
    # 5. CASH FLOW ANALYSIS
    # ============================================================================
    
    # Monthly cash flow
    monthly_cashflow = []
    for i in range(11, -1, -1):
        month_date = today - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        if i > 0:
            next_month = month_date.replace(day=28) + timedelta(days=4)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        else:
            month_end = today
        
        # Cash in (allocations)
        cash_in = BudgetReallocation.objects.filter(
            status='APPROVED',
            approved_at__gte=month_start,
            approved_at__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Cash out (payments)
        cash_out = Payment.objects.filter(
            payment_date__gte=month_start,
            payment_date__lte=month_end,
            status='COMPLETED'
        ).aggregate(total=Sum('payment_amount'))['total'] or 0
        
        monthly_cashflow.append({
            'month': month_start.strftime('%b %Y'),
            'cash_in': cash_in,
            'cash_out': cash_out,
            'net_flow': cash_in - cash_out
        })
    
    # ============================================================================
    # 6. AUDIT & COMPLIANCE DATA
    # ============================================================================
    
    # Audit trail summary
    audit_summary = AuditLog.objects.filter(
        timestamp__gte=start_date,
        timestamp__lte=end_date
    ).values('action').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # High-value transactions
    high_value_payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED',
        payment_amount__gte=1000000  # Payments over 1M
    ).select_related(
        'invoice', 'invoice__supplier'
    ).order_by('-payment_amount')[:10]
    
    # Compliance checks
    compliance_data = {
        'three_way_matched': Invoice.objects.filter(
            is_three_way_matched=True,
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count(),
        'unmatched_invoices': Invoice.objects.filter(
            is_three_way_matched=False,
            status='PAID',
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count(),
        'suppliers_with_expired_docs': Supplier.objects.filter(
            Q(tax_compliance_expiry__lt=today) | 
            Q(registration_expiry__lt=today),
            status='APPROVED'
        ).count(),
    }
    
    # ============================================================================
    # 7. PROCUREMENT PERFORMANCE
    # ============================================================================
    
    # Requisition to PO timeline
    timeline_data = PurchaseOrder.objects.filter(
        created_at__gte=start_date,
        created_at__lte=end_date
    ).annotate(
        timeline=ExpressionWrapper(
            F('created_at') - F('requisition__created_at'),
            output_field=DurationField()
        )
    ).aggregate(
        avg_days=Avg('timeline'),
        min_days=Min('timeline'),
        max_days=Max('timeline')
    )
    
    # Tender success rate
    tender_stats = Tender.objects.filter(
        created_at__gte=start_date,
        created_at__lte=end_date
    ).aggregate(
        total=Count('id'),
        awarded=Count('id', filter=Q(status='AWARDED')),
        cancelled=Count('id', filter=Q(status='CANCELLED'))
    )
    
    if tender_stats['total']:
        tender_success_rate = (tender_stats['awarded'] / tender_stats['total']) * 100
    else:
        tender_success_rate = 0
    
    # ============================================================================
    # 8. REPORT TYPES WITH DYNAMIC DATA
    # ============================================================================
    
    # Get all budget years for filter
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    
    reports = {
        'budget_utilization': {
            'title': 'Budget Utilization Report',
            'description': f'Budget analysis for {selected_year.name if selected_year else "current year"}',
            'icon': 'chart-pie',
            'data_available': budget_summary['count'] > 0,
            'summary': {
                'allocated': budget_summary['total_allocated'] or 0,
                'spent': budget_summary['total_spent'] or 0,
                'utilization': utilization_rate
            }
        },
        'expenditure_analysis': {
            'title': 'Expenditure Analysis',
            'description': f'Expenditure breakdown from {start_date} to {end_date}',
            'icon': 'graph-up',
            'data_available': monthly_expenditure.count() > 0,
            'summary': {
                'total': Payment.objects.filter(
                    payment_date__gte=start_date,
                    payment_date__lte=end_date,
                    status='COMPLETED'
                ).aggregate(total=Sum('payment_amount'))['total'] or 0,
                'count': Payment.objects.filter(
                    payment_date__gte=start_date,
                    payment_date__lte=end_date,
                    status='COMPLETED'
                ).count()
            }
        },
        'supplier_payments': {
            'title': 'Supplier Payment Report',
            'description': f'Supplier payments from {start_date} to {end_date}',
            'icon': 'people',
            'data_available': supplier_payments.count() > 0,
            'summary': {
                'total_suppliers': supplier_payments.count(),
                'total_paid': sum(p['total'] for p in supplier_payments),
                'avg_per_supplier': sum(p['total'] for p in supplier_payments) / supplier_payments.count() if supplier_payments.count() > 0 else 0
            }
        },
        'invoice_aging': {
            'title': 'Invoice Aging Report',
            'description': 'Invoice status and aging analysis',
            'icon': 'clock-history',
            'data_available': aging_data and any(d['count'] > 0 for d in aging_data),
            'summary': {
                'overdue_total': overdue_invoices['total'] or 0,
                'overdue_count': overdue_invoices['count'] or 0,
                'current_count': aging_data[0]['count'] if aging_data else 0
            }
        },
        'cashflow': {
            'title': 'Cash Flow Report',
            'description': f'Monthly cash flow from {start_date} to {end_date}',
            'icon': 'currency-exchange',
            'data_available': len(monthly_cashflow) > 0,
            'summary': {
                'total_in': sum(c['cash_in'] for c in monthly_cashflow),
                'total_out': sum(c['cash_out'] for c in monthly_cashflow),
                'net_flow': sum(c['net_flow'] for c in monthly_cashflow)
            }
        },
        'audit_trail': {
            'title': 'Audit Trail Report',
            'description': f'System audit from {start_date} to {end_date}',
            'icon': 'shield-check',
            'data_available': audit_summary.count() > 0,
            'summary': {
                'total_actions': sum(a['count'] for a in audit_summary),
                'top_action': audit_summary[0]['action'] if audit_summary else 'None',
                'users_involved': AuditLog.objects.filter(
                    timestamp__gte=start_date,
                    timestamp__lte=end_date
                ).values('user').distinct().count()
            }
        },
        'compliance': {
            'title': 'Compliance Report',
            'description': 'Procurement compliance analysis',
            'icon': 'file-check',
            'data_available': True,
            'summary': {
                'matched_invoices': compliance_data['three_way_matched'],
                'unmatched_invoices': compliance_data['unmatched_invoices'],
                'suppliers_expired': compliance_data['suppliers_with_expired_docs']
            }
        },
        'performance': {
            'title': 'Procurement Performance',
            'description': 'System performance metrics',
            'icon': 'speedometer',
            'data_available': True,
            'summary': {
                'avg_timeline': timeline_data['avg_days'].days if timeline_data['avg_days'] else 0,
                'tender_success': tender_success_rate,
                'high_value_count': high_value_payments.count()
            }
        }
    }
    
    # Prepare chart data for JavaScript
    context = {
        'current_year': current_year,
        'selected_year': selected_year,
        'budget_years': budget_years,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
        
        # Report data
        'reports': reports,
        
        # Chart data for quick preview
        'chart_data': {
            # Budget Utilization Charts
            'budget_utilization': {
                'labels': [d['department__name'] for d in dept_budgets],
                'allocated': [float(d['allocated'] or 0) for d in dept_budgets],
                'spent': [float(d['spent'] or 0) for d in dept_budgets],
            },
            'category_budget': {
                'labels': [c['category__name'] for c in category_budgets],
                'allocated': [float(c['allocated'] or 0) for c in category_budgets],
                'utilization': [float(c['utilization'] or 0) for c in category_budgets],
            },
            
            # Expenditure Charts
            'monthly_expenditure': {
                'labels': [m['month'].strftime('%b %Y') for m in monthly_expenditure],
                'amounts': [float(m['amount'] or 0) for m in monthly_expenditure],
                'counts': [m['count'] for m in monthly_expenditure],
            },
            'dept_expenditure': {
                'labels': [d['purchase_order__requisition__department__name'] or 'Unknown' for d in dept_expenditure],
                'amounts': [float(d['total'] or 0) for d in dept_expenditure],
            },
            
            # Supplier Charts
            'supplier_payments': {
                'labels': [s['invoice__supplier__name'] for s in supplier_payments[:8]],
                'amounts': [float(s['total'] or 0) for s in supplier_payments[:8]],
            },
            'payment_methods': {
                'labels': [p['payment_method'] for p in payment_methods],
                'amounts': [float(p['total'] or 0) for p in payment_methods],
            },
            'supplier_performance': {
                'labels': [s['supplier__name'] for s in supplier_performance],
                'ratings': [float(s['avg_rating'] or 0) for s in supplier_performance],
            },
            
            # Invoice Aging Charts
            'invoice_aging': {
                'labels': [a['bucket'] for a in aging_data],
                'amounts': [float(a['total'] or 0) for a in aging_data],
                'counts': [a['count'] for a in aging_data],
            },
            
            # Cash Flow Charts
            'cash_flow': {
                'labels': [c['month'] for c in monthly_cashflow],
                'cash_in': [float(c['cash_in'] or 0) for c in monthly_cashflow],
                'cash_out': [float(c['cash_out'] or 0) for c in monthly_cashflow],
                'net_flow': [float(c['net_flow'] or 0) for c in monthly_cashflow],
            },
            
            # Audit Charts
            'audit_actions': {
                'labels': [a['action'] for a in audit_summary[:10]],
                'counts': [a['count'] for a in audit_summary[:10]],
            },
            
            # Compliance Charts
            'compliance_status': {
                'labels': ['3-Way Matched', 'Unmatched Paid', 'Suppliers Expired'],
                'values': [
                    compliance_data['three_way_matched'],
                    compliance_data['unmatched_invoices'],
                    compliance_data['suppliers_with_expired_docs']
                ],
            },
        },
        
        # Summary statistics for display
        'summary_stats': {
            'budget_allocated': budget_summary['total_allocated'] or 0,
            'budget_spent': budget_summary['total_spent'] or 0,
            'budget_utilization': round(utilization_rate, 2),
            'total_payments': Payment.objects.filter(
                payment_date__gte=start_date,
                payment_date__lte=end_date,
                status='COMPLETED'
            ).aggregate(total=Sum('payment_amount'))['total'] or 0,
            'total_invoices': Invoice.objects.filter(
                created_at__gte=start_date,
                created_at__lte=end_date
            ).count(),
            'total_suppliers': Supplier.objects.filter(status='APPROVED').count(),
            'overdue_invoices': overdue_invoices['count'] or 0,
            'overdue_amount': overdue_invoices['total'] or 0,
            'avg_supplier_rating': Supplier.objects.filter(
                status='APPROVED',
                rating__gt=0
            ).aggregate(avg=Avg('rating'))['avg'] or 0,
            'procurement_timeline': timeline_data['avg_days'].days if timeline_data['avg_days'] else 0,
        },
        
        # Table data for quick view
        'table_data': {
            'top_suppliers': list(supplier_payments[:5]),
            'high_value_payments': high_value_payments,
            'recent_audits': AuditLog.objects.filter(
                timestamp__gte=start_date,
                timestamp__lte=end_date
            ).select_related('user').order_by('-timestamp')[:5],
            'department_budgets': list(dept_budgets[:5]),
        }
    }
    
    return render(request, 'finance/reports.html', context)

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal

@login_required
def expenditure_report(request):
    """Detailed expenditure analysis with dynamic charts"""

    current_year = BudgetYear.objects.filter(is_active=True).first()

    # Date range filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().replace(month=1, day=1).date()

    if end_date:
        end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()

    # =========================
    # Department Expenditure
    # =========================
    dept_expenditure = Budget.objects.filter(
        budget_year=current_year
    ).values('department__name').annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        committed=Sum('committed_amount')
    ).order_by('-spent')[:10]

    # =========================
    # Category Expenditure
    # =========================
    category_expenditure = Budget.objects.filter(
        budget_year=current_year
    ).values('category__name').annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent')
    ).order_by('-spent')[:8]

    # =========================
    # Monthly Expenditure Trend
    # =========================
    monthly_payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('month')

    # =========================
    # Payment Methods
    # =========================
    payment_methods = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).values('payment_method').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('-total')

    # =========================
    # Top Suppliers
    # =========================
    top_suppliers = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).values('invoice__supplier__name').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('-total')[:10]

    # =========================
    # Budget Utilization
    # =========================
    budget_utilization = Budget.objects.filter(
        budget_year=current_year
    ).aggregate(
        total_allocated=Sum('allocated_amount'),
        total_spent=Sum('actual_spent'),
        total_committed=Sum('committed_amount')
    )

    utilization_rate = (
        (budget_utilization['total_spent'] / budget_utilization['total_allocated']) * 100
        if budget_utilization['total_allocated'] else 0
    )

    # =========================
    # Invoice Aging (FIXED)
    # =========================
    invoice_aging = Invoice.objects.filter(
        status__in=['SUBMITTED', 'VERIFYING', 'MATCHED', 'APPROVED']
    ).annotate(
        age=ExpressionWrapper(
            timezone.now().date() - F('created_at'),
            output_field=DurationField()
        )
    ).values('status').annotate(
        total_amount=Sum('total_amount'),
        count=Count('id'),
        avg_age=Avg('age')
    )

    # =========================
    # Payment Status Summary
    # =========================
    payment_status = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    ).values('status').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    )

    # =========================
    # Recent Payments
    # =========================
    recent_payments = Payment.objects.filter(
        status='COMPLETED'
    ).select_related(
        'invoice', 'invoice__supplier'
    ).order_by('-payment_date')[:10]

    # =========================
    # Context
    # =========================
    context = {
        'current_year': current_year,
        'start_date': start_date,
        'end_date': end_date,

        'monthly_chart': {
            'labels': [m['month'].strftime('%b %Y') for m in monthly_payments],
            'values': [float(m['total']) for m in monthly_payments],
            'counts': [m['count'] for m in monthly_payments],
        },

        'category_chart': {
            'labels': [c['category__name'] for c in category_expenditure],
            'values': [float(c['spent']) for c in category_expenditure],
            'allocated': [float(c['allocated']) for c in category_expenditure],
        },

        'dept_chart': {
            'labels': [d['department__name'] for d in dept_expenditure],
            'allocated': [float(d['allocated']) for d in dept_expenditure],
            'spent': [float(d['spent']) for d in dept_expenditure],
            'committed': [float(d['committed']) for d in dept_expenditure],
        },

        'payment_method_chart': {
            'labels': [p['payment_method'] for p in payment_methods],
            'values': [float(p['total']) for p in payment_methods],
            'counts': [p['count'] for p in payment_methods],
        },

        'supplier_chart': {
            'labels': [s['invoice__supplier__name'] for s in top_suppliers],
            'values': [float(s['total']) for s in top_suppliers],
            'counts': [s['count'] for s in top_suppliers],
        },

        'payment_status_chart': {
            'labels': [p['status'] for p in payment_status],
            'values': [float(p['total']) for p in payment_status],
            'counts': [p['count'] for p in payment_status],
        },

        'stats': {
            'total_allocated': budget_utilization['total_allocated'] or 0,
            'total_spent': budget_utilization['total_spent'] or 0,
            'total_committed': budget_utilization['total_committed'] or 0,
            'utilization_rate': round(utilization_rate, 2),
        },

        'recent_payments': recent_payments,
        'invoice_aging': invoice_aging,
        'category_expenditure': category_expenditure,
        'dept_expenditure': dept_expenditure,
    }

    return render(request, 'finance/expenditure_report.html', context)


"""
views.py - Add this view to your finance views
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q, F, ExpressionWrapper, DecimalField
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal

@login_required
def budget_utilization_report(request):
    """Budget Utilization Report with detailed breakdown"""
    
    # Get filters from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    budget_year_id = request.GET.get('budget_year')
    department_id = request.GET.get('department')
    
    # Default to current year
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    if budget_year_id:
        selected_year = BudgetYear.objects.filter(id=budget_year_id).first()
    else:
        selected_year = current_year
    
    # Set date range
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = selected_year.start_date if selected_year else timezone.now().date().replace(month=1, day=1)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = selected_year.end_date if selected_year else timezone.now().date()
    
    # Base query
    budget_query = Budget.objects.filter(budget_year=selected_year)
    
    # Apply department filter if selected
    if department_id:
        budget_query = budget_query.filter(department_id=department_id)
    
    # ============================================================================
    # 1. OVERALL BUDGET SUMMARY
    # ============================================================================
    overall_summary = budget_query.aggregate(
        total_allocated=Sum('allocated_amount'),
        total_committed=Sum('committed_amount'),
        total_spent=Sum('actual_spent'),
        count=Count('id')
    )
    
    # Calculate available balance and utilization
    total_allocated = overall_summary['total_allocated'] or 0
    total_committed = overall_summary['total_committed'] or 0
    total_spent = overall_summary['total_spent'] or 0
    total_used = total_committed + total_spent
    available_balance = total_allocated - total_used
    
    if total_allocated > 0:
        utilization_rate = (total_spent / total_allocated) * 100
        commitment_rate = (total_committed / total_allocated) * 100
        total_utilization = ((total_committed + total_spent) / total_allocated) * 100
    else:
        utilization_rate = 0
        commitment_rate = 0
        total_utilization = 0
    
    # ============================================================================
    # 2. BUDGET BY DEPARTMENT
    # ============================================================================
    dept_budgets = budget_query.values(
        'department__id',
        'department__name',
        'department__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent'),
        available=ExpressionWrapper(
            Sum('allocated_amount') - Sum('committed_amount') - Sum('actual_spent'),
            output_field=DecimalField()
        )
    ).order_by('-allocated')
    
    # Calculate utilization for each department
    dept_budgets_list = []
    for dept in dept_budgets:
        allocated = dept['allocated'] or 0
        spent = dept['spent'] or 0
        committed = dept['committed'] or 0
        available = dept['available'] or 0
        
        if allocated > 0:
            dept_utilization = (spent / allocated) * 100
            dept_commitment = (committed / allocated) * 100
        else:
            dept_utilization = 0
            dept_commitment = 0
        
        dept_budgets_list.append({
            'id': dept['department__id'],
            'name': dept['department__name'],
            'code': dept['department__code'],
            'allocated': allocated,
            'committed': committed,
            'spent': spent,
            'available': available,
            'utilization': dept_utilization,
            'commitment_rate': dept_commitment
        })
    
    # ============================================================================
    # 3. BUDGET BY CATEGORY
    # ============================================================================
    category_budgets = budget_query.values(
        'category__id',
        'category__name',
        'category__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent'),
        available=ExpressionWrapper(
            Sum('allocated_amount') - Sum('committed_amount') - Sum('actual_spent'),
            output_field=DecimalField()
        )
    ).order_by('-allocated')
    
    # Calculate utilization for each category
    category_budgets_list = []
    for cat in category_budgets:
        allocated = cat['allocated'] or 0
        spent = cat['spent'] or 0
        committed = cat['committed'] or 0
        available = cat['available'] or 0
        
        if allocated > 0:
            cat_utilization = (spent / allocated) * 100
            cat_commitment = (committed / allocated) * 100
        else:
            cat_utilization = 0
            cat_commitment = 0
        
        category_budgets_list.append({
            'id': cat['category__id'],
            'name': cat['category__name'],
            'code': cat['category__code'],
            'allocated': allocated,
            'committed': committed,
            'spent': spent,
            'available': available,
            'utilization': cat_utilization,
            'commitment_rate': cat_commitment
        })
    
    # ============================================================================
    # 4. BUDGET BY TYPE
    # ============================================================================
    budget_by_type = budget_query.values('budget_type').annotate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent'),
        count=Count('id')
    ).order_by('-allocated')
    
    budget_types_list = []
    for bt in budget_by_type:
        allocated = bt['allocated'] or 0
        spent = bt['spent'] or 0
        
        if allocated > 0:
            type_utilization = (spent / allocated) * 100
        else:
            type_utilization = 0
        
        budget_types_list.append({
            'type': bt['budget_type'],
            'type_display': dict(Budget.BUDGET_TYPE).get(bt['budget_type'], bt['budget_type']),
            'allocated': allocated,
            'committed': bt['committed'] or 0,
            'spent': spent,
            'count': bt['count'],
            'utilization': type_utilization
        })
    
    # ============================================================================
    # 5. OVER/UNDER UTILIZED BUDGETS
    # ============================================================================
    # Get individual budgets with utilization
    all_budgets = budget_query.select_related(
        'department', 'category', 'budget_year'
    ).annotate(
        available=ExpressionWrapper(
            F('allocated_amount') - F('committed_amount') - F('actual_spent'),
            output_field=DecimalField()
        )
    )
    
    over_utilized = []
    under_utilized = []
    critical_budgets = []
    
    for budget in all_budgets:
        if budget.allocated_amount > 0:
            util_rate = (budget.actual_spent / budget.allocated_amount) * 100
            
            budget_data = {
                'id': budget.id,
                'department': budget.department.name,
                'category': budget.category.name,
                'allocated': budget.allocated_amount,
                'spent': budget.actual_spent,
                'committed': budget.committed_amount,
                'available': budget.available_balance,
                'utilization': util_rate
            }
            
            # Over-utilized (>100%)
            if util_rate > 100:
                over_utilized.append(budget_data)
            
            # Critical (>90% but <=100%)
            elif util_rate > 90:
                critical_budgets.append(budget_data)
            
            # Under-utilized (<50%)
            elif util_rate < 50:
                under_utilized.append(budget_data)
    
    # Sort lists
    over_utilized.sort(key=lambda x: x['utilization'], reverse=True)
    under_utilized.sort(key=lambda x: x['utilization'])
    critical_budgets.sort(key=lambda x: x['utilization'], reverse=True)
    
    # ============================================================================
    # 6. MONTHLY TREND
    # ============================================================================
    # Get monthly spending data
    monthly_data = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        status='COMPLETED'
    ).extra(
        select={'month': "EXTRACT(month FROM payment_date)", 'year': "EXTRACT(year FROM payment_date)"}
    ).values('month', 'year').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('year', 'month')
    
    # ============================================================================
    # 7. BUDGET REALLOCATIONS
    # ============================================================================
    reallocations = BudgetReallocation.objects.filter(
        status='APPROVED',
        approved_at__gte=start_date,
        approved_at__lte=end_date
    ).select_related(
        'from_budget__department',
        'from_budget__category',
        'to_budget__department',
        'to_budget__category',
        'requested_by',
        'approved_by'
    ).order_by('-approved_at')[:10]
    
    # ============================================================================
    # 8. PREPARE CHART DATA
    # ============================================================================
    chart_data = {
        # Department Budget Chart
        'dept_budget': {
            'labels': [d['name'] for d in dept_budgets_list[:10]],
            'allocated': [float(d['allocated']) for d in dept_budgets_list[:10]],
            'spent': [float(d['spent']) for d in dept_budgets_list[:10]],
            'committed': [float(d['committed']) for d in dept_budgets_list[:10]],
            'available': [float(d['available']) for d in dept_budgets_list[:10]],
        },
        
        # Category Budget Chart
        'category_budget': {
            'labels': [c['name'] for c in category_budgets_list[:8]],
            'allocated': [float(c['allocated']) for c in category_budgets_list[:8]],
            'spent': [float(c['spent']) for c in category_budgets_list[:8]],
        },
        
        # Budget Type Chart
        'budget_type': {
            'labels': [bt['type_display'] for bt in budget_types_list],
            'values': [float(bt['allocated']) for bt in budget_types_list],
        },
        
        # Utilization Chart
        'utilization': {
            'labels': [d['name'] for d in dept_budgets_list[:10]],
            'values': [float(d['utilization']) for d in dept_budgets_list[:10]],
        },
    }
    
    # Get all departments and budget years for filters
    all_departments = Department.objects.filter(is_active=True).order_by('name')
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    
    context = {
        'current_year': current_year,
        'selected_year': selected_year,
        'budget_years': budget_years,
        'all_departments': all_departments,
        'selected_department': department_id,
        'start_date': start_date,
        'end_date': end_date,
        
        # Summary Stats
        'stats': {
            'total_allocated': total_allocated,
            'total_committed': total_committed,
            'total_spent': total_spent,
            'available_balance': available_balance,
            'utilization_rate': round(utilization_rate, 2),
            'commitment_rate': round(commitment_rate, 2),
            'total_utilization': round(total_utilization, 2),
            'budget_count': overall_summary['count'] or 0,
        },
        
        # Detailed Data
        'dept_budgets': dept_budgets_list,
        'category_budgets': category_budgets_list,
        'budget_types': budget_types_list,
        
        # Special Categories
        'over_utilized': over_utilized[:5],
        'under_utilized': under_utilized[:5],
        'critical_budgets': critical_budgets[:5],
        
        # Other Data
        'monthly_data': monthly_data,
        'reallocations': reallocations,
        
        # Chart Data
        'chart_data': chart_data,
    }
    
    return render(request, 'finance/budget_utilization_report.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.utils import timezone
from decimal import Decimal
from datetime import datetime, timedelta

from .models import (
    Store, GoodsReceivedNote, GRNItem, StockItem, StockMovement,
    StockIssue, StockIssueItem, Asset, PurchaseOrder, PurchaseOrderItem,
    Department, Item, User
)


# ============================================================================
# GOODS RECEIVED NOTES
# ============================================================================

@login_required
def grn_list(request):
    """List all Goods Received Notes with filters"""
    grns = GoodsReceivedNote.objects.select_related(
        'purchase_order', 'purchase_order__supplier', 'store',
        'received_by', 'inspected_by'
    ).all()
    
    # Filters
    filters = {
        'search': request.GET.get('search', ''),
        'status': request.GET.get('status', ''),
        'store': request.GET.get('store', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }
    
    # Apply search
    if filters['search']:
        grns = grns.filter(
            Q(grn_number__icontains=filters['search']) |
            Q(purchase_order__po_number__icontains=filters['search']) |
            Q(delivery_note_number__icontains=filters['search']) |
            Q(purchase_order__supplier__name__icontains=filters['search'])
        )
    
    # Apply status filter
    if filters['status']:
        grns = grns.filter(status=filters['status'])
    
    # Apply store filter
    if filters['store']:
        grns = grns.filter(store_id=filters['store'])
    
    # Apply date range filter
    if filters['date_from']:
        grns = grns.filter(received_date__gte=filters['date_from'])
    if filters['date_to']:
        grns = grns.filter(received_date__lte=filters['date_to'])
    
    # Get filter options
    stores = Store.objects.filter(is_active=True)
    status_choices = GoodsReceivedNote.STATUS_CHOICES
    
    # Calculate statistics
    total_received = grns.count()
    pending_inspection = grns.filter(status='INSPECTING').count()
    accepted = grns.filter(status='ACCEPTED').count()
    rejected = grns.filter(status='REJECTED').count()
    
    context = {
        'grns': grns.order_by('-created_at'),
        'filters': filters,
        'stores': stores,
        'status_choices': status_choices,
        'total_received': total_received,
        'pending_inspection': pending_inspection,
        'accepted': accepted,
        'rejected': rejected,
    }
    
    return render(request, 'inventory/grn_list.html', context)


@login_required
def grn_detail(request, grn_id):
    """View GRN details"""
    grn = get_object_or_404(
        GoodsReceivedNote.objects.select_related(
            'purchase_order', 'purchase_order__supplier', 'store',
            'received_by', 'inspected_by'
        ),
        id=grn_id
    )
    
    items = grn.items.select_related('po_item', 'po_item__requisition_item')
    
    context = {
        'grn': grn,
        'items': items,
    }
    
    return render(request, 'inventory/grn_detail.html', context)


@login_required
def grn_create(request):
    """Create new GRN"""
    if request.method == 'POST':
        po_id = request.POST.get('purchase_order')
        store_id = request.POST.get('store')
        delivery_note = request.POST.get('delivery_note_number')
        delivery_date = request.POST.get('delivery_date')
        
        po = get_object_or_404(PurchaseOrder, id=po_id)
        store = get_object_or_404(Store, id=store_id)
        
        # Create GRN
        grn = GoodsReceivedNote.objects.create(
            purchase_order=po,
            store=store,
            delivery_note_number=delivery_note,
            delivery_date=delivery_date,
            received_by=request.user,
            status='DRAFT'
        )
        
        # Create GRN items from PO items
        po_items = po.items.all()
        for po_item in po_items:
            qty_delivered = request.POST.get(f'qty_delivered_{po_item.id}', 0)
            if Decimal(qty_delivered) > 0:
                GRNItem.objects.create(
                    grn=grn,
                    po_item=po_item,
                    quantity_ordered=po_item.quantity,
                    quantity_delivered=qty_delivered,
                    quantity_accepted=0,
                    item_status='ACCEPTED'
                )
        
        messages.success(request, f'GRN {grn.grn_number} created successfully!')
        return redirect('grn_detail', grn_id=grn.id)
    
    # GET request - Preserve form state
    po_id = request.GET.get('po')
    selected_store_id = request.GET.get('store', '')
    delivery_note_number = request.GET.get('delivery_note', '')
    delivery_date = request.GET.get('delivery_date', '')
    
    po = None
    if po_id:
        po = get_object_or_404(
            PurchaseOrder.objects.select_related('supplier').prefetch_related('items'),
            id=po_id,
            status__in=['SENT', 'ACKNOWLEDGED']
        )
    
    # Collect quantities from query params
    quantities_delivered = {}
    for key, value in request.GET.items():
        if key.startswith('qty_delivered_'):
            item_id = key.replace('qty_delivered_', '')
            quantities_delivered[int(item_id)] = value
    
    stores = Store.objects.filter(is_active=True)
    pending_pos = PurchaseOrder.objects.filter(
        status__in=['SENT', 'ACKNOWLEDGED']
    ).select_related('supplier')
    
    # Get today's date for default
    from datetime import date
    today = date.today()
    
    context = {
        'po': po,
        'stores': stores,
        'pending_pos': pending_pos,
        'selected_po_id': po_id or '',
        'selected_store_id': selected_store_id,
        'delivery_note_number': delivery_note_number,
        'delivery_date': delivery_date,
        'quantities_delivered': quantities_delivered,
        'today': today,
    }
    
    return render(request, 'inventory/grn_create.html', context)


@login_required
def grn_inspect(request, grn_id):
    """Inspect and accept/reject GRN items"""
    grn = get_object_or_404(GoodsReceivedNote, id=grn_id)
    
    if request.method == 'POST':
        grn.status = 'INSPECTING'
        grn.inspected_by = request.user
        grn.inspection_date = timezone.now().date()
        
        all_accepted = True
        any_accepted = False
        
        for item in grn.items.all():
            qty_accepted = Decimal(request.POST.get(f'qty_accepted_{item.id}', 0))
            qty_rejected = Decimal(request.POST.get(f'qty_rejected_{item.id}', 0))
            item_status = request.POST.get(f'status_{item.id}', 'ACCEPTED')
            remarks = request.POST.get(f'remarks_{item.id}', '')
            
            item.quantity_accepted = qty_accepted
            item.quantity_rejected = qty_rejected
            item.item_status = item_status
            item.remarks = remarks
            item.save()
            
            if qty_rejected > 0 or item_status != 'ACCEPTED':
                all_accepted = False
            if qty_accepted > 0:
                any_accepted = True
            
            # Update stock if accepted
            if qty_accepted > 0:
                # Get the item from the catalog - handle case where it might be None
                catalog_item = item.po_item.requisition_item.item
                
                if catalog_item is None:
                    # If no catalog item exists, skip stock update or create a warning
                    messages.warning(
                        request, 
                        f'Item "{item.po_item.item_description}" is not linked to the master catalog. '
                        f'Stock was not updated. Please link this item to the catalog first.'
                    )
                    continue  # Skip to next item
                
                stock_item, created = StockItem.objects.get_or_create(
                    store=grn.store,
                    item=catalog_item,
                    defaults={'quantity_on_hand': 0, 'average_unit_cost': 0}
                )
                
                # Update stock quantity
                balance_before = stock_item.quantity_on_hand
                stock_item.quantity_on_hand += qty_accepted
                stock_item.last_restock_date = timezone.now().date()
                
                # Update average cost (weighted average)
                if stock_item.quantity_on_hand > 0:
                    total_cost = (balance_before * stock_item.average_unit_cost) + \
                                 (qty_accepted * item.po_item.unit_price)
                    stock_item.average_unit_cost = total_cost / stock_item.quantity_on_hand
                    stock_item.total_value = stock_item.quantity_on_hand * stock_item.average_unit_cost
                else:
                    # First receipt
                    stock_item.average_unit_cost = item.po_item.unit_price
                    stock_item.total_value = qty_accepted * item.po_item.unit_price
                
                stock_item.save()
                
                # Record stock movement
                StockMovement.objects.create(
                    stock_item=stock_item,
                    movement_type='RECEIPT',
                    reference_number=grn.grn_number,
                    reference_type='GRN',
                    quantity=qty_accepted,
                    unit_cost=item.po_item.unit_price,
                    balance_before=balance_before,
                    balance_after=stock_item.quantity_on_hand,
                    to_store=grn.store,
                    performed_by=request.user
                )
        
        # Set final GRN status
        if all_accepted:
            grn.status = 'ACCEPTED'
        elif any_accepted:
            grn.status = 'PARTIAL'
        else:
            grn.status = 'REJECTED'
            grn.rejection_reason = request.POST.get('rejection_reason', '')
        
        grn.general_condition = request.POST.get('general_condition', '')
        grn.save()
        
        messages.success(request, f'GRN {grn.grn_number} inspection completed!')
        return redirect('grn_detail', grn_id=grn.id)
    
    context = {
        'grn': grn,
        'items': grn.items.all(),
    }
    
    return render(request, 'inventory/grn_inspect.html', context)


# ============================================================================
# STOCK ITEMS
# ============================================================================

@login_required
def stock_list(request):
    """List all stock items with filters"""
    stocks = StockItem.objects.select_related('store', 'item', 'item__category').all()
    
    # Filters
    filters = {
        'search': request.GET.get('search', ''),
        'store': request.GET.get('store', ''),
        'category': request.GET.get('category', ''),
        'low_stock': request.GET.get('low_stock', ''),
    }
    
    # Apply search
    if filters['search']:
        stocks = stocks.filter(
            Q(item__name__icontains=filters['search']) |
            Q(item__code__icontains=filters['search'])
        )
    
    # Apply store filter
    if filters['store']:
        stocks = stocks.filter(store_id=filters['store'])
    
    # Apply category filter
    if filters['category']:
        stocks = stocks.filter(item__category_id=filters['category'])
    
    # Apply low stock filter
    if filters['low_stock']:
        stocks = stocks.filter(quantity_on_hand__lte=F('reorder_level'))
    
    # Get filter options
    stores = Store.objects.filter(is_active=True)
    from .models import ItemCategory
    categories = ItemCategory.objects.filter(is_active=True, category_type='GOODS')
    
    # Calculate statistics
    total_items = stocks.count()
    low_stock_count = stocks.filter(quantity_on_hand__lte=F('reorder_level')).count()
    total_value = stocks.aggregate(
        total=Coalesce(Sum('total_value'), Decimal('0'))
    )['total']
    
    context = {
        'stocks': stocks.order_by('item__name'),
        'filters': filters,
        'stores': stores,
        'categories': categories,
        'total_items': total_items,
        'low_stock_count': low_stock_count,
        'total_value': total_value,
    }
    
    return render(request, 'inventory/stock_list.html', context)


@login_required
def stock_detail(request, stock_id):
    """View stock item details and movements"""
    stock = get_object_or_404(
        StockItem.objects.select_related('store', 'item', 'item__category'),
        id=stock_id
    )
    
    movements = stock.movements.select_related(
        'from_store', 'to_store', 'performed_by'
    ).order_by('-movement_date')[:50]
    
    # Recent issues
    recent_issues = StockIssueItem.objects.filter(
        stock_item=stock
    ).select_related('stock_issue', 'stock_issue__department')[:10]
    
    context = {
        'stock': stock,
        'movements': movements,
        'recent_issues': recent_issues,
    }
    
    return render(request, 'inventory/stock_detail.html', context)


@login_required
def stock_adjustment(request, stock_id):
    """Adjust stock quantity"""
    stock = get_object_or_404(StockItem, id=stock_id)
    
    if request.method == 'POST':
        adjustment_type = request.POST.get('adjustment_type')
        quantity = Decimal(request.POST.get('quantity', 0))
        reason = request.POST.get('reason')
        
        if quantity <= 0:
            messages.error(request, 'Quantity must be greater than zero!')
            return redirect('stock_detail', stock_id=stock_id)
        
        balance_before = stock.quantity_on_hand
        
        if adjustment_type == 'ADD':
            stock.quantity_on_hand += quantity
        else:  # SUBTRACT
            if quantity > stock.quantity_on_hand:
                messages.error(request, 'Cannot subtract more than available stock!')
                return redirect('stock_detail', stock_id=stock_id)
            stock.quantity_on_hand -= quantity
        
        stock.total_value = stock.quantity_on_hand * stock.average_unit_cost
        stock.save()
        
        # Record movement
        StockMovement.objects.create(
            stock_item=stock,
            movement_type='ADJUSTMENT',
            reference_number=f'ADJ-{timezone.now().strftime("%Y%m%d%H%M%S")}',
            reference_type='ADJUSTMENT',
            quantity=quantity if adjustment_type == 'ADD' else -quantity,
            unit_cost=stock.average_unit_cost,
            balance_before=balance_before,
            balance_after=stock.quantity_on_hand,
            remarks=reason,
            performed_by=request.user
        )
        
        messages.success(request, 'Stock adjusted successfully!')
        return redirect('stock_detail', stock_id=stock_id)
    
    context = {
        'stock': stock,
    }
    
    return render(request, 'inventory/stock_adjustment.html', context)


# ============================================================================
# STOCK ISSUES
# ============================================================================

@login_required
def issue_list(request):
    """List all stock issues"""
    issues = StockIssue.objects.select_related(
        'store', 'department', 'requested_by', 'issued_by'
    ).all()
    
    # Filters
    filters = {
        'search': request.GET.get('search', ''),
        'status': request.GET.get('status', ''),
        'department': request.GET.get('department', ''),
        'store': request.GET.get('store', ''),
    }
    
    # Apply filters
    if filters['search']:
        issues = issues.filter(
            Q(issue_number__icontains=filters['search']) |
            Q(purpose__icontains=filters['search'])
        )
    
    if filters['status']:
        issues = issues.filter(status=filters['status'])
    
    if filters['department']:
        issues = issues.filter(department_id=filters['department'])
    
    if filters['store']:
        issues = issues.filter(store_id=filters['store'])
    
    # Get filter options
    departments = Department.objects.filter(is_active=True)
    stores = Store.objects.filter(is_active=True)
    status_choices = StockIssue.STATUS_CHOICES
    
    # Statistics
    total_issues = issues.count()
    pending = issues.filter(status='PENDING').count()
    issued = issues.filter(status='ISSUED').count()
    
    context = {
        'issues': issues.order_by('-created_at'),
        'filters': filters,
        'departments': departments,
        'stores': stores,
        'status_choices': status_choices,
        'total_issues': total_issues,
        'pending': pending,
        'issued': issued,
    }
    
    return render(request, 'inventory/issue_list.html', context)


@login_required
def issue_create(request):
    """Create new stock issue request"""
    if request.method == 'POST':
        store_id = request.POST.get('store')
        department_id = request.POST.get('department')
        purpose = request.POST.get('purpose')
        
        store = get_object_or_404(Store, id=store_id)
        department = get_object_or_404(Department, id=department_id)
        
        # Create issue
        issue = StockIssue.objects.create(
            store=store,
            department=department,
            requested_by=request.user,
            purpose=purpose,
            status='PENDING'
        )
        
        # Add items
        item_ids = request.POST.getlist('item_id[]')
        quantities = request.POST.getlist('quantity[]')
        
        for item_id, qty in zip(item_ids, quantities):
            if item_id and Decimal(qty) > 0:
                stock_item = get_object_or_404(StockItem, id=item_id)
                StockIssueItem.objects.create(
                    stock_issue=issue,
                    stock_item=stock_item,
                    quantity_requested=qty,
                    quantity_issued=0
                )
        
        messages.success(request, f'Stock issue {issue.issue_number} created successfully!')
        return redirect('issue_detail', issue_id=issue.id)
    
    # GET request
    stores = Store.objects.filter(is_active=True)
    departments = Department.objects.filter(is_active=True)
    
    context = {
        'stores': stores,
        'departments': departments,
    }
    
    return render(request, 'inventory/issue_create.html', context)


@login_required
def issue_detail(request, issue_id):
    """View stock issue details"""
    issue = get_object_or_404(
        StockIssue.objects.select_related(
            'store', 'department', 'requested_by', 'issued_by'
        ),
        id=issue_id
    )
    
    items = issue.items.select_related('stock_item', 'stock_item__item')
    
    context = {
        'issue': issue,
        'items': items,
    }
    
    return render(request, 'inventory/issue_detail.html', context)


@login_required
def issue_process(request, issue_id):
    """Process and issue stock"""
    issue = get_object_or_404(StockIssue, id=issue_id, status='PENDING')
    
    if request.method == 'POST':
        issue.status = 'ISSUED'
        issue.issued_by = request.user
        issue.issue_date = timezone.now().date()
        
        for item in issue.items.all():
            qty_issued = Decimal(request.POST.get(f'qty_issued_{item.id}', 0))
            
            if qty_issued > 0:
                # Check available stock
                if qty_issued > item.stock_item.quantity_on_hand:
                    messages.error(
                        request,
                        f'Insufficient stock for {item.stock_item.item.name}!'
                    )
                    return redirect('issue_process', issue_id=issue_id)
                
                # Update issue item
                item.quantity_issued = qty_issued
                item.save()
                
                # Update stock
                balance_before = item.stock_item.quantity_on_hand
                item.stock_item.quantity_on_hand -= qty_issued
                item.stock_item.last_issue_date = timezone.now().date()
                item.stock_item.total_value = \
                    item.stock_item.quantity_on_hand * item.stock_item.average_unit_cost
                item.stock_item.save()
                
                # Record movement
                StockMovement.objects.create(
                    stock_item=item.stock_item,
                    movement_type='ISSUE',
                    reference_number=issue.issue_number,
                    reference_type='ISSUE',
                    quantity=qty_issued,
                    unit_cost=item.stock_item.average_unit_cost,
                    balance_before=balance_before,
                    balance_after=item.stock_item.quantity_on_hand,
                    from_store=issue.store,
                    performed_by=request.user
                )
        
        issue.save()
        messages.success(request, f'Stock issue {issue.issue_number} processed successfully!')
        return redirect('issue_detail', issue_id=issue.id)
    
    context = {
        'issue': issue,
        'items': issue.items.all(),
    }
    
    return render(request, 'inventory/issue_process.html', context)


# ============================================================================
# ASSETS
# ============================================================================

@login_required
def asset_list(request):
    """List all assets"""
    assets = Asset.objects.select_related(
        'item', 'department', 'custodian', 'grn'
    ).all()
    
    # Filters
    filters = {
        'search': request.GET.get('search', ''),
        'status': request.GET.get('status', ''),
        'department': request.GET.get('department', ''),
    }
    
    # Apply filters
    if filters['search']:
        assets = assets.filter(
            Q(asset_number__icontains=filters['search']) |
            Q(asset_tag__icontains=filters['search']) |
            Q(description__icontains=filters['search']) |
            Q(serial_number__icontains=filters['search'])
        )
    
    if filters['status']:
        assets = assets.filter(status=filters['status'])
    
    if filters['department']:
        assets = assets.filter(department_id=filters['department'])
    
    # Get filter options
    departments = Department.objects.filter(is_active=True)
    status_choices = Asset.ASSET_STATUS
    
    # Statistics
    total_assets = assets.count()
    total_value = assets.aggregate(
        total=Coalesce(Sum('current_value'), Decimal('0'))
    )['total']
    active_assets = assets.filter(status='ACTIVE').count()
    
    context = {
        'assets': assets.order_by('-created_at'),
        'filters': filters,
        'departments': departments,
        'status_choices': status_choices,
        'total_assets': total_assets,
        'total_value': total_value,
        'active_assets': active_assets,
    }
    
    return render(request, 'inventory/asset_list.html', context)


@login_required
def asset_detail(request, asset_id):
    """View asset details"""
    asset = get_object_or_404(
        Asset.objects.select_related(
            'item', 'department', 'custodian', 'grn'
        ),
        id=asset_id
    )
    
    context = {
        'asset': asset,
    }
    
    return render(request, 'inventory/asset_detail.html', context)


# ============================================================================
# API ENDPOINTS
# ============================================================================

@login_required
def get_stock_items_by_store(request, store_id):
    """API: Get stock items for a specific store"""
    stock_items = StockItem.objects.filter(
        store_id=store_id,
        quantity_on_hand__gt=0
    ).select_related('item').values(
        'id', 'item__name', 'item__code', 'quantity_on_hand',
        'unit_of_measure', 'average_unit_cost'
    )
    
    return JsonResponse(list(stock_items), safe=False)


@login_required
def get_po_items(request, po_id):
    """API: Get PO items for GRN creation"""
    po_items = PurchaseOrderItem.objects.filter(
        purchase_order_id=po_id
    ).select_related('requisition_item').values(
        'id', 'item_description', 'quantity', 'unit_of_measure',
        'unit_price', 'quantity_delivered', 'quantity_pending'
    )
    
    return JsonResponse(list(po_items), safe=False)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.utils import timezone
from .models import (
    User, Permission, RolePermission, AuditLog, Faculty, Department,
    SystemConfiguration, ProcurementPolicy, Notification
)
from .forms import (
    UserForm, DepartmentForm, FacultyForm, SystemConfigForm,
    ProcurementPolicyForm
)


# ============================================================================
# USER MANAGEMENT VIEWS
# ============================================================================

@login_required
def user_list(request):
    """List all users with filtering"""
    if request.user.role not in ['ADMIN', 'PROCUREMENT']:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')
    
    users = User.objects.all().select_related('department', 'department__faculty')
    
    # Filters
    search = request.GET.get('search', '')
    role = request.GET.get('role', '')
    department = request.GET.get('department', '')
    status = request.GET.get('status', '')
    
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(employee_id__icontains=search)
        )
    
    if role:
        users = users.filter(role=role)
    
    if department:
        users = users.filter(department_id=department)
    
    if status:
        is_active = status == 'active'
        users = users.filter(is_active_user=is_active)
    
    # Statistics
    total_users = users.count()
    active_users = users.filter(is_active_user=True).count()
    inactive_users = users.filter(is_active_user=False).count()
    
    # Role distribution
    role_counts = users.values('role').annotate(count=Count('id'))
    
    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'users': page_obj,
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'role_counts': role_counts,
        'role_choices': User.ROLE_CHOICES,
        'departments': Department.objects.filter(is_active=True),
        'filters': {
            'search': search,
            'role': role,
            'department': department,
            'status': status,
        }
    }
    return render(request, 'administration/user_list.html', context)


@login_required
def user_create(request):
    """Create new user"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to create users.')
        return redirect('user_list')
    
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            
            # Log audit
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model_name='User',
                object_id=str(user.id),
                object_repr=str(user),
                changes={'created': 'New user created'}
            )
            
            messages.success(request, f'User {user.username} created successfully.')
            return redirect('user_detail', user.id)
    else:
        form = UserForm()
    
    context = {
        'form': form,
        'departments': Department.objects.filter(is_active=True),
        'role_choices': User.ROLE_CHOICES,
    }
    return render(request, 'administration/user_form.html', context)


@login_required
def user_detail(request, user_id):
    """View user details"""
    if request.user.role not in ['ADMIN', 'PROCUREMENT']:
        messages.error(request, 'You do not have permission to view user details.')
        return redirect('dashboard')
    
    user_obj = get_object_or_404(User, id=user_id)
    
    # Get user activity
    recent_activity = AuditLog.objects.filter(user=user_obj).order_by('-timestamp')[:20]
    
    # Get statistics
    total_requisitions = user_obj.requisitions_created.count()
    total_approvals = user_obj.approvals_made.count()
    
    context = {
        'user_obj': user_obj,
        'recent_activity': recent_activity,
        'total_requisitions': total_requisitions,
        'total_approvals': total_approvals,
    }
    return render(request, 'administration/user_detail.html', context)


@login_required
def user_edit(request, user_id):
    """Edit user"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to edit users.')
        return redirect('user_list')
    
    user_obj = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=user_obj)
        if form.is_valid():
            user = form.save()
            
            # Log audit
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model_name='User',
                object_id=str(user.id),
                object_repr=str(user),
                changes={'updated': 'User information updated'}
            )
            
            messages.success(request, f'User {user.username} updated successfully.')
            return redirect('user_detail', user.id)
    else:
        form = UserForm(instance=user_obj)
    
    context = {
        'form': form,
        'user_obj': user_obj,
        'departments': Department.objects.filter(is_active=True),
        'role_choices': User.ROLE_CHOICES,
    }
    return render(request, 'administration/user_form.html', context)


@login_required
def user_toggle_status(request, user_id):
    """Toggle user active status"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to modify user status.')
        return redirect('user_list')
    
    user_obj = get_object_or_404(User, id=user_id)
    
    # Prevent self-deactivation
    if user_obj == request.user:
        messages.error(request, 'You cannot deactivate your own account.')
        return redirect('user_detail', user_id)
    
    user_obj.is_active_user = not user_obj.is_active_user
    user_obj.save()
    
    status = 'activated' if user_obj.is_active_user else 'deactivated'
    
    # Log audit
    AuditLog.objects.create(
        user=request.user,
        action='UPDATE',
        model_name='User',
        object_id=str(user_obj.id),
        object_repr=str(user_obj),
        changes={'status_changed': status}
    )
    
    messages.success(request, f'User {user_obj.username} has been {status}.')
    return redirect('user_detail', user_id)


# ============================================================================
# ROLES & PERMISSIONS VIEWS
# ============================================================================

@login_required
def role_permissions_list(request):
    """List roles and their permissions"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')
    
    # Get all roles with their permissions
    roles_data = []
    for role_code, role_name in User.ROLE_CHOICES:
        permissions = RolePermission.objects.filter(role=role_code).select_related('permission')
        roles_data.append({
            'code': role_code,
            'name': role_name,
            'permissions': permissions,
            'permission_count': permissions.count(),
        })
    
    # Get all available permissions grouped by module
    all_permissions = Permission.objects.all().order_by('module', 'name')
    modules = {}
    for perm in all_permissions:
        if perm.module not in modules:
            modules[perm.module] = []
        modules[perm.module].append(perm)
    
    context = {
        'roles_data': roles_data,
        'modules': modules,
        'all_permissions': all_permissions,
    }
    return render(request, 'administration/role_permissions_list.html', context)


@login_required
def role_permissions_edit(request, role_code):
    """Edit permissions for a specific role"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to edit role permissions.')
        return redirect('role_permissions_list')
    
    role_name = dict(User.ROLE_CHOICES).get(role_code, role_code)
    
    if request.method == 'POST':
        permission_ids = request.POST.getlist('permissions')
        
        # Remove existing permissions for this role
        RolePermission.objects.filter(role=role_code).delete()
        
        # Add new permissions
        for perm_id in permission_ids:
            permission = Permission.objects.get(id=perm_id)
            RolePermission.objects.create(role=role_code, permission=permission)
        
        # Log audit
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='RolePermission',
            object_id=role_code,
            object_repr=f'{role_name} permissions',
            changes={'permissions_updated': f'{len(permission_ids)} permissions assigned'}
        )
        
        messages.success(request, f'Permissions for {role_name} updated successfully.')
        return redirect('role_permissions_list')
    
    # Get current permissions for this role
    current_permissions = RolePermission.objects.filter(role=role_code).values_list('permission_id', flat=True)
    
    # Get all permissions grouped by module
    all_permissions = Permission.objects.all().order_by('module', 'name')
    modules = {}
    for perm in all_permissions:
        if perm.module not in modules:
            modules[perm.module] = []
        modules[perm.module].append(perm)
    
    context = {
        'role_code': role_code,
        'role_name': role_name,
        'modules': modules,
        'current_permissions': list(current_permissions),
    }
    return render(request, 'administration/role_permissions_edit.html', context)


# ============================================================================
# DEPARTMENT MANAGEMENT VIEWS
# ============================================================================

@login_required
def department_list(request):
    """List all departments"""
    if request.user.role not in ['ADMIN', 'PROCUREMENT']:
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')
    
    departments = Department.objects.all().select_related('faculty', 'hod')
    
    # Filters
    search = request.GET.get('search', '')
    faculty = request.GET.get('faculty', '')
    dept_type = request.GET.get('type', '')
    status = request.GET.get('status', '')
    
    if search:
        departments = departments.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search)
        )
    
    if faculty:
        departments = departments.filter(faculty_id=faculty)
    
    if dept_type:
        departments = departments.filter(department_type=dept_type)
    
    if status:
        is_active = status == 'active'
        departments = departments.filter(is_active=is_active)
    
    # Statistics
    total_depts = departments.count()
    active_depts = departments.filter(is_active=True).count()
    
    # Type distribution
    type_counts = departments.values('department_type').annotate(count=Count('id'))
    
    context = {
        'departments': departments,
        'total_depts': total_depts,
        'active_depts': active_depts,
        'type_counts': type_counts,
        'faculties': Faculty.objects.filter(is_active=True),
        'type_choices': Department.DEPARTMENT_TYPE,
        'filters': {
            'search': search,
            'faculty': faculty,
            'type': dept_type,
            'status': status,
        }
    }
    return render(request, 'administration/department_list.html', context)


@login_required
def department_create(request):
    """Create new department"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to create departments.')
        return redirect('department_list')
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            
            # Log audit
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model_name='Department',
                object_id=str(department.id),
                object_repr=str(department),
                changes={'created': 'New department created'}
            )
            
            messages.success(request, f'Department {department.name} created successfully.')
            return redirect('department_detail', department.id)
    else:
        form = DepartmentForm()
    
    context = {
        'form': form,
        'faculties': Faculty.objects.filter(is_active=True),
        'type_choices': Department.DEPARTMENT_TYPE,
        'hods': User.objects.filter(role='HOD', is_active_user=True),
    }
    return render(request, 'administration/department_form.html', context)


@login_required
def department_detail(request, dept_id):
    """View department details"""
    if request.user.role not in ['ADMIN', 'PROCUREMENT', 'HOD']:
        messages.error(request, 'You do not have permission to view department details.')
        return redirect('dashboard')
    
    department = get_object_or_404(Department, id=dept_id)
    
    # Statistics
    total_staff = User.objects.filter(department=department, is_active_user=True).count()
    total_requisitions = department.requisitions.count()
    total_budgets = department.budgets.count()
    
    # Recent activity
    recent_requisitions = department.requisitions.order_by('-created_at')[:5]
    
    context = {
        'department': department,
        'total_staff': total_staff,
        'total_requisitions': total_requisitions,
        'total_budgets': total_budgets,
        'recent_requisitions': recent_requisitions,
    }
    return render(request, 'administration/department_detail.html', context)


# ============================================================================
# SYSTEM SETTINGS VIEWS
# ============================================================================

@login_required
def system_settings(request):
    """System configuration settings"""
    if request.user.role not in ['ADMIN']:
        messages.error(request, 'You do not have permission to access system settings.')
        return redirect('dashboard')
    
    configs = SystemConfiguration.objects.all().order_by('key')
    
    if request.method == 'POST':
        for config in configs:
            if config.is_editable:
                new_value = request.POST.get(f'config_{config.id}')
                if new_value is not None:
                    config.value = new_value
                    config.updated_by = request.user
                    config.save()
        
        # Log audit
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            model_name='SystemConfiguration',
            object_id='bulk',
            object_repr='System Settings',
            changes={'updated': 'System configuration updated'}
        )
        
        messages.success(request, 'System settings updated successfully.')
        return redirect('system_settings')
    
    context = {
        'configs': configs,
    }
    return render(request, 'administration/system_settings.html', context)


# ============================================================================
# AUDIT TRAIL VIEWS
# ============================================================================

@login_required
def audit_trail(request):
    """View audit trail"""
    if request.user.role not in ['ADMIN', 'AUDITOR']:
        messages.error(request, 'You do not have permission to access audit trail.')
        return redirect('dashboard')
    
    logs = AuditLog.objects.all().select_related('user')
    
    # Filters
    search = request.GET.get('search', '')
    user = request.GET.get('user', '')
    action = request.GET.get('action', '')
    model = request.GET.get('model', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if search:
        logs = logs.filter(
            Q(object_repr__icontains=search) |
            Q(object_id__icontains=search)
        )
    
    if user:
        logs = logs.filter(user_id=user)
    
    if action:
        logs = logs.filter(action=action)
    
    if model:
        logs = logs.filter(model_name=model)
    
    if date_from:
        logs = logs.filter(timestamp__gte=date_from)
    
    if date_to:
        logs = logs.filter(timestamp__lte=date_to)
    
    # Statistics
    total_logs = logs.count()
    action_counts = logs.values('action').annotate(count=Count('id'))
    
    # Pagination
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'logs': page_obj,
        'total_logs': total_logs,
        'action_counts': action_counts,
        'users': User.objects.filter(is_active_user=True),
        'action_choices': AuditLog.ACTION_TYPES,
        'filters': {
            'search': search,
            'user': user,
            'action': action,
            'model': model,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    return render(request, 'administration/audit_trail.html', context)


# ============================================================================
# NOTIFICATIONS MANAGEMENT
# ============================================================================

@login_required
def notifications_list(request):
    """List all notifications"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    
    # Mark as read if requested
    if request.GET.get('mark_read'):
        notification_id = request.GET.get('mark_read')
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save()
        return redirect('notifications_list')
    
    # Mark all as read
    if request.GET.get('mark_all_read'):
        notifications.filter(is_read=False).update(is_read=True, read_at=timezone.now())
        messages.success(request, 'All notifications marked as read.')
        return redirect('notifications_list')
    
    # Statistics
    unread_count = notifications.filter(is_read=False).count()
    
    # Pagination
    paginator = Paginator(notifications, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'notifications': page_obj,
        'unread_count': unread_count,
    }
    return render(request, 'administration/notifications_list.html', context)


# ============================================================================
# PROCUREMENT POLICIES
# ============================================================================

@login_required
def policy_list(request):
    """List procurement policies"""
    policies = ProcurementPolicy.objects.all().order_by('-effective_date')
    
    # Filter by active status
    status = request.GET.get('status', '')
    if status == 'active':
        policies = policies.filter(is_active=True)
    elif status == 'inactive':
        policies = policies.filter(is_active=False)
    
    context = {
        'policies': policies,
        'filters': {'status': status},
    }
    return render(request, 'administration/policy_list.html', context)


@login_required
def policy_detail(request, policy_id):
    """View policy details"""
    policy = get_object_or_404(ProcurementPolicy, id=policy_id)
    
    context = {
        'policy': policy,
    }
    return render(request, 'administration/policy_detail.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg
from django.http import JsonResponse
from django.utils import timezone
from decimal import Decimal
from datetime import datetime, timedelta

from .models import (
    Supplier, SupplierDocument, SupplierPerformance,
    ItemCategory, PurchaseOrder, Bid, Contract
)


# ============================================================================
# SUPPLIER VIEWS
# ============================================================================

@login_required
def supplier_list(request):
    """List all suppliers with filtering"""
    
    # Get filter parameters
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    rating_min = request.GET.get('rating_min', '')
    
    # Base queryset
    suppliers = Supplier.objects.all().select_related().prefetch_related('categories')
    
    # Apply filters
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(supplier_number__icontains=search) |
            Q(registration_number__icontains=search) |
            Q(email__icontains=search)
        )
    
    if status:
        suppliers = suppliers.filter(status=status)
    
    if category:
        suppliers = suppliers.filter(categories__id=category)
    
    if rating_min:
        suppliers = suppliers.filter(rating__gte=rating_min)
    
    # Get statistics
    total_suppliers = suppliers.count()
    approved_suppliers = suppliers.filter(status='APPROVED').count()
    pending_suppliers = suppliers.filter(status='PENDING').count()
    
    # Get categories for filter
    categories = ItemCategory.objects.filter(is_active=True)
    
    context = {
        'suppliers': suppliers.order_by('-created_at'),
        'total_suppliers': total_suppliers,
        'approved_suppliers': approved_suppliers,
        'pending_suppliers': pending_suppliers,
        'categories': categories,
        'status_choices': Supplier.STATUS_CHOICES,
        'filters': {
            'search': search,
            'status': status,
            'category': category,
            'rating_min': rating_min,
        }
    }
    
    return render(request, 'suppliers/supplier_list.html', context)


@login_required
def supplier_detail(request, supplier_id):
    """View supplier details"""
    
    supplier = get_object_or_404(
        Supplier.objects.prefetch_related('categories', 'documents', 'performances'),
        id=supplier_id
    )
    
    # Get supplier statistics
    total_pos = PurchaseOrder.objects.filter(supplier=supplier).count()
    total_value = PurchaseOrder.objects.filter(
        supplier=supplier,
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED', 'CLOSED']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Get recent purchase orders
    recent_pos = PurchaseOrder.objects.filter(
        supplier=supplier
    ).select_related('requisition__department').order_by('-created_at')[:5]
    
    # Get performance reviews
    performances = supplier.performances.select_related(
        'purchase_order', 'reviewed_by'
    ).order_by('-reviewed_at')[:10]
    
    # Calculate average ratings
    avg_ratings = supplier.performances.aggregate(
        avg_quality=Avg('quality_rating'),
        avg_delivery=Avg('delivery_rating'),
        avg_service=Avg('service_rating'),
        avg_overall=Avg('overall_rating')
    )
    
    # Get active contracts
    active_contracts = Contract.objects.filter(
        supplier=supplier,
        status='ACTIVE'
    ).order_by('-start_date')[:5]
    
    # Check document expiry
    documents = supplier.documents.filter(is_verified=True)
    expiring_docs = documents.filter(
        expiry_date__lte=timezone.now().date() + timedelta(days=30),
        expiry_date__gte=timezone.now().date()
    )
    expired_docs = documents.filter(expiry_date__lt=timezone.now().date())
    
    context = {
        'supplier': supplier,
        'total_pos': total_pos,
        'total_value': total_value,
        'recent_pos': recent_pos,
        'performances': performances,
        'avg_ratings': avg_ratings,
        'active_contracts': active_contracts,
        'documents': documents,
        'expiring_docs': expiring_docs,
        'expired_docs': expired_docs,
    }
    
    return render(request, 'suppliers/supplier_detail.html', context)


@login_required
def supplier_create(request):
    """Create new supplier"""
    
    if request.method == 'POST':
        try:
            # Generate supplier number
            year = timezone.now().year
            last_supplier = Supplier.objects.filter(
                supplier_number__startswith=f'SUP-{year}'
            ).order_by('-supplier_number').first()
            
            if last_supplier:
                last_number = int(last_supplier.supplier_number.split('-')[-1])
                new_number = last_number + 1
            else:
                new_number = 1
            
            supplier_number = f'SUP-{year}-{new_number:06d}'
            
            # Create supplier
            supplier = Supplier.objects.create(
                supplier_number=supplier_number,
                name=request.POST.get('name'),
                registration_number=request.POST.get('registration_number'),
                tax_id=request.POST.get('tax_id', ''),
                email=request.POST.get('email'),
                phone_number=request.POST.get('phone_number'),
                physical_address=request.POST.get('physical_address'),
                postal_address=request.POST.get('postal_address', ''),
                website=request.POST.get('website', ''),
                contact_person=request.POST.get('contact_person'),
                contact_person_phone=request.POST.get('contact_person_phone'),
                contact_person_email=request.POST.get('contact_person_email'),
                bank_name=request.POST.get('bank_name'),
                bank_branch=request.POST.get('bank_branch'),
                account_number=request.POST.get('account_number'),
                account_name=request.POST.get('account_name'),
                swift_code=request.POST.get('swift_code', ''),
                notes=request.POST.get('notes', ''),
                created_by=request.user
            )
            
            # Add categories
            category_ids = request.POST.getlist('categories')
            if category_ids:
                supplier.categories.set(category_ids)
            
            messages.success(request, f'Supplier {supplier.name} created successfully!')
            return redirect('supplier_detail', supplier_id=supplier.id)
            
        except Exception as e:
            messages.error(request, f'Error creating supplier: {str(e)}')
    
    categories = ItemCategory.objects.filter(is_active=True)
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'suppliers/supplier_create.html', context)


@login_required
def supplier_edit(request, supplier_id):
    """Edit supplier information"""
    
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    if request.method == 'POST':
        try:
            supplier.name = request.POST.get('name')
            supplier.registration_number = request.POST.get('registration_number')
            supplier.tax_id = request.POST.get('tax_id', '')
            supplier.email = request.POST.get('email')
            supplier.phone_number = request.POST.get('phone_number')
            supplier.physical_address = request.POST.get('physical_address')
            supplier.postal_address = request.POST.get('postal_address', '')
            supplier.website = request.POST.get('website', '')
            supplier.contact_person = request.POST.get('contact_person')
            supplier.contact_person_phone = request.POST.get('contact_person_phone')
            supplier.contact_person_email = request.POST.get('contact_person_email')
            supplier.bank_name = request.POST.get('bank_name')
            supplier.bank_branch = request.POST.get('bank_branch')
            supplier.account_number = request.POST.get('account_number')
            supplier.account_name = request.POST.get('account_name')
            supplier.swift_code = request.POST.get('swift_code', '')
            supplier.notes = request.POST.get('notes', '')
            supplier.save()
            
            # Update categories
            category_ids = request.POST.getlist('categories')
            supplier.categories.set(category_ids)
            
            messages.success(request, 'Supplier updated successfully!')
            return redirect('supplier_detail', supplier_id=supplier.id)
            
        except Exception as e:
            messages.error(request, f'Error updating supplier: {str(e)}')
    
    categories = ItemCategory.objects.filter(is_active=True)
    
    context = {
        'supplier': supplier,
        'categories': categories,
    }
    
    return render(request, 'suppliers/supplier_edit.html', context)


@login_required
def supplier_update_status(request, supplier_id):
    """Update supplier status (approve/suspend/blacklist)"""
    
    if request.method == 'POST':
        supplier = get_object_or_404(Supplier, id=supplier_id)
        new_status = request.POST.get('status')
        reason = request.POST.get('reason', '')
        
        if new_status in dict(Supplier.STATUS_CHOICES).keys():
            old_status = supplier.status
            supplier.status = new_status
            
            if reason:
                supplier.notes = f"{supplier.notes}\n\n[{timezone.now()}] Status changed from {old_status} to {new_status}: {reason}"
            
            supplier.save()
            
            messages.success(request, f'Supplier status updated to {supplier.get_status_display()}')
        else:
            messages.error(request, 'Invalid status')
    
    return redirect('supplier_detail', supplier_id=supplier_id)


@login_required
def supplier_documents(request, supplier_id):
    """Manage supplier documents"""
    
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    if request.method == 'POST':
        try:
            document = SupplierDocument.objects.create(
                supplier=supplier,
                document_type=request.POST.get('document_type'),
                document_name=request.POST.get('document_name'),
                file=request.FILES.get('file'),
                issue_date=request.POST.get('issue_date') or None,
                expiry_date=request.POST.get('expiry_date') or None,
            )
            
            messages.success(request, 'Document uploaded successfully!')
            return redirect('supplier_detail', supplier_id=supplier.id)
            
        except Exception as e:
            messages.error(request, f'Error uploading document: {str(e)}')
    
    documents = supplier.documents.all().order_by('-uploaded_at')
    
    context = {
        'supplier': supplier,
        'documents': documents,
        'document_types': SupplierDocument.DOCUMENT_TYPES,
    }
    
    return render(request, 'suppliers/supplier_documents.html', context)


@login_required
def supplier_verify_document(request, document_id):
    """Verify supplier document"""
    
    if request.method == 'POST':
        document = get_object_or_404(SupplierDocument, id=document_id)
        
        document.is_verified = True
        document.verified_by = request.user
        document.verified_at = timezone.now()
        document.save()
        
        messages.success(request, f'Document {document.document_name} verified successfully!')
    
    return redirect('supplier_detail', supplier_id=document.supplier.id)


# ============================================================================
# VENDOR MANAGEMENT VIEWS
# ============================================================================

@login_required
def vendor_dashboard(request):
    """Vendor management dashboard"""
    
    # Get statistics
    total_vendors = Supplier.objects.count()
    approved_vendors = Supplier.objects.filter(status='APPROVED').count()
    pending_vendors = Supplier.objects.filter(status='PENDING').count()
    suspended_vendors = Supplier.objects.filter(status='SUSPENDED').count()
    blacklisted_vendors = Supplier.objects.filter(status='BLACKLISTED').count()
    
    # Top performing vendors
    top_vendors = Supplier.objects.filter(
        status='APPROVED'
    ).order_by('-rating')[:10]
    
    # Recent vendor registrations
    recent_vendors = Supplier.objects.order_by('-created_at')[:10]
    
    # Vendors with expiring documents
    expiring_date = timezone.now().date() + timedelta(days=30)
    vendors_expiring_docs = Supplier.objects.filter(
        documents__expiry_date__lte=expiring_date,
        documents__expiry_date__gte=timezone.now().date(),
        documents__is_verified=True
    ).distinct()
    
    # Vendors with expired documents
    vendors_expired_docs = Supplier.objects.filter(
        documents__expiry_date__lt=timezone.now().date(),
        documents__is_verified=True
    ).distinct()
    
    # Performance distribution
    excellent_vendors = Supplier.objects.filter(rating__gte=4.5).count()
    good_vendors = Supplier.objects.filter(rating__gte=3.5, rating__lt=4.5).count()
    average_vendors = Supplier.objects.filter(rating__gte=2.5, rating__lt=3.5).count()
    poor_vendors = Supplier.objects.filter(rating__lt=2.5).count()
    
    # Category distribution
    category_stats = ItemCategory.objects.annotate(
        vendor_count=Count('suppliers')
    ).order_by('-vendor_count')[:10]
    
    context = {
        'total_vendors': total_vendors,
        'approved_vendors': approved_vendors,
        'pending_vendors': pending_vendors,
        'suspended_vendors': suspended_vendors,
        'blacklisted_vendors': blacklisted_vendors,
        'top_vendors': top_vendors,
        'recent_vendors': recent_vendors,
        'vendors_expiring_docs': vendors_expiring_docs,
        'vendors_expired_docs': vendors_expired_docs,
        'excellent_vendors': excellent_vendors,
        'good_vendors': good_vendors,
        'average_vendors': average_vendors,
        'poor_vendors': poor_vendors,
        'category_stats': category_stats,
    }
    
    return render(request, 'vendors/vendor_dashboard.html', context)


@login_required
def vendor_performance_list(request):
    """List all vendor performance reviews"""
    
    # Get filter parameters
    supplier_id = request.GET.get('supplier', '')
    rating_min = request.GET.get('rating_min', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    performances = SupplierPerformance.objects.select_related(
        'supplier', 'purchase_order', 'reviewed_by'
    )
    
    # Apply filters
    if supplier_id:
        performances = performances.filter(supplier_id=supplier_id)
    
    if rating_min:
        performances = performances.filter(overall_rating__gte=rating_min)
    
    if date_from:
        performances = performances.filter(reviewed_at__gte=date_from)
    
    if date_to:
        performances = performances.filter(reviewed_at__lte=date_to)
    
    # Get statistics
    total_reviews = performances.count()
    avg_quality = performances.aggregate(avg=Avg('quality_rating'))['avg'] or 0
    avg_delivery = performances.aggregate(avg=Avg('delivery_rating'))['avg'] or 0
    avg_service = performances.aggregate(avg=Avg('service_rating'))['avg'] or 0
    avg_overall = performances.aggregate(avg=Avg('overall_rating'))['avg'] or 0
    
    # Get suppliers for filter
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    
    context = {
        'performances': performances.order_by('-reviewed_at'),
        'total_reviews': total_reviews,
        'avg_quality': avg_quality,
        'avg_delivery': avg_delivery,
        'avg_service': avg_service,
        'avg_overall': avg_overall,
        'suppliers': suppliers,
        'filters': {
            'supplier': supplier_id,
            'rating_min': rating_min,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    return render(request, 'vendors/vendor_performance_list.html', context)


@login_required
def vendor_performance_create(request, po_id):
    """Create vendor performance review"""
    
    po = get_object_or_404(
        PurchaseOrder.objects.select_related('supplier'),
        id=po_id
    )
    
    # Check if already reviewed
    existing_review = SupplierPerformance.objects.filter(purchase_order=po).first()
    if existing_review:
        messages.warning(request, 'This purchase order has already been reviewed.')
        return redirect('po_detail', po_id=po.id)
    
    if request.method == 'POST':
        try:
            quality_rating = int(request.POST.get('quality_rating'))
            delivery_rating = int(request.POST.get('delivery_rating'))
            service_rating = int(request.POST.get('service_rating'))
            
            performance = SupplierPerformance.objects.create(
                supplier=po.supplier,
                purchase_order=po,
                quality_rating=quality_rating,
                delivery_rating=delivery_rating,
                service_rating=service_rating,
                comments=request.POST.get('comments', ''),
                reviewed_by=request.user
            )
            
            # Update supplier overall rating
            avg_rating = SupplierPerformance.objects.filter(
                supplier=po.supplier
            ).aggregate(avg=Avg('overall_rating'))['avg']
            
            po.supplier.rating = avg_rating
            po.supplier.save()
            
            messages.success(request, 'Performance review submitted successfully!')
            return redirect('po_detail', po_id=po.id)
            
        except Exception as e:
            messages.error(request, f'Error creating review: {str(e)}')
    
    context = {
        'po': po,
    }
    
    return render(request, 'vendors/vendor_performance_create.html', context)


@login_required
def vendor_comparison(request):
    """Compare vendors for selection"""
    
    supplier_ids = request.GET.getlist('suppliers')
    
    if not supplier_ids:
        messages.warning(request, 'Please select suppliers to compare.')
        return redirect('supplier_list')
    
    suppliers = Supplier.objects.filter(
        id__in=supplier_ids
    ).prefetch_related('performances', 'categories')
    
    # Get comparison data
    comparison_data = []
    for supplier in suppliers:
        performances = supplier.performances.aggregate(
            avg_quality=Avg('quality_rating'),
            avg_delivery=Avg('delivery_rating'),
            avg_service=Avg('service_rating'),
            avg_overall=Avg('overall_rating'),
            total_reviews=Count('id')
        )
        
        total_pos = PurchaseOrder.objects.filter(supplier=supplier).count()
        total_value = PurchaseOrder.objects.filter(
            supplier=supplier,
            status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'DELIVERED', 'CLOSED']
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        comparison_data.append({
            'supplier': supplier,
            'performances': performances,
            'total_pos': total_pos,
            'total_value': total_value,
        })
    
    context = {
        'comparison_data': comparison_data,
    }
    
    return render(request, 'vendors/vendor_comparison.html', context)


@login_required
def vendor_compliance(request):
    """Vendor compliance tracking"""
    
    # Get compliance statistics
    total_vendors = Supplier.objects.filter(status='APPROVED').count()
    
    # Tax compliance
    tax_compliant = Supplier.objects.filter(
        status='APPROVED',
        tax_compliance_expiry__gte=timezone.now().date()
    ).count()
    
    tax_expiring = Supplier.objects.filter(
        status='APPROVED',
        tax_compliance_expiry__lte=timezone.now().date() + timedelta(days=30),
        tax_compliance_expiry__gte=timezone.now().date()
    ).count()
    
    tax_expired = Supplier.objects.filter(
        status='APPROVED',
        tax_compliance_expiry__lt=timezone.now().date()
    ).count()
    
    # Registration compliance
    reg_compliant = Supplier.objects.filter(
        status='APPROVED',
        registration_expiry__gte=timezone.now().date()
    ).count()
    
    reg_expiring = Supplier.objects.filter(
        status='APPROVED',
        registration_expiry__lte=timezone.now().date() + timedelta(days=30),
        registration_expiry__gte=timezone.now().date()
    ).count()
    
    reg_expired = Supplier.objects.filter(
        status='APPROVED',
        registration_expiry__lt=timezone.now().date()
    ).count()
    
    # Get non-compliant vendors
    non_compliant_vendors = Supplier.objects.filter(
        status='APPROVED'
    ).filter(
        Q(tax_compliance_expiry__lt=timezone.now().date()) |
        Q(registration_expiry__lt=timezone.now().date())
    ).distinct()
    
    # Expiring soon
    expiring_soon = Supplier.objects.filter(
        status='APPROVED'
    ).filter(
        Q(tax_compliance_expiry__lte=timezone.now().date() + timedelta(days=30),
          tax_compliance_expiry__gte=timezone.now().date()) |
        Q(registration_expiry__lte=timezone.now().date() + timedelta(days=30),
          registration_expiry__gte=timezone.now().date())
    ).distinct()
    
    context = {
        'total_vendors': total_vendors,
        'tax_compliant': tax_compliant,
        'tax_expiring': tax_expiring,
        'tax_expired': tax_expired,
        'reg_compliant': reg_compliant,
        'reg_expiring': reg_expiring,
        'reg_expired': reg_expired,
        'non_compliant_vendors': non_compliant_vendors,
        'expiring_soon': expiring_soon,
    }
    
    return render(request, 'vendors/vendor_compliance.html', context)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q

from .models import User, AuditLog


@login_required
def help_center(request):
    """Help Center page with FAQs and support options"""
    
    # Common FAQs organized by category
    faqs = {
        'Requisitions': [
            {
                'question': 'How do I create a purchase requisition?',
                'answer': 'Navigate to Requisitions > Create New. Fill in all required fields including item details, justification, and budget information. Submit for approval once complete.'
            },
            {
                'question': 'What is the approval process for requisitions?',
                'answer': 'Requisitions follow a multi-level approval workflow: HOD approval, Faculty approval (if required), Budget verification, and Procurement approval. The specific workflow depends on the requisition amount.'
            },
            {
                'question': 'How long does requisition approval take?',
                'answer': 'Standard requisitions are typically approved within 5-7 business days. Urgent requisitions may be expedited with proper justification.'
            },
        ],
        'Purchase Orders': [
            {
                'question': 'Who can create purchase orders?',
                'answer': 'Only Procurement Officers and System Administrators can create purchase orders. POs are generated after requisitions are fully approved.'
            },
            {
                'question': 'How do I track my purchase order?',
                'answer': 'Go to Purchase Orders > My Orders. Click on any PO number to view detailed status, delivery tracking, and related documents.'
            },
            {
                'question': 'Can a purchase order be amended?',
                'answer': 'Yes, PO amendments can be requested through the system. The amendment must be justified and approved before implementation.'
            },
        ],
        'Suppliers': [
            {
                'question': 'How do suppliers register in the system?',
                'answer': 'Suppliers can submit registration through the Supplier Portal. They must provide all required documents including business registration, tax compliance, and bank details.'
            },
            {
                'question': 'How are suppliers evaluated?',
                'answer': 'Suppliers are evaluated based on quality, delivery performance, and service. Performance ratings are recorded after each transaction.'
            },
        ],
        'Budget': [
            {
                'question': 'How do I check my department budget?',
                'answer': 'Navigate to Budget > Department Budget. You can view allocated amounts, committed amounts, actual spent, and available balance.'
            },
            {
                'question': 'Can budget be reallocated between line items?',
                'answer': 'Yes, budget reallocations require justification and approval from authorized officers.'
            },
        ],
        'Inventory': [
            {
                'question': 'How do I request items from store?',
                'answer': 'Create a stock issue request specifying items and quantities needed. The store keeper will process approved requests.'
            },
            {
                'question': 'How do I check stock availability?',
                'answer': 'Go to Stores > Inventory. Search for items to view current stock levels across all stores.'
            },
        ],
    }
    
    # System guides
    guides = [
        {
            'title': 'Getting Started Guide',
            'description': 'Learn the basics of using the procurement system',
            'icon': 'fa-rocket',
            'url': 'documentation#getting-started'
        },
        {
            'title': 'Requisition Process',
            'description': 'Step-by-step guide to creating requisitions',
            'icon': 'fa-file-alt',
            'url': 'documentation#requisitions'
        },
        {
            'title': 'Budget Management',
            'description': 'Understanding budget tracking and reporting',
            'icon': 'fa-chart-line',
            'url': 'documentation#budget'
        },
        {
            'title': 'User Roles & Permissions',
            'description': 'Learn about different user roles in the system',
            'icon': 'fa-users-cog',
            'url': 'documentation#roles'
        },
    ]
    
    # Contact information
    support_contacts = {
        'Technical Support': {
            'email': 'ict.support@university.ac.ke',
            'phone': '+254 700 000 001',
            'hours': 'Mon-Fri: 8:00 AM - 5:00 PM'
        },
        'Procurement Office': {
            'email': 'procurement@university.ac.ke',
            'phone': '+254 700 000 002',
            'hours': 'Mon-Fri: 8:00 AM - 5:00 PM'
        },
        'Finance Office': {
            'email': 'finance@university.ac.ke',
            'phone': '+254 700 000 003',
            'hours': 'Mon-Fri: 8:00 AM - 4:00 PM'
        },
    }
    
    context = {
        'faqs': faqs,
        'guides': guides,
        'support_contacts': support_contacts,
    }
    
    return render(request, 'support/help_center.html', context)


@login_required
def documentation(request):
    """System documentation and user guides"""
    
    # Documentation sections
    documentation_sections = [
        {
            'id': 'getting-started',
            'title': 'Getting Started',
            'icon': 'fa-rocket',
            'content': [
                {
                    'subtitle': 'System Overview',
                    'text': 'The University Procurement System is a comprehensive platform for managing all procurement activities from requisition to payment. It ensures transparency, compliance, and efficiency in the procurement process.'
                },
                {
                    'subtitle': 'Login and Dashboard',
                    'text': 'Access the system using your university credentials. The dashboard provides an overview of your pending tasks, recent activities, and quick access to frequently used features.'
                },
                {
                    'subtitle': 'Navigation',
                    'text': 'Use the sidebar menu to navigate between different modules. The top bar provides quick search, notifications, and access to your profile settings.'
                },
            ]
        },
        {
            'id': 'requisitions',
            'title': 'Purchase Requisitions',
            'icon': 'fa-file-alt',
            'content': [
                {
                    'subtitle': 'Creating a Requisition',
                    'text': '1. Navigate to Requisitions > Create New\n2. Select your department and budget category\n3. Add items with detailed specifications\n4. Provide justification for the purchase\n5. Attach supporting documents (quotations, specifications)\n6. Submit for approval'
                },
                {
                    'subtitle': 'Approval Workflow',
                    'text': 'Requisitions go through multiple approval stages:\n- Head of Department (HOD) approval\n- Faculty/School Dean approval (for amounts above threshold)\n- Budget verification by Finance\n- Procurement Officer review\n- Final approval based on amount'
                },
                {
                    'subtitle': 'Tracking Status',
                    'text': 'Monitor your requisition status in real-time. You will receive email notifications at each approval stage. Access the requisition detail page to view approval history and comments.'
                },
            ]
        },
        {
            'id': 'purchase-orders',
            'title': 'Purchase Orders',
            'icon': 'fa-shopping-cart',
            'content': [
                {
                    'subtitle': 'PO Generation',
                    'text': 'Purchase Orders are automatically generated by Procurement Officers after requisitions are fully approved and suppliers are selected through the tendering process.'
                },
                {
                    'subtitle': 'PO Tracking',
                    'text': 'Track PO status including:\n- Approval status\n- Supplier acknowledgment\n- Delivery progress\n- Goods receipt\n- Invoice matching\n- Payment status'
                },
                {
                    'subtitle': 'PO Amendments',
                    'text': 'If changes are needed, submit a PO amendment request with justification. Amendments require approval before implementation.'
                },
            ]
        },
        {
            'id': 'suppliers',
            'title': 'Supplier Management',
            'icon': 'fa-building',
            'content': [
                {
                    'subtitle': 'Supplier Registration',
                    'text': 'Suppliers must register and provide:\n- Business registration certificate\n- Tax compliance certificate\n- Bank details\n- Product/service categories\n- Contact information'
                },
                {
                    'subtitle': 'Supplier Evaluation',
                    'text': 'Suppliers are evaluated on:\n- Quality of goods/services\n- Delivery timeliness\n- Service and communication\nRatings affect future tender opportunities.'
                },
            ]
        },
        {
            'id': 'budget',
            'title': 'Budget Management',
            'icon': 'fa-chart-line',
            'content': [
                {
                    'subtitle': 'Budget Tracking',
                    'text': 'Monitor your department budget:\n- Allocated amount\n- Committed (pending requisitions)\n- Actual spent\n- Available balance'
                },
                {
                    'subtitle': 'Budget Reports',
                    'text': 'Generate budget utilization reports by:\n- Department\n- Category\n- Time period\n- Budget year'
                },
            ]
        },
        {
            'id': 'inventory',
            'title': 'Inventory & Stores',
            'icon': 'fa-warehouse',
            'content': [
                {
                    'subtitle': 'Goods Receipt',
                    'text': 'When goods are delivered:\n- Store keeper receives and inspects\n- GRN (Goods Received Note) is generated\n- Items are added to inventory\n- Supplier invoice is matched'
                },
                {
                    'subtitle': 'Stock Issues',
                    'text': 'Request items from store:\n- Create stock issue request\n- Specify items and quantities\n- Provide purpose/justification\n- Store keeper approves and issues items'
                },
            ]
        },
        {
            'id': 'roles',
            'title': 'User Roles & Permissions',
            'icon': 'fa-users-cog',
            'content': [
                {
                    'subtitle': 'System Roles',
                    'text': '- Requesting Staff: Create requisitions\n- Head of Department: Approve department requisitions\n- Procurement Officer: Manage tenders and POs\n- Finance Officer: Budget verification and payments\n- Stores Officer: Manage inventory\n- Supplier: Submit bids and track orders\n- Auditor: View audit trails\n- Administrator: System configuration'
                },
            ]
        },
    ]
    
    context = {
        'documentation_sections': documentation_sections,
    }
    
    return render(request, 'support/documentation.html', context)


@login_required
def submit_support_ticket(request):
    """Submit a support ticket"""
    
    if request.method == 'POST':
        try:
            subject = request.POST.get('subject')
            category = request.POST.get('category')
            priority = request.POST.get('priority')
            description = request.POST.get('description')
            
            # Send email to support team
            email_subject = f'[Support Ticket] {category} - {subject}'
            email_body = f"""
Support Ticket Details:

From: {request.user.get_full_name()} ({request.user.email})
Employee ID: {request.user.employee_id}
Department: {request.user.department.name if request.user.department else 'N/A'}
Role: {request.user.get_role_display()}

Category: {category}
Priority: {priority}
Subject: {subject}

Description:
{description}

---
Submitted via University Procurement System
            """
            
            send_mail(
                email_subject,
                email_body,
                settings.DEFAULT_FROM_EMAIL,
                ['ict.support@university.ac.ke'],
                fail_silently=False,
            )
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model_name='SupportTicket',
                object_id='TICKET',
                object_repr=subject,
                changes={'category': category, 'priority': priority}
            )
            
            messages.success(
                request, 
                'Your support ticket has been submitted. Our team will respond within 24 hours.'
            )
            return redirect('help_center')
            
        except Exception as e:
            messages.error(request, f'Error submitting support ticket: {str(e)}')
    
    return redirect('help_center')


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from .models import (
    Requisition, RequisitionApproval, User, Notification, 
    AuditLog, Budget, Department
)
import json


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_user_approval_stage(user):
    """Get the approval stage based on user role"""
    role_to_stage = {
        'HOD': 'HOD',
        'FINANCE': 'BUDGET',
        'PROCUREMENT': 'PROCUREMENT',
    }
    return role_to_stage.get(user.role)


def can_approve_stage(user, approval_stage):
    """Check if user can approve at a specific stage"""
    if user.role == 'ADMIN':
        return True  # Admin can approve at all levels
    
    if user.role == 'HOD' and approval_stage == 'HOD':
        return True
    
    if user.role == 'FINANCE' and approval_stage == 'BUDGET':
        return True
    
    if user.role == 'PROCUREMENT' and approval_stage == 'PROCUREMENT':
        return True
    
    return False


def get_pending_approvals_for_user(user):
    """Get all pending approvals for a specific user"""
    if user.role == 'ADMIN':
        # Admin can see all pending approvals
        return RequisitionApproval.objects.filter(
            status='PENDING'
        ).select_related('requisition', 'approver')
    
    # Get approvals where user is the approver
    return RequisitionApproval.objects.filter(
        approver=user,
        status='PENDING'
    ).select_related('requisition')


def check_budget_availability(requisition):
    """Check if budget is available for requisition"""
    if not requisition.budget:
        return {
            'available': False,
            'message': 'No budget line assigned',
            'details': None
        }
    
    budget = requisition.budget
    available = budget.available_balance
    required = requisition.estimated_amount
    
    if available >= required:
        return {
            'available': True,
            'message': 'Sufficient budget available',
            'details': {
                'allocated': float(budget.allocated_amount),
                'committed': float(budget.committed_amount),
                'spent': float(budget.actual_spent),
                'available': float(available),
                'required': float(required),
                'balance_after': float(available - required)
            }
        }
    else:
        return {
            'available': False,
            'message': f'Insufficient budget. Available: {available}, Required: {required}',
            'details': {
                'allocated': float(budget.allocated_amount),
                'committed': float(budget.committed_amount),
                'spent': float(budget.actual_spent),
                'available': float(available),
                'required': float(required),
                'shortfall': float(required - available)
            }
        }


def update_requisition_status(requisition):
    """Update requisition status based on approval workflow"""
    approvals = requisition.approvals.all().order_by('sequence')
    
    # Check if any approval is rejected
    if approvals.filter(status='REJECTED').exists():
        requisition.status = 'REJECTED'
        requisition.save()
        return
    
    # Check if all approvals are approved
    if all(approval.status == 'APPROVED' for approval in approvals):
        requisition.status = 'APPROVED'
        requisition.save()
        
        # Commit budget if approved
        if requisition.budget:
            budget = requisition.budget
            budget.committed_amount += requisition.estimated_amount
            budget.save()
        
        return
    
    # Update to intermediate status
    for approval in approvals:
        if approval.status == 'APPROVED':
            if approval.approval_stage == 'HOD':
                requisition.status = 'HOD_APPROVED'
            elif approval.approval_stage == 'BUDGET':
                requisition.status = 'BUDGET_APPROVED'
            elif approval.approval_stage == 'PROCUREMENT':
                requisition.status = 'PROCUREMENT_APPROVED'
        elif approval.status == 'PENDING':
            break  # Stop at first pending approval
    
    requisition.save()


# ============================================================================
# API ENDPOINTS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_approval_stats(request):
    """API endpoint to get approval statistics"""
    user = request.user
    
    if user.role == 'ADMIN':
        pending = RequisitionApproval.objects.filter(status='PENDING').count()
        approved = RequisitionApproval.objects.filter(status='APPROVED').count()
        rejected = RequisitionApproval.objects.filter(status='REJECTED').count()
    else:
        pending = RequisitionApproval.objects.filter(
            approver=user, 
            status='PENDING'
        ).count()
        approved = RequisitionApproval.objects.filter(
            approver=user, 
            status='APPROVED'
        ).count()
        rejected = RequisitionApproval.objects.filter(
            approver=user, 
            status='REJECTED'
        ).count()
    
    return JsonResponse({
        'success': True,
        'data': {
            'pending': pending,
            'approved': approved,
            'rejected': rejected,
            'total': pending + approved + rejected
        }
    })


@login_required
@require_http_methods(["GET"])
def api_check_budget(request, requisition_id):
    """API endpoint to check budget availability"""
    requisition = get_object_or_404(Requisition, pk=requisition_id)
    
    # Check permissions
    if not can_approve_stage(request.user, 'BUDGET') and request.user.role != 'ADMIN':
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to check budget'
        }, status=403)
    
    result = check_budget_availability(requisition)
    
    return JsonResponse({
        'success': True,
        'data': result
    })


@login_required
@require_http_methods(["GET"])
def api_approval_details(request, approval_id):
    """API endpoint to get approval details"""
    approval = get_object_or_404(RequisitionApproval, pk=approval_id)
    
    # Check permissions
    if approval.approver != request.user and request.user.role != 'ADMIN':
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to view this approval'
        }, status=403)
    
    requisition = approval.requisition
    
    data = {
        'approval': {
            'id': str(approval.id),
            'stage': approval.approval_stage,
            'stage_display': approval.get_approval_stage_display(),
            'status': approval.status,
            'approver': approval.approver.get_full_name(),
            'comments': approval.comments,
            'sequence': approval.sequence,
            'created_at': approval.created_at.isoformat(),
        },
        'requisition': {
            'id': str(requisition.id),
            'number': requisition.requisition_number,
            'title': requisition.title,
            'department': requisition.department.name,
            'requested_by': requisition.requested_by.get_full_name(),
            'amount': float(requisition.estimated_amount),
            'priority': requisition.priority,
            'status': requisition.status,
            'items_count': requisition.items.count(),
        }
    }
    
    # Add budget check if Finance/Budget stage
    if approval.approval_stage == 'BUDGET':
        data['budget_check'] = check_budget_availability(requisition)
    
    return JsonResponse({
        'success': True,
        'data': data
    })


# ============================================================================
# PENDING APPROVALS VIEW
# ============================================================================

@login_required
def pending_approvals(request):
    """Display pending approvals for the user"""
    user = request.user
    
    # Get pending approvals based on role
    if user.role == 'ADMIN':
        approvals_list = RequisitionApproval.objects.filter(
            status='PENDING',
            requisition__status__in=['SUBMITTED', 'HOD_APPROVED', 'BUDGET_APPROVED', 'PROCUREMENT_APPROVED']
        ).select_related(
            'requisition', 
            'requisition__department', 
            'requisition__requested_by',
            'approver'
        ).order_by('-created_at')
    else:
        approvals_list = RequisitionApproval.objects.filter(
            approver=user,
            status='PENDING',
            requisition__status__in=['SUBMITTED', 'HOD_APPROVED', 'BUDGET_APPROVED', 'PROCUREMENT_APPROVED']
        ).select_related(
            'requisition', 
            'requisition__department', 
            'requisition__requested_by'
        ).order_by('-created_at')
    
    # Filters
    priority = request.GET.get('priority')
    department = request.GET.get('department')
    stage = request.GET.get('stage')
    
    if priority:
        approvals_list = approvals_list.filter(requisition__priority=priority)
    
    if department:
        approvals_list = approvals_list.filter(requisition__department_id=department)
    
    if stage:
        approvals_list = approvals_list.filter(approval_stage=stage)
    
    # Pagination
    paginator = Paginator(approvals_list, 20)
    page_number = request.GET.get('page')
    approvals = paginator.get_page(page_number)
    
    # Get filter options
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Calculate statistics
    total_pending = approvals_list.count()
    urgent_count = approvals_list.filter(requisition__priority='URGENT').count()
    high_count = approvals_list.filter(requisition__priority='HIGH').count()
    total_amount = approvals_list.aggregate(
        total=Sum('requisition__estimated_amount')
    )['total'] or 0
    
    context = {
        'approvals': approvals,
        'departments': departments,
        'selected_priority': priority,
        'selected_department': department,
        'selected_stage': stage,
        'total_pending': total_pending,
        'urgent_count': urgent_count,
        'high_count': high_count,
        'total_amount': total_amount,
        'user_role': user.role,
    }
    
    return render(request, 'approvals/pending_approvals.html', context)


# ============================================================================
# APPROVED REQUISITIONS VIEW
# ============================================================================

@login_required
def approved_requisitions(request):
    """Display approved requisitions"""
    user = request.user
    
    # Get approved items based on role
    if user.role == 'ADMIN':
        approvals_list = RequisitionApproval.objects.filter(
            status='APPROVED'
        ).select_related(
            'requisition', 
            'requisition__department', 
            'requisition__requested_by',
            'approver'
        ).order_by('-approval_date')
    else:
        approvals_list = RequisitionApproval.objects.filter(
            approver=user,
            status='APPROVED'
        ).select_related(
            'requisition', 
            'requisition__department', 
            'requisition__requested_by'
        ).order_by('-approval_date')
    
    # Filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department = request.GET.get('department')
    
    if date_from:
        approvals_list = approvals_list.filter(approval_date__gte=date_from)
    
    if date_to:
        approvals_list = approvals_list.filter(approval_date__lte=date_to)
    
    if department:
        approvals_list = approvals_list.filter(requisition__department_id=department)
    
    # Pagination
    paginator = Paginator(approvals_list, 20)
    page_number = request.GET.get('page')
    approvals = paginator.get_page(page_number)
    
    # Get filter options
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Calculate statistics
    total_approved = approvals_list.count()
    total_amount = approvals_list.aggregate(
        total=Sum('requisition__estimated_amount')
    )['total'] or 0
    
    context = {
        'approvals': approvals,
        'departments': departments,
        'selected_department': department,
        'date_from': date_from,
        'date_to': date_to,
        'total_approved': total_approved,
        'total_amount': total_amount,
        'user_role': user.role,
    }
    
    return render(request, 'approvals/approved_requisitions.html', context)


# ============================================================================
# REJECTED REQUISITIONS VIEW
# ============================================================================

@login_required
def rejected_requisitions(request):
    """Display rejected requisitions"""
    user = request.user
    
    # Get rejected items based on role
    if user.role == 'ADMIN':
        approvals_list = RequisitionApproval.objects.filter(
            status='REJECTED'
        ).select_related(
            'requisition', 
            'requisition__department', 
            'requisition__requested_by',
            'approver'
        ).order_by('-approval_date')
    else:
        approvals_list = RequisitionApproval.objects.filter(
            approver=user,
            status='REJECTED'
        ).select_related(
            'requisition', 
            'requisition__department', 
            'requisition__requested_by'
        ).order_by('-approval_date')
    
    # Filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    department = request.GET.get('department')
    
    if date_from:
        approvals_list = approvals_list.filter(approval_date__gte=date_from)
    
    if date_to:
        approvals_list = approvals_list.filter(approval_date__lte=date_to)
    
    if department:
        approvals_list = approvals_list.filter(requisition__department_id=department)
    
    # Pagination
    paginator = Paginator(approvals_list, 20)
    page_number = request.GET.get('page')
    approvals = paginator.get_page(page_number)
    
    # Get filter options
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Calculate statistics
    total_rejected = approvals_list.count()
    
    context = {
        'approvals': approvals,
        'departments': departments,
        'selected_department': department,
        'date_from': date_from,
        'date_to': date_to,
        'total_rejected': total_rejected,
        'user_role': user.role,
    }
    
    return render(request, 'approvals/rejected_requisitions.html', context)


# ============================================================================
# APPROVAL DETAIL & ACTION VIEW
# ============================================================================

@login_required
def approval_detail(request, approval_id):
    """Display approval detail and handle approval/rejection"""
    approval = get_object_or_404(
        RequisitionApproval.objects.select_related(
            'requisition',
            'requisition__department',
            'requisition__requested_by',
            'requisition__budget',
            'approver'
        ),
        pk=approval_id
    )
    
    requisition = approval.requisition
    
    # Check permissions
    if approval.approver != request.user and request.user.role != 'ADMIN':
        messages.error(request, 'You do not have permission to view this approval.')
        return redirect('pending_approvals')
    
    # Check if user can approve this stage
    can_approve = can_approve_stage(request.user, approval.approval_stage)
    
    # Get budget check if finance stage
    budget_check = None
    if approval.approval_stage == 'BUDGET':
        budget_check = check_budget_availability(requisition)
    
    # Get all approvals for this requisition
    all_approvals = requisition.approvals.all().order_by('sequence')
    
    # Get requisition items
    items = requisition.items.all()
    items_total = sum(item.estimated_total for item in items)
    
    context = {
        'approval': approval,
        'requisition': requisition,
        'can_approve': can_approve,
        'budget_check': budget_check,
        'all_approvals': all_approvals,
        'items': items,
        'items_total': items_total,
    }
    
    return render(request, 'approvals/approval_detail.html', context)


# views.py

@login_required
def process_approval(request, requisition_id):
    """Process single or bulk approval/rejection"""
    requisition = get_object_or_404(Requisition, pk=requisition_id)
    
    # Get pending approvals for this user
    pending_approvals = RequisitionApproval.objects.filter(
        requisition=requisition,
        status='PENDING'
    ).order_by('sequence')
    
    # Check if user is admin and can bulk approve
    is_admin = request.user.role == 'ADMIN'
    can_bulk_approve = is_admin and pending_approvals.exists()
    
    if request.method == 'GET':
        # Show approval form
        context = {
            'requisition': requisition,
            'pending_approvals': pending_approvals,
            'can_bulk_approve': can_bulk_approve,
            'is_admin': is_admin,
        }
        return render(request, 'procurement/process_approval.html', context)
    
    elif request.method == 'POST':
        action = request.POST.get('action')  # 'approve' or 'reject'
        comments = request.POST.get('comments', '')
        bulk_approve = request.POST.get('bulk_approve') == 'true'
        
        if action not in ['approve', 'reject']:
            messages.error(request, 'Invalid action.')
            return redirect('requisition_detail', pk=requisition_id)
        
        try:
            with transaction.atomic():
                # Bulk approval for admin
                if bulk_approve and is_admin and action == 'approve':
                    approved_count = 0
                    
                    for approval in pending_approvals:
                        # For budget stage, check availability
                        if approval.approval_stage == 'BUDGET':
                            budget_check = check_budget_availability(requisition)
                            if not budget_check['available']:
                                messages.error(request, f"Budget check failed: {budget_check['message']}")
                                raise Exception("Insufficient budget")
                        
                        # Approve this stage
                        approval.status = 'APPROVED'
                        approval.approver = request.user
                        approval.comments = comments or f'Bulk approved by {request.user.get_full_name()}'
                        approval.approval_date = timezone.now()
                        approval.save()
                        approved_count += 1
                        
                        # Create audit log
                        AuditLog.objects.create(
                            user=request.user,
                            action='APPROVE',
                            model_name='RequisitionApproval',
                            object_id=str(approval.id),
                            object_repr=f'{requisition.requisition_number} - {approval.get_approval_stage_display()}',
                            changes={
                                'status': 'APPROVED',
                                'comments': approval.comments,
                                'bulk_approved': True
                            }
                        )
                    
                    # Update requisition to fully approved
                    requisition.status = 'APPROVED'
                    requisition.save()
                    
                    # Notify requester
                    Notification.objects.create(
                        user=requisition.requested_by,
                        notification_type='APPROVAL',
                        priority='HIGH',
                        title='Requisition Fully Approved',
                        message=f'Your requisition {requisition.requisition_number} has been fully approved (bulk approval).',
                        link_url=f'/requisitions/{requisition.id}/'
                    )
                    
                    messages.success(request, f'Requisition {requisition.requisition_number} has been fully approved ({approved_count} stages).')
                
                else:
                    # Single approval
                    approval_id = request.POST.get('approval_id')
                    if not approval_id:
                        messages.error(request, 'No approval ID provided.')
                        return redirect('requisition_detail', pk=requisition_id)
                    
                    approval = get_object_or_404(RequisitionApproval, pk=approval_id)
                    
                    if approval.status != 'PENDING':
                        messages.error(request, 'This approval has already been processed.')
                        return redirect('requisition_detail', pk=requisition_id)
                    
                    # For budget stage, check budget availability
                    if approval.approval_stage == 'BUDGET' and action == 'approve':
                        budget_check = check_budget_availability(requisition)
                        if not budget_check['available']:
                            messages.error(request, f"Budget check failed: {budget_check['message']}")
                            return redirect('process_approval', requisition_id=requisition_id)
                    
                    # Update approval
                    approval.status = 'APPROVED' if action == 'approve' else 'REJECTED'
                    approval.approver = request.user
                    approval.comments = comments
                    approval.approval_date = timezone.now()
                    approval.save()
                    
                    # If rejected, update requisition
                    if action == 'reject':
                        requisition.status = 'REJECTED'
                        requisition.rejection_reason = comments
                        requisition.save()
                    else:
                        # Update requisition status based on workflow
                        update_requisition_status(requisition)
                    
                    # Notify requester
                    Notification.objects.create(
                        user=requisition.requested_by,
                        notification_type='APPROVAL',
                        priority='HIGH',
                        title=f'Requisition {action.title()}d',
                        message=f'Your requisition {requisition.requisition_number} has been {action}d at {approval.get_approval_stage_display()} stage.',
                        link_url=f'/requisitions/{requisition.id}/'
                    )
                    
                    # Create audit log
                    AuditLog.objects.create(
                        user=request.user,
                        action='APPROVE' if action == 'approve' else 'REJECT',
                        model_name='RequisitionApproval',
                        object_id=str(approval.id),
                        object_repr=f'{requisition.requisition_number} - {approval.get_approval_stage_display()}',
                        changes={
                            'status': approval.status,
                            'comments': comments
                        }
                    )
                    
                    messages.success(request, f'Requisition {requisition.requisition_number} has been {action}d successfully.')
                
                return redirect('requisition_detail', pk=requisition_id)
                
        except Exception as e:
            messages.error(request, f'Error processing approval: {str(e)}')
            return redirect('process_approval', requisition_id=requisition_id)


def update_requisition_status(requisition):
    """Update requisition status based on approval workflow"""
    approvals = RequisitionApproval.objects.filter(
        requisition=requisition
    ).order_by('sequence')
    
    # Check if all approved
    if all(approval.status == 'APPROVED' for approval in approvals):
        requisition.status = 'APPROVED'
    # Check for any rejections
    elif any(approval.status == 'REJECTED' for approval in approvals):
        requisition.status = 'REJECTED'
    # Otherwise, update to current stage
    else:
        for approval in approvals:
            if approval.status == 'APPROVED':
                # Map approval stage to requisition status
                status_map = {
                    'HOD': 'HOD_APPROVED',
                    'FACULTY': 'FACULTY_APPROVED',
                    'BUDGET': 'BUDGET_APPROVED',
                    'PROCUREMENT': 'PROCUREMENT_APPROVED',
                }
                requisition.status = status_map.get(approval.approval_stage, requisition.status)
            elif approval.status == 'PENDING':
                break  # Stop at first pending
    
    requisition.save()


# ============================================================================
# BULK APPROVAL ACTION
# ============================================================================

@login_required
@require_http_methods(["POST"])
def bulk_approve(request):
    """Bulk approve multiple requisitions"""
    if request.user.role != 'ADMIN':
        return JsonResponse({
            'success': False,
            'message': 'Only administrators can perform bulk approvals.'
        }, status=403)
    
    approval_ids = request.POST.getlist('approval_ids[]')
    comments = request.POST.get('comments', 'Bulk approval')
    
    if not approval_ids:
        return JsonResponse({
            'success': False,
            'message': 'No approvals selected.'
        }, status=400)
    
    approved_count = 0
    failed_count = 0
    errors = []
    
    for approval_id in approval_ids:
        try:
            with transaction.atomic():
                approval = RequisitionApproval.objects.select_for_update().get(
                    pk=approval_id,
                    status='PENDING'
                )
                
                requisition = approval.requisition
                
                # Check budget if needed
                if approval.approval_stage == 'BUDGET':
                    budget_check = check_budget_availability(requisition)
                    if not budget_check['available']:
                        failed_count += 1
                        errors.append(f'{requisition.requisition_number}: {budget_check["message"]}')
                        continue
                
                # Approve
                approval.status = 'APPROVED'
                approval.comments = comments
                approval.approval_date = timezone.now()
                approval.save()
                
                update_requisition_status(requisition)
                
                # Notification
                Notification.objects.create(
                    user=requisition.requested_by,
                    notification_type='APPROVAL',
                    priority='HIGH',
                    title='Requisition Approved',
                    message=f'Your requisition {requisition.requisition_number} has been approved.',
                    link_url=f'/requisitions/{requisition.id}/'
                )
                
                approved_count += 1
                
        except RequisitionApproval.DoesNotExist:
            failed_count += 1
            errors.append(f'Approval {approval_id} not found or already processed')
        except Exception as e:
            failed_count += 1
            errors.append(f'Approval {approval_id}: {str(e)}')
    
    return JsonResponse({
        'success': True,
        'message': f'Bulk approval complete. Approved: {approved_count}, Failed: {failed_count}',
        'approved_count': approved_count,
        'failed_count': failed_count,
        'errors': errors
    })
    

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from decimal import Decimal

from .models import (
    User, Supplier, SupplierDocument, SupplierPerformance,
    Tender, TenderDocument, Bid, BidItem, BidDocument,
    PurchaseOrder, PurchaseOrderItem,
    Invoice, InvoiceItem, InvoiceDocument, Payment,
    GoodsReceivedNote, GRNItem, Notification
)
from .forms import (
    SupplierProfileForm, SupplierDocumentForm, BidForm, 
    BidItemFormSet, BidDocumentForm, InvoiceForm, 
    InvoiceItemFormSet, InvoiceDocumentForm
)


# ============================================================================
# SUPPLIER DASHBOARD
# ============================================================================

@login_required
def supplier_dashboard(request):
    """Supplier portal dashboard with key metrics and notifications"""
    
    # Get supplier profile
    try:
        supplier = request.user.supplier_profile
    except:
        messages.error(request, "Supplier profile not found. Please contact administrator.")
        return redirect('home')
    
    # Date ranges
    today = timezone.now().date()
    month_start = today.replace(day=1)
    year_start = today.replace(month=1, day=1)
    
    # Key Metrics
    # Active tenders
    active_tenders = Tender.objects.filter(
        status='PUBLISHED',
        closing_date__gte=timezone.now()
    ).count()
    
    # My bids
    my_bids = Bid.objects.filter(supplier=supplier)
    total_bids = my_bids.count()
    pending_bids = my_bids.filter(status='SUBMITTED').count()
    awarded_bids = my_bids.filter(status='AWARDED').count()
    
    # Purchase orders
    purchase_orders = PurchaseOrder.objects.filter(supplier=supplier)
    pending_pos = purchase_orders.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED']
    ).count()
    
    # Invoices
    invoices = Invoice.objects.filter(supplier=supplier)
    pending_invoices = invoices.filter(
        status__in=['SUBMITTED', 'VERIFYING', 'MATCHED', 'APPROVED']
    ).count()
    
    # Financial metrics
    total_invoiced = invoices.aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    
    total_paid = invoices.filter(
        status='PAID'
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    pending_payments = total_invoiced - total_paid
    
    # This month's orders
    month_orders = purchase_orders.filter(
        po_date__gte=month_start
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Performance rating
    avg_rating = SupplierPerformance.objects.filter(
        supplier=supplier
    ).aggregate(avg=Avg('overall_rating'))['avg'] or 0
    
    # Recent tenders (open for bidding)
    recent_tenders = Tender.objects.filter(
        Q(status='PUBLISHED') &
        Q(closing_date__gte=timezone.now())
    ).order_by('-publish_date')[:5]
    
    # Recent bids
    recent_bids = Bid.objects.filter(
        supplier=supplier
    ).select_related('tender').order_by('-submitted_at')[:5]
    
    # Active purchase orders
    active_pos = PurchaseOrder.objects.filter(
        supplier=supplier,
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY']
    ).order_by('-po_date')[:5]
    
    # Recent invoices
    recent_invoices = Invoice.objects.filter(
        supplier=supplier
    ).order_by('-created_at')[:5]
    
    # Notifications
    notifications = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).order_by('-created_at')[:10]
    
    # Documents expiring soon (30 days)
    expiring_docs = SupplierDocument.objects.filter(
        supplier=supplier,
        expiry_date__lte=today + timedelta(days=30),
        expiry_date__gte=today
    ).order_by('expiry_date')
    
    context = {
        'supplier': supplier,
        'active_tenders': active_tenders,
        'total_bids': total_bids,
        'pending_bids': pending_bids,
        'awarded_bids': awarded_bids,
        'pending_pos': pending_pos,
        'pending_invoices': pending_invoices,
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'pending_payments': pending_payments,
        'month_orders': month_orders,
        'avg_rating': avg_rating,
        'recent_tenders': recent_tenders,
        'recent_bids': recent_bids,
        'active_pos': active_pos,
        'recent_invoices': recent_invoices,
        'notifications': notifications,
        'expiring_docs': expiring_docs,
    }
    
    return render(request, 'supplier/dashboard.html', context)


# ============================================================================
# TENDERS & OPPORTUNITIES
# ============================================================================

@login_required
def supplier_tenders_list(request):
    """List all available tenders"""
    
    supplier = request.user.supplier_profile
    
    # Filter options
    status = request.GET.get('status', 'all')
    tender_type = request.GET.get('type', 'all')
    search = request.GET.get('search', '')
    
    # Base query
    tenders = Tender.objects.filter(
        status='PUBLISHED',
        closing_date__gte=timezone.now()
    )
    
    # Apply filters
    if tender_type != 'all':
        tenders = tenders.filter(tender_type=tender_type)
    
    if search:
        tenders = tenders.filter(
            Q(tender_number__icontains=search) |
            Q(title__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Check if supplier has bid
    for tender in tenders:
        tender.has_bid = Bid.objects.filter(
            tender=tender,
            supplier=supplier
        ).exists()
    
    # Pagination
    paginator = Paginator(tenders, 10)
    page_number = request.GET.get('page')
    tenders_page = paginator.get_page(page_number)
    
    context = {
        'tenders': tenders_page,
        'tender_types': Tender.TENDER_TYPES,
        'current_status': status,
        'current_type': tender_type,
        'search': search,
    }
    
    return render(request, 'supplier/tenders_list.html', context)


@login_required
def supplier_tender_detail(request, tender_id):
    """View tender details"""
    
    supplier = request.user.supplier_profile
    tender = get_object_or_404(Tender, id=tender_id)
    
    # Check if supplier has already bid
    existing_bid = Bid.objects.filter(
        tender=tender,
        supplier=supplier
    ).first()
    
    # Get tender documents
    documents = TenderDocument.objects.filter(tender=tender)
    
    # Get evaluation criteria
    criteria = tender.evaluation_criteria.all()
    
    # Calculate time remaining
    time_remaining = tender.closing_date - timezone.now()
    
    context = {
        'tender': tender,
        'existing_bid': existing_bid,
        'documents': documents,
        'criteria': criteria,
        'time_remaining': time_remaining,
        'can_bid': tender.closing_date > timezone.now() and not existing_bid,
    }
    
    return render(request, 'supplier/tender_detail.html', context)


# ============================================================================
# BIDS MANAGEMENT
# ============================================================================

@login_required
def supplier_bids_list(request):
    """List all supplier's bids"""
    
    supplier = request.user.supplier_profile
    
    # Filter options
    status = request.GET.get('status', 'all')
    search = request.GET.get('search', '')
    
    # Base query
    bids = Bid.objects.filter(supplier=supplier).select_related('tender')
    
    # Apply filters
    if status != 'all':
        bids = bids.filter(status=status)
    
    if search:
        bids = bids.filter(
            Q(bid_number__icontains=search) |
            Q(tender__tender_number__icontains=search) |
            Q(tender__title__icontains=search)
        )
    
    # Order by submission date
    bids = bids.order_by('-submitted_at')
    
    # Statistics
    stats = {
        'total': bids.count(),
        'submitted': bids.filter(status='SUBMITTED').count(),
        'evaluating': bids.filter(status='EVALUATING').count(),
        'qualified': bids.filter(status='QUALIFIED').count(),
        'awarded': bids.filter(status='AWARDED').count(),
        'rejected': bids.filter(status='REJECTED').count(),
    }
    
    # Pagination
    paginator = Paginator(bids, 15)
    page_number = request.GET.get('page')
    bids_page = paginator.get_page(page_number)
    
    context = {
        'bids': bids_page,
        'stats': stats,
        'status_choices': Bid.STATUS_CHOICES,
        'current_status': status,
        'search': search,
    }
    
    return render(request, 'supplier/bids_list.html', context)


@login_required
def supplier_bid_detail(request, bid_id):
    """View bid details"""
    
    supplier = request.user.supplier_profile
    bid = get_object_or_404(Bid, id=bid_id, supplier=supplier)
    
    # Get bid items
    items = BidItem.objects.filter(bid=bid).select_related('requisition_item')
    
    # Get bid documents
    documents = BidDocument.objects.filter(bid=bid)
    
    # Get evaluation if exists
    evaluation = bid.evaluations.first()
    
    context = {
        'bid': bid,
        'items': items,
        'documents': documents,
        'evaluation': evaluation,
    }
    
    return render(request, 'supplier/bid_detail.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum
from decimal import Decimal
from .models import (
    Tender, Bid, BidItem, BidDocument, Supplier,
    TechnicalEvaluationCriteria, BidTechnicalResponse,
    RequisitionItem
)


@login_required
def supplier_submit_bid(request, tender_id):
    """Submit a new bid for a tender with technical responses"""
    
    # Get supplier profile
    try:
        supplier = request.user.supplier_profile
    except Supplier.DoesNotExist:
        messages.error(request, "You must be registered as a supplier to submit bids.")
        return redirect('supplier_dashboard')
    
    # Check supplier status
    if supplier.status != 'APPROVED':
        messages.error(request, "Your supplier account is not approved. Please contact procurement office.")
        return redirect('supplier_dashboard')
    
    tender = get_object_or_404(Tender, id=tender_id)
    
    # Check if tender is still open
    if tender.closing_date < timezone.now():
        messages.error(request, "This tender has closed.")
        return redirect('supplier_tender_detail', tender_id=tender_id)
    
    # Check if tender is published
    if tender.status != 'PUBLISHED':
        messages.error(request, "This tender is not open for bidding.")
        return redirect('supplier_tender_detail', tender_id=tender_id)
    
    # Check if already submitted
    if Bid.objects.filter(tender=tender, supplier=supplier).exists():
        messages.error(request, "You have already submitted a bid for this tender.")
        return redirect('supplier_tender_detail', tender_id=tender_id)
    
    # Get requisition items and technical criteria
    requisition_items = tender.requisition.items.all()
    technical_criteria = tender.technical_criteria.all().order_by('sequence')
    
    # Get required documents from tender
    tender_documents = tender.documents.filter(is_mandatory=True)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Validate bid amount
                bid_amount = Decimal(request.POST.get('bid_amount', 0))
                if bid_amount <= 0:
                    raise ValueError("Bid amount must be greater than zero")
                
                # Validate delivery period
                delivery_period = int(request.POST.get('delivery_period_days', 0))
                if delivery_period < 1:
                    raise ValueError("Delivery period must be at least 1 day")
                
                # Validate validity period
                validity_period = int(request.POST.get('validity_period_days', 0))
                if validity_period < 30:
                    raise ValueError("Validity period must be at least 30 days")
                
                # Create the bid
                bid = Bid.objects.create(
                    tender=tender,
                    supplier=supplier,
                    bid_amount=bid_amount,
                    bid_bond_amount=Decimal(request.POST.get('bid_bond_amount', 0)),
                    validity_period_days=validity_period,
                    delivery_period_days=delivery_period,
                    status='SUBMITTED',
                    notes=request.POST.get('notes', ''),
                    preliminary_evaluation_status='PENDING',
                    technical_evaluation_status='PENDING',
                    financial_evaluation_status='PENDING'
                )
                
                # Calculate total from items
                total_calculated = Decimal('0')
                
                # Save bid items
                for i, req_item in enumerate(requisition_items):
                    quoted_unit_price = Decimal(request.POST.get(f'items-{i}-quoted_unit_price', 0))
                    
                    if quoted_unit_price <= 0:
                        raise ValueError(f"Please provide valid unit price for item {i+1}")
                    
                    quoted_total = quoted_unit_price * req_item.quantity
                    total_calculated += quoted_total
                    
                    delivery_days = int(request.POST.get(f'items-{i}-delivery_period_days', delivery_period))
                    warranty_months = int(request.POST.get(f'items-{i}-warranty_period_months', 0))
                    
                    BidItem.objects.create(
                        bid=bid,
                        requisition_item=req_item,
                        quoted_unit_price=quoted_unit_price,
                        quoted_total=quoted_total,
                        delivery_period_days=delivery_days,
                        brand=request.POST.get(f'items-{i}-brand', ''),
                        model=request.POST.get(f'items-{i}-model', ''),
                        specifications=request.POST.get(f'items-{i}-specifications', ''),
                        warranty_period_months=warranty_months,
                        notes=request.POST.get(f'items-{i}-notes', '')
                    )
                
                # Verify calculated total matches submitted bid amount
                if abs(total_calculated - bid_amount) > Decimal('0.01'):
                    raise ValueError(
                        f"Bid amount mismatch. Calculated: {total_calculated}, Submitted: {bid_amount}"
                    )
                
                # Save technical responses if tender requires technical evaluation
                if tender.requires_technical_evaluation and technical_criteria.exists():
                    for i, criterion in enumerate(technical_criteria):
                        response_text = request.POST.get(f'tech-{i}-response_text', '')
                        response_value = request.POST.get(f'tech-{i}-response_value', '')
                        supplier_remarks = request.POST.get(f'tech-{i}-remarks', '')
                        
                        # Handle file upload for this criterion
                        doc_file = request.FILES.get(f'tech-{i}-document')
                        
                        # Create technical response
                        tech_response = BidTechnicalResponse.objects.create(
                            bid=bid,
                            criterion=criterion,
                            response_text=response_text,
                            response_value=response_value,
                            supplier_remarks=supplier_remarks,
                            compliance_status='PENDING'
                        )
                        
                        # Save supporting document if uploaded
                        if doc_file:
                            tech_response.supporting_document = doc_file
                            tech_response.save()
                
                # Handle mandatory bid documents
                doc_count = 0
                uploaded_doc_types = set()
                
                # Get all uploaded files and their metadata
                documents = request.FILES.getlist('documents')
                doc_types = request.POST.getlist('doc_types')
                doc_names = request.POST.getlist('doc_names')
                doc_descriptions = request.POST.getlist('doc_descriptions')
                
                if len(documents) != len(doc_types) or len(documents) != len(doc_names):
                    raise ValueError("Document upload data is inconsistent")
                
                # Save each document
                for i, doc_file in enumerate(documents):
                    if doc_file:
                        doc_type = doc_types[i] if i < len(doc_types) else 'OTHER'
                        doc_name = doc_names[i] if i < len(doc_names) else doc_file.name
                        doc_desc = doc_descriptions[i] if i < len(doc_descriptions) else ''
                        
                        BidDocument.objects.create(
                            bid=bid,
                            document_type=doc_type,
                            document_name=doc_name,
                            file=doc_file,
                            description=doc_desc
                        )
                        
                        uploaded_doc_types.add(doc_type)
                        doc_count += 1
                
                # Verify all mandatory document types are uploaded
                required_doc_types = ['TECHNICAL', 'FINANCIAL', 'COMPLIANCE']
                missing_docs = [dt for dt in required_doc_types if dt not in uploaded_doc_types]
                
                if missing_docs:
                    raise ValueError(
                        f"Missing required documents: {', '.join(missing_docs)}"
                    )
                
                if doc_count < 3:
                    raise ValueError("Please upload all required documents (Technical, Financial, Compliance)")
                
                # Create notification for procurement team
                from .models import Notification, User
                procurement_users = User.objects.filter(role='PROCUREMENT', is_active=True)
                
                for user in procurement_users:
                    Notification.objects.create(
                        user=user,
                        notification_type='TENDER',
                        priority='MEDIUM',
                        title=f'New Bid Submitted - {tender.tender_number}',
                        message=f'{supplier.name} has submitted a bid for {tender.title}. '
                                f'Bid Amount: KES {bid_amount:,.2f}',
                        link_url=f'/procurement/bids/{bid.id}/'
                    )
                
                # Log the submission
                from .models import AuditLog
                AuditLog.objects.create(
                    user=request.user,
                    action='CREATE',
                    model_name='Bid',
                    object_id=str(bid.id),
                    object_repr=str(bid),
                    changes={
                        'bid_number': bid.bid_number,
                        'tender': tender.tender_number,
                        'supplier': supplier.name,
                        'bid_amount': str(bid_amount),
                        'items_count': requisition_items.count(),
                        'documents_count': doc_count,
                        'technical_responses': technical_criteria.count()
                    },
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
                )
                
                messages.success(
                    request, 
                    f"Bid {bid.bid_number} submitted successfully! "
                    f"You will be notified when the evaluation is complete."
                )
                return redirect('supplier_bid_detail', bid_id=bid.id)
                
        except ValueError as e:
            messages.error(request, f"Validation Error: {str(e)}")
        except Exception as e:
            messages.error(request, f"Error submitting bid: {str(e)}")
            # Log the error
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Bid submission error: {str(e)}", exc_info=True)
    
    context = {
        'tender': tender,
        'requisition_items': requisition_items,
        'technical_criteria': technical_criteria,
        'tender_documents': tender_documents,
        'supplier': supplier,
    }
    
    return render(request, 'supplier/submit_bid.html', context)




@login_required
def supplier_awarded_contracts(request):
    """List awarded contracts/bids"""
    
    supplier = request.user.supplier_profile
    
    # Get awarded bids with purchase orders prefetched
    awarded_bids = Bid.objects.filter(
        supplier=supplier,
        status='AWARDED'
    ).select_related(
        'tender',
        'tender__requisition',
        'tender__requisition__department'
    ).prefetch_related(
        'purchase_order'  # This will prefetch the related PurchaseOrder
    ).order_by('-submitted_at')
    
    context = {
        'awarded_bids': awarded_bids,
    }
    
    return render(request, 'supplier/awarded_contracts.html', context)


# ============================================================================
# PURCHASE ORDERS
# ============================================================================

@login_required
def supplier_purchase_orders_list(request):
    """List all purchase orders"""
    
    supplier = request.user.supplier_profile
    
    # Filter options
    status = request.GET.get('status', 'all')
    search = request.GET.get('search', '')
    
    # Base query
    pos = PurchaseOrder.objects.filter(supplier=supplier)
    
    # Apply filters
    if status != 'all':
        pos = pos.filter(status=status)
    
    if search:
        pos = pos.filter(
            Q(po_number__icontains=search) |
            Q(requisition__requisition_number__icontains=search)
        )
    
    pos = pos.order_by('-po_date')
    
    # Statistics
    stats = {
        'total': pos.count(),
        'pending': pos.filter(status__in=['APPROVED', 'SENT']).count(),
        'acknowledged': pos.filter(status='ACKNOWLEDGED').count(),
        'partial': pos.filter(status='PARTIAL_DELIVERY').count(),
        'delivered': pos.filter(status='DELIVERED').count(),
        'total_value': pos.aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
    }
    
    # Pagination
    paginator = Paginator(pos, 15)
    page_number = request.GET.get('page')
    pos_page = paginator.get_page(page_number)
    
    context = {
        'pos': pos_page,
        'stats': stats,
        'status_choices': PurchaseOrder.STATUS_CHOICES,
        'current_status': status,
        'search': search,
    }
    
    return render(request, 'supplier/purchase_orders_list.html', context)


@login_required
def supplier_purchase_order_detail(request, po_id):
    """View purchase order details"""
    
    supplier = request.user.supplier_profile
    po = get_object_or_404(PurchaseOrder, id=po_id, supplier=supplier)
    
    # Get PO items
    items = PurchaseOrderItem.objects.filter(purchase_order=po)
    
    # Get amendments
    amendments = po.amendments.all()
    
    # Get related GRNs
    grns = GoodsReceivedNote.objects.filter(purchase_order=po)
    
    # Get related invoices
    invoices = Invoice.objects.filter(purchase_order=po)
    
    context = {
        'po': po,
        'items': items,
        'amendments': amendments,
        'grns': grns,
        'invoices': invoices,
        'can_acknowledge': po.status == 'SENT',
    }
    
    return render(request, 'supplier/purchase_order_detail.html', context)


@login_required
def supplier_acknowledge_po(request, po_id):
    """Acknowledge receipt of purchase order"""
    
    supplier = request.user.supplier_profile
    po = get_object_or_404(PurchaseOrder, id=po_id, supplier=supplier)
    
    if po.status != 'SENT':
        messages.error(request, "This PO cannot be acknowledged.")
        return redirect('supplier_purchase_order_detail', po_id=po_id)
    
    if request.method == 'POST':
        po.status = 'ACKNOWLEDGED'
        po.acknowledged_at = timezone.now()
        po.save()
        
        messages.success(request, f"Purchase Order {po.po_number} acknowledged successfully!")
        return redirect('supplier_purchase_order_detail', po_id=po_id)
    
    return render(request, 'supplier/acknowledge_po.html', {'po': po})


@login_required
def supplier_pending_orders(request):
    """List pending purchase orders requiring action"""
    
    supplier = request.user.supplier_profile
    
    pending_pos = PurchaseOrder.objects.filter(
        supplier=supplier,
        status__in=['SENT', 'ACKNOWLEDGED']
    ).order_by('delivery_date')
    
    context = {
        'pending_pos': pending_pos,
    }
    
    return render(request, 'supplier/pending_orders.html', context)


@login_required
def supplier_deliveries(request):
    """Track deliveries and GRNs"""
    
    supplier = request.user.supplier_profile
    
    # Get all GRNs for supplier's POs
    grns = GoodsReceivedNote.objects.filter(
        purchase_order__supplier=supplier
    ).select_related('purchase_order', 'store').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(grns, 15)
    page_number = request.GET.get('page')
    grns_page = paginator.get_page(page_number)
    
    context = {
        'grns': grns_page,
    }
    
    return render(request, 'supplier/deliveries.html', context)


@login_required
def supplier_completed_orders(request):
    """List completed orders"""
    
    supplier = request.user.supplier_profile
    
    completed_pos = PurchaseOrder.objects.filter(
        supplier=supplier,
        status__in=['DELIVERED', 'CLOSED']
    ).order_by('-po_date')
    
    # Pagination
    paginator = Paginator(completed_pos, 15)
    page_number = request.GET.get('page')
    pos_page = paginator.get_page(page_number)
    
    context = {
        'pos': pos_page,
    }
    
    return render(request, 'supplier/completed_orders.html', context)


# ============================================================================
# INVOICES & PAYMENTS
# ============================================================================

@login_required
def supplier_invoices_list(request):
    """List all supplier invoices"""
    
    supplier = request.user.supplier_profile
    
    # Filter options
    status = request.GET.get('status', 'all')
    search = request.GET.get('search', '')
    
    # Base query
    invoices = Invoice.objects.filter(supplier=supplier)
    
    # Apply filters
    if status != 'all':
        invoices = invoices.filter(status=status)
    
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(supplier_invoice_number__icontains=search) |
            Q(purchase_order__po_number__icontains=search)
        )
    
    invoices = invoices.order_by('-created_at')
    
    # Statistics
    stats = {
        'total': invoices.count(),
        'draft': invoices.filter(status='DRAFT').count(),
        'submitted': invoices.filter(status='SUBMITTED').count(),
        'approved': invoices.filter(status='APPROVED').count(),
        'paid': invoices.filter(status='PAID').count(),
        'total_invoiced': invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
        'total_paid': invoices.filter(status='PAID').aggregate(total=Sum('total_amount'))['total'] or Decimal('0'),
    }
    
    # Pagination
    paginator = Paginator(invoices, 15)
    page_number = request.GET.get('page')
    invoices_page = paginator.get_page(page_number)
    
    context = {
        'invoices': invoices_page,
        'stats': stats,
        'status_choices': Invoice.STATUS_CHOICES,
        'current_status': status,
        'search': search,
    }
    
    return render(request, 'supplier/invoices_list.html', context)


@login_required
def supplier_invoice_detail(request, invoice_id):
    """View invoice details"""
    
    supplier = request.user.supplier_profile
    invoice = get_object_or_404(Invoice, id=invoice_id, supplier=supplier)
    
    # Get invoice items
    items = InvoiceItem.objects.filter(invoice=invoice)
    
    # Get invoice documents
    documents = InvoiceDocument.objects.filter(invoice=invoice)
    
    # Get payments
    payments = Payment.objects.filter(invoice=invoice)
    
    context = {
        'invoice': invoice,
        'items': items,
        'documents': documents,
        'payments': payments,
    }
    
    return render(request, 'supplier/invoice_detail.html', context)


@login_required
def supplier_submit_invoice(request, po_id=None):
    """Submit a new invoice"""
    
    supplier = request.user.supplier_profile
    
    # If PO specified, get it
    po = None
    if po_id:
        po = get_object_or_404(PurchaseOrder, id=po_id, supplier=supplier)
    
    if request.method == 'POST':
        invoice_form = InvoiceForm(request.POST, supplier=supplier)
        item_formset = InvoiceItemFormSet(request.POST, prefix='items')
        
        if invoice_form.is_valid() and item_formset.is_valid():
            # Create invoice
            invoice = invoice_form.save(commit=False)
            invoice.supplier = supplier
            invoice.status = 'SUBMITTED'
            invoice.submitted_by = request.user
            invoice.save()
            
            # Save items
            items = item_formset.save(commit=False)
            total = Decimal('0')
            tax_total = Decimal('0')
            
            for item in items:
                item.invoice = invoice
                item.save()
                total += item.total_price
                tax_total += item.tax_amount
            
            # Update invoice totals
            invoice.subtotal = total
            invoice.tax_amount = tax_total
            invoice.total_amount = total + tax_total + invoice.other_charges
            invoice.save()
            
            # Handle document uploads
            documents = request.FILES.getlist('documents')
            doc_names = request.POST.getlist('doc_names')
            
            for i, doc_file in enumerate(documents):
                InvoiceDocument.objects.create(
                    invoice=invoice,
                    document_name=doc_names[i] if i < len(doc_names) else doc_file.name,
                    file=doc_file
                )
            
            messages.success(request, f"Invoice {invoice.invoice_number} submitted successfully!")
            return redirect('supplier_invoice_detail', invoice_id=invoice.id)
    else:
        initial = {}
        if po:
            initial = {
                'purchase_order': po,
                'due_date': timezone.now().date() + timedelta(days=30)
            }
        
        invoice_form = InvoiceForm(initial=initial, supplier=supplier)
        
        # If PO specified, pre-fill items
        if po:
            po_items = PurchaseOrderItem.objects.filter(purchase_order=po)
            item_formset = InvoiceItemFormSet(
                prefix='items',
                queryset=InvoiceItem.objects.none(),
                initial=[{
                    'po_item': item,
                    'description': item.item_description,
                    'quantity': item.quantity,
                    'unit_price': item.unit_price,
                } for item in po_items]
            )
        else:
            item_formset = InvoiceItemFormSet(prefix='items', queryset=InvoiceItem.objects.none())
    
    context = {
        'invoice_form': invoice_form,
        'item_formset': item_formset,
        'po': po,
    }
    
    return render(request, 'supplier/submit_invoice.html', context)


@login_required
def supplier_payments(request):
    """View payment information"""
    
    supplier = request.user.supplier_profile
    
    # Get all payments for supplier's invoices
    payments = Payment.objects.filter(
        invoice__supplier=supplier
    ).select_related('invoice').order_by('-payment_date')
    
    # Statistics
    stats = {
        'total_paid': payments.filter(status='COMPLETED').aggregate(
            total=Sum('payment_amount')
        )['total'] or Decimal('0'),
        'pending': payments.filter(status='PENDING').aggregate(
            total=Sum('payment_amount')
        )['total'] or Decimal('0'),
        'this_month': payments.filter(
            status='COMPLETED',
            payment_date__gte=timezone.now().date().replace(day=1)
        ).aggregate(total=Sum('payment_amount'))['total'] or Decimal('0'),
    }
    
    # Pagination
    paginator = Paginator(payments, 15)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)
    
    context = {
        'payments': payments_page,
        'stats': stats,
    }
    
    return render(request, 'supplier/payments.html', context)


@login_required
def supplier_payment_history(request):
    """Detailed payment history with filters"""
    
    supplier = request.user.supplier_profile
    
    # Date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status', 'all')
    
    # Base query
    payments = Payment.objects.filter(
        invoice__supplier=supplier
    ).select_related('invoice', 'invoice__purchase_order')
    
    # Apply filters
    if start_date:
        payments = payments.filter(payment_date__gte=start_date)
    if end_date:
        payments = payments.filter(payment_date__lte=end_date)
    if status != 'all':
        payments = payments.filter(status=status)
    
    payments = payments.order_by('-payment_date')
    
    # Calculate totals
    totals = payments.aggregate(
        total=Sum('payment_amount'),
        count=Count('id')
    )
    
    # Pagination
    paginator = Paginator(payments, 20)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)
    
    context = {
        'payments': payments_page,
        'totals': totals,
        'status_choices': Payment.PAYMENT_STATUS,
        'current_status': status,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'supplier/payment_history.html', context)


# ============================================================================
# SUPPLIER PROFILE & DOCUMENTS
# ============================================================================

@login_required
def supplier_company_profile(request):
    """View and edit company profile"""
    
    supplier = request.user.supplier_profile
    
    if request.method == 'POST':
        form = SupplierProfileForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, "Company profile updated successfully!")
            return redirect('supplier_company_profile')
    else:
        form = SupplierProfileForm(instance=supplier)
    
    # Performance statistics
    performance_stats = SupplierPerformance.objects.filter(
        supplier=supplier
    ).aggregate(
        avg_quality=Avg('quality_rating'),
        avg_delivery=Avg('delivery_rating'),
        avg_service=Avg('service_rating'),
        avg_overall=Avg('overall_rating'),
        count=Count('id')
    )
    
    context = {
        'form': form,
        'supplier': supplier,
        'performance_stats': performance_stats,
    }
    
    return render(request, 'supplier/company_profile.html', context)

@login_required
def supplier_documents(request):
    """Manage supplier documents"""
    
    supplier = request.user.supplier_profile
    
    # Get all documents
    documents = SupplierDocument.objects.filter(
        supplier=supplier
    ).order_by('-uploaded_at')
    
    # Check for expiring documents
    today = timezone.now().date()
    expiring_threshold = today + timedelta(days=30)
    
    expiring_soon = documents.filter(
        expiry_date__lte=expiring_threshold,
        expiry_date__gte=today
    )
    
    expired = documents.filter(expiry_date__lt=today)
    
    # Count verified documents
    verified_count = documents.filter(is_verified=True).count()
    
    # Check which document types exist (for checklist)
    has_registration = documents.filter(document_type='REGISTRATION').exists()
    has_tax = documents.filter(document_type='TAX').exists()
    has_bank = documents.filter(document_type='BANK').exists()
    has_license = documents.filter(document_type='LICENSE').exists()
    has_insurance = documents.filter(document_type='INSURANCE').exists()
    
    if request.method == 'POST':
        form = SupplierDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.supplier = supplier
            doc.save()
            messages.success(request, "Document uploaded successfully!")
            return redirect('supplier_documents')
    else:
        form = SupplierDocumentForm()
    
    context = {
        'documents': documents,
        'expiring_soon': expiring_soon,
        'expired': expired,
        'form': form,
        'today': today,
        'expiring_threshold': expiring_threshold,
        'verified_count': verified_count,
        'has_registration': has_registration,
        'has_tax': has_tax,
        'has_bank': has_bank,
        'has_license': has_license,
        'has_insurance': has_insurance,
    }
    
    return render(request, 'supplier/documents.html', context)

@login_required
def supplier_certifications(request):
    """View certifications and compliance status"""
    
    supplier = request.user.supplier_profile
    
    # Get certification documents
    certifications = SupplierDocument.objects.filter(
        supplier=supplier,
        document_type__in=['REGISTRATION', 'LICENSE', 'INSURANCE']
    ).order_by('-uploaded_at')

    # Calculate dates
    today = timezone.now().date()
    expiring_threshold = today + timedelta(days=30)
    
    # Compliance status
    compliance = {
        'tax_compliant': supplier.tax_compliance_expiry and supplier.tax_compliance_expiry >= today,
        'registration_valid': supplier.registration_expiry and supplier.registration_expiry >= today,
        'documents_verified': certifications.filter(is_verified=True).count(),
        'total_documents': certifications.count(),
    }
    
    # Check which certificate types exist
    has_license = certifications.filter(document_type='LICENSE').exists()
    has_insurance = certifications.filter(document_type='INSURANCE').exists()

    context = {
        'certifications': certifications,
        'compliance': compliance,
        'supplier': supplier,
        'today': today,
        'expiring_threshold': expiring_threshold,
        'has_license': has_license,
        'has_insurance': has_insurance,
    }

    return render(request, 'supplier/certifications.html', context)

#============================================================================
#SUPPORT & HELP
#============================================================================
@login_required
def supplier_help_center(request):
    """Help center and FAQs"""
    faqs = [
        {
            'question': 'How do I submit a bid?',
            'answer': 'Navigate to "Available Tenders", select a tender, and click "Submit Bid". Fill in all required information and upload necessary documents.'
        },
        {
            'question': 'When will I receive payment?',
            'answer': 'Payments are typically processed within 30 days of invoice approval. You can track payment status in the "Payments" section.'
        },
        # Add more FAQs
    ]

    context = {
        'faqs': faqs,
    }

    return render(request, 'supplier/help_center.html', context)

@login_required
def supplier_contact_support(request):
    """Contact support form"""
    if request.method == 'POST':
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        
        # Send email or create support ticket
        # Implementation depends on your requirements
        
        messages.success(request, "Your message has been sent. We'll get back to you soon!")
        return redirect('supplier_dashboard')

    return render(request, 'supplier/contact_support.html')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum, F
from django.utils import timezone
from datetime import timedelta
from django.http import JsonResponse
from django.core.paginator import Paginator
from decimal import Decimal

from .models import (
    Requisition, RequisitionItem, RequisitionAttachment,
    RequisitionApproval, Department, ItemCategory, Item,
    Budget, BudgetYear, User
)
from .forms import (
    RequisitionForm, RequisitionItemFormSet, 
    RequisitionAttachmentFormSet, RequisitionFilterForm
)


@login_required
def staff_dashboard(request):
    """Staff dashboard with overview of requisitions and statistics"""
    
    user = request.user
    
    # Get requisitions created by this user
    user_requisitions = Requisition.objects.filter(requested_by=user)
    
    # Calculate statistics
    total_requisitions = user_requisitions.count()
    draft_count = user_requisitions.filter(status='DRAFT').count()
    submitted_count = user_requisitions.filter(status='SUBMITTED').count()
    pending_count = user_requisitions.filter(
        status__in=['SUBMITTED', 'HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED', 'PROCUREMENT_APPROVED']
    ).count()
    approved_count = user_requisitions.filter(status='APPROVED').count()
    rejected_count = user_requisitions.filter(status='REJECTED').count()
    
    # Recent requisitions
    recent_requisitions = user_requisitions.order_by('-created_at')[:5]
    
    # Monthly statistics
    today = timezone.now().date()
    current_month_start = today.replace(day=1)
    
    monthly_requisitions = user_requisitions.filter(
        created_at__gte=current_month_start
    ).count()
    
    # Total estimated amount for approved requisitions this year
    total_approved_amount = user_requisitions.filter(
        status='APPROVED',
        created_at__year=today.year
    ).aggregate(
        total=Sum('estimated_amount')
    )['total'] or Decimal('0')
    
    # Requisitions by priority
    urgent_count = user_requisitions.filter(
        priority='URGENT',
        status__in=['SUBMITTED', 'HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED', 'PROCUREMENT_APPROVED']
    ).count()
    
    context = {
        'total_requisitions': total_requisitions,
        'draft_count': draft_count,
        'submitted_count': submitted_count,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'recent_requisitions': recent_requisitions,
        'monthly_requisitions': monthly_requisitions,
        'total_approved_amount': total_approved_amount,
        'urgent_count': urgent_count,
        'user': user,
    }
    
    return render(request, 'staff/dashboard.html', context)


@login_required
def staff_requisitions_list(request):
    """List all requisitions for the staff member"""
    
    user = request.user
    
    # Get requisitions created by this user
    requisitions = Requisition.objects.filter(requested_by=user)
    
    # Apply filters from GET parameters
    status_filter = request.GET.get('status')
    if status_filter:
        requisitions = requisitions.filter(status=status_filter)
    
    priority_filter = request.GET.get('priority')
    if priority_filter:
        requisitions = requisitions.filter(priority=priority_filter)
    
    search_query = request.GET.get('search')
    if search_query:
        requisitions = requisitions.filter(
            Q(requisition_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(justification__icontains=search_query)
        )
    
    # Date filters
    date_from = request.GET.get('date_from')
    if date_from:
        requisitions = requisitions.filter(created_at__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        requisitions = requisitions.filter(created_at__lte=date_to)
    
    # Order by most recent
    requisitions = requisitions.select_related(
        'department', 'budget'
    ).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(requisitions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter form
    filter_form = RequisitionFilterForm(request.GET)
    
    context = {
        'requisitions': page_obj,
        'filter_form': filter_form,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'search_query': search_query,
        'total_count': requisitions.count(),
    }
    
    return render(request, 'staff/requisitions_list.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum
from decimal import Decimal
from .models import (
    Requisition, RequisitionItem, RequisitionAttachment,
    RequisitionApproval, ProcurementPlan, ProcurementPlanItem,
    ProcurementPlanAmendment, Budget
)
from .forms import (
    RequisitionForm, RequisitionItemFormSet, 
    RequisitionAttachmentFormSet
)


"""
Fixed views.py for staff requisition creation
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q
from decimal import Decimal
import json

from .models import (
    Requisition, RequisitionItem, RequisitionAttachment,
    ProcurementPlanItem, BudgetYear, Budget, Item, ItemCategory
)
from .forms import (
    RequisitionForm, RequisitionItemFormSet, RequisitionAttachmentFormSet
)


"""
Fixed views.py for staff requisition creation
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q
from decimal import Decimal
import json

from .models import (
    Requisition, RequisitionItem, RequisitionAttachment,
    ProcurementPlanItem, BudgetYear, Budget, Item, ItemCategory
)
from .forms import (
    RequisitionForm, RequisitionItemFormSet, RequisitionAttachmentFormSet
)


@login_required
def staff_requisition_create(request):
    """Create a new requisition with plan enforcement and enhanced UX"""
    
    # ========================================================================
    # POST: Handle form submission
    # ========================================================================
    if request.method == 'POST':
        form = RequisitionForm(request.POST, user=request.user)
        formset = RequisitionItemFormSet(request.POST)
        attachment_formset = RequisitionAttachmentFormSet(request.POST, request.FILES)
        
        if form.is_valid() and formset.is_valid() and attachment_formset.is_valid():
            requisition = form.save(commit=False)
            requisition.requested_by = request.user
            
            # Set department from user's department
            if request.user.department:
                requisition.department = request.user.department
            else:
                messages.error(request, 'You must be assigned to a department to create requisitions.')
                return redirect('staff_requisitions_list')
            
            requisition.status = 'DRAFT'
            
            # PLAN ENFORCEMENT LOGIC
            is_planned = form.cleaned_data.get('is_planned', True)
            procurement_plan_item = form.cleaned_data.get('procurement_plan_item')
            is_emergency = form.cleaned_data.get('is_emergency', False)
            
            # Validate plan compliance
            if is_planned and not procurement_plan_item:
                messages.error(
                    request, 
                    'Please select a procurement plan item for planned requisitions.'
                )
                context = _get_context_data(request.user, form, formset, attachment_formset)
                return render(request, 'staff/requisition_form.html', context)
            
            # Validate unplanned requisitions
            if not is_planned:
                if not is_emergency:
                    messages.error(
                        request, 
                        'Unplanned requisitions must be marked as emergency.'
                    )
                    context = _get_context_data(request.user, form, formset, attachment_formset)
                    return render(request, 'staff/requisition_form.html', context)
                
                if not requisition.emergency_justification:
                    messages.error(
                        request, 
                        'Emergency justification is required for unplanned requisitions.'
                    )
                    context = _get_context_data(request.user, form, formset, attachment_formset)
                    return render(request, 'staff/requisition_form.html', context)
                
                # Mark as requiring plan amendment
                requisition.requires_plan_amendment = True
            
            # CALCULATE ESTIMATED AMOUNT BEFORE SAVING
            # Process items first to get total
            items = formset.save(commit=False)
            total_estimated = Decimal('0')
            
            for item in items:
                # Calculate item total if not already set
                item_qty = Decimal(str(item.quantity))
                item_price = Decimal(str(item.estimated_unit_price))
                item.estimated_total = item_qty * item_price
                total_estimated += item.estimated_total
            
            # Set estimated amount BEFORE first save
            requisition.estimated_amount = total_estimated
            
            # Now save the requisition
            requisition.save()
            
            # Save requisition items with requisition reference
            for item in items:
                item.requisition = requisition
                item.save()
            
            # VALIDATE AGAINST PLAN ITEM
            if is_planned and procurement_plan_item:
                # Check quantity (if applicable)
                total_qty = sum(Decimal(str(i.quantity)) for i in items)
                
                # Check budget
                if total_estimated > procurement_plan_item.remaining_budget:
                    messages.warning(
                        request,
                        f'Warning: Estimated amount (KES {total_estimated:,.2f}) exceeds '
                        f'remaining plan budget (KES {procurement_plan_item.remaining_budget:,.2f}). '
                        f'This may require additional approvals.'
                    )
                
                # Update plan item tracking
                procurement_plan_item.quantity_requisitioned += total_qty
                procurement_plan_item.amount_committed += total_estimated
                procurement_plan_item.save()
            
            # No need to save requisition again - estimated_amount already set
            
            # Save attachments
            attachments = attachment_formset.save(commit=False)
            for attachment in attachments:
                attachment.requisition = requisition
                attachment.uploaded_by = request.user
                attachment.save()
            
            messages.success(
                request, 
                f'Requisition {requisition.requisition_number} created successfully!'
            )
            
            if not is_planned:
                messages.info(
                    request,
                    'This is an unplanned/emergency requisition. '
                    'It will require HOD emergency approval and procurement plan amendment.'
                )
            
            return redirect('staff_requisition_detail', pk=requisition.pk)
        else:
            # Show validation errors
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
            if formset.errors:
                messages.error(request, 'Please correct errors in the items section.')
            if attachment_formset.errors:
                messages.error(request, 'Please correct errors in the attachments section.')
    
    # ========================================================================
    # GET: Display empty form
    # ========================================================================
    else:
        form = RequisitionForm(user=request.user)
        formset = RequisitionItemFormSet()
        attachment_formset = RequisitionAttachmentFormSet()
    
    # Get context data
    context = _get_context_data(request.user, form, formset, attachment_formset)
    
    return render(request, 'staff/requisition_form.html', context)


def _get_context_data(user, form, formset, attachment_formset):
    """Helper function to get context data for the template"""
    
    # Get available procurement plan items for the user's department
    available_plan_items = []
    current_budget_year = None
    debug_info = {
        'has_department': bool(user.department),
        'department': str(user.department) if user.department else None,
        'budget_year': None,
        'total_plans': 0,
        'active_plans': 0,
        'plan_items_count': 0,
    }
    
    if user.department:
        current_budget_year = BudgetYear.objects.filter(is_active=True).first()
        debug_info['budget_year'] = str(current_budget_year) if current_budget_year else None
        
        if current_budget_year:
            # Count all plans for this department and year
            from .models import ProcurementPlan
            all_plans = ProcurementPlan.objects.filter(
                department=user.department,
                budget_year=current_budget_year
            )
            debug_info['total_plans'] = all_plans.count()
            
            # Count active plans
            active_plans = all_plans.filter(status='ACTIVE')
            debug_info['active_plans'] = active_plans.count()
            
            # FIXED: Changed from 'APPROVED' to 'ACTIVE'
            available_plan_items = ProcurementPlanItem.objects.filter(
                procurement_plan__department=user.department,
                procurement_plan__budget_year=current_budget_year,
                procurement_plan__status__in=['ACTIVE', 'APPROVED'],  # Allow both ACTIVE and APPROVED
                status='PLANNED'
            ).select_related('procurement_plan', 'budget', 'item').order_by(
                'planned_quarter', 'sequence'
            )
            
            debug_info['plan_items_count'] = available_plan_items.count()
            
            # Add debug message if no items found
            if not available_plan_items:
                print(f"\n{'='*60}")
                print(f"DEBUG: No procurement plan items found")
                print(f"Department: {user.department.name} (ID: {user.department.id})")
                print(f"Budget Year: {current_budget_year.name} (ID: {current_budget_year.id})")
                print(f"Total Plans: {debug_info['total_plans']}")
                print(f"Active Plans: {debug_info['active_plans']}")
                
                # Show all plans with their status
                if all_plans.exists():
                    print("\nAll Plans for this department:")
                    for plan in all_plans:
                        items_count = plan.items.count()
                        print(f"  - {plan.plan_number}: {plan.status} ({items_count} items)")
                else:
                    print("\nNo plans exist for this department and year")
                print(f"{'='*60}\n")
    
    context = {
        'form': form,
        'formset': formset,
        'attachment_formset': attachment_formset,
        'page_title': 'Create New Requisition',
        'available_plan_items': available_plan_items,
        'current_budget_year': current_budget_year,
        'debug_info': debug_info,  # Add debug info to context
    }
    
    return context


@login_required
def api_search_items(request):
    """
    API endpoint to search items for requisition
    Returns JSON with item details
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'items': []})
    
    items = Item.objects.filter(
        Q(name__icontains=query) | 
        Q(code__icontains=query) | 
        Q(description__icontains=query),
        is_active=True
    ).select_related('category')[:20]
    
    items_data = [{
        'id': str(item.id),
        'code': item.code,
        'name': item.name,
        'description': item.description,
        'category': item.category.name,
        'unit_of_measure': item.unit_of_measure,
        'standard_price': str(item.standard_price) if item.standard_price else '0.00',
        'specifications': item.specifications
    } for item in items]
    
    return JsonResponse({'items': items_data})


@login_required
def api_get_plan_item_details(request, plan_item_id):
    """
    API endpoint to get procurement plan item details
    """
    try:
        plan_item = ProcurementPlanItem.objects.select_related(
            'procurement_plan', 'budget', 'item'
        ).get(id=plan_item_id)
        
        # Check if user has access
        if plan_item.procurement_plan.department != request.user.department:
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        data = {
            'id': str(plan_item.id),
            'description': plan_item.description,
            'specifications': plan_item.specifications,
            'quantity': str(plan_item.quantity),
            'unit_of_measure': plan_item.unit_of_measure,
            'estimated_cost': str(plan_item.estimated_cost),
            'remaining_budget': str(plan_item.remaining_budget),
            'remaining_quantity': str(plan_item.remaining_quantity),
            'planned_quarter': plan_item.get_planned_quarter_display(),
            'procurement_method': plan_item.get_procurement_method_display(),
            'budget_code': f"{plan_item.budget.category.code} - {plan_item.budget.category.name}" if plan_item.budget else '',
            'item_type': plan_item.get_item_type_display(),
        }
        
        return JsonResponse(data)
        
    except ProcurementPlanItem.DoesNotExist:
        return JsonResponse({'error': 'Plan item not found'}, status=404)


@login_required
def api_validate_budget(request):
    """
    API endpoint to validate if budget has sufficient balance
    """
    budget_id = request.GET.get('budget_id')
    amount = request.GET.get('amount', '0')
    
    try:
        budget = Budget.objects.get(id=budget_id)
        amount_decimal = Decimal(amount)
        
        available = budget.available_balance
        sufficient = available >= amount_decimal
        
        return JsonResponse({
            'sufficient': sufficient,
            'available_balance': str(available),
            'allocated_amount': str(budget.allocated_amount),
            'committed_amount': str(budget.committed_amount),
            'actual_spent': str(budget.actual_spent),
            'budget_code': f"{budget.category.code} - {budget.category.name}",
        })
        
    except Budget.DoesNotExist:
        return JsonResponse({'error': 'Budget not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def staff_requisition_detail(request, pk):
    """View requisition details with plan information"""
    
    requisition = get_object_or_404(
        Requisition.objects.select_related(
            'department', 'budget', 'requested_by',
            'procurement_plan_item__procurement_plan',
            'plan_amendment'
        ).prefetch_related('items', 'attachments', 'approvals'),
        pk=pk
    )
    
    # Check permissions - only creator can view their own requisitions
    if requisition.requested_by != request.user:
        messages.error(request, 'You do not have permission to view this requisition.')
        return redirect('staff_requisitions_list')
    
    # Get items and attachments
    items = requisition.items.all()
    attachments = requisition.attachments.all()
    
    # Get approval history
    approvals = requisition.approvals.select_related('approver').order_by('sequence')
    
    # Calculate totals
    total_items = items.count()
    
    # Plan compliance check
    plan_warnings = []
    if requisition.is_planned and requisition.procurement_plan_item:
        plan_item = requisition.procurement_plan_item
        
        if requisition.estimated_amount > plan_item.estimated_cost:
            plan_warnings.append(
                f'Requisition amount exceeds planned budget by '
                f'KES {(requisition.estimated_amount - plan_item.estimated_cost):,.2f}'
            )
        
        if plan_item.remaining_budget < 0:
            plan_warnings.append('Procurement plan item is over budget')
    
    context = {
        'requisition': requisition,
        'items': items,
        'attachments': attachments,
        'approvals': approvals,
        'total_items': total_items,
        'plan_warnings': plan_warnings,
        'can_edit': requisition.status == 'DRAFT',
        'can_submit': requisition.status == 'DRAFT' and items.exists(),
        'can_cancel': requisition.status in ['DRAFT', 'SUBMITTED'],
    }
    
    return render(request, 'staff/requisition_detail.html', context)

@login_required
def staff_requisition_edit(request, pk):
    """Edit a requisition (only if in DRAFT status)"""
    
    requisition = get_object_or_404(
        Requisition.objects.select_related(
            'procurement_plan_item__procurement_plan'
        ),
        pk=pk
    )
    
    # Check permissions
    if requisition.requested_by != request.user:
        messages.error(request, 'You can only edit your own requisitions.')
        return redirect('staff_requisitions_list')
    
    if requisition.status != 'DRAFT':
        messages.error(request, 'You can only edit draft requisitions.')
        return redirect('staff_requisition_detail', pk=pk)
    
    if request.method == 'POST':
        form = RequisitionForm(request.POST, instance=requisition, user=request.user)
        formset = RequisitionItemFormSet(request.POST, instance=requisition)
        attachment_formset = RequisitionAttachmentFormSet(
            request.POST, 
            request.FILES, 
            instance=requisition
        )
        
        if form.is_valid() and formset.is_valid() and attachment_formset.is_valid():
            # Track old values for plan item updates
            old_plan_item = requisition.procurement_plan_item
            old_total = requisition.estimated_amount
            
            requisition = form.save()
            
            # Save items and recalculate total
            items = formset.save()
            total_estimated = sum(
                item.estimated_total for item in requisition.items.all()
            )
            requisition.estimated_amount = total_estimated
            
            # Update plan item tracking
            new_plan_item = requisition.procurement_plan_item
            
            # Revert old plan item
            if old_plan_item:
                old_plan_item.amount_committed -= old_total
                old_plan_item.save()
            
            # Update new plan item
            if new_plan_item and requisition.is_planned:
                new_plan_item.amount_committed += total_estimated
                new_plan_item.save()
                
                # Validate budget
                if total_estimated > new_plan_item.remaining_budget:
                    messages.warning(
                        request,
                        f'Warning: Amount exceeds remaining budget by '
                        f'KES {(total_estimated - new_plan_item.remaining_budget):,.2f}'
                    )
            
            requisition.save()
            
            # Save attachments
            attachments = attachment_formset.save(commit=False)
            for attachment in attachments:
                if not attachment.uploaded_by:
                    attachment.uploaded_by = request.user
                attachment.save()
            
            messages.success(request, 'Requisition updated successfully!')
            return redirect('staff_requisition_detail', pk=pk)
    else:
        form = RequisitionForm(instance=requisition, user=request.user)
        formset = RequisitionItemFormSet(instance=requisition)
        attachment_formset = RequisitionAttachmentFormSet(instance=requisition)
    
    # Get available plan items
    available_plan_items = None
    if request.user.department:
        current_budget_year = BudgetYear.objects.filter(is_active=True).first()
        if current_budget_year:
            available_plan_items = ProcurementPlanItem.objects.filter(
                procurement_plan__department=request.user.department,
                procurement_plan__budget_year=current_budget_year,
                procurement_plan__status='APPROVED',
                status='PLANNED'
            ).select_related('procurement_plan', 'budget')
    
    context = {
        'form': form,
        'formset': formset,
        'attachment_formset': attachment_formset,
        'requisition': requisition,
        'page_title': f'Edit Requisition {requisition.requisition_number}',
        'available_plan_items': available_plan_items,
    }
    
    return render(request, 'staff/requisition_form.html', context)


@login_required
def staff_requisition_submit(request, pk):
    """Submit a requisition for approval with plan validation"""
    
    requisition = get_object_or_404(
        Requisition.objects.select_related(
            'procurement_plan_item__procurement_plan'
        ),
        pk=pk
    )
    
    # Check permissions
    if requisition.requested_by != request.user:
        messages.error(request, 'You can only submit your own requisitions.')
        return redirect('staff_requisitions_list')
    
    if requisition.status != 'DRAFT':
        messages.error(request, 'This requisition has already been submitted.')
        return redirect('staff_requisition_detail', pk=pk)
    
    # Validate that requisition has items
    if not requisition.items.exists():
        messages.error(request, 'Cannot submit a requisition without items.')
        return redirect('staff_requisition_edit', pk=pk)
    
    # Check if budget is allocated
    if not requisition.budget:
        messages.error(request, 'Please select a budget before submitting.')
        return redirect('staff_requisition_edit', pk=pk)
    
    # PLAN VALIDATION
    validation_errors = []
    
    if requisition.is_planned:
        if not requisition.procurement_plan_item:
            validation_errors.append(
                'Planned requisition must be linked to a procurement plan item'
            )
        else:
            plan_item = requisition.procurement_plan_item
            
            # Check if plan is approved
            if plan_item.procurement_plan.status != 'APPROVED':
                validation_errors.append(
                    'Procurement plan is not approved. Status: '
                    f'{plan_item.procurement_plan.get_status_display()}'
                )
            
            # Check budget
            if requisition.estimated_amount > plan_item.remaining_budget:
                validation_errors.append(
                    f'Amount (KES {requisition.estimated_amount:,.2f}) exceeds '
                    f'remaining plan budget (KES {plan_item.remaining_budget:,.2f})'
                )
    else:
        # Unplanned requisition validation
        if not requisition.is_emergency:
            validation_errors.append(
                'Unplanned requisition must be marked as emergency'
            )
        
        if not requisition.emergency_justification:
            validation_errors.append(
                'Emergency justification is required for unplanned requisitions'
            )
        
        if not requisition.emergency_type:
            validation_errors.append('Emergency type must be specified')
        
        if requisition.requires_plan_amendment and not requisition.plan_amendment:
            messages.warning(
                request,
                'This unplanned requisition will require procurement plan amendment. '
                'HOD and procurement officer will create the amendment during approval.'
            )
    
    # If there are validation errors, don't submit
    if validation_errors:
        for error in validation_errors:
            messages.error(request, error)
        return redirect('staff_requisition_edit', pk=pk)
    
    # Change status to submitted
    requisition.status = 'SUBMITTED'
    requisition.submitted_at = timezone.now()
    requisition.save()
    
    # Create approval workflow (HOD approval first)
    RequisitionApproval.objects.create(
        requisition=requisition,
        approval_stage='HOD',
        sequence=1,
        status='PENDING'
    )
    
    success_msg = f'Requisition {requisition.requisition_number} submitted for approval!'
    
    if not requisition.is_planned:
        success_msg += ' (Emergency/Unplanned - requires HOD emergency approval)'
    
    messages.success(request, success_msg)
    
    return redirect('staff_requisition_detail', pk=pk)


@login_required
def staff_requisition_cancel(request, pk):
    """Cancel a requisition and revert plan commitments"""
    
    requisition = get_object_or_404(
        Requisition.objects.select_related('procurement_plan_item'),
        pk=pk
    )
    
    # Check permissions
    if requisition.requested_by != request.user:
        messages.error(request, 'You can only cancel your own requisitions.')
        return redirect('staff_requisitions_list')
    
    if requisition.status not in ['DRAFT', 'SUBMITTED']:
        messages.error(request, 'Cannot cancel this requisition at its current stage.')
        return redirect('staff_requisition_detail', pk=pk)
    
    if request.method == 'POST':
        # Revert plan item commitments
        if requisition.is_planned and requisition.procurement_plan_item:
            plan_item = requisition.procurement_plan_item
            plan_item.amount_committed -= requisition.estimated_amount
            
            # Revert quantity if tracked
            total_qty = sum(
                item.quantity for item in requisition.items.all()
            )
            plan_item.quantity_requisitioned -= total_qty
            plan_item.save()
        
        requisition.status = 'CANCELLED'
        requisition.save()
        
        messages.success(request, 'Requisition cancelled successfully.')
        return redirect('staff_requisitions_list')
    
    context = {
        'requisition': requisition,
    }
    
    return render(request, 'staff/requisition_cancel_confirm.html', context)


@login_required
def get_plan_item_details(request, plan_item_id):
    """AJAX endpoint to get plan item details"""
    try:
        plan_item = ProcurementPlanItem.objects.select_related(
            'procurement_plan', 'budget'
        ).get(pk=plan_item_id)
        
        return JsonResponse({
            'success': True,
            'data': {
                'description': plan_item.description,
                'estimated_cost': float(plan_item.estimated_cost),
                'remaining_budget': float(plan_item.remaining_budget),
                'quantity': float(plan_item.quantity),
                'remaining_quantity': float(plan_item.remaining_quantity),
                'unit_of_measure': plan_item.unit_of_measure,
                'planned_quarter': plan_item.get_planned_quarter_display(),
                'procurement_method': plan_item.get_procurement_method_display(),
                'budget_name': plan_item.budget.category.name if plan_item.budget else 'N/A',
            }
        })
    except ProcurementPlanItem.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Plan item not found'
        }, status=404)

@login_required
def staff_requisitions_pending(request):
    """List all pending requisitions"""
    
    user = request.user
    
    # Get pending requisitions (all statuses except DRAFT, APPROVED, REJECTED, CANCELLED)
    requisitions = Requisition.objects.filter(
        requested_by=user,
        status__in=['SUBMITTED', 'HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED', 'PROCUREMENT_APPROVED']
    ).select_related('department', 'budget').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(requisitions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'requisitions': page_obj,
        'count': requisitions.count(),
        'page_title': 'Pending Requisitions',
    }
    
    return render(request, 'staff/requisitions_status.html', context)


@login_required
def staff_requisitions_approved(request):
    """List all approved requisitions"""
    
    user = request.user
    
    requisitions = Requisition.objects.filter(
        requested_by=user,
        status='APPROVED'
    ).select_related('department', 'budget').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(requisitions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'requisitions': page_obj,
        'count': requisitions.count(),
        'page_title': 'Approved Requisitions',
    }
    
    return render(request, 'staff/requisitions_status.html', context)


@login_required
def staff_requisitions_rejected(request):
    """List all rejected requisitions"""
    
    user = request.user
    
    requisitions = Requisition.objects.filter(
        requested_by=user,
        status='REJECTED'
    ).select_related('department', 'budget').order_by('-updated_at')
    
    # Pagination
    paginator = Paginator(requisitions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'requisitions': page_obj,
        'count': requisitions.count(),
        'page_title': 'Rejected Requisitions',
    }
    
    return render(request, 'staff/requisitions_status.html', context)


@login_required
def staff_help_center(request):
    """Help center with FAQs and support information"""
    
    faqs = [
        {
            'category': 'Getting Started',
            'questions': [
                {
                    'question': 'How do I create a new requisition?',
                    'answer': 'Click on "New Requisition" from the sidebar, fill in the required details including title, justification, and required date. Then add items with specifications and quantities. You can also attach supporting documents. Save as draft or submit directly for approval.'
                },
                {
                    'question': 'What information do I need to provide?',
                    'answer': 'You need to provide: requisition title, justification, required date, budget allocation, and detailed item descriptions with specifications, quantities, and estimated prices. Supporting documents like quotations or specifications are recommended.'
                },
            ]
        },
        {
            'category': 'Approval Process',
            'questions': [
                {
                    'question': 'How long does approval take?',
                    'answer': 'Approval typically takes 3-7 business days depending on the amount and complexity. The requisition goes through multiple approval stages: HOD, Faculty (for amounts above threshold), Budget/Finance, and Procurement.'
                },
                {
                    'question': 'Can I track my requisition status?',
                    'answer': 'Yes! Go to "My Requisitions" and click on any requisition to view its detailed status, approval history, and comments from approvers.'
                },
                {
                    'question': 'What if my requisition is rejected?',
                    'answer': 'Check the rejection reason provided by the approver. You can create a new requisition addressing the concerns raised. Contact the approver if you need clarification.'
                },
            ]
        },
        {
            'category': 'Editing & Cancellation',
            'questions': [
                {
                    'question': 'Can I edit a submitted requisition?',
                    'answer': 'No, once submitted you cannot edit. However, you can cancel it (if not yet approved) and create a new one with the correct information.'
                },
                {
                    'question': 'How do I cancel a requisition?',
                    'answer': 'Open the requisition detail page and click the "Cancel" button. You can only cancel requisitions in DRAFT or SUBMITTED status.'
                },
            ]
        },
        {
            'category': 'Budget & Amounts',
            'questions': [
                {
                    'question': 'How do I know my department\'s budget?',
                    'answer': 'Contact your Head of Department or the Finance office for information about available budget allocations and codes.'
                },
                {
                    'question': 'What happens after approval?',
                    'answer': 'Once fully approved, the procurement team will initiate the tendering or quotation process, issue purchase orders, and process your request according to procurement guidelines.'
                },
            ]
        },
    ]
    
    context = {
        'faqs': faqs,
        'support_email': 'procurement@university.ac.ke',
        'support_phone': '+254 XXX XXX XXX',
        'office_hours': 'Monday - Friday, 8:00 AM - 5:00 PM',
    }
    
    return render(request, 'staff/help_center.html', context)


@login_required
def staff_guidelines(request):
    """Procurement guidelines and policies"""
    
    guidelines = [
        {
            'title': 'Requisition Submission Guidelines',
            'icon': 'bi-file-earmark-text',
            'items': [
                'Submit requisitions at least 3 weeks before the required date to allow for processing time.',
                'Provide detailed specifications to ensure accurate procurement.',
                'Include market research or quotations to support estimated costs.',
                'Ensure all mandatory fields are completed before submission.',
                'Attach relevant supporting documents (specifications, quotations, TOR).',
            ]
        },
        {
            'title': 'Budget & Financial Guidelines',
            'icon': 'bi-cash-stack',
            'items': [
                'All requisitions must be charged to an approved budget line.',
                'Ensure sufficient budget balance before submitting requisitions.',
                'Departmental budget limits apply - seek HOD approval for large amounts.',
                'Emergency purchases require special justification and additional approvals.',
                'Budget reallocations must be approved by Finance before requisition submission.',
            ]
        },
        {
            'title': 'Approval Thresholds',
            'icon': 'bi-diagram-3',
            'items': [
                'Up to KES 50,000: HOD and Procurement approval required.',
                'KES 50,001 - 500,000: Requires Faculty and Finance approval.',
                'Above KES 500,000: Additional Director approval and competitive tendering required.',
                'Emergency purchases follow expedited approval process with proper justification.',
            ]
        },
        {
            'title': 'Item Specifications',
            'icon': 'bi-list-check',
            'items': [
                'Provide clear, detailed technical specifications for all items.',
                'Avoid specifying brand names unless justified (use "or equivalent").',
                'Include quantity, unit of measure, and quality standards.',
                'For technical items, attach manufacturer specifications or datasheets.',
                'Specify any special delivery or installation requirements.',
            ]
        },
        {
            'title': 'Supplier Selection',
            'icon': 'bi-people',
            'items': [
                'Procurement team handles supplier selection based on value for money.',
                'Three quotations required for purchases above KES 50,000.',
                'Competitive tendering mandatory for purchases above KES 500,000.',
                'Supplier must be registered and compliant with tax requirements.',
                'You may suggest preferred suppliers but final selection rests with procurement.',
            ]
        },
        {
            'title': 'Documentation Requirements',
            'icon': 'bi-folder',
            'items': [
                'Attach justification documents explaining need and urgency.',
                'Include quotations or price estimates from potential suppliers.',
                'Provide technical specifications or terms of reference.',
                'For specialized items, attach supporting research or documentation.',
                'Ensure all attachments are clearly labeled and readable.',
            ]
        },
        {
            'title': 'Compliance & Ethics',
            'icon': 'bi-shield-check',
            'items': [
                'All procurement must follow university policies and government regulations.',
                'Declare any conflict of interest with suppliers or vendors.',
                'Do not split requisitions to circumvent approval thresholds.',
                'Report any suspected fraud or misconduct to the appropriate office.',
                'Maintain confidentiality of procurement information.',
            ]
        },
        {
            'title': 'Delivery & Receipt',
            'icon': 'bi-truck',
            'items': [
                'Specify accurate delivery location and contact person.',
                'Be available to inspect and receive goods when delivered.',
                'Report any discrepancies or damages immediately to Stores.',
                'Sign Goods Received Notes only after thorough inspection.',
                'Follow up with procurement team if delivery is delayed beyond agreed date.',
            ]
        },
    ]
    
    context = {
        'guidelines': guidelines,
        'policy_document_url': '#',  # Link to full procurement policy document
        'last_updated': '2024',
    }
    
    return render(request, 'staff/guidelines.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg, F
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from decimal import Decimal

from .models import (
    User, Department, Requisition, RequisitionItem, RequisitionApproval,
    Budget, BudgetYear, PurchaseOrder, Invoice, Asset, AuditLog,
    Notification, ProcurementPolicy
)
from .forms import (
    RequisitionForm, RequisitionItemFormSet, RequisitionAttachmentFormSet,
    RequisitionFilterForm
)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def check_hod_permission(user):
    """Check if user is HOD"""
    return user.role == 'HOD' and hasattr(user, 'head_of') and user.head_of.exists()


def log_action(user, action, model_name, object_id, object_repr, changes=None, request=None):
    """Helper function to log audit trail"""
    ip_address = None
    user_agent = ''
    
    if request:
        ip_address = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    AuditLog.objects.create(
        user=user,
        action=action,
        model_name=model_name,
        object_id=str(object_id),
        object_repr=object_repr,
        changes=changes,
        ip_address=ip_address,
        user_agent=user_agent
    )


# ============================================================================
# DASHBOARD & ANALYTICS
# ============================================================================
@login_required
def hod_dashboard(request):
    """HOD Dashboard with department overview"""
    
    # Simple permission check
    if request.user.role != 'HOD':
        messages.error(request, 'You do not have permission to access the HOD dashboard.')
        return redirect('login')
    
    # Get HOD's department
    department = request.user.head_of.first()
    
    if not department:
        messages.error(request, 'You are not assigned as Head of any department. Please contact the administrator.')
        return redirect('login')
    
    # Current budget year
    current_budget_year = BudgetYear.objects.filter(is_active=True).first()
    
    # Dashboard statistics
    stats = {
        'pending_approvals': Requisition.objects.filter(
            department=department,
            status='SUBMITTED'
        ).count(),
        
        'total_requisitions_month': Requisition.objects.filter(
            department=department,
            created_at__month=timezone.now().month,
            created_at__year=timezone.now().year
        ).count(),
        
        'approved_this_month': Requisition.objects.filter(
            department=department,
            status__in=['HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED', 'APPROVED'],
            approvals__approval_stage='HOD',
            approvals__approval_date__month=timezone.now().month,
            approvals__approval_date__year=timezone.now().year
        ).distinct().count(),
        
        'department_staff': User.objects.filter(
            department=department,
            is_active_user=True
        ).count(),
    }
    
    # Budget overview
    budget_overview = None
    if current_budget_year:
        budgets = Budget.objects.filter(
            department=department,
            budget_year=current_budget_year,
            is_active=True
        ).aggregate(
            total_allocated=Sum('allocated_amount'),
            total_committed=Sum('committed_amount'),
            total_spent=Sum('actual_spent')
        )
        
        total_allocated = budgets['total_allocated'] or Decimal('0')
        total_committed = budgets['total_committed'] or Decimal('0')
        total_spent = budgets['total_spent'] or Decimal('0')
        
        budget_overview = {
            'allocated': total_allocated,
            'committed': total_committed,
            'spent': total_spent,
            'available': total_allocated - total_committed - total_spent,
            'utilization_percentage': (
                (total_spent / total_allocated * 100) if total_allocated > 0 else 0
            )
        }
    
    # Recent requisitions
    recent_requisitions = Requisition.objects.filter(
        department=department
    ).select_related('requested_by', 'budget').order_by('-created_at')[:10]
    
    # Pending approvals
    pending_approvals = Requisition.objects.filter(
        department=department,
        status='SUBMITTED'
    ).select_related('requested_by', 'budget').order_by('-submitted_at')[:5]
    
    # Monthly requisition trend (last 6 months)
    monthly_trend = []
    for i in range(6):
        month_start = timezone.now() - timedelta(days=30 * i)
        count = Requisition.objects.filter(
            department=department,
            created_at__month=month_start.month,
            created_at__year=month_start.year
        ).count()
        monthly_trend.append({
            'month': month_start.strftime('%b %Y'),
            'count': count
        })
    monthly_trend.reverse()
    
    # Department performance metrics
    performance = {
        'avg_approval_time': RequisitionApproval.objects.filter(
            requisition__department=department,
            approval_stage='HOD',
            status='APPROVED'
        ).aggregate(
            avg_time=Avg(F('approval_date') - F('created_at'))
        )['avg_time'],
        
        'approval_rate': 0,
        'rejection_rate': 0,
    }
    
    total_processed = RequisitionApproval.objects.filter(
        requisition__department=department,
        approval_stage='HOD',
        status__in=['APPROVED', 'REJECTED']
    ).count()
    
    if total_processed > 0:
        approved = RequisitionApproval.objects.filter(
            requisition__department=department,
            approval_stage='HOD',
            status='APPROVED'
        ).count()
        performance['approval_rate'] = (approved / total_processed) * 100
        performance['rejection_rate'] = 100 - performance['approval_rate']
    
    context = {
        'department': department,
        'stats': stats,
        'budget_overview': budget_overview,
        'recent_requisitions': recent_requisitions,
        'pending_approvals': pending_approvals,
        'monthly_trend': monthly_trend,
        'performance': performance,
        'current_budget_year': current_budget_year,
    }
    
    return render(request, 'hod/dashboard.html', context)

@login_required
def hod_analytics_view(request):
    """Department analytics and reports"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Default to current year
    if not date_from:
        date_from = timezone.now().replace(month=1, day=1).date()
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    
    if not date_to:
        date_to = timezone.now().date()
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Requisition analytics
    requisitions = Requisition.objects.filter(
        department=department,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )
    
    requisition_stats = {
        'total': requisitions.count(),
        'by_status': requisitions.values('status').annotate(count=Count('id')),
        'by_priority': requisitions.values('priority').annotate(count=Count('id')),
        'total_value': requisitions.aggregate(total=Sum('estimated_amount'))['total'] or Decimal('0'),
        'avg_value': requisitions.aggregate(avg=Avg('estimated_amount'))['avg'] or Decimal('0'),
    }
    
    # Budget utilization by category
    budget_utilization = Budget.objects.filter(
        department=department,
        is_active=True
    ).select_related('category').values(
        'category__name', 'category__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        committed=Sum('committed_amount')
    )
    
    # Top requesters
    top_requesters = User.objects.filter(
        requisitions_created__department=department,
        requisitions_created__created_at__date__gte=date_from,
        requisitions_created__created_at__date__lte=date_to
    ).annotate(
        req_count=Count('requisitions_created'),
        total_value=Sum('requisitions_created__estimated_amount')
    ).order_by('-req_count')[:10]
    
    # Purchase order analytics
    po_stats = PurchaseOrder.objects.filter(
        requisition__department=department,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).aggregate(
        total_pos=Count('id'),
        total_value=Sum('total_amount'),
        avg_value=Avg('total_amount')
    )
    
    # Supplier distribution
    supplier_distribution = PurchaseOrder.objects.filter(
        requisition__department=department,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).values('supplier__name').annotate(
        po_count=Count('id'),
        total_value=Sum('total_amount')
    ).order_by('-total_value')[:10]
    
    context = {
        'department': department,
        'date_from': date_from,
        'date_to': date_to,
        'requisition_stats': requisition_stats,
        'budget_utilization': budget_utilization,
        'top_requesters': top_requesters,
        'po_stats': po_stats,
        'supplier_distribution': supplier_distribution,
    }
    
    return render(request, 'hod/analytics.html', context)


# ============================================================================
# REQUISITION MANAGEMENT
# ============================================================================

@login_required
def hod_my_requisitions_view(request):
    """HOD's own requisitions"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    # Filter form
    filter_form = RequisitionFilterForm(request.GET)
    
    # Base queryset
    requisitions = Requisition.objects.filter(
        requested_by=request.user
    ).select_related('department', 'budget').order_by('-created_at')
    
    # Apply filters
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('status'):
            requisitions = requisitions.filter(status=filter_form.cleaned_data['status'])
        
        if filter_form.cleaned_data.get('priority'):
            requisitions = requisitions.filter(priority=filter_form.cleaned_data['priority'])
        
        if filter_form.cleaned_data.get('search'):
            search = filter_form.cleaned_data['search']
            requisitions = requisitions.filter(
                Q(requisition_number__icontains=search) |
                Q(title__icontains=search)
            )
        
        if filter_form.cleaned_data.get('date_from'):
            requisitions = requisitions.filter(created_at__date__gte=filter_form.cleaned_data['date_from'])
        
        if filter_form.cleaned_data.get('date_to'):
            requisitions = requisitions.filter(created_at__date__lte=filter_form.cleaned_data['date_to'])
    
    # Pagination
    paginator = Paginator(requisitions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'filter_form': filter_form,
        'total_count': requisitions.count(),
    }
    
    return render(request, 'hod/my_requisitions.html', context)


@login_required
def hod_new_requisition_view(request):
    """Create new requisition"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    if request.method == 'POST':
        form = RequisitionForm(request.POST, user=request.user)
        item_formset = RequisitionItemFormSet(request.POST, prefix='items')
        attachment_formset = RequisitionAttachmentFormSet(request.POST, request.FILES, prefix='attachments')
        
        if form.is_valid() and item_formset.is_valid() and attachment_formset.is_valid():
            # Create requisition
            requisition = form.save(commit=False)
            requisition.requested_by = request.user
            requisition.department = department
            requisition.status = 'DRAFT'
            
            # Calculate estimated amount from items
            total_amount = Decimal('0')
            for item_form in item_formset:
                if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE'):
                    qty = item_form.cleaned_data.get('quantity', 0)
                    price = item_form.cleaned_data.get('estimated_unit_price', 0)
                    total_amount += qty * price
            
            requisition.estimated_amount = total_amount
            requisition.save()
            
            # Save items
            items = item_formset.save(commit=False)
            for item in items:
                item.requisition = requisition
                item.save()
            
            # Save attachments
            attachments = attachment_formset.save(commit=False)
            for attachment in attachments:
                attachment.requisition = requisition
                attachment.uploaded_by = request.user
                attachment.save()
            
            # Log action
            log_action(
                request.user, 'CREATE', 'Requisition',
                requisition.id, str(requisition),
                request=request
            )
            
            messages.success(request, f'Requisition {requisition.requisition_number} created successfully!')
            return redirect('hod_requisition_detail', pk=requisition.pk)
    else:
        form = RequisitionForm(user=request.user)
        item_formset = RequisitionItemFormSet(prefix='items')
        attachment_formset = RequisitionAttachmentFormSet(prefix='attachments')
    
    context = {
        'form': form,
        'item_formset': item_formset,
        'attachment_formset': attachment_formset,
        'department': department,
    }
    
    return render(request, 'hod/new_requisition.html', context)


@login_required
def hod_department_requests_view(request):
    """View all department requisitions"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Filter form
    filter_form = RequisitionFilterForm(request.GET)
    
    # Base queryset - all department requisitions
    requisitions = Requisition.objects.filter(
        department=department
    ).select_related('requested_by', 'budget').order_by('-created_at')
    
    # Apply filters
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('status'):
            requisitions = requisitions.filter(status=filter_form.cleaned_data['status'])
        
        if filter_form.cleaned_data.get('priority'):
            requisitions = requisitions.filter(priority=filter_form.cleaned_data['priority'])
        
        if filter_form.cleaned_data.get('search'):
            search = filter_form.cleaned_data['search']
            requisitions = requisitions.filter(
                Q(requisition_number__icontains=search) |
                Q(title__icontains=search) |
                Q(requested_by__first_name__icontains=search) |
                Q(requested_by__last_name__icontains=search)
            )
        
        if filter_form.cleaned_data.get('date_from'):
            requisitions = requisitions.filter(created_at__date__gte=filter_form.cleaned_data['date_from'])
        
        if filter_form.cleaned_data.get('date_to'):
            requisitions = requisitions.filter(created_at__date__lte=filter_form.cleaned_data['date_to'])
    
    # Pagination
    paginator = Paginator(requisitions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Summary statistics
    stats = {
        'total': requisitions.count(),
        'pending': requisitions.filter(status='SUBMITTED').count(),
        'approved': requisitions.filter(status__in=['HOD_APPROVED', 'APPROVED']).count(),
        'rejected': requisitions.filter(status='REJECTED').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'filter_form': filter_form,
        'stats': stats,
        'department': department,
    }
    
    return render(request, 'hod/department_requests.html', context)


@login_required
def hod_requisition_detail_view(request, pk):
    """View requisition details"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    requisition = get_object_or_404(
        Requisition.objects.select_related('requested_by', 'department', 'budget'),
        pk=pk,
        department=department
    )
    
    # Get items and attachments
    items = requisition.items.all()
    attachments = requisition.attachments.all()
    
    # Get approval history
    approvals = requisition.approvals.select_related('approver').order_by('sequence')
    
    # Check if HOD can approve
    can_approve = (
        requisition.status == 'SUBMITTED' and
        request.user == department.hod
    )
    
    context = {
        'requisition': requisition,
        'items': items,
        'attachments': attachments,
        'approvals': approvals,
        'can_approve': can_approve,
    }
    
    return render(request, 'hod/requisition_detail.html', context)


# ============================================================================
# APPROVALS
# ============================================================================

@login_required
def hod_pending_approvals_view(request):
    """View pending requisitions for HOD approval"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Get pending requisitions
    pending = Requisition.objects.filter(
        department=department,
        status='SUBMITTED'
    ).select_related('requested_by', 'budget').order_by('-submitted_at')
    
    # Pagination
    paginator = Paginator(pending, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_pending': pending.count(),
    }
    
    return render(request, 'hod/pending_approvals.html', context)


@login_required
def hod_approve_requisition_view(request, pk):
    """Approve a requisition"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    requisition = get_object_or_404(
        Requisition,
        pk=pk,
        department=department,
        status='SUBMITTED'
    )
    
    if request.method == 'POST':
        comments = request.POST.get('comments', '')
        
        # Create or update approval record
        approval, created = RequisitionApproval.objects.get_or_create(
            requisition=requisition,
            approval_stage='HOD',
            defaults={
                'approver': request.user,
                'sequence': 1
            }
        )
        
        approval.status = 'APPROVED'
        approval.comments = comments
        approval.approval_date = timezone.now()
        approval.approver = request.user
        approval.save()
        
        # Update requisition status
        requisition.status = 'HOD_APPROVED'
        requisition.save()
        
        # Log action
        log_action(
            request.user, 'APPROVE', 'Requisition',
            requisition.id, str(requisition),
            changes={'status': 'HOD_APPROVED', 'comments': comments},
            request=request
        )
        
        # Create notification for requester
        Notification.objects.create(
            user=requisition.requested_by,
            notification_type='APPROVAL',
            priority='MEDIUM',
            title='Requisition Approved by HOD',
            message=f'Your requisition {requisition.requisition_number} has been approved by HOD.',
            link_url=f'/requisitions/{requisition.pk}/'
        )
        
        messages.success(request, f'Requisition {requisition.requisition_number} approved successfully!')
        return redirect('hod_pending_approvals')
    
    return redirect('hod_requisition_detail', pk=pk)


@login_required
def hod_reject_requisition_view(request, pk):
    """Reject a requisition"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    requisition = get_object_or_404(
        Requisition,
        pk=pk,
        department=department,
        status='SUBMITTED'
    )
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        if not rejection_reason:
            messages.error(request, 'Please provide a reason for rejection.')
            return redirect('hod_requisition_detail', pk=pk)
        
        # Create or update approval record
        approval, created = RequisitionApproval.objects.get_or_create(
            requisition=requisition,
            approval_stage='HOD',
            defaults={
                'approver': request.user,
                'sequence': 1
            }
        )
        
        approval.status = 'REJECTED'
        approval.comments = rejection_reason
        approval.approval_date = timezone.now()
        approval.approver = request.user
        approval.save()
        
        # Update requisition
        requisition.status = 'REJECTED'
        requisition.rejection_reason = rejection_reason
        requisition.save()
        
        # Log action
        log_action(
            request.user, 'REJECT', 'Requisition',
            requisition.id, str(requisition),
            changes={'status': 'REJECTED', 'reason': rejection_reason},
            request=request
        )
        
        # Notify requester
        Notification.objects.create(
            user=requisition.requested_by,
            notification_type='APPROVAL',
            priority='HIGH',
            title='Requisition Rejected by HOD',
            message=f'Your requisition {requisition.requisition_number} has been rejected. Reason: {rejection_reason}',
            link_url=f'/requisitions/{requisition.pk}/'
        )
        
        messages.success(request, f'Requisition {requisition.requisition_number} rejected.')
        return redirect('hod_pending_approvals')
    
    return redirect('hod_requisition_detail', pk=pk)


@login_required
def hod_approved_requisitions_view(request):
    """View approved requisitions"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Get approved requisitions
    approved = Requisition.objects.filter(
        department=department,
        status__in=['HOD_APPROVED', 'FACULTY_APPROVED', 'BUDGET_APPROVED', 'APPROVED']
    ).select_related('requested_by', 'budget').order_by('-updated_at')
    
    # Pagination
    paginator = Paginator(approved, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_approved': approved.count(),
    }
    
    return render(request, 'hod/approved_requisitions.html', context)


@login_required
def hod_rejected_requisitions_view(request):
    """View rejected requisitions"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Get rejected requisitions
    rejected = Requisition.objects.filter(
        department=department,
        status='REJECTED'
    ).select_related('requested_by', 'budget').order_by('-updated_at')
    
    # Pagination
    paginator = Paginator(rejected, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_rejected': rejected.count(),
    }
    
    return render(request, 'hod/rejected_requisitions.html', context)


@login_required
def hod_approval_history_view(request):
    """View approval history"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Get all approvals made by this HOD
    approvals = RequisitionApproval.objects.filter(
        approver=request.user,
        approval_stage='HOD'
    ).select_related('requisition', 'requisition__requested_by').order_by('-approval_date')
    
    # Pagination
    paginator = Paginator(approvals, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total': approvals.count(),
        'approved': approvals.filter(status='APPROVED').count(),
        'rejected': approvals.filter(status='REJECTED').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
    }
    
    return render(request, 'hod/approval_history.html', context)


# ============================================================================
# BUDGET MANAGEMENT
# ============================================================================

@login_required
def hod_budget_overview_view(request):
    """Department budget overview"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Get current budget year
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    # Get all budget years for selection
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    
    # Selected year from query params
    selected_year_id = request.GET.get('year')
    if selected_year_id:
        selected_year = get_object_or_404(BudgetYear, pk=selected_year_id)
    else:
        selected_year = current_year
    
    # Get budgets for selected year
    budgets = None
    budget_summary = None
    
    if selected_year:
        budgets = Budget.objects.filter(
            department=department,
            budget_year=selected_year,
            is_active=True
        ).select_related('category').order_by('category__code')
        
        # Calculate summary
        summary = budgets.aggregate(
            total_allocated=Sum('allocated_amount'),
            total_committed=Sum('committed_amount'),
            total_spent=Sum('actual_spent')
        )
        
        total_allocated = summary['total_allocated'] or Decimal('0')
        total_committed = summary['total_committed'] or Decimal('0')
        total_spent = summary['total_spent'] or Decimal('0')
        
        budget_summary = {
            'allocated': total_allocated,
            'committed': total_committed,
            'spent': total_spent,
            'available': total_allocated - total_committed - total_spent,
            'utilization': (total_spent / total_allocated * 100) if total_allocated > 0 else 0,
        }
    
    context = {
        'department': department,
        'budget_years': budget_years,
        'selected_year': selected_year,
        'budgets': budgets,
        'budget_summary': budget_summary,
    }
    
    return render(request, 'hod/budget_overview.html', context)


@login_required
def hod_expenditure_reports_view(request):
    """Department expenditure reports"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = timezone.now().replace(month=1, day=1).date()
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    
    if not date_to:
        date_to = timezone.now().date()
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Requisitions expenditure
    requisitions = Requisition.objects.filter(
        department=department,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).aggregate(
        total_requested=Sum('estimated_amount'),
        count=Count('id')
    )
    
    # Purchase orders
    purchase_orders = PurchaseOrder.objects.filter(
        requisition__department=department,
        po_date__gte=date_from,
        po_date__lte=date_to
    ).aggregate(
        total_po_value=Sum('total_amount'),
        count=Count('id')
    )
    
    # Invoices and payments
    invoices = Invoice.objects.filter(
        purchase_order__requisition__department=department,
        invoice_date__gte=date_from,
        invoice_date__lte=date_to
    ).aggregate(
        total_invoiced=Sum('total_amount'),
        total_paid=Sum('amount_paid'),
        count=Count('id')
    )
    
    # Expenditure by budget category
    category_expenditure = Budget.objects.filter(
        department=department,
        is_active=True
    ).select_related('category').values(
        'category__name',
        'category__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        committed=Sum('committed_amount')
    ).order_by('-spent')
    
    # Monthly expenditure trend
    monthly_expenditure = []
    current = date_from
    while current <= date_to:
        month_end = (current.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        if month_end > date_to:
            month_end = date_to
        
        month_data = PurchaseOrder.objects.filter(
            requisition__department=department,
            po_date__gte=current,
            po_date__lte=month_end
        ).aggregate(
            total=Sum('total_amount')
        )
        
        monthly_expenditure.append({
            'month': current.strftime('%b %Y'),
            'amount': month_data['total'] or Decimal('0')
        })
        
        current = (month_end + timedelta(days=1)).replace(day=1)
    
    # Top suppliers by expenditure
    top_suppliers = PurchaseOrder.objects.filter(
        requisition__department=department,
        po_date__gte=date_from,
        po_date__lte=date_to
    ).values('supplier__name').annotate(
        total_value=Sum('total_amount'),
        po_count=Count('id')
    ).order_by('-total_value')[:10]
    
    # Assets acquired
    assets = Asset.objects.filter(
        department=department,
        acquisition_date__gte=date_from,
        acquisition_date__lte=date_to
    ).aggregate(
        total_value=Sum('acquisition_cost'),
        count=Count('id')
    )
    
    context = {
        'department': department,
        'date_from': date_from,
        'date_to': date_to,
        'requisitions': requisitions,
        'purchase_orders': purchase_orders,
        'invoices': invoices,
        'category_expenditure': category_expenditure,
        'monthly_expenditure': monthly_expenditure,
        'top_suppliers': top_suppliers,
        'assets': assets,
    }
    
    return render(request, 'hod/expenditure_reports.html', context)


@login_required
def hod_staff_management_view(request):
    """Manage department staff"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    # Get all department staff
    staff = User.objects.filter(
        department=department,
        is_active_user=True
    ).exclude(
        role='SUPPLIER'
    ).order_by('first_name', 'last_name')
    
    # Staff statistics
    stats = {
        'total_staff': staff.count(),
        'by_role': staff.values('role').annotate(count=Count('id')),
    }
    
    # Staff requisition activity
    staff_activity = []
    for user in staff:
        activity = {
            'user': user,
            'requisitions_count': Requisition.objects.filter(
                requested_by=user,
                created_at__year=timezone.now().year
            ).count(),
            'total_value': Requisition.objects.filter(
                requested_by=user,
                created_at__year=timezone.now().year
            ).aggregate(total=Sum('estimated_amount'))['total'] or Decimal('0'),
            'pending': Requisition.objects.filter(
                requested_by=user,
                status__in=['DRAFT', 'SUBMITTED']
            ).count()
        }
        staff_activity.append(activity)
    
    # Sort by requisitions count
    staff_activity.sort(key=lambda x: x['requisitions_count'], reverse=True)
    
    # Pagination
    paginator = Paginator(staff_activity, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'department': department,
        'stats': stats,
        'page_obj': page_obj,
        'total_staff': staff.count(),
    }
    
    return render(request, 'hod/staff_management.html', context)


@login_required
def hod_staff_detail_view(request, user_id):
    """View staff member details and activity"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    department = request.user.head_of.first()
    
    staff_member = get_object_or_404(
        User,
        id=user_id,
        department=department,
        is_active_user=True
    )
    
    # Staff requisitions
    requisitions = Requisition.objects.filter(
        requested_by=staff_member
    ).select_related('budget').order_by('-created_at')[:20]
    
    # Activity summary
    activity = {
        'total_requisitions': Requisition.objects.filter(
            requested_by=staff_member
        ).count(),
        'this_year': Requisition.objects.filter(
            requested_by=staff_member,
            created_at__year=timezone.now().year
        ).count(),
        'total_value': Requisition.objects.filter(
            requested_by=staff_member
        ).aggregate(total=Sum('estimated_amount'))['total'] or Decimal('0'),
        'approved': Requisition.objects.filter(
            requested_by=staff_member,
            status__in=['APPROVED', 'HOD_APPROVED', 'FACULTY_APPROVED']
        ).count(),
        'rejected': Requisition.objects.filter(
            requested_by=staff_member,
            status='REJECTED'
        ).count(),
    }
    
    context = {
        'staff_member': staff_member,
        'requisitions': requisitions,
        'activity': activity,
        'department': department,
    }
    
    return render(request, 'hod/staff_detail.html', context)


# ============================================================================
# SUPPORT & HELP
# ============================================================================

@login_required
def hod_help_center_view(request):
    """Help center for HODs"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    # FAQ categories
    faqs = [
        {
            'category': 'Requisition Approval',
            'questions': [
                {
                    'question': 'How long do I have to approve a requisition?',
                    'answer': 'You should approve or reject requisitions within 3 working days of submission. Urgent requisitions should be processed within 24 hours.'
                },
                {
                    'question': 'What should I check before approving a requisition?',
                    'answer': 'Verify: 1) Budget availability, 2) Justification validity, 3) Item specifications, 4) Estimated costs are reasonable, 5) Compliance with procurement policies.'
                },
                {
                    'question': 'Can I approve a requisition that exceeds my department budget?',
                    'answer': 'No. Requisitions must be charged to an active budget line with sufficient funds. If needed, request budget reallocation first.'
                },
            ]
        },
        {
            'category': 'Budget Management',
            'questions': [
                {
                    'question': 'How do I check my department\'s budget balance?',
                    'answer': 'Go to Department > Budget Overview to see real-time budget balances, commitments, and expenditure for all budget lines.'
                },
                {
                    'question': 'What is the difference between committed and spent amounts?',
                    'answer': 'Committed = funds reserved for approved requisitions. Spent = actual payments made. Available = Allocated - Committed - Spent.'
                },
            ]
        },
        {
            'category': 'Reporting',
            'questions': [
                {
                    'question': 'How do I generate expenditure reports?',
                    'answer': 'Navigate to Department > Expenditure Reports, select date range, and view or export detailed spending analysis.'
                },
            ]
        },
    ]
    
    # Quick links
    quick_links = [
        {
            'title': 'Procurement Guidelines',
            'icon': 'book',
            'url': 'hod_guidelines',
            'description': 'Complete procurement policies and procedures'
        },
        {
            'title': 'Video Tutorials',
            'icon': 'video',
            'url': '#',
            'description': 'Step-by-step video guides'
        },
        {
            'title': 'Contact Support',
            'icon': 'envelope',
            'url': '#',
            'description': 'Email: procurement@university.ac.ke'
        },
    ]
    
    context = {
        'faqs': faqs,
        'quick_links': quick_links,
    }
    
    return render(request, 'hod/help_center.html', context)


@login_required
def hod_guidelines_view(request):
    """View procurement guidelines and policies"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    # Get active policies
    policies = ProcurementPolicy.objects.filter(
        is_active=True
    ).order_by('-effective_date')
    
    # Get approval thresholds
    from .models import ApprovalThreshold
    thresholds = ApprovalThreshold.objects.filter(
        is_active=True
    ).order_by('min_amount')
    
    context = {
        'policies': policies,
        'thresholds': thresholds,
    }
    
    return render(request, 'hod/guidelines.html', context)


# ============================================================================
# AJAX/API ENDPOINTS
# ============================================================================

@login_required
def hod_quick_stats_api(request):
    """API endpoint for dashboard quick stats"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    department = request.user.head_of.first()
    
    stats = {
        'pending_approvals': Requisition.objects.filter(
            department=department,
            status='SUBMITTED'
        ).count(),
        'approved_today': RequisitionApproval.objects.filter(
            requisition__department=department,
            approval_stage='HOD',
            status='APPROVED',
            approval_date__date=timezone.now().date()
        ).count(),
        'total_staff': User.objects.filter(
            department=department,
            is_active_user=True
        ).count(),
    }
    
    return JsonResponse(stats)


@login_required
def hod_budget_check_api(request):
    """API to check budget availability"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    budget_id = request.GET.get('budget_id')
    amount = request.GET.get('amount')
    
    if not budget_id or not amount:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    
    try:
        budget = Budget.objects.get(pk=budget_id)
        amount = Decimal(amount)
        
        available = budget.available_balance
        is_sufficient = available >= amount
        
        return JsonResponse({
            'budget': {
                'name': str(budget),
                'allocated': float(budget.allocated_amount),
                'committed': float(budget.committed_amount),
                'spent': float(budget.actual_spent),
                'available': float(available),
            },
            'requested_amount': float(amount),
            'is_sufficient': is_sufficient,
            'shortage': float(amount - available) if not is_sufficient else 0,
        })
    except Budget.DoesNotExist:
        return JsonResponse({'error': 'Budget not found'}, status=404)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid amount'}, status=400)


@login_required
def hod_notifications_api(request):
    """API to get recent notifications"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    notifications = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).order_by('-created_at')[:10]
    
    data = [{
        'id': str(n.id),
        'title': n.title,
        'message': n.message,
        'type': n.notification_type,
        'priority': n.priority,
        'created_at': n.created_at.isoformat(),
        'link_url': n.link_url,
    } for n in notifications]
    
    return JsonResponse({'notifications': data, 'count': len(data)})


@login_required
def hod_mark_notification_read_api(request, notification_id):
    """Mark notification as read"""
    if not check_hod_permission(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        notification = Notification.objects.get(
            id=notification_id,
            user=request.user
        )
        notification.is_read = True
        notification.read_at = timezone.now()
        notification.save()
        
        return JsonResponse({'success': True})
    except Notification.DoesNotExist:
        return JsonResponse({'error': 'Notification not found'}, status=404)


# ============================================================================
# EXPORT/DOWNLOAD FUNCTIONS
# ============================================================================

@login_required
def hod_export_requisitions_csv(request):
    """Export requisitions to CSV"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    import csv
    from django.utils.text import slugify
    
    department = request.user.head_of.first()
    
    # Get requisitions
    requisitions = Requisition.objects.filter(
        department=department
    ).select_related('requested_by', 'budget').order_by('-created_at')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="requisitions_{slugify(department.name)}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Requisition Number', 'Title', 'Requested By', 'Status',
        'Priority', 'Estimated Amount', 'Required Date',
        'Submitted Date', 'Budget Line'
    ])
    
    for req in requisitions:
        writer.writerow([
            req.requisition_number,
            req.title,
            req.requested_by.get_full_name(),
            req.get_status_display(),
            req.get_priority_display(),
            req.estimated_amount,
            req.required_date,
            req.submitted_at.strftime('%Y-%m-%d %H:%M') if req.submitted_at else '',
            str(req.budget) if req.budget else '',
        ])
    
    return response


@login_required
def hod_export_budget_pdf(request):
    """Export budget overview to PDF"""
    if not check_hod_permission(request.user):
        messages.error(request, 'You do not have HOD permissions.')
        return redirect('dashboard')
    
    # This would require a PDF library like reportlab or weasyprint
    # Placeholder for now
    messages.info(request, 'PDF export feature coming soon.')
    return redirect('hod_budget_overview')


"""
Finance Officer Views - Complete Implementation
All views for Finance role functionality
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg, F
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from datetime import datetime, timedelta
import csv
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from .models import (
    User, Budget, BudgetYear, BudgetCategory, BudgetReallocation,
    Invoice, InvoiceItem, InvoiceDocument, Payment,
    Requisition, RequisitionApproval, PurchaseOrder,
    Department, Supplier
)
from .forms import PaymentForm, BudgetForm, BudgetReallocationForm


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def check_finance_permission(user):
    """Check if user has finance officer permissions"""
    return user.role == 'FINANCE' or user.is_superuser


# ============================================================================
# DASHBOARD & ANALYTICS
# ============================================================================

@login_required
def finance_dashboard_view(request):
    """Finance Officer Dashboard"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    # Get current budget year
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    # Budget Statistics
    if current_year:
        budgets = Budget.objects.filter(
            budget_year=current_year,
            is_active=True
        )
        
        total_allocated = budgets.aggregate(Sum('allocated_amount'))['allocated_amount__sum'] or 0
        total_committed = budgets.aggregate(Sum('committed_amount'))['committed_amount__sum'] or 0
        total_spent = budgets.aggregate(Sum('actual_spent'))['actual_spent__sum'] or 0
        available_balance = total_allocated - total_committed - total_spent
    else:
        total_allocated = total_committed = total_spent = available_balance = 0
    
    # Invoice Statistics
    pending_invoices_count = Invoice.objects.filter(
        status__in=['SUBMITTED', 'VERIFYING']
    ).count()
    
    pending_invoices_value = Invoice.objects.filter(
        status__in=['SUBMITTED', 'VERIFYING']
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    approved_invoices_count = Invoice.objects.filter(
        status='APPROVED'
    ).count()
    
    approved_invoices_value = Invoice.objects.filter(
        status='APPROVED'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Payment Statistics
    pending_payments = Payment.objects.filter(
        status='PENDING'
    ).count()
    
    completed_payments = Payment.objects.filter(
        status='COMPLETED',
        created_at__gte=timezone.now() - timedelta(days=30)
    ).count()
    
    total_paid_this_month = Payment.objects.filter(
        status='COMPLETED',
        payment_date__gte=timezone.now().replace(day=1)
    ).aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
    
    # Pending Approvals (Budget Check Stage)
    pending_approvals = RequisitionApproval.objects.filter(
        approval_stage='BUDGET',
        status='PENDING'
    ).select_related('requisition', 'requisition__department')[:5]
    
    # Recent Activities
    recent_invoices = Invoice.objects.filter(
        status__in=['SUBMITTED', 'VERIFYING', 'APPROVED']
    ).select_related('supplier', 'purchase_order').order_by('-created_at')[:5]
    
    recent_payments = Payment.objects.filter(
        status__in=['PENDING', 'PROCESSING', 'COMPLETED']
    ).select_related('invoice', 'invoice__supplier').order_by('-created_at')[:5]
    
    # Budget Utilization by Department
    department_utilization = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).values(
        'department__name', 'department__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        committed=Sum('committed_amount')
    ).order_by('-allocated')[:10]
    
    context = {
        'current_year': current_year,
        'total_allocated': total_allocated,
        'total_committed': total_committed,
        'total_spent': total_spent,
        'available_balance': available_balance,
        'budget_utilization_percentage': (total_spent / total_allocated * 100) if total_allocated > 0 else 0,
        
        'pending_invoices_count': pending_invoices_count,
        'pending_invoices_value': pending_invoices_value,
        'approved_invoices_count': approved_invoices_count,
        'approved_invoices_value': approved_invoices_value,
        
        'pending_payments': pending_payments,
        'completed_payments': completed_payments,
        'total_paid_this_month': total_paid_this_month,
        
        'pending_approvals': pending_approvals,
        'recent_invoices': recent_invoices,
        'recent_payments': recent_payments,
        'department_utilization': department_utilization,
    }
    
    return render(request, 'finance/finance_module/dashboard.html', context)


@login_required
def finance_analytics_view(request):
    """Finance Analytics Dashboard"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    # Monthly spending trend (last 12 months)
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_spending = Payment.objects.filter(
        status='COMPLETED',
        payment_date__gte=twelve_months_ago
    ).extra(
        select={'month': "EXTRACT(month FROM payment_date)", 'year': "EXTRACT(year FROM payment_date)"}
    ).values('month', 'year').annotate(
        total=Sum('payment_amount')
    ).order_by('year', 'month')
    
    # Category-wise spending
    category_spending = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).values('category__name').annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent')
    ).order_by('-spent')[:10]
    
    # Department-wise spending
    department_spending = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).values('department__name').annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent')
    ).order_by('-spent')[:10]
    
    # Top suppliers by payment value
    top_suppliers = Payment.objects.filter(
        status='COMPLETED',
        payment_date__year=timezone.now().year
    ).values('invoice__supplier__name').annotate(
        total_paid=Sum('payment_amount'),
        payment_count=Count('id')
    ).order_by('-total_paid')[:10]
    
    # Payment trends
    payment_by_method = Payment.objects.filter(
        status='COMPLETED',
        payment_date__year=timezone.now().year
    ).values('payment_method').annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('-total')
    
    context = {
        'current_year': current_year,
        'monthly_spending': list(monthly_spending),
        'category_spending': list(category_spending),
        'department_spending': list(department_spending),
        'top_suppliers': list(top_suppliers),
        'payment_by_method': list(payment_by_method),
    }
    
    return render(request, 'finance/finance_module/analytics.html', context)


# ============================================================================
# BUDGET MANAGEMENT
# ============================================================================

@login_required
def finance_budgets_list_view(request):
    """List all budgets"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    # Filters
    year_id = request.GET.get('year')
    department_id = request.GET.get('department')
    category_id = request.GET.get('category')
    budget_type = request.GET.get('budget_type')
    search = request.GET.get('search', '')
    
    budgets = Budget.objects.select_related(
        'budget_year', 'department', 'category', 'created_by'
    ).all()
    
    if year_id:
        budgets = budgets.filter(budget_year_id=year_id)
    if department_id:
        budgets = budgets.filter(department_id=department_id)
    if category_id:
        budgets = budgets.filter(category_id=category_id)
    if budget_type:
        budgets = budgets.filter(budget_type=budget_type)
    if search:
        budgets = budgets.filter(
            Q(department__name__icontains=search) |
            Q(category__name__icontains=search) |
            Q(reference_number__icontains=search)
        )
    
    budgets = budgets.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(budgets, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    # For filters
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    departments = Department.objects.filter(is_active=True).order_by('name')
    categories = BudgetCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'budget_years': budget_years,
        'departments': departments,
        'categories': categories,
        'total_count': budgets.count(),
    }
    
    return render(request, 'finance/finance_module/budgets_list.html', context)


@login_required
def finance_budget_detail_view(request, pk):
    """Budget detail view"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    budget = get_object_or_404(
        Budget.objects.select_related(
            'budget_year', 'department', 'category', 'created_by'
        ),
        pk=pk
    )
    
    # Get requisitions linked to this budget
    requisitions = Requisition.objects.filter(
        budget=budget
    ).select_related('requested_by', 'department').order_by('-created_at')
    
    # Budget reallocations
    reallocations_from = BudgetReallocation.objects.filter(
        from_budget=budget
    ).select_related('to_budget', 'requested_by', 'approved_by')
    
    reallocations_to = BudgetReallocation.objects.filter(
        to_budget=budget
    ).select_related('from_budget', 'requested_by', 'approved_by')
    
    context = {
        'budget': budget,
        'requisitions': requisitions,
        'reallocations_from': reallocations_from,
        'reallocations_to': reallocations_to,
    }
    
    return render(request, 'finance/finance_module/budget_detail.html', context)


@login_required
def finance_budget_create_view(request):
    """Create new budget"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            budget = form.save(commit=False)
            budget.created_by = request.user
            budget.save()
            
            log_action(request.user, 'CREATE', 'Budget', budget.id, str(budget), request=request)
            messages.success(request, f'Budget created successfully for {budget.department.name}.')
            return redirect('finance_budget_detail', pk=budget.pk)
    else:
        form = BudgetForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'finance/finance_module/budget_form.html', context)


@login_required
def finance_budget_edit_view(request, pk):
    """Edit existing budget"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    budget = get_object_or_404(Budget, pk=pk)
    
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            budget = form.save()
            
            log_action(request.user, 'UPDATE', 'Budget', budget.id, str(budget), request=request)
            messages.success(request, f'Budget updated successfully.')
            return redirect('finance_budget_detail', pk=budget.pk)
    else:
        form = BudgetForm(instance=budget)
    
    context = {
        'form': form,
        'budget': budget,
    }
    
    return render(request, 'finance/finance_module/budget_form.html', context)


@login_required
def finance_budget_allocation_view(request):
    """View and manage budget allocations"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    if not current_year:
        messages.warning(request, 'No active budget year found.')
        return redirect('finance_dashboard')
    
    # Get all budgets for current year
    budgets = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).select_related('department', 'category').order_by('department__name', 'category__name')
    
    # Summary by department
    dept_summary = budgets.values('department__name').annotate(
        total_allocated=Sum('allocated_amount'),
        total_spent=Sum('actual_spent'),
        total_committed=Sum('committed_amount')
    ).order_by('department__name')
    
    context = {
        'current_year': current_year,
        'budgets': budgets,
        'dept_summary': dept_summary,
    }
    
    return render(request, 'finance/finance_module/budget_allocation.html', context)


@login_required
def finance_budget_tracking_view(request):
    """Budget tracking and utilization"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    if not current_year:
        messages.warning(request, 'No active budget year found.')
        return redirect('finance_dashboard')
    
    # Department-wise budget summary
    department_summary = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).values(
        'department__id',
        'department__name',
        'department__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent')
    ).order_by('department__name')
    
    # Calculate available balance for each
    for dept in department_summary:
        dept['available'] = dept['allocated'] - dept['committed'] - dept['spent']
        dept['utilization_percentage'] = (
            (dept['spent'] / dept['allocated'] * 100) if dept['allocated'] > 0 else 0
        )
    
    # Category-wise budget summary
    category_summary = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).values(
        'category__id',
        'category__name',
        'category__code'
    ).annotate(
        allocated=Sum('allocated_amount'),
        committed=Sum('committed_amount'),
        spent=Sum('actual_spent')
    ).order_by('category__name')
    
    for cat in category_summary:
        cat['available'] = cat['allocated'] - cat['committed'] - cat['spent']
        cat['utilization_percentage'] = (
            (cat['spent'] / cat['allocated'] * 100) if cat['allocated'] > 0 else 0
        )
    
    context = {
        'current_year': current_year,
        'department_summary': department_summary,
        'category_summary': category_summary,
    }
    
    return render(request, 'finance/finance_module/budget_tracking.html', context)


@login_required
def finance_budget_reallocate_view(request):
    """Budget reallocation/virement"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = BudgetReallocationForm(request.POST)
        if form.is_valid():
            reallocation = form.save(commit=False)
            reallocation.requested_by = request.user
            reallocation.approved_by = request.user
            reallocation.approved_at = timezone.now()
            reallocation.status = 'APPROVED'
            
            # Update budgets
            from_budget = reallocation.from_budget
            to_budget = reallocation.to_budget
            amount = reallocation.amount
            
            from_budget.allocated_amount -= amount
            from_budget.save()
            
            to_budget.allocated_amount += amount
            to_budget.save()
            
            reallocation.save()
            
            log_action(request.user, 'CREATE', 'BudgetReallocation', reallocation.id, str(reallocation), request=request)
            messages.success(request, f'Budget reallocation of {amount} completed successfully.')
            return redirect('finance_budget_tracking')
    else:
        form = BudgetReallocationForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'finance/finance_module/budget_reallocate.html', context)


# ============================================================================
# INVOICE MANAGEMENT
# ============================================================================

@login_required
def finance_invoices_list_view(request):
    """List all invoices"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    # Filters
    status = request.GET.get('status')
    supplier_id = request.GET.get('supplier')
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    invoices = Invoice.objects.select_related(
        'supplier', 'purchase_order', 'grn', 'verified_by', 'approved_by'
    ).all()
    
    if status:
        invoices = invoices.filter(status=status)
    if supplier_id:
        invoices = invoices.filter(supplier_id=supplier_id)
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(supplier_invoice_number__icontains=search) |
            Q(supplier__name__icontains=search)
        )
    if date_from:
        invoices = invoices.filter(invoice_date__gte=date_from)
    if date_to:
        invoices = invoices.filter(invoice_date__lte=date_to)
    
    invoices = invoices.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(invoices, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    # Statistics
    total_value = invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': invoices.count(),
        'total_value': total_value,
    }
    
    return render(request, 'finance/finance_module/invoices_list.html', context)


@login_required
def finance_invoice_detail_view(request, pk):
    """Invoice detail view"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    invoice = get_object_or_404(
        Invoice.objects.select_related(
            'supplier', 'purchase_order', 'grn',
            'verified_by', 'approved_by', 'submitted_by'
        ),
        pk=pk
    )
    
    # Get invoice items
    items = invoice.items.select_related('po_item').all()
    
    # Get invoice documents
    documents = invoice.documents.all()
    
    # Get payments for this invoice
    payments = invoice.payments.select_related(
        'processed_by', 'approved_by'
    ).order_by('-created_at')
    
    context = {
        'invoice': invoice,
        'items': items,
        'documents': documents,
        'payments': payments,
    }
    
    return render(request, 'finance/finance_module/invoice_detail.html', context)


@login_required
def finance_invoice_verify_view(request, pk):
    """Verify invoice (3-way matching)"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if invoice.status not in ['SUBMITTED', 'VERIFYING']:
        messages.warning(request, 'This invoice cannot be verified.')
        return redirect('finance_invoice_detail', pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        comments = request.POST.get('comments', '')
        
        if action == 'approve':
            invoice.status = 'MATCHED'
            invoice.verified_by = request.user
            invoice.verified_at = timezone.now()
            invoice.matching_notes = comments
            invoice.is_three_way_matched = True
            invoice.save()
            
            log_action(request.user, 'APPROVE', 'Invoice', invoice.id, str(invoice), request=request)
            messages.success(request, f'Invoice {invoice.invoice_number} verified successfully.')
            
        elif action == 'dispute':
            invoice.status = 'DISPUTED'
            invoice.dispute_reason = comments
            invoice.verified_by = request.user
            invoice.verified_at = timezone.now()
            invoice.save()
            
            log_action(request.user, 'REJECT', 'Invoice', invoice.id, str(invoice), request=request)
            messages.info(request, f'Invoice {invoice.invoice_number} marked as disputed.')
        
        return redirect('finance_invoice_detail', pk=pk)
    
    # Get PO and GRN for comparison
    po = invoice.purchase_order
    grn = invoice.grn
    
    context = {
        'invoice': invoice,
        'po': po,
        'grn': grn,
    }
    
    return render(request, 'finance/finance_module/invoice_verify.html', context)


@login_required
def finance_invoice_approve_view(request, pk):
    """Approve invoice for payment"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    invoice = get_object_or_404(Invoice, pk=pk)
    
    if invoice.status != 'MATCHED':
        messages.warning(request, 'Invoice must be verified before approval.')
        return redirect('finance_invoice_detail', pk=pk)
    
    if request.method == 'POST':
        comments = request.POST.get('comments', '')
        
        invoice.status = 'APPROVED'
        invoice.approved_by = request.user
        invoice.approved_at = timezone.now()
        invoice.balance_due = invoice.total_amount  # Set initial balance
        if comments:
            invoice.notes = f"{invoice.notes}\n\nApproval Notes: {comments}" if invoice.notes else f"Approval Notes: {comments}"
        invoice.save()
        
        log_action(request.user, 'APPROVE', 'Invoice', invoice.id, str(invoice), request=request)
        messages.success(request, f'Invoice {invoice.invoice_number} approved for payment.')
        
        return redirect('finance_invoice_detail', pk=pk)
    
    context = {
        'invoice': invoice,
    }
    
    return render(request, 'finance/finance_module/invoice_approve.html', context)


@login_required
def finance_pending_invoices_view(request):
    """List pending invoices"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    invoices = Invoice.objects.filter(
        status__in=['SUBMITTED', 'VERIFYING']
    ).select_related('supplier', 'purchase_order').order_by('-created_at')
    
    paginator = Paginator(invoices, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_value = invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': invoices.count(),
        'total_value': total_value,
    }
    
    return render(request, 'finance/finance_module/pending_invoices.html', context)


@login_required
def finance_paid_invoices_view(request):
    """List paid invoices"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    invoices = Invoice.objects.filter(
        status='PAID'
    ).select_related('supplier', 'purchase_order').order_by('-payment_date')
    
    paginator = Paginator(invoices, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_paid = invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': invoices.count(),
        'total_paid': total_paid,
    }
    
    return render(request, 'finance/finance_module/paid_invoices.html', context)


@login_required
def finance_overdue_invoices_view(request):
    """List overdue invoices"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    today = timezone.now().date()
    
    invoices = Invoice.objects.filter(
        status__in=['APPROVED', 'MATCHED'],
        due_date__lt=today
    ).select_related('supplier', 'purchase_order').order_by('due_date')
    
    paginator = Paginator(invoices, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_overdue = invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': invoices.count(),
        'total_overdue': total_overdue,
    }
    
    return render(request, 'finance/finance_module/overdue_invoices.html', context)


# ============================================================================
# PAYMENT MANAGEMENT
# ============================================================================

@login_required
def finance_payments_list_view(request):
    """List all payments"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    # Filters
    status = request.GET.get('status')
    search = request.GET.get('search', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    payments = Payment.objects.select_related(
        'invoice', 'invoice__supplier', 'processed_by', 'approved_by'
    ).all()
    
    if status:
        payments = payments.filter(status=status)
    if search:
        payments = payments.filter(
            Q(payment_number__icontains=search) |
            Q(payment_reference__icontains=search) |
            Q(invoice__invoice_number__icontains=search) |
            Q(invoice__supplier__name__icontains=search)
        )
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
    
    payments = payments.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(payments, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_amount = payments.aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': payments.count(),
        'total_amount': total_amount,
    }
    
    return render(request, 'finance/finance_module/payments_list.html', context)


@login_required
def finance_payment_detail_view(request, pk):
    """Payment detail view"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    payment = get_object_or_404(
        Payment.objects.select_related(
            'invoice', 'invoice__supplier', 'invoice__purchase_order',
            'processed_by', 'approved_by'
        ),
        pk=pk
    )
    
    context = {
        'payment': payment,
    }
    
    return render(request, 'finance/finance_module/payment_detail.html', context)


@login_required
def finance_process_payment_view(request, invoice_id):
    """Process payment for an invoice"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    invoice = get_object_or_404(
        Invoice.objects.select_related('supplier', 'purchase_order'),
        pk=invoice_id
    )
    
    if invoice.status != 'APPROVED':
        messages.error(request, 'Invoice must be approved before payment.')
        return redirect('finance_invoice_detail', pk=invoice_id)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.invoice = invoice
            payment.processed_by = request.user
            payment.status = 'PENDING'
            payment.save()
            
            # Update invoice
            invoice.update_payment_status()
            
            log_action(request.user, 'CREATE', 'Payment', payment.id, str(payment), request=request)
            messages.success(request, f'Payment {payment.payment_number} created successfully.')
            
            return redirect('finance_payment_detail', pk=payment.pk)
    else:
        form = PaymentForm(initial={
            'payment_date': timezone.now().date(),
            'payment_amount': invoice.balance_due,
        })
    
    context = {
        'form': form,
        'invoice': invoice,
    }
    
    return render(request, 'finance/finance_module/process_payment.html', context)


@login_required
def finance_approve_payment_view(request, pk):
    """Approve a payment"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    payment = get_object_or_404(Payment, pk=pk)
    
    if payment.status != 'PENDING':
        messages.warning(request, 'Payment is not pending approval.')
        return redirect('finance_payment_detail', pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            payment.status = 'COMPLETED'
            payment.approved_by = request.user
            payment.save()
            
            # Update invoice payment status
            payment.invoice.update_payment_status()
            
            log_action(request.user, 'APPROVE', 'Payment', payment.id, str(payment), request=request)
            messages.success(request, f'Payment {payment.payment_number} approved.')
            
        elif action == 'reject':
            payment.status = 'CANCELLED'
            payment.notes = request.POST.get('notes', 'Rejected by finance')
            payment.save()
            
            log_action(request.user, 'REJECT', 'Payment', payment.id, str(payment), request=request)
            messages.info(request, f'Payment {payment.payment_number} rejected.')
        
        return redirect('finance_payment_detail', pk=pk)
    
    context = {
        'payment': payment,
    }
    
    return render(request, 'finance/finance_module/approve_payment.html', context)


@login_required
def finance_payment_schedule_view(request):
    """Payment schedule view"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    # Get approved invoices awaiting payment
    upcoming_payments = Invoice.objects.filter(
        status='APPROVED'
    ).select_related('supplier', 'purchase_order').order_by('due_date')
    
    paginator = Paginator(upcoming_payments, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_scheduled = upcoming_payments.aggregate(Sum('balance_due'))['balance_due__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': upcoming_payments.count(),
        'total_scheduled': total_scheduled,
    }
    
    return render(request, 'finance/finance_module/payment_schedule.html', context)


@login_required
def finance_payment_history_view(request):
    """Payment history"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    payments = Payment.objects.filter(
        status='COMPLETED'
    ).select_related(
        'invoice', 'invoice__supplier', 'processed_by', 'approved_by'
    ).order_by('-payment_date')
    
    paginator = Paginator(payments, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_paid = payments.aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': payments.count(),
        'total_paid': total_paid,
    }
    
    return render(request, 'finance/finance_module/payment_history.html', context)


# ============================================================================
# APPROVALS
# ============================================================================

@login_required
def finance_pending_approvals_view(request):
    """Pending budget approvals"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    approvals = RequisitionApproval.objects.filter(
        approval_stage='BUDGET',
        status='PENDING'
    ).select_related(
        'requisition', 'requisition__department', 'requisition__requested_by'
    ).order_by('-created_at')
    
    paginator = Paginator(approvals, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': approvals.count(),
    }
    
    return render(request, 'finance/finance_module/pending_approvals.html', context)


@login_required
def finance_approve_requisition_view(request, requisition_id):
    """Approve or reject requisition (budget check)"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    requisition = get_object_or_404(
        Requisition.objects.select_related('department', 'budget', 'requested_by'),
        pk=requisition_id
    )
    
    approval = RequisitionApproval.objects.filter(
        requisition=requisition,
        approval_stage='BUDGET',
        status='PENDING'
    ).first()
    
    if not approval:
        messages.error(request, 'Budget approval not found or already processed.')
        return redirect('finance_pending_approvals')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        comments = request.POST.get('comments', '')
        
        if action == 'approve':
            approval.status = 'APPROVED'
            approval.approver = request.user
            approval.approval_date = timezone.now()
            approval.comments = comments
            approval.save()
            
            requisition.status = 'BUDGET_APPROVED'
            requisition.save()
            
            log_action(request.user, 'APPROVE', 'Requisition', requisition.id, str(requisition), request=request)
            messages.success(request, f'Requisition {requisition.requisition_number} approved.')
            
        elif action == 'reject':
            approval.status = 'REJECTED'
            approval.approver = request.user
            approval.approval_date = timezone.now()
            approval.comments = comments
            approval.save()
            
            requisition.status = 'REJECTED'
            requisition.rejection_reason = comments
            requisition.save()
            
            log_action(request.user, 'REJECT', 'Requisition', requisition.id, str(requisition), request=request)
            messages.info(request, f'Requisition {requisition.requisition_number} rejected.')
        
        return redirect('finance_pending_approvals')
    
    context = {
        'requisition': requisition,
        'approval': approval,
    }
    
    return render(request, 'finance/finance_module/approve_requisition.html', context)


@login_required
def finance_approved_requisitions_view(request):
    """List approved requisitions"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    requisitions = Requisition.objects.filter(
        status__in=['BUDGET_APPROVED', 'APPROVED']
    ).select_related('department', 'requested_by', 'budget').order_by('-updated_at')
    
    paginator = Paginator(requisitions, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': requisitions.count(),
    }
    
    return render(request, 'finance/finance_module/approved_requisitions.html', context)


@login_required
def finance_rejected_requisitions_view(request):
    """List rejected requisitions"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    requisitions = Requisition.objects.filter(
        status='REJECTED'
    ).select_related(
        'department', 
        'requested_by', 
        'budget'
    ).prefetch_related('approvals').order_by('-updated_at')
    
    # Add rejection info from approvals
    for req in requisitions:
        rejection_approval = req.approvals.filter(status='REJECTED').first()
        if rejection_approval:
            req.rejected_by = rejection_approval.approver
            req.rejected_by_role = rejection_approval.get_approval_stage_display()
            req.rejection_comments = rejection_approval.comments
        else:
            req.rejected_by = None
            req.rejected_by_role = ''
            req.rejection_comments = ''
    
    paginator = Paginator(requisitions, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    # Calculate summary stats
    total_amount = sum(req.estimated_amount for req in requisitions)
    departments = requisitions.values_list('department', flat=True).distinct()
    
    # Calculate average review days
    avg_days = 0
    if requisitions:
        total_days = sum(
            (req.updated_at - req.created_at).days 
            for req in requisitions
        )
        avg_days = total_days / requisitions.count()
    
    context = {
        'page_obj': page_obj,
        'total_count': requisitions.count(),
        'total_amount': total_amount,
        'departments': Department.objects.filter(id__in=departments),
        'avg_days': avg_days,
    }
    
    return render(request, 'finance/finance_module/rejected_requisitions.html', context)

# ============================================================================
# REPORTS
# ============================================================================

@login_required
def finance_expenditure_report_view(request):
    """Expenditure report"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    # Get filter parameters
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    department_id = request.GET.get('department')
    
    payments = Payment.objects.filter(
        status='COMPLETED',
        payment_date__gte=date_from,
        payment_date__lte=date_to
    ).select_related('invoice', 'invoice__supplier', 'invoice__purchase_order__requisition__department')
    
    if department_id:
        payments = payments.filter(invoice__purchase_order__requisition__department_id=department_id)
    
    # Summary by department
    dept_expenditure = payments.values(
        'invoice__purchase_order__requisition__department__name'
    ).annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Summary by supplier
    supplier_expenditure = payments.values(
        'invoice__supplier__name'
    ).annotate(
        total=Sum('payment_amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    total_expenditure = payments.aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
    
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'payments': payments[:50],  # Limit display
        'dept_expenditure': dept_expenditure,
        'supplier_expenditure': supplier_expenditure,
        'total_expenditure': total_expenditure,
        'departments': departments,
    }
    
    return render(request, 'finance/finance_module/expenditure_report.html', context)


@login_required
def finance_budget_vs_actual_view(request):
    """Budget vs Actual report"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    if not current_year:
        messages.warning(request, 'No active budget year found.')
        return redirect('finance_dashboard')
    
    budgets = Budget.objects.filter(
        budget_year=current_year,
        is_active=True
    ).select_related('department', 'category').annotate(
        variance=F('allocated_amount') - F('actual_spent'),
        variance_percentage=(F('actual_spent') / F('allocated_amount') * 100)
    ).order_by('department__name', 'category__name')
    
    context = {
        'current_year': current_year,
        'budgets': budgets,
    }
    
    return render(request, 'finance/finance_module/budget_vs_actual.html', context)


@login_required
def finance_financial_statements_view(request):
    """Financial statements"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    current_year = BudgetYear.objects.filter(is_active=True).first()
    
    # Income statement components
    total_payments = Payment.objects.filter(
        status='COMPLETED',
        payment_date__year=timezone.now().year
    ).aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
    
    # Balance sheet components
    outstanding_invoices = Invoice.objects.filter(
        status='APPROVED'
    ).aggregate(Sum('balance_due'))['balance_due__sum'] or 0
    
    context = {
        'current_year': current_year,
        'total_payments': total_payments,
        'outstanding_invoices': outstanding_invoices,
    }
    
    return render(request, 'finance/finance_module/financial_statements.html', context)


@login_required
def finance_transaction_report_view(request):
    """Transaction report"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    
    payments = Payment.objects.filter(
        payment_date__gte=date_from,
        payment_date__lte=date_to
    ).select_related('invoice', 'invoice__supplier').order_by('-payment_date')
    
    paginator = Paginator(payments, 50)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_amount = payments.aggregate(Sum('payment_amount'))['payment_amount__sum'] or 0
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'page_obj': page_obj,
        'total_amount': total_amount,
    }
    
    return render(request, 'finance/finance_module/transaction_report.html', context)


@login_required
def finance_export_report_view(request):
    """Export reports to CSV/PDF"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    report_type = request.GET.get('type', 'payments')
    export_format = request.GET.get('format', 'csv')
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    
    if report_type == 'payments':
        data = Payment.objects.filter(
            status='COMPLETED',
            payment_date__gte=date_from,
            payment_date__lte=date_to
        ).select_related('invoice', 'invoice__supplier')
        
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="payments_{timezone.now().strftime("%Y%m%d")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Payment Number', 'Date', 'Supplier', 'Invoice Number', 'Amount', 'Method', 'Reference'])
            
            for payment in data:
                writer.writerow([
                    payment.payment_number,
                    payment.payment_date,
                    payment.invoice.supplier.name,
                    payment.invoice.invoice_number,
                    payment.payment_amount,
                    payment.get_payment_method_display(),
                    payment.payment_reference
                ])
            
            return response
    
    messages.warning(request, 'Export format not implemented yet.')
    return redirect('finance_dashboard')


@login_required
def finance_help_view(request):
    """Finance help and documentation"""
    if not check_finance_permission(request.user):
        messages.error(request, 'You do not have finance officer permissions.')
        return redirect('dashboard')
    
    context = {}
    return render(request, 'finance/finance_module/help.html', context)

# procurement/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal

from .models import (
    Requisition, Tender, Bid, PurchaseOrder, Supplier, Contract,
    RequisitionApproval, BidEvaluation, Department, ItemCategory
)


def check_procurement_permission(user):
    """Check if user has procurement officer permissions"""
    return user.role in ['PROCUREMENT', 'ADMIN']


# ============================================================================
# DASHBOARD & ANALYTICS
# ============================================================================

@login_required
def procurement_dashboard_view(request):
    """Procurement officer dashboard"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    # Get statistics
    pending_requisitions = Requisition.objects.filter(
        status='PROCUREMENT_APPROVED'
    ).count()
    
    active_tenders = Tender.objects.filter(
        status='PUBLISHED',
        closing_date__gte=timezone.now()
    ).count()
    
    pending_pos = PurchaseOrder.objects.filter(
        status__in=['DRAFT', 'PENDING_APPROVAL']
    ).count()
    
    active_suppliers = Supplier.objects.filter(
        status='APPROVED'
    ).count()
    
    # Recent activities
    recent_requisitions = Requisition.objects.filter(
        status__in=['SUBMITTED', 'HOD_APPROVED', 'BUDGET_APPROVED']
    ).select_related('department', 'requested_by').order_by('-created_at')[:5]
    
    recent_tenders = Tender.objects.all().select_related(
        'requisition'
    ).order_by('-created_at')[:5]
    
    recent_pos = PurchaseOrder.objects.all().select_related(
        'supplier', 'requisition'
    ).order_by('-created_at')[:5]
    
    # Monthly trends
    current_month = timezone.now().month
    monthly_spend = PurchaseOrder.objects.filter(
        po_date__month=current_month,
        status__in=['APPROVED', 'SENT', 'DELIVERED']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'pending_requisitions': pending_requisitions,
        'active_tenders': active_tenders,
        'pending_pos': pending_pos,
        'active_suppliers': active_suppliers,
        'recent_requisitions': recent_requisitions,
        'recent_tenders': recent_tenders,
        'recent_pos': recent_pos,
        'monthly_spend': monthly_spend,
    }
    
    return render(request, 'procurement/procurement_module/dashboard.html', context)


@login_required
def procurement_analytics_view(request):
    """Procurement analytics and insights"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    # Date range filter
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=90)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    
    # Spend analysis
    total_spend = PurchaseOrder.objects.filter(
        po_date__gte=date_from,
        po_date__lte=date_to,
        status__in=['APPROVED', 'SENT', 'DELIVERED']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Department spending
    dept_spending = PurchaseOrder.objects.filter(
        po_date__gte=date_from,
        po_date__lte=date_to,
        status__in=['APPROVED', 'SENT', 'DELIVERED']
    ).values(
        'requisition__department__name'
    ).annotate(
        total=Sum('total_amount')
    ).order_by('-total')[:10]
    
    # Supplier performance
    supplier_stats = PurchaseOrder.objects.filter(
        po_date__gte=date_from,
        po_date__lte=date_to
    ).values(
        'supplier__name'
    ).annotate(
        order_count=Count('id'),
        total_value=Sum('total_amount')
    ).order_by('-total_value')[:10]
    
    # Processing times
    avg_processing_time = Requisition.objects.filter(
        status='APPROVED',
        submitted_at__gte=date_from,
        updated_at__lte=date_to
    ).annotate(
        processing_days=(timezone.now() - timezone.now())  # Placeholder
    ).count()
    
    context = {
        'date_from': date_from,
        'date_to': date_to,
        'total_spend': total_spend,
        'dept_spending': dept_spending,
        'supplier_stats': supplier_stats,
        'avg_processing_time': avg_processing_time,
    }
    
    return render(request, 'procurement/procurement_module/analytics.html', context)


# ============================================================================
# REQUISITIONS MANAGEMENT
# ============================================================================

@login_required
def procurement_all_requisitions_view(request):
    """List all requisitions"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    # Filters
    status = request.GET.get('status', '')
    department = request.GET.get('department', '')
    search = request.GET.get('search', '')
    
    requisitions = Requisition.objects.all().select_related(
        'department', 'requested_by', 'budget'
    ).order_by('-created_at')
    
    if status:
        requisitions = requisitions.filter(status=status)
    if department:
        requisitions = requisitions.filter(department_id=department)
    if search:
        requisitions = requisitions.filter(
            Q(requisition_number__icontains=search) |
            Q(title__icontains=search)
        )
    
    paginator = Paginator(requisitions, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    departments = Department.objects.all()
    
    context = {
        'page_obj': page_obj,
        'departments': departments,
        'total_count': requisitions.count(),
        'status_choices': Requisition.STATUS_CHOICES,
    }
    
    return render(request, 'procurement/procurement_module/all_requisitions.html', context)


@login_required
def procurement_pending_requisitions_view(request):
    """List pending requisitions for procurement processing"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    requisitions = Requisition.objects.filter(
        status__in=['SUBMITTED', 'HOD_APPROVED', 'BUDGET_APPROVED']
    ).select_related(
        'department', 'requested_by', 'budget'
    ).order_by('-created_at')
    
    paginator = Paginator(requisitions, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_value = requisitions.aggregate(Sum('estimated_amount'))['estimated_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': requisitions.count(),
        'total_value': total_value,
    }
    
    return render(request, 'procurement/procurement_module/pending_requisitions.html', context)


@login_required
def procurement_processed_requisitions_view(request):
    """List processed requisitions"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    requisitions = Requisition.objects.filter(
        status__in=['APPROVED', 'PROCUREMENT_APPROVED']
    ).select_related(
        'department', 'requested_by', 'budget'
    ).order_by('-updated_at')
    
    paginator = Paginator(requisitions, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': requisitions.count(),
    }
    
    return render(request, 'procurement/procurement_module/processed_requisitions.html', context)


# ============================================================================
# TENDERS MANAGEMENT
# ============================================================================

@login_required
def procurement_active_tenders_view(request):
    """List active tenders"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    tenders = Tender.objects.filter(
        status='PUBLISHED',
        closing_date__gte=timezone.now()
    ).select_related('requisition').order_by('closing_date')
    
    paginator = Paginator(tenders, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': tenders.count(),
    }
    
    return render(request, 'procurement/procurement_module/active_tenders.html', context)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from .models import (
    Tender, Requisition, Supplier, TenderDocument,
    Notification
)
from django.utils import timezone


@login_required
def procurement_create_tender_view(request):
    """Create new tender with document uploads"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Extract form data
                requisition_id = request.POST.get('requisition')
                title = request.POST.get('title')
                tender_type = request.POST.get('tender_type')
                procurement_method = request.POST.get('procurement_method')
                description = request.POST.get('description')
                estimated_budget = request.POST.get('estimated_budget')
                closing_date = request.POST.get('closing_date')
                bid_opening_date = request.POST.get('bid_opening_date')
                supplier_ids = request.POST.getlist('suppliers')
                
                # Create tender
                requisition = Requisition.objects.get(id=requisition_id)
                tender = Tender.objects.create(
                    requisition=requisition,
                    title=title,
                    tender_type=tender_type,
                    procurement_method=procurement_method,
                    description=description,
                    estimated_budget=estimated_budget,
                    closing_date=closing_date,
                    bid_opening_date=bid_opening_date,
                    status='DRAFT',
                    created_by=request.user
                )
                
                # Add invited suppliers if any
                if supplier_ids:
                    suppliers = Supplier.objects.filter(id__in=supplier_ids)
                    tender.invited_suppliers.set(suppliers)
                
                # Handle document uploads
                uploaded_files = request.FILES.getlist('tender_documents')
                document_names = request.POST.getlist('document_names')
                document_descriptions = request.POST.getlist('document_descriptions')
                mandatory_flags = request.POST.getlist('document_mandatory')
                
                for i, file in enumerate(uploaded_files):
                    doc_name = document_names[i] if i < len(document_names) else file.name
                    doc_desc = document_descriptions[i] if i < len(document_descriptions) else ''
                    is_mandatory = str(i) in mandatory_flags
                    
                    TenderDocument.objects.create(
                        tender=tender,
                        document_name=doc_name,
                        file=file,
                        description=doc_desc,
                        is_mandatory=is_mandatory
                    )
                
                # Create notifications for invited suppliers
                if supplier_ids:
                    for supplier_id in supplier_ids:
                        supplier = Supplier.objects.get(id=supplier_id)
                        if supplier.user:
                            Notification.objects.create(
                                user=supplier.user,
                                notification_type='TENDER',
                                priority='HIGH',
                                title=f'New Tender Invitation: {tender.title}',
                                message=f'You have been invited to submit a bid for tender {tender.tender_number}. Closing date: {tender.closing_date.strftime("%d %b %Y %H:%M")}',
                                link_url=f'/suppliers/tenders/{tender.id}/'
                            )
                
                messages.success(
                    request, 
                    f'Tender {tender.tender_number} created successfully with {len(uploaded_files)} document(s).'
                )
                return redirect('procurement_active_tenders')
                
        except Requisition.DoesNotExist:
            messages.error(request, 'Selected requisition does not exist.')
        except Exception as e:
            messages.error(request, f'Error creating tender: {str(e)}')
            return redirect('procurement_create_tender')
    
    # GET request - show form
    requisitions = Requisition.objects.filter(
        status='APPROVED'
    ).exclude(
        tenders__isnull=False
    ).select_related('department')
    
    suppliers = Supplier.objects.filter(status='APPROVED')
    
    # Common tender document types
    document_types = [
        'Invitation to Bid (ITB)',
        'Terms of Reference (TOR)',
        'Technical Specifications',
        'Bill of Quantities (BOQ)',
        'Bid Data Sheet',
        'General Conditions of Contract',
        'Special Conditions of Contract',
        'Sample Forms (Bid Form, Performance Security)',
        'Evaluation Criteria',
        'Tender Drawings/Plans',
    ]
    
    context = {
        'requisitions': requisitions,
        'suppliers': suppliers,
        'tender_types': Tender.TENDER_TYPES,
        'methods': Tender.METHOD_CHOICES,
        'document_types': document_types,
    }
    
    return render(request, 'procurement/procurement_module/create_tender.html', context)


@login_required
def procurement_closed_tenders_view(request):
    """List closed tenders"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    tenders = Tender.objects.filter(
        status__in=['CLOSED', 'AWARDED', 'CANCELLED']
    ).select_related('requisition').order_by('-closing_date')
    
    paginator = Paginator(tenders, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': tenders.count(),
    }
    
    return render(request, 'procurement/procurement_module/closed_tenders.html', context)


@login_required
def procurement_tender_evaluation_view(request):
    """Tender evaluation dashboard"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    tenders = Tender.objects.filter(
        status__in=['CLOSED', 'EVALUATING']
    ).select_related('requisition').order_by('-closing_date')
    
    # Get tenders with bids
    tenders_with_bids = []
    for tender in tenders:
        bid_count = tender.bids.count()
        if bid_count > 0:
            tenders_with_bids.append({
                'tender': tender,
                'bid_count': bid_count,
                'evaluated_count': tender.bids.filter(evaluations__isnull=False).distinct().count()
            })
    
    context = {
        'tenders_with_bids': tenders_with_bids,
    }
    
    return render(request, 'procurement/procurement_module/tender_evaluation.html', context)


# ============================================================================
# BIDS MANAGEMENT
# ============================================================================

@login_required
def procurement_bids_management_view(request):
    """Manage all bids"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    status = request.GET.get('status', '')
    tender_id = request.GET.get('tender', '')
    
    bids = Bid.objects.all().select_related(
        'tender', 'supplier'
    ).order_by('-submitted_at')
    
    if status:
        bids = bids.filter(status=status)
    if tender_id:
        bids = bids.filter(tender_id=tender_id)
    
    paginator = Paginator(bids, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    tenders = Tender.objects.filter(status__in=['PUBLISHED', 'CLOSED', 'EVALUATING'])
    
    context = {
        'page_obj': page_obj,
        'tenders': tenders,
        'status_choices': Bid.STATUS_CHOICES,
        'total_count': bids.count(),
    }
    
    return render(request, 'procurement/procurement_module/bids_management.html', context)


# ============================================================================
# PURCHASE ORDERS
# ============================================================================

@login_required
def procurement_all_orders_view(request):
    """List all purchase orders"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    orders = PurchaseOrder.objects.all().select_related(
        'supplier', 'requisition'
    ).order_by('-created_at')
    
    if status:
        orders = orders.filter(status=status)
    if search:
        orders = orders.filter(
            Q(po_number__icontains=search) |
            Q(supplier__name__icontains=search)
        )
    
    paginator = Paginator(orders, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_value = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    context = {
        'page_obj': page_obj,
        'total_count': orders.count(),
        'total_value': total_value,
        'status_choices': PurchaseOrder.STATUS_CHOICES,
    }
    
    return render(request, 'procurement/procurement_module/all_orders.html', context)


@login_required
def procurement_create_order_view(request):
    """Create new purchase order"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Handle PO creation
        messages.success(request, 'Purchase order created successfully.')
        return redirect('procurement_all_orders')
    
    # Get awarded bids without POs
    awarded_bids = Bid.objects.filter(
        status='AWARDED',
        purchase_order__isnull=True
    ).select_related('tender', 'supplier')
    
    suppliers = Supplier.objects.filter(status='APPROVED')
    
    context = {
        'awarded_bids': awarded_bids,
        'suppliers': suppliers,
    }
    
    return render(request, 'procurement/procurement_module/create_order.html', context)


@login_required
def procurement_pending_orders_view(request):
    """List pending purchase orders"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    orders = PurchaseOrder.objects.filter(
        status__in=['DRAFT', 'PENDING_APPROVAL']
    ).select_related('supplier', 'requisition').order_by('-created_at')
    
    paginator = Paginator(orders, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': orders.count(),
    }
    
    return render(request, 'procurement/procurement_module/pending_orders.html', context)


@login_required
def procurement_completed_orders_view(request):
    """List completed purchase orders"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    orders = PurchaseOrder.objects.filter(
        status__in=['DELIVERED', 'CLOSED']
    ).select_related('supplier', 'requisition').order_by('-updated_at')
    
    paginator = Paginator(orders, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': orders.count(),
    }
    
    return render(request, 'procurement/procurement_module/completed_orders.html', context)


# ============================================================================
# CONTRACTS
# ============================================================================

@login_required
def procurement_contracts_view(request):
    """List all contracts"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    status = request.GET.get('status', '')
    contract_type = request.GET.get('type', '')
    
    contracts = Contract.objects.all().select_related(
        'supplier', 'purchase_order'
    ).order_by('-created_at')
    
    if status:
        contracts = contracts.filter(status=status)
    if contract_type:
        contracts = contracts.filter(contract_type=contract_type)
    
    paginator = Paginator(contracts, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    total_value = contracts.aggregate(Sum('contract_value'))['contract_value__sum'] or 0
    active_count = contracts.filter(status='ACTIVE').count()
    
    context = {
        'page_obj': page_obj,
        'total_count': contracts.count(),
        'total_value': total_value,
        'active_count': active_count,
        'status_choices': Contract.STATUS_CHOICES,
        'type_choices': Contract.CONTRACT_TYPES,
    }
    
    return render(request, 'procurement/procurement_module/contracts.html', context)


# ============================================================================
# SUPPLIERS
# ============================================================================

@login_required
def procurement_all_suppliers_view(request):
    """List all suppliers"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    suppliers = Supplier.objects.all().prefetch_related('categories').order_by('name')
    
    if status:
        suppliers = suppliers.filter(status=status)
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(supplier_number__icontains=search) |
            Q(email__icontains=search)
        )
    
    paginator = Paginator(suppliers, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': suppliers.count(),
        'status_choices': Supplier.STATUS_CHOICES,
    }
    
    return render(request, 'procurement/procurement_module/all_suppliers.html', context)


@login_required
def procurement_add_supplier_view(request):
    """Add new supplier"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        # Handle supplier creation
        messages.success(request, 'Supplier added successfully.')
        return redirect('procurement_all_suppliers')
    
    categories = ItemCategory.objects.filter(is_active=True)
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'procurement/procurement_module/add_supplier.html', context)


@login_required
def procurement_supplier_evaluation_view(request):
    """Supplier performance evaluation"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    suppliers = Supplier.objects.filter(
        status='APPROVED'
    ).annotate(
        order_count=Count('purchase_orders'),
        total_value=Sum('purchase_orders__total_amount'),
        avg_rating=Avg('performances__overall_rating')
    ).order_by('-avg_rating')
    
    paginator = Paginator(suppliers, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': suppliers.count(),
    }
    
    return render(request, 'procurement/procurement_module/supplier_evaluation.html', context)


@login_required
def procurement_blacklisted_suppliers_view(request):
    """List blacklisted suppliers"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    suppliers = Supplier.objects.filter(
        status='BLACKLISTED'
    ).order_by('-updated_at')
    
    paginator = Paginator(suppliers, 20)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    
    context = {
        'page_obj': page_obj,
        'total_count': suppliers.count(),
    }
    
    return render(request, 'procurement/procurement_module/blacklisted_suppliers.html', context)


# ============================================================================
# REPORTS
# ============================================================================
@login_required
def procurement_reports_view(request):
    """Procurement reports dashboard with comprehensive analytics"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    from django.db.models import Sum, Count, Q, Avg, F
    from django.db.models.functions import TruncMonth, TruncQuarter
    from decimal import Decimal
    import json
    
    # Date filters
    current_year = timezone.now().year
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Base querysets
    requisitions_qs = Requisition.objects.all()
    tenders_qs = Tender.objects.all()
    pos_qs = PurchaseOrder.objects.all()
    invoices_qs = Invoice.objects.all()
    
    # Apply date filters if provided
    if start_date:
        requisitions_qs = requisitions_qs.filter(created_at__gte=start_date)
        tenders_qs = tenders_qs.filter(created_at__gte=start_date)
        pos_qs = pos_qs.filter(created_at__gte=start_date)
        invoices_qs = invoices_qs.filter(created_at__gte=start_date)
    
    if end_date:
        requisitions_qs = requisitions_qs.filter(created_at__lte=end_date)
        tenders_qs = tenders_qs.filter(created_at__lte=end_date)
        pos_qs = pos_qs.filter(created_at__lte=end_date)
        invoices_qs = invoices_qs.filter(created_at__lte=end_date)
    
    # ==========================================
    # 1. SUMMARY STATISTICS
    # ==========================================
    total_requisitions = requisitions_qs.count()
    total_tenders = tenders_qs.count()
    total_pos = pos_qs.count()
    total_suppliers = Supplier.objects.filter(status='APPROVED').count()
    
    total_spend = pos_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    total_budget = Budget.objects.filter(
        budget_year__is_active=True
    ).aggregate(Sum('allocated_amount'))['allocated_amount__sum'] or Decimal('0')
    
    budget_utilization = (total_spend / total_budget * 100) if total_budget > 0 else 0
    
    # Pending approvals
    pending_requisitions = requisitions_qs.exclude(
        status__in=['APPROVED', 'REJECTED', 'CANCELLED']
    ).count()
    
    pending_invoices = invoices_qs.exclude(
        status__in=['PAID', 'REJECTED']
    ).count()
    
    # ==========================================
    # 2. LINE CHART - Monthly Spend Trend
    # ==========================================
    monthly_spend = pos_qs.filter(
        created_at__year=current_year
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('month')
    
    spend_trend_labels = []
    spend_trend_amounts = []
    spend_trend_count = []
    
    for item in monthly_spend:
        spend_trend_labels.append(item['month'].strftime('%B'))
        spend_trend_amounts.append(float(item['total'] or 0))
        spend_trend_count.append(item['count'])
    
    # ==========================================
    # 3. BAR CHART - Spend by Department
    # ==========================================
    department_spend = pos_qs.values(
        'requisition__department__name'
    ).annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')[:10]  # Top 10 departments
    
    dept_labels = []
    dept_amounts = []
    dept_colors = []
    
    # Color palette for departments
    color_palette = [
        '#2563EB', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
        '#EC4899', '#14B8A6', '#F97316', '#06B6D4', '#84CC16'
    ]
    
    for idx, item in enumerate(department_spend):
        dept_labels.append(item['requisition__department__name'] or 'Unknown')
        dept_amounts.append(float(item['total'] or 0))
        dept_colors.append(color_palette[idx % len(color_palette)])
    
    # ==========================================
    # 4. DONUT CHART - Requisition Status Distribution
    # ==========================================
    req_status_data = requisitions_qs.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    req_status_labels = []
    req_status_values = []
    req_status_colors = {
        'DRAFT': '#94A3B8',
        'SUBMITTED': '#3B82F6',
        'HOD_APPROVED': '#8B5CF6',
        'FACULTY_APPROVED': '#EC4899',
        'BUDGET_APPROVED': '#14B8A6',
        'PROCUREMENT_APPROVED': '#10B981',
        'APPROVED': '#22C55E',
        'REJECTED': '#EF4444',
        'CANCELLED': '#6B7280',
    }
    
    donut_colors = []
    for item in req_status_data:
        status_display = dict(Requisition.STATUS_CHOICES).get(
            item['status'], item['status']
        )
        req_status_labels.append(status_display)
        req_status_values.append(item['count'])
        donut_colors.append(req_status_colors.get(item['status'], '#64748B'))
    
    # ==========================================
    # 5. RADAR CHART - Supplier Performance Metrics
    # ==========================================
    # Get top 5 suppliers by number of orders
    top_suppliers = pos_qs.values(
        'supplier__name'
    ).annotate(
        order_count=Count('id'),
        total_value=Sum('total_amount'),
        avg_delivery_rating=Avg('performance_reviews__delivery_rating'),
        avg_quality_rating=Avg('performance_reviews__quality_rating'),
        avg_service_rating=Avg('performance_reviews__service_rating')
    ).order_by('-order_count')[:5]
    
    radar_labels = ['Quality', 'Delivery', 'Service', 'Value', 'Reliability']
    radar_datasets = []
    
    for supplier in top_suppliers:
        # Normalize value score (0-5 scale)
        max_value = float(top_suppliers[0]['total_value'] or 1)
        value_score = (float(supplier['total_value'] or 0) / max_value) * 5
        
        # Calculate reliability score based on order completion
        reliability_score = min(5.0, (supplier['order_count'] / 10) * 5)
        
        dataset = {
            'label': supplier['supplier__name'] or 'Unknown',
            'data': [
                float(supplier['avg_quality_rating'] or 3.0),
                float(supplier['avg_delivery_rating'] or 3.0),
                float(supplier['avg_service_rating'] or 3.0),
                value_score,
                reliability_score
            ],
            'backgroundColor': f'rgba({", ".join(map(str, [
                hash(supplier["supplier__name"]) % 255,
                (hash(supplier["supplier__name"]) * 2) % 255,
                (hash(supplier["supplier__name"]) * 3) % 255
            ]))}, 0.2)',
            'borderColor': f'rgb({", ".join(map(str, [
                hash(supplier["supplier__name"]) % 255,
                (hash(supplier["supplier__name"]) * 2) % 255,
                (hash(supplier["supplier__name"]) * 3) % 255
            ]))})',
            'borderWidth': 2
        }
        radar_datasets.append(dataset)
    
    # ==========================================
    # 6. 3D DONUT - Budget Utilization by Category
    # ==========================================
    budget_by_category = Budget.objects.filter(
        budget_year__is_active=True
    ).values(
        'category__name'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent'),
        committed=Sum('committed_amount')
    ).order_by('-allocated')[:8]  # Top 8 categories
    
    budget_3d_labels = []
    budget_3d_allocated = []
    budget_3d_spent = []
    budget_3d_available = []
    
    for item in budget_by_category:
        budget_3d_labels.append(item['category__name'] or 'Unknown')
        allocated = float(item['allocated'] or 0)
        spent = float(item['spent'] or 0)
        committed = float(item['committed'] or 0)
        
        budget_3d_allocated.append(allocated)
        budget_3d_spent.append(spent)
        budget_3d_available.append(max(0, allocated - spent - committed))
    
    # ==========================================
    # 7. PROCUREMENT METHOD BREAKDOWN
    # ==========================================
    procurement_methods = tenders_qs.values(
        'procurement_method'
    ).annotate(
        count=Count('id'),
        total_value=Sum('estimated_budget')
    ).order_by('-count')
    
    method_labels = []
    method_values = []
    method_colors_list = ['#2563EB', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6']
    
    for idx, item in enumerate(procurement_methods):
        method_display = dict(Tender.METHOD_CHOICES).get(
            item['procurement_method'], item['procurement_method']
        )
        method_labels.append(method_display)
        method_values.append(item['count'])
    
    # ==========================================
    # 8. QUARTERLY COMPARISON
    # ==========================================
    quarterly_data = pos_qs.filter(
        created_at__year=current_year
    ).annotate(
        quarter=TruncQuarter('created_at')
    ).values('quarter').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('quarter')
    
    quarterly_labels = []
    quarterly_amounts = []
    
    for item in quarterly_data:
        quarter_num = ((item['quarter'].month - 1) // 3) + 1
        quarterly_labels.append(f'Q{quarter_num} {item["quarter"].year}')
        quarterly_amounts.append(float(item['total'] or 0))
    
    # ==========================================
    # 9. PAYMENT STATUS OVERVIEW
    # ==========================================
    payment_status = invoices_qs.values('status').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('-count')
    
    payment_labels = []
    payment_counts = []
    payment_amounts = []
    
    for item in payment_status:
        status_display = dict(Invoice.STATUS_CHOICES).get(
            item['status'], item['status']
        )
        payment_labels.append(status_display)
        payment_counts.append(item['count'])
        payment_amounts.append(float(item['total'] or 0))
    
    # ==========================================
    # CONTEXT PREPARATION
    # ==========================================
    context = {
        # Summary stats
        'total_requisitions': total_requisitions,
        'total_tenders': total_tenders,
        'total_pos': total_pos,
        'total_suppliers': total_suppliers,
        'total_spend': total_spend,
        'total_budget': total_budget,
        'budget_utilization': round(budget_utilization, 2),
        'pending_requisitions': pending_requisitions,
        'pending_invoices': pending_invoices,
        
        # Chart data - JSON encoded for JavaScript
        'spend_trend_data': json.dumps({
            'labels': spend_trend_labels,
            'amounts': spend_trend_amounts,
            'count': spend_trend_count
        }),
        
        'department_data': json.dumps({
            'labels': dept_labels,
            'amounts': dept_amounts,
            'colors': dept_colors
        }),
        
        'requisition_status_data': json.dumps({
            'labels': req_status_labels,
            'values': req_status_values,
            'colors': donut_colors
        }),
        
        'radar_data': json.dumps({
            'labels': radar_labels,
            'datasets': radar_datasets
        }),
        
        'budget_3d_data': json.dumps({
            'labels': budget_3d_labels,
            'allocated': budget_3d_allocated,
            'spent': budget_3d_spent,
            'available': budget_3d_available
        }),
        
        'procurement_method_data': json.dumps({
            'labels': method_labels,
            'values': method_values,
            'colors': method_colors_list[:len(method_values)]
        }),
        
        'quarterly_data': json.dumps({
            'labels': quarterly_labels,
            'amounts': quarterly_amounts
        }),
        
        'payment_status_data': json.dumps({
            'labels': payment_labels,
            'counts': payment_counts,
            'amounts': payment_amounts
        }),
        
        # Filters
        'current_year': current_year,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'procurement/procurement_module/reports.html', context)


@login_required
def procurement_spend_analysis_view(request):
    """Comprehensive spend analysis report with advanced analytics"""
    if not check_procurement_permission(request.user):
        messages.error(request, 'You do not have procurement officer permissions.')
        return redirect('dashboard')
    
    from django.db.models import Sum, Count, Avg, Q, F, Case, When, DecimalField
    from django.db.models.functions import TruncMonth, TruncWeek, TruncQuarter
    from decimal import Decimal
    import json
    
    # Date filters with defaults
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=365)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    
    # Additional filters
    selected_department = request.GET.get('department', '')
    selected_category = request.GET.get('category', '')
    selected_supplier = request.GET.get('supplier', '')
    
    # Base queryset
    pos_qs = PurchaseOrder.objects.filter(
        po_date__gte=date_from,
        po_date__lte=date_to
    )
    
    # Apply additional filters
    if selected_department:
        pos_qs = pos_qs.filter(requisition__department_id=selected_department)
    if selected_category:
        pos_qs = pos_qs.filter(requisition__budget__category_id=selected_category)
    if selected_supplier:
        pos_qs = pos_qs.filter(supplier_id=selected_supplier)
    
    # ==========================================
    # 1. SUMMARY STATISTICS
    # ==========================================
    total_spend = pos_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    total_orders = pos_qs.count()
    avg_order_value = pos_qs.aggregate(Avg('total_amount'))['total_amount__avg'] or Decimal('0')
    
    # Compare with previous period
    period_days = (datetime.strptime(date_to, '%Y-%m-%d') - datetime.strptime(date_from, '%Y-%m-%d')).days
    prev_period_start = (datetime.strptime(date_from, '%Y-%m-%d') - timedelta(days=period_days)).strftime('%Y-%m-%d')
    prev_period_end = date_from
    
    prev_spend = PurchaseOrder.objects.filter(
        po_date__gte=prev_period_start,
        po_date__lt=prev_period_end
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    
    spend_change = 0
    if prev_spend > 0:
        spend_change = ((total_spend - prev_spend) / prev_spend * 100)
    
    # Unique suppliers and departments
    unique_suppliers = pos_qs.values('supplier').distinct().count()
    unique_departments = pos_qs.values('requisition__department').distinct().count()
    
    # Order status breakdown
    completed_orders = pos_qs.filter(status__in=['DELIVERED', 'CLOSED']).count()
    pending_orders = pos_qs.exclude(status__in=['DELIVERED', 'CLOSED', 'CANCELLED']).count()
    
    # ==========================================
    # 2. LINE CHART - Spend Trend Over Time
    # ==========================================
    # Monthly trend
    monthly_trend = pos_qs.annotate(
        month=TruncMonth('po_date')
    ).values('month').annotate(
        total=Sum('total_amount'),
        count=Count('id'),
        avg=Avg('total_amount')
    ).order_by('month')
    
    trend_labels = []
    trend_amounts = []
    trend_counts = []
    trend_averages = []
    
    for item in monthly_trend:
        trend_labels.append(item['month'].strftime('%b %Y'))
        trend_amounts.append(float(item['total'] or 0))
        trend_counts.append(item['count'])
        trend_averages.append(float(item['avg'] or 0))
    
    # ==========================================
    # 3. BAR CHART - Top Departments by Spend
    # ==========================================
    dept_spending = pos_qs.values(
        'requisition__department__name',
        'requisition__department__id'
    ).annotate(
        total=Sum('total_amount'),
        count=Count('id'),
        avg=Avg('total_amount')
    ).order_by('-total')[:15]
    
    dept_labels = []
    dept_amounts = []
    dept_counts = []
    dept_colors = []
    
    color_palette = [
        '#2563EB', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
        '#EC4899', '#14B8A6', '#F97316', '#06B6D4', '#84CC16',
        '#6366F1', '#F43F5E', '#22D3EE', '#A855F7', '#EAB308'
    ]
    
    for idx, item in enumerate(dept_spending):
        dept_labels.append(item['requisition__department__name'] or 'Unknown')
        dept_amounts.append(float(item['total'] or 0))
        dept_counts.append(item['count'])
        dept_colors.append(color_palette[idx % len(color_palette)])
    
    # ==========================================
    # 4. HORIZONTAL BAR - Top Categories
    # ==========================================
    category_spending = pos_qs.values(
        'requisition__budget__category__name'
    ).annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    category_labels = []
    category_amounts = []
    category_counts = []
    
    for item in category_spending:
        category_labels.append(item['requisition__budget__category__name'] or 'Unbudgeted')
        category_amounts.append(float(item['total'] or 0))
        category_counts.append(item['count'])
    
    # ==========================================
    # 5. DONUT CHART - Supplier Concentration
    # ==========================================
    supplier_spending = pos_qs.values(
        'supplier__name'
    ).annotate(
        total=Sum('total_amount')
    ).order_by('-total')[:10]
    
    supplier_labels = []
    supplier_amounts = []
    supplier_colors = [
        '#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6',
        '#EC4899', '#14B8A6', '#F97316', '#06B6D4', '#84CC16'
    ]
    
    # Calculate "Others" if there are more suppliers
    top_10_total = Decimal('0')
    for item in supplier_spending:
        supplier_labels.append(item['supplier__name'])
        amount = float(item['total'] or 0)
        supplier_amounts.append(amount)
        top_10_total += item['total'] or Decimal('0')
    
    # Add "Others" category if applicable
    if total_spend > top_10_total:
        others_amount = float(total_spend - top_10_total)
        if others_amount > 0:
            supplier_labels.append('Others')
            supplier_amounts.append(others_amount)
            supplier_colors.append('#94A3B8')
    
    # ==========================================
    # 6. RADAR CHART - Spending Dimensions
    # ==========================================
    # Analyze spending across different dimensions
    
    # Emergency vs Planned spending
    emergency_spend = pos_qs.filter(
        requisition__is_emergency=True
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    
    planned_spend = total_spend - emergency_spend
    
    # Local vs International suppliers (assuming this is tracked)
    local_spend = pos_qs.filter(
        supplier__categories__name__icontains='Local'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    
    # Goods vs Services vs Works
    goods_spend = pos_qs.filter(
        requisition__items__item__category__category_type='GOODS'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    
    services_spend = pos_qs.filter(
        requisition__items__item__category__category_type='SERVICES'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    
    works_spend = pos_qs.filter(
        requisition__items__item__category__category_type='WORKS'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0')
    
    # Normalize values for radar chart (0-100 scale)
    max_value = float(total_spend) if total_spend > 0 else 1
    
    radar_data = {
        'labels': ['Planned Procurement', 'Emergency Procurement', 'Goods', 'Services', 'Works', 'Compliance Score'],
        'datasets': [{
            'label': 'Spending Pattern',
            'data': [
                (float(planned_spend) / max_value) * 100,
                (float(emergency_spend) / max_value) * 100,
                (float(goods_spend) / max_value) * 100,
                (float(services_spend) / max_value) * 100,
                (float(works_spend) / max_value) * 100,
                85  # Placeholder compliance score
            ],
            'backgroundColor': 'rgba(37, 99, 235, 0.2)',
            'borderColor': 'rgb(37, 99, 235)',
            'borderWidth': 2
        }]
    }
    
    # ==========================================
    # 7. STACKED BAR - Monthly Spend by Category Type
    # ==========================================
    monthly_by_type = pos_qs.annotate(
        month=TruncMonth('po_date')
    ).values('month').annotate(
        goods=Sum(Case(
            When(requisition__items__item__category__category_type='GOODS', 
                 then='total_amount'),
            default=0,
            output_field=DecimalField()
        )),
        services=Sum(Case(
            When(requisition__items__item__category__category_type='SERVICES', 
                 then='total_amount'),
            default=0,
            output_field=DecimalField()
        )),
        works=Sum(Case(
            When(requisition__items__item__category__category_type='WORKS', 
                 then='total_amount'),
            default=0,
            output_field=DecimalField()
        ))
    ).order_by('month')
    
    stacked_labels = []
    stacked_goods = []
    stacked_services = []
    stacked_works = []
    
    for item in monthly_by_type:
        stacked_labels.append(item['month'].strftime('%b %Y'))
        stacked_goods.append(float(item['goods'] or 0))
        stacked_services.append(float(item['services'] or 0))
        stacked_works.append(float(item['works'] or 0))
    
    # ==========================================
    # 8. BUBBLE CHART - Supplier Performance vs Spend
    # ==========================================
    supplier_performance = pos_qs.values(
        'supplier__name',
        'supplier__id'
    ).annotate(
        total_spend=Sum('total_amount'),
        order_count=Count('id'),
        avg_rating=Avg('performance_reviews__overall_rating')
    ).order_by('-total_spend')[:20]
    
    bubble_data = []
    for supplier in supplier_performance:
        bubble_data.append({
            'x': float(supplier['total_spend'] or 0),
            'y': float(supplier['avg_rating'] or 3.0),
            'r': min(supplier['order_count'] * 2, 30),  # Bubble size
            'label': supplier['supplier__name']
        })
    
    # ==========================================
    # 9. POLAR AREA - Procurement Methods Distribution
    # ==========================================
    method_distribution = Tender.objects.filter(
        requisition__purchase_orders__po_date__gte=date_from,
        requisition__purchase_orders__po_date__lte=date_to
    ).values('procurement_method').annotate(
        total=Sum('requisition__purchase_orders__total_amount'),
        count=Count('id')
    ).order_by('-total')
    
    polar_labels = []
    polar_values = []
    polar_colors = ['#2563EB', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6']
    
    for idx, item in enumerate(method_distribution):
        method_display = dict(Tender.METHOD_CHOICES).get(
            item['procurement_method'], item['procurement_method']
        )
        polar_labels.append(method_display)
        polar_values.append(float(item['total'] or 0))
    
    # ==========================================
    # 10. PIE CHART - Order Status Distribution
    # ==========================================
    status_distribution = pos_qs.values('status').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('-count')
    
    status_labels = []
    status_counts = []
    status_amounts = []
    status_colors = {
        'DRAFT': '#94A3B8',
        'PENDING_APPROVAL': '#3B82F6',
        'APPROVED': '#10B981',
        'SENT': '#8B5CF6',
        'ACKNOWLEDGED': '#EC4899',
        'PARTIAL_DELIVERY': '#F59E0B',
        'DELIVERED': '#22C55E',
        'CLOSED': '#6B7280',
        'CANCELLED': '#EF4444',
    }
    
    pie_colors = []
    for item in status_distribution:
        status_display = dict(PurchaseOrder.STATUS_CHOICES).get(
            item['status'], item['status']
        )
        status_labels.append(status_display)
        status_counts.append(item['count'])
        status_amounts.append(float(item['total'] or 0))
        pie_colors.append(status_colors.get(item['status'], '#64748B'))
    
    # ==========================================
    # 11. TOP SPENDING ANALYSIS
    # ==========================================
    # Top 10 individual purchase orders
    top_pos = pos_qs.order_by('-total_amount')[:10].values(
        'po_number',
        'supplier__name',
        'total_amount',
        'po_date',
        'status'
    )
    
    # Average order value by department
    dept_avg_orders = pos_qs.values(
        'requisition__department__name'
    ).annotate(
        avg_order=Avg('total_amount'),
        max_order=Max('total_amount'),
        min_order=Min('total_amount'),
        count=Count('id')
    ).order_by('-avg_order')[:10]
    
    # ==========================================
    # 12. SAVINGS ANALYSIS
    # ==========================================
    # Compare requisition estimates with actual PO amounts
    savings_data = pos_qs.annotate(
        estimated=F('requisition__estimated_amount'),
        actual=F('total_amount')
    ).aggregate(
        total_estimated=Sum('requisition__estimated_amount'),
        total_actual=Sum('total_amount')
    )
    
    total_estimated = savings_data['total_estimated'] or Decimal('0')
    total_actual = savings_data['total_actual'] or Decimal('0')
    total_savings = total_estimated - total_actual
    savings_percentage = 0
    if total_estimated > 0:
        savings_percentage = (total_savings / total_estimated * 100)
    
    # ==========================================
    # FILTER OPTIONS
    # ==========================================
    departments = Department.objects.filter(is_active=True).order_by('name')
    categories = BudgetCategory.objects.filter(is_active=True).order_by('name')
    suppliers = Supplier.objects.filter(status='APPROVED').order_by('name')
    
    # ==========================================
    # CONTEXT PREPARATION
    # ==========================================
    context = {
        # Filters
        'date_from': date_from,
        'date_to': date_to,
        'selected_department': selected_department,
        'selected_category': selected_category,
        'selected_supplier': selected_supplier,
        'departments': departments,
        'categories': categories,
        'suppliers': suppliers,
        
        # Summary Statistics
        'total_spend': total_spend,
        'total_orders': total_orders,
        'avg_order_value': avg_order_value,
        'spend_change': round(spend_change, 2),
        'unique_suppliers': unique_suppliers,
        'unique_departments': unique_departments,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'total_savings': total_savings,
        'savings_percentage': round(savings_percentage, 2),
        
        # Chart Data - JSON encoded
        'trend_data': json.dumps({
            'labels': trend_labels,
            'amounts': trend_amounts,
            'counts': trend_counts,
            'averages': trend_averages
        }),
        
        'department_data': json.dumps({
            'labels': dept_labels,
            'amounts': dept_amounts,
            'counts': dept_counts,
            'colors': dept_colors
        }),
        
        'category_data': json.dumps({
            'labels': category_labels,
            'amounts': category_amounts,
            'counts': category_counts
        }),
        
        'supplier_data': json.dumps({
            'labels': supplier_labels,
            'amounts': supplier_amounts,
            'colors': supplier_colors
        }),
        
        'radar_data': json.dumps(radar_data),
        
        'stacked_data': json.dumps({
            'labels': stacked_labels,
            'goods': stacked_goods,
            'services': stacked_services,
            'works': stacked_works
        }),
        
        'bubble_data': json.dumps(bubble_data),
        
        'polar_data': json.dumps({
            'labels': polar_labels,
            'values': polar_values,
            'colors': polar_colors
        }),
        
        'status_data': json.dumps({
            'labels': status_labels,
            'counts': status_counts,
            'amounts': status_amounts,
            'colors': pie_colors
        }),
        
        # Table Data
        'dept_spending': dept_spending,
        'category_spending': category_spending,
        'top_pos': top_pos,
        'dept_avg_orders': dept_avg_orders,
    }
    
    return render(request, 'procurement/procurement_module/spend_analysis.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from datetime import datetime
import json

from .models import (
    ProcurementPlan, ProcurementPlanItem, ProcurementPlanAmendment,
    BudgetYear, Department, Budget, Item, ItemCategory, User
)
from .forms import (
    ProcurementPlanForm, ProcurementPlanItemForm, 
    ProcurementPlanAmendmentForm
)


# ============================================================================
# PERMISSION DECORATORS
# ============================================================================

def procurement_or_admin_required(user):
    """Check if user is procurement officer or admin"""
    return user.role in ['PROCUREMENT', 'ADMIN']


def admin_required(user):
    """Check if user is admin"""
    return user.role == 'ADMIN'


# ============================================================================
# PROCUREMENT PLAN LIST & DASHBOARD
# ============================================================================

@login_required
@user_passes_test(procurement_or_admin_required)
def procurement_plan_list(request):
    """List all procurement plans with filters"""
    
    # Get filter parameters
    budget_year_id = request.GET.get('budget_year')
    department_id = request.GET.get('department')
    status = request.GET.get('status')
    search = request.GET.get('search', '').strip()
    
    # Base queryset
    plans = ProcurementPlan.objects.select_related(
        'budget_year', 'department', 'submitted_by', 'approved_by'
    ).annotate(
        total_items=Count('items'),
        total_estimated_cost=Sum('items__estimated_cost'),
        total_committed=Sum('items__amount_committed')
    ).order_by('-created_at')
    
    # Apply filters
    if budget_year_id:
        plans = plans.filter(budget_year_id=budget_year_id)
    
    if department_id:
        plans = plans.filter(department_id=department_id)
    
    if status:
        plans = plans.filter(status=status)
    
    if search:
        plans = plans.filter(
            Q(plan_number__icontains=search) |
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(department__name__icontains=search)
        )
    
    # Get filter options
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    # Statistics
    total_plans = plans.count()
    approved_plans = plans.filter(status='APPROVED').count()
    active_plans = plans.filter(status='ACTIVE').count()
    draft_plans = plans.filter(status='DRAFT').count()
    
    context = {
        'plans': plans,
        'budget_years': budget_years,
        'departments': departments,
        'total_plans': total_plans,
        'approved_plans': approved_plans,
        'active_plans': active_plans,
        'draft_plans': draft_plans,
        'page_title': 'Procurement Plans Management',
        'selected_budget_year': budget_year_id,
        'selected_department': department_id,
        'selected_status': status,
        'search_query': search,
    }
    
    return render(request, 'procurement/plan_list.html', context)


@login_required
@user_passes_test(procurement_or_admin_required)
def procurement_plan_detail(request, pk):
    """View detailed procurement plan with items"""
    
    plan = get_object_or_404(
        ProcurementPlan.objects.select_related(
            'budget_year', 'department', 'submitted_by', 'approved_by'
        ).prefetch_related('items__budget', 'items__item', 'amendments'),
        pk=pk
    )
    
    # Get plan items grouped by quarter
    items_q1 = plan.items.filter(planned_quarter='Q1').select_related('budget', 'item')
    items_q2 = plan.items.filter(planned_quarter='Q2').select_related('budget', 'item')
    items_q3 = plan.items.filter(planned_quarter='Q3').select_related('budget', 'item')
    items_q4 = plan.items.filter(planned_quarter='Q4').select_related('budget', 'item')
    
    # Calculate statistics
    total_estimated = plan.items.aggregate(
        total=Sum('estimated_cost')
    )['total'] or Decimal('0')
    
    total_committed = plan.items.aggregate(
        total=Sum('amount_committed')
    )['total'] or Decimal('0')
    
    total_items = plan.items.count()
    completed_items = plan.items.filter(status='COMPLETED').count()
    in_progress_items = plan.items.filter(status='IN_PROGRESS').count()
    
    # Get amendments
    amendments = plan.amendments.select_related(
        'requested_by', 'approved_by', 'plan_item'
    ).order_by('-requested_at')
    
    # Permissions
    can_edit = plan.status in ['DRAFT', 'SUBMITTED'] and request.user.role in ['PROCUREMENT', 'ADMIN']
    can_approve = plan.status == 'SUBMITTED' and request.user.role == 'ADMIN'
    can_activate = plan.status == 'APPROVED' and request.user.role == 'ADMIN'
    
    context = {
        'plan': plan,
        'items_q1': items_q1,
        'items_q2': items_q2,
        'items_q3': items_q3,
        'items_q4': items_q4,
        'total_estimated': total_estimated,
        'total_committed': total_committed,
        'total_items': total_items,
        'completed_items': completed_items,
        'in_progress_items': in_progress_items,
        'amendments': amendments,
        'can_edit': can_edit,
        'can_approve': can_approve,
        'can_activate': can_activate,
        'page_title': f'Procurement Plan: {plan.plan_number}',
    }
    
    return render(request, 'procurement/plan_detail.html', context)


# ============================================================================
# CREATE & EDIT PROCUREMENT PLAN
# ============================================================================

@login_required
@user_passes_test(procurement_or_admin_required)
def procurement_plan_create(request):
    """Create new procurement plan"""
    
    if request.method == 'POST':
        plan_title = request.POST.get('title')
        plan_description = request.POST.get('description', '')
        budget_year_id = request.POST.get('budget_year')
        department_id = request.POST.get('department')
        
        # Validation
        if not all([plan_title, budget_year_id, department_id]):
            messages.error(request, 'Please fill in all required fields.')
            return redirect('procurement_plan_create')
        
        # Check if plan already exists for this department/year
        existing_plan = ProcurementPlan.objects.filter(
            budget_year_id=budget_year_id,
            department_id=department_id
        ).exists()
        
        if existing_plan:
            messages.error(
                request,
                'A procurement plan already exists for this department and budget year. '
                'Please edit the existing plan instead.'
            )
            return redirect('procurement_plan_list')
        
        # Create plan
        plan = ProcurementPlan.objects.create(
            budget_year_id=budget_year_id,
            department_id=department_id,
            title=plan_title,
            description=plan_description,
            status='DRAFT',
            submitted_by=request.user
        )
        
        messages.success(
            request,
            f'Procurement plan {plan.plan_number} created successfully! '
            f'Now add items to the plan.'
        )
        return redirect('procurement_plan_edit', pk=plan.pk)
    
    # GET request
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'budget_years': budget_years,
        'departments': departments,
        'page_title': 'Create Procurement Plan',
    }
    
    return render(request, 'procurement/plan_create.html', context)


@login_required
@user_passes_test(procurement_or_admin_required)
def procurement_plan_edit(request, pk):
    """Edit procurement plan and manage items"""
    
    plan = get_object_or_404(
        ProcurementPlan.objects.select_related('budget_year', 'department'),
        pk=pk
    )
    
    # Check if plan can be edited
    if plan.status not in ['DRAFT', 'SUBMITTED']:
        messages.error(request, 'Cannot edit an approved or active plan. Create an amendment instead.')
        return redirect('procurement_plan_detail', pk=pk)
    
    # Get plan items
    plan_items = plan.items.select_related('budget', 'item').order_by('sequence')
    
    # Get available budgets for this department
    available_budgets = Budget.objects.filter(
        department=plan.department,
        budget_year=plan.budget_year,
        is_active=True
    ).select_related('category')
    
    # Get item catalog
    item_categories = ItemCategory.objects.filter(is_active=True).prefetch_related('items')
    
    context = {
        'plan': plan,
        'plan_items': plan_items,
        'available_budgets': available_budgets,
        'item_categories': item_categories,
        'page_title': f'Edit Plan: {plan.plan_number}',
    }
    
    return render(request, 'procurement/plan_edit.html', context)


# ============================================================================
# PLAN ITEM MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(procurement_or_admin_required)
def plan_item_add(request, plan_pk):
    """Add item to procurement plan (AJAX)"""
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    plan = get_object_or_404(ProcurementPlan, pk=plan_pk)
    
    # Check if plan can be edited
    if plan.status not in ['DRAFT', 'SUBMITTED']:
        return JsonResponse({
            'success': False,
            'error': 'Cannot add items to approved/active plan'
        }, status=400)
    
    try:
        # Get form data
        item_type = request.POST.get('item_type')
        description = request.POST.get('description')
        specifications = request.POST.get('specifications', '')
        quantity = Decimal(request.POST.get('quantity'))
        unit_of_measure = request.POST.get('unit_of_measure')
        estimated_cost = Decimal(request.POST.get('estimated_cost'))
        budget_id = request.POST.get('budget')
        procurement_method = request.POST.get('procurement_method')
        planned_quarter = request.POST.get('planned_quarter')
        source_of_funds = request.POST.get('source_of_funds')
        item_id = request.POST.get('item_id') or None  # Optional catalog item
        
        # Validation
        if not all([item_type, description, quantity, unit_of_measure, 
                   estimated_cost, procurement_method, planned_quarter, source_of_funds]):
            return JsonResponse({
                'success': False,
                'error': 'Please fill in all required fields'
            }, status=400)
        
        # Get next sequence number
        last_sequence = plan.items.aggregate(
            max_seq=Count('sequence')
        )['max_seq'] or 0
        
        # Create plan item
        plan_item = ProcurementPlanItem.objects.create(
            procurement_plan=plan,
            item_id=item_id,
            item_type=item_type,
            description=description,
            specifications=specifications,
            quantity=quantity,
            unit_of_measure=unit_of_measure,
            estimated_cost=estimated_cost,
            budget_id=budget_id if budget_id else None,
            procurement_method=procurement_method,
            planned_quarter=planned_quarter,
            source_of_funds=source_of_funds,
            status='PLANNED',
            sequence=last_sequence + 1
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Item added successfully',
            'item': {
                'id': str(plan_item.id),
                'description': plan_item.description,
                'estimated_cost': float(plan_item.estimated_cost),
                'quarter': plan_item.get_planned_quarter_display()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@user_passes_test(procurement_or_admin_required)
def plan_item_edit(request, item_pk):
    """Edit plan item (AJAX)"""
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    plan_item = get_object_or_404(
        ProcurementPlanItem.objects.select_related('procurement_plan'),
        pk=item_pk
    )
    
    # Check if plan can be edited
    if plan_item.procurement_plan.status not in ['DRAFT', 'SUBMITTED']:
        return JsonResponse({
            'success': False,
            'error': 'Cannot edit items in approved/active plan'
        }, status=400)
    
    try:
        # Update fields
        plan_item.item_type = request.POST.get('item_type', plan_item.item_type)
        plan_item.description = request.POST.get('description', plan_item.description)
        plan_item.specifications = request.POST.get('specifications', plan_item.specifications)
        plan_item.quantity = Decimal(request.POST.get('quantity', plan_item.quantity))
        plan_item.unit_of_measure = request.POST.get('unit_of_measure', plan_item.unit_of_measure)
        plan_item.estimated_cost = Decimal(request.POST.get('estimated_cost', plan_item.estimated_cost))
        plan_item.procurement_method = request.POST.get('procurement_method', plan_item.procurement_method)
        plan_item.planned_quarter = request.POST.get('planned_quarter', plan_item.planned_quarter)
        plan_item.source_of_funds = request.POST.get('source_of_funds', plan_item.source_of_funds)
        
        budget_id = request.POST.get('budget')
        if budget_id:
            plan_item.budget_id = budget_id
        
        plan_item.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Item updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@user_passes_test(procurement_or_admin_required)
def plan_item_delete(request, item_pk):
    """Delete plan item"""
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    plan_item = get_object_or_404(
        ProcurementPlanItem.objects.select_related('procurement_plan'),
        pk=item_pk
    )
    
    plan_id = plan_item.procurement_plan.id
    
    # Check if plan can be edited
    if plan_item.procurement_plan.status not in ['DRAFT', 'SUBMITTED']:
        return JsonResponse({
            'success': False,
            'error': 'Cannot delete items from approved/active plan'
        }, status=400)
    
    # Check if item has been used in requisitions
    if plan_item.requisitions.exists():
        return JsonResponse({
            'success': False,
            'error': 'Cannot delete item that has been used in requisitions'
        }, status=400)
    
    plan_item.delete()
    
    return JsonResponse({
        'success': True,
        'message': 'Item deleted successfully'
    })


# ============================================================================
# PLAN SUBMISSION & APPROVAL WORKFLOW
# ============================================================================

@login_required
@user_passes_test(procurement_or_admin_required)
def procurement_plan_submit(request, pk):
    """Submit plan for approval"""
    
    plan = get_object_or_404(ProcurementPlan, pk=pk)
    
    # Check status
    if plan.status != 'DRAFT':
        messages.error(request, 'This plan has already been submitted.')
        return redirect('procurement_plan_detail', pk=pk)
    
    # Validate plan has items
    if not plan.items.exists():
        messages.error(request, 'Cannot submit a plan without any items.')
        return redirect('procurement_plan_edit', pk=pk)
    
    # Check budget allocation
    items_without_budget = plan.items.filter(budget__isnull=True).count()
    if items_without_budget > 0:
        messages.warning(
            request,
            f'{items_without_budget} items do not have budget allocation. '
            f'Please assign budgets before submission.'
        )
        return redirect('procurement_plan_edit', pk=pk)
    
    # Submit plan
    plan.status = 'SUBMITTED'
    plan.submitted_at = timezone.now()
    plan.submitted_by = request.user
    plan.save()
    
    messages.success(
        request,
        f'Procurement plan {plan.plan_number} submitted for approval!'
    )
    
    # TODO: Send notification to admin for approval
    
    return redirect('procurement_plan_detail', pk=pk)


@login_required
@user_passes_test(admin_required)
def procurement_plan_approve(request, pk):
    """Approve procurement plan (Admin only)"""
    
    plan = get_object_or_404(ProcurementPlan, pk=pk)
    
    # Check status
    if plan.status != 'SUBMITTED':
        messages.error(request, 'Only submitted plans can be approved.')
        return redirect('procurement_plan_detail', pk=pk)
    
    if request.method == 'POST':
        comments = request.POST.get('comments', '')
        
        plan.status = 'APPROVED'
        plan.approved_by = request.user
        plan.approved_at = timezone.now()
        plan.save()
        
        messages.success(
            request,
            f'Procurement plan {plan.plan_number} approved successfully!'
        )
        
        # TODO: Send notification to department
        
        return redirect('procurement_plan_detail', pk=pk)
    
    context = {
        'plan': plan,
        'page_title': f'Approve Plan: {plan.plan_number}',
    }
    
    return render(request, 'procurement/plan_approve.html', context)


@login_required
@user_passes_test(admin_required)
def procurement_plan_reject(request, pk):
    """Reject procurement plan (Admin only)"""
    
    plan = get_object_or_404(ProcurementPlan, pk=pk)
    
    # Check status
    if plan.status != 'SUBMITTED':
        messages.error(request, 'Only submitted plans can be rejected.')
        return redirect('procurement_plan_detail', pk=pk)
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        if not rejection_reason:
            messages.error(request, 'Please provide a rejection reason.')
            return redirect('procurement_plan_reject', pk=pk)
        
        plan.status = 'DRAFT'
        plan.notes = f"REJECTED: {rejection_reason}\n\n{plan.notes}"
        plan.save()
        
        messages.success(
            request,
            f'Procurement plan {plan.plan_number} rejected. '
            f'Department can revise and resubmit.'
        )
        
        # TODO: Send notification to department
        
        return redirect('procurement_plan_detail', pk=pk)
    
    context = {
        'plan': plan,
        'page_title': f'Reject Plan: {plan.plan_number}',
    }
    
    return render(request, 'procurement/plan_reject.html', context)


@login_required
@user_passes_test(admin_required)
def procurement_plan_activate(request, pk):
    """Activate approved procurement plan"""
    
    plan = get_object_or_404(ProcurementPlan, pk=pk)
    
    # Check status
    if plan.status != 'APPROVED':
        messages.error(request, 'Only approved plans can be activated.')
        return redirect('procurement_plan_detail', pk=pk)
    
    plan.status = 'ACTIVE'
    plan.save()
    
    messages.success(
        request,
        f'Procurement plan {plan.plan_number} is now active! '
        f'Departments can create requisitions from this plan.'
    )
    
    return redirect('procurement_plan_detail', pk=pk)


# ============================================================================
# PLAN AMENDMENTS
# ============================================================================

@login_required
@user_passes_test(procurement_or_admin_required)
def plan_amendment_list(request):
    """List all plan amendments"""
    
    amendments = ProcurementPlanAmendment.objects.select_related(
        'procurement_plan__department',
        'procurement_plan__budget_year',
        'plan_item',
        'requested_by',
        'approved_by'
    ).order_by('-requested_at')
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        amendments = amendments.filter(status=status)
    
    # Statistics
    total_amendments = amendments.count()
    pending_amendments = amendments.filter(status='PENDING').count()
    approved_amendments = amendments.filter(status='APPROVED').count()
    rejected_amendments = amendments.filter(status='REJECTED').count()
    
    context = {
        'amendments': amendments,
        'total_amendments': total_amendments,
        'pending_amendments': pending_amendments,
        'approved_amendments': approved_amendments,
        'rejected_amendments': rejected_amendments,
        'page_title': 'Plan Amendments',
        'selected_status': status,
    }
    
    return render(request, 'procurement/amendment_list.html', context)


@login_required
@user_passes_test(procurement_or_admin_required)
def plan_amendment_create(request, plan_pk):
    """Create plan amendment"""
    
    plan = get_object_or_404(ProcurementPlan, pk=plan_pk)
    
    # Only allow amendments for approved/active plans
    if plan.status not in ['APPROVED', 'ACTIVE']:
        messages.error(request, 'Can only amend approved or active plans.')
        return redirect('procurement_plan_detail', pk=plan_pk)
    
    if request.method == 'POST':
        amendment_type = request.POST.get('amendment_type')
        plan_item_id = request.POST.get('plan_item')
        justification = request.POST.get('justification')
        old_values = request.POST.get('old_values', '{}')
        new_values = request.POST.get('new_values', '{}')
        
        # Validation
        if not all([amendment_type, justification]):
            messages.error(request, 'Please fill in all required fields.')
            return redirect('plan_amendment_create', plan_pk=plan_pk)
        
        if len(justification) < 100:
            messages.error(request, 'Justification must be at least 100 characters.')
            return redirect('plan_amendment_create', plan_pk=plan_pk)
        
        # Parse JSON fields
        try:
            old_values_json = json.loads(old_values) if old_values != '{}' else None
            new_values_json = json.loads(new_values) if new_values != '{}' else None
        except json.JSONDecodeError:
            messages.error(request, 'Invalid JSON format in old/new values.')
            return redirect('plan_amendment_create', plan_pk=plan_pk)
        
        # Create amendment
        amendment = ProcurementPlanAmendment.objects.create(
            procurement_plan=plan,
            amendment_type=amendment_type,
            plan_item_id=plan_item_id if plan_item_id else None,
            justification=justification,
            old_values=old_values_json,
            new_values=new_values_json,
            status='PENDING',
            requested_by=request.user
        )
        
        # Update plan amendment count
        plan.amendment_count += 1
        plan.is_amended = True
        plan.save()
        
        messages.success(
            request,
            f'Amendment {amendment.amendment_number} created successfully and pending approval.'
        )
        
        return redirect('procurement_plan_detail', pk=plan_pk)
    
    # GET request
    plan_items = plan.items.all()
    
    context = {
        'plan': plan,
        'plan_items': plan_items,
        'page_title': f'Create Amendment for {plan.plan_number}',
    }
    
    return render(request, 'procurement/amendment_create.html', context)


@login_required
@user_passes_test(admin_required)
def plan_amendment_approve(request, pk):
    """Approve plan amendment (Admin only)"""
    
    amendment = get_object_or_404(
        ProcurementPlanAmendment.objects.select_related('procurement_plan'),
        pk=pk
    )
    
    if amendment.status != 'PENDING':
        messages.error(request, 'Only pending amendments can be approved.')
        return redirect('plan_amendment_list')
    
    if request.method == 'POST':
        amendment.status = 'APPROVED'
        amendment.approved_by = request.user
        amendment.approved_at = timezone.now()
        amendment.save()
        
        # Apply the amendment based on type
        if amendment.amendment_type == 'ADD_ITEM' and amendment.new_values:
            # Create new plan item
            # Implementation depends on new_values structure
            pass
        elif amendment.amendment_type == 'MODIFY_ITEM' and amendment.plan_item:
            # Update plan item with new values
            # Implementation depends on new_values structure
            pass
        
        messages.success(
            request,
            f'Amendment {amendment.amendment_number} approved successfully!'
        )
        
        return redirect('plan_amendment_list')
    
    context = {
        'amendment': amendment,
        'page_title': f'Approve Amendment: {amendment.amendment_number}',
    }
    
    return render(request, 'procurement/amendment_approve.html', context)


@login_required
@user_passes_test(admin_required)
def plan_amendment_reject(request, pk):
    """Reject plan amendment (Admin only)"""
    
    amendment = get_object_or_404(ProcurementPlanAmendment, pk=pk)
    
    if amendment.status != 'PENDING':
        messages.error(request, 'Only pending amendments can be rejected.')
        return redirect('plan_amendment_list')
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        if not rejection_reason:
            messages.error(request, 'Please provide a rejection reason.')
            return redirect('plan_amendment_reject', pk=pk)
        
        amendment.status = 'REJECTED'
        amendment.rejection_reason = rejection_reason
        amendment.save()
        
        messages.success(
            request,
            f'Amendment {amendment.amendment_number} rejected.'
        )
        
        return redirect('plan_amendment_list')
    
    context = {
        'amendment': amendment,
        'page_title': f'Reject Amendment: {amendment.amendment_number}',
    }
    
    return render(request, 'procurement/amendment_reject.html', context)


# ============================================================================
# REPORTS & ANALYTICS
# ============================================================================

@login_required
@user_passes_test(procurement_or_admin_required)
def procurement_plan_reports(request):
    """Procurement plan reports and analytics"""
    
    budget_year_id = request.GET.get('budget_year')
    department_id = request.GET.get('department')
    
    # Get active budget year if not specified
    if not budget_year_id:
        active_year = BudgetYear.objects.filter(is_active=True).first()
        budget_year_id = active_year.id if active_year else None
    
    # Base query
    plans = ProcurementPlan.objects.all()
    if budget_year_id:
        plans = plans.filter(budget_year_id=budget_year_id)
    if department_id:
        plans = plans.filter(department_id=department_id)
    
    # Overall statistics
    total_plans = plans.count()
    approved_plans = plans.filter(status='APPROVED').count()
    total_estimated = ProcurementPlanItem.objects.filter(
        procurement_plan__in=plans
    ).aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    
    total_committed = ProcurementPlanItem.objects.filter(
        procurement_plan__in=plans
    ).aggregate(total=Sum('amount_committed'))['total'] or Decimal('0')
    
    # By department
    by_department = plans.values(
        'department__name'
    ).annotate(
        total_items=Count('items'),
        total_cost=Sum('items__estimated_cost'),
        total_committed=Sum('items__amount_committed')
    ).order_by('-total_cost')
    
    # By quarter
    by_quarter = ProcurementPlanItem.objects.filter(
        procurement_plan__in=plans
    ).values('planned_quarter').annotate(
        total_items=Count('id'),
        total_cost=Sum('estimated_cost')
    ).order_by('planned_quarter')
    
    # By procurement method
    by_method = ProcurementPlanItem.objects.filter(
        procurement_plan__in=plans
    ).values('procurement_method').annotate(
        total_items=Count('id'),
        total_cost=Sum('estimated_cost')
    ).order_by('-total_cost')
    
    # Get filter options
    budget_years = BudgetYear.objects.all().order_by('-start_date')
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'total_plans': total_plans,
        'approved_plans': approved_plans,
        'total_estimated': total_estimated,
        'total_committed': total_committed,
        'utilization_rate': (total_committed / total_estimated * 100) if total_estimated > 0 else 0,
        'by_department': by_department,
        'by_quarter': by_quarter,
        'by_method': by_method,
        'budget_years': budget_years,
        'departments': departments,
        'selected_budget_year': budget_year_id,
        'selected_department': department_id,
        'page_title': 'Procurement Plan Reports',
    }
    return render(request, 'procurement/plan_reports.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.http import JsonResponse
from decimal import Decimal
from datetime import datetime, timedelta

from .models import (
    Store, GoodsReceivedNote, GRNItem, PurchaseOrder, PurchaseOrderItem,
    StockItem, StockMovement, StockIssue, StockIssueItem, Item,
    Department, User, Asset
)


# ============================================================================
# GOODS RECEIPT VIEWS
# ============================================================================

@login_required
def store_pending_deliveries_view(request):
    """View pending deliveries awaiting receipt"""
    
    # Get all POs that are approved/sent but not fully delivered
    pending_pos = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY']
    ).select_related(
        'supplier', 'requisition', 'created_by'
    ).prefetch_related('items').order_by('-delivery_date')
    
    # Calculate delivery statistics
    stats = {
        'total_pending': pending_pos.count(),
        'overdue': pending_pos.filter(delivery_date__lt=timezone.now().date()).count(),
        'due_today': pending_pos.filter(delivery_date=timezone.now().date()).count(),
        'due_this_week': pending_pos.filter(
            delivery_date__range=[
                timezone.now().date(),
                timezone.now().date() + timedelta(days=7)
            ]
        ).count()
    }
    
    context = {
        'pending_pos': pending_pos,
        'stats': stats,
    }
    
    return render(request, 'stores/pending_deliveries.html', context)


@login_required
def store_receive_goods_view(request, po_id=None):
    """Receive goods and create GRN"""
    
    stores = Store.objects.filter(is_active=True)
    
    if po_id:
        po = get_object_or_404(PurchaseOrder, id=po_id)
        po_items = po.items.all()
    else:
        po = None
        po_items = []
        
    if request.method == 'POST':
        try:
            po = get_object_or_404(PurchaseOrder, id=request.POST.get('purchase_order'))
            store = get_object_or_404(Store, id=request.POST.get('store'))
            
            # Create GRN
            grn = GoodsReceivedNote.objects.create(
                purchase_order=po,
                store=store,
                delivery_note_number=request.POST.get('delivery_note_number'),
                delivery_date=request.POST.get('delivery_date'),
                received_by=request.user,
                general_condition=request.POST.get('general_condition', ''),
                notes=request.POST.get('notes', ''),
                status='DRAFT'
            )
            
            # Create GRN Items
            for po_item in po.items.all():
                qty_delivered = Decimal(request.POST.get(f'qty_delivered_{po_item.id}', '0'))
                qty_accepted = Decimal(request.POST.get(f'qty_accepted_{po_item.id}', '0'))
                qty_rejected = Decimal(request.POST.get(f'qty_rejected_{po_item.id}', '0'))
                item_status = request.POST.get(f'item_status_{po_item.id}', 'ACCEPTED')
                
                if qty_delivered > 0:
                    GRNItem.objects.create(
                        grn=grn,
                        po_item=po_item,
                        quantity_ordered=po_item.quantity,
                        quantity_delivered=qty_delivered,
                        quantity_accepted=qty_accepted,
                        quantity_rejected=qty_rejected,
                        item_status=item_status,
                        remarks=request.POST.get(f'remarks_{po_item.id}', '')
                    )
                    
                    # Update PO item delivered quantity
                    po_item.quantity_delivered += qty_accepted
                    po_item.save()
            
            messages.success(request, f'GRN {grn.grn_number} created successfully!')
            return redirect('store_receipt_history')
            
        except Exception as e:
            messages.error(request, f'Error creating GRN: {str(e)}')
    
    # Get pending POs for selection
    pending_pos = PurchaseOrder.objects.filter(
        status__in=['APPROVED', 'SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY']
    ).select_related('supplier')
    
    context = {
        'stores': stores,
        'po': po,
        'po_items': po_items,
        'pending_pos': pending_pos,
    }
    
    return render(request, 'stores/receive_goods.html', context)


@login_required
def store_receipt_history_view(request):
    """View GRN history"""
    
    grns = GoodsReceivedNote.objects.all().select_related(
        'purchase_order', 'purchase_order__supplier', 'store', 'received_by'
    ).prefetch_related('items').order_by('-created_at')
    
    # Filters
    status_filter = request.GET.get('status')
    store_filter = request.GET.get('store')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if status_filter:
        grns = grns.filter(status=status_filter)
    if store_filter:
        grns = grns.filter(store_id=store_filter)
    if date_from:
        grns = grns.filter(received_date__gte=date_from)
    if date_to:
        grns = grns.filter(received_date__lte=date_to)
    
    stores = Store.objects.filter(is_active=True)
    
    context = {
        'grns': grns,
        'stores': stores,
    }
    
    return render(request, 'stores/receipt_history.html', context)


@login_required
def store_grn_detail_view(request, grn_id):
    """View GRN details and perform inspection"""
    
    grn = get_object_or_404(
        GoodsReceivedNote.objects.select_related(
            'purchase_order', 'purchase_order__supplier', 'store', 'received_by', 'inspected_by'
        ).prefetch_related('items__po_item'),
        id=grn_id
    )
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'inspect':
            grn.status = request.POST.get('status', 'ACCEPTED')
            grn.inspected_by = request.user
            grn.inspection_date = timezone.now().date()
            grn.general_condition = request.POST.get('general_condition', '')
            grn.rejection_reason = request.POST.get('rejection_reason', '')
            grn.notes = request.POST.get('notes', '')
            grn.save()
            
            # If accepted, update stock
            if grn.status == 'ACCEPTED':
                for grn_item in grn.items.all():
                    if grn_item.quantity_accepted > 0:
                        # Update or create stock item
                        stock_item, created = StockItem.objects.get_or_create(
                            store=grn.store,
                            item=grn_item.po_item.requisition_item.item,
                            defaults={
                                'quantity_on_hand': grn_item.quantity_accepted,
                                'average_unit_cost': grn_item.po_item.unit_price,
                                'last_restock_date': timezone.now().date()
                            }
                        )
                        
                        if not created:
                            # Update existing stock
                            balance_before = stock_item.quantity_on_hand
                            stock_item.quantity_on_hand += grn_item.quantity_accepted
                            stock_item.last_restock_date = timezone.now().date()
                            
                            # Recalculate average cost
                            total_cost = (stock_item.average_unit_cost * balance_before) + \
                                       (grn_item.po_item.unit_price * grn_item.quantity_accepted)
                            stock_item.average_unit_cost = total_cost / stock_item.quantity_on_hand
                            stock_item.save()
                            
                            # Record stock movement
                            StockMovement.objects.create(
                                stock_item=stock_item,
                                movement_type='RECEIPT',
                                reference_number=grn.grn_number,
                                reference_type='GRN',
                                quantity=grn_item.quantity_accepted,
                                unit_cost=grn_item.po_item.unit_price,
                                balance_before=balance_before,
                                balance_after=stock_item.quantity_on_hand,
                                to_store=grn.store,
                                performed_by=request.user
                            )
            
            messages.success(request, 'Inspection completed successfully!')
            return redirect('store_receipt_history')
    
    context = {
        'grn': grn,
    }
    
    return render(request, 'stores/grn_detail.html', context)


@login_required
def store_quality_issues_view(request):
    """View rejected or problematic deliveries"""
    
    quality_issues = GoodsReceivedNote.objects.filter(
        Q(status='REJECTED') | Q(status='PARTIAL')
    ).select_related(
        'purchase_order', 'purchase_order__supplier', 'store'
    ).order_by('-created_at')
    
    # Get items with quality issues
    problem_items = GRNItem.objects.filter(
        Q(item_status='REJECTED') | Q(item_status='DAMAGED')
    ).select_related(
        'grn', 'grn__purchase_order', 'grn__purchase_order__supplier', 'po_item'
    ).order_by('-grn__created_at')
    
    context = {
        'quality_issues': quality_issues,
        'problem_items': problem_items,
    }
    
    return render(request, 'stores/quality_issues.html', context)


# ============================================================================
# STOCK MANAGEMENT VIEWS
# ============================================================================

@login_required
def store_all_stock_view(request):
    """View all stock items"""
    
    stock_items = StockItem.objects.all().select_related(
        'store', 'item', 'item__category'
    ).order_by('store', 'item__name')
    
    # Filters
    store_filter = request.GET.get('store')
    category_filter = request.GET.get('category')
    search = request.GET.get('search')
    
    if store_filter:
        stock_items = stock_items.filter(store_id=store_filter)
    if category_filter:
        stock_items = stock_items.filter(item__category_id=category_filter)
    if search:
        stock_items = stock_items.filter(
            Q(item__name__icontains=search) | 
            Q(item__code__icontains=search)
        )
    
    # Calculate statistics
    stats = {
        'total_items': stock_items.count(),
        'total_value': stock_items.aggregate(total=Sum('total_value'))['total'] or 0,
        'low_stock': stock_items.filter(quantity_on_hand__lte=F('reorder_level')).count(),
        'out_of_stock': stock_items.filter(quantity_on_hand=0).count(),
    }
    
    stores = Store.objects.filter(is_active=True)
    
    context = {
        'stock_items': stock_items,
        'stores': stores,
        'stats': stats,
    }
    
    return render(request, 'stores/all_stock.html', context)


@login_required
def store_add_stock_item_view(request):
    """Add new stock item"""
    
    if request.method == 'POST':
        try:
            store = get_object_or_404(Store, id=request.POST.get('store'))
            item = get_object_or_404(Item, id=request.POST.get('item'))
            
            stock_item = StockItem.objects.create(
                store=store,
                item=item,
                quantity_on_hand=Decimal(request.POST.get('quantity_on_hand', '0')),
                reorder_level=Decimal(request.POST.get('reorder_level', '0')),
                max_stock_level=Decimal(request.POST.get('max_stock_level', '0')) if request.POST.get('max_stock_level') else None,
                average_unit_cost=Decimal(request.POST.get('average_unit_cost', '0'))
            )
            
            stock_item.total_value = stock_item.quantity_on_hand * stock_item.average_unit_cost
            stock_item.save()
            
            messages.success(request, 'Stock item added successfully!')
            return redirect('store_all_stock')
            
        except Exception as e:
            messages.error(request, f'Error adding stock item: {str(e)}')
    
    stores = Store.objects.filter(is_active=True)
    items = Item.objects.filter(is_active=True)
    
    context = {
        'stores': stores,
        'items': items,
    }
    
    return render(request, 'stores/add_stock_item.html', context)


@login_required
def store_low_stock_alert_view(request):
    """View items at or below reorder level"""
    
    low_stock_items = StockItem.objects.filter(
        quantity_on_hand__lte=F('reorder_level')
    ).select_related('store', 'item').order_by('quantity_on_hand')
    
    context = {
        'low_stock_items': low_stock_items,
    }
    
    return render(request, 'stores/low_stock_alert.html', context)


@login_required
def store_out_of_stock_view(request):
    """View out of stock items"""
    
    out_of_stock_items = StockItem.objects.filter(
        quantity_on_hand=0
    ).select_related('store', 'item', 'item__category').order_by('-last_issue_date')
    
    context = {
        'out_of_stock_items': out_of_stock_items,
    }
    
    return render(request, 'stores/out_of_stock.html', context)


# ============================================================================
# STOCK ISSUES VIEWS
# ============================================================================

@login_required
def store_pending_issues_view(request):
    """View pending stock issue requests"""
    
    pending_issues = StockIssue.objects.filter(
        status='PENDING'
    ).select_related(
        'store', 'department', 'requested_by'
    ).prefetch_related('items').order_by('-created_at')
    
    context = {
        'pending_issues': pending_issues,
    }
    
    return render(request, 'stores/pending_issues.html', context)


@login_required
def store_issue_stock_view(request, issue_id=None):
    """Issue stock to departments"""
    
    if issue_id:
        issue = get_object_or_404(StockIssue, id=issue_id)
        
        if request.method == 'POST':
            try:
                # Process the issue
                for issue_item in issue.items.all():
                    qty_to_issue = Decimal(request.POST.get(f'qty_issue_{issue_item.id}', '0'))
                    
                    if qty_to_issue > 0:
                        stock_item = issue_item.stock_item
                        
                        # Check availability
                        if qty_to_issue > stock_item.quantity_on_hand:
                            messages.error(request, f'Insufficient stock for {stock_item.item.name}')
                            continue
                        
                        # Update stock
                        balance_before = stock_item.quantity_on_hand
                        stock_item.quantity_on_hand -= qty_to_issue
                        stock_item.last_issue_date = timezone.now().date()
                        stock_item.save()
                        
                        # Update issue item
                        issue_item.quantity_issued = qty_to_issue
                        issue_item.save()
                        
                        # Record movement
                        StockMovement.objects.create(
                            stock_item=stock_item,
                            movement_type='ISSUE',
                            reference_number=issue.issue_number,
                            reference_type='Issue',
                            quantity=qty_to_issue,
                            unit_cost=stock_item.average_unit_cost,
                            balance_before=balance_before,
                            balance_after=stock_item.quantity_on_hand,
                            from_store=issue.store,
                            performed_by=request.user
                        )
                
                # Update issue status
                issue.status = 'ISSUED'
                issue.issued_by = request.user
                issue.issue_date = timezone.now().date()
                issue.save()
                
                messages.success(request, f'Stock issued successfully! Issue Number: {issue.issue_number}')
                return redirect('store_issue_history')
                
            except Exception as e:
                messages.error(request, f'Error issuing stock: {str(e)}')
        
        context = {
            'issue': issue,
        }
        return render(request, 'stores/process_issue.html', context)
    
    else:
        # Create new issue
        if request.method == 'POST':
            try:
                issue = StockIssue.objects.create(
                    store=get_object_or_404(Store, id=request.POST.get('store')),
                    department=get_object_or_404(Department, id=request.POST.get('department')),
                    requested_by=request.user,
                    purpose=request.POST.get('purpose'),
                    notes=request.POST.get('notes', ''),
                    status='PENDING'
                )
                
                # Add items
                item_count = int(request.POST.get('item_count', 0))
                for i in range(item_count):
                    stock_item_id = request.POST.get(f'stock_item_{i}')
                    quantity = request.POST.get(f'quantity_{i}')
                    
                    if stock_item_id and quantity:
                        StockIssueItem.objects.create(
                            stock_issue=issue,
                            stock_item=get_object_or_404(StockItem, id=stock_item_id),
                            quantity_requested=Decimal(quantity)
                        )
                
                messages.success(request, f'Stock issue request created: {issue.issue_number}')
                return redirect('store_pending_issues')
                
            except Exception as e:
                messages.error(request, f'Error creating issue: {str(e)}')
        
        stores = Store.objects.filter(is_active=True)
        departments = Department.objects.filter(is_active=True)
        
        context = {
            'stores': stores,
            'departments': departments,
        }
        return render(request, 'stores/create_issue.html', context)


@login_required
def store_issue_history_view(request):
    """View stock issue history"""
    
    issues = StockIssue.objects.all().select_related(
        'store', 'department', 'requested_by', 'issued_by'
    ).prefetch_related('items').order_by('-created_at')
    
    # Filters
    status_filter = request.GET.get('status')
    department_filter = request.GET.get('department')
    date_from = request.GET.get('date_from')
    
    if status_filter:
        issues = issues.filter(status=status_filter)
    if department_filter:
        issues = issues.filter(department_id=department_filter)
    if date_from:
        issues = issues.filter(created_at__gte=date_from)
    
    departments = Department.objects.filter(is_active=True)
    
    context = {
        'issues': issues,
        'departments': departments,
    }
    
    return render(request, 'stores/issue_history.html', context)


@login_required
def store_returns_view(request):
    """Handle stock returns"""
    
    # Get issues that might have returns
    issued_items = StockIssue.objects.filter(
        status='ISSUED'
    ).select_related('store', 'department').order_by('-issue_date')
    
    if request.method == 'POST':
        try:
            issue = get_object_or_404(StockIssue, id=request.POST.get('issue_id'))
            
            for issue_item in issue.items.all():
                qty_returned = Decimal(request.POST.get(f'qty_return_{issue_item.id}', '0'))
                
                if qty_returned > 0:
                    stock_item = issue_item.stock_item
                    
                    # Update stock
                    balance_before = stock_item.quantity_on_hand
                    stock_item.quantity_on_hand += qty_returned
                    stock_item.save()
                    
                    # Record movement
                    StockMovement.objects.create(
                        stock_item=stock_item,
                        movement_type='RETURN',
                        reference_number=f'RET-{issue.issue_number}',
                        reference_type='Return',
                        quantity=qty_returned,
                        unit_cost=stock_item.average_unit_cost,
                        balance_before=balance_before,
                        balance_after=stock_item.quantity_on_hand,
                        to_store=issue.store,
                        performed_by=request.user,
                        remarks=request.POST.get(f'return_reason_{issue_item.id}', '')
                    )
            
            messages.success(request, 'Stock return processed successfully!')
            return redirect('store_returns')
            
        except Exception as e:
            messages.error(request, f'Error processing return: {str(e)}')
    
    context = {
        'issued_items': issued_items,
    }
    
    return render(request, 'stores/returns.html', context)


# ============================================================================
# OTHER STOCK VIEWS
# ============================================================================

@login_required
def store_stock_takes_view(request):
    """Stock taking/physical verification"""
    
    if request.method == 'POST':
        # Process stock take
        store_id = request.POST.get('store')
        stock_items = StockItem.objects.filter(store_id=store_id)
        
        discrepancies = []
        
        for stock_item in stock_items:
            physical_count = Decimal(request.POST.get(f'physical_{stock_item.id}', '0'))
            system_count = stock_item.quantity_on_hand
            
            if physical_count != system_count:
                discrepancies.append({
                    'item': stock_item,
                    'system': system_count,
                    'physical': physical_count,
                    'variance': physical_count - system_count
                })
                
                # Create adjustment movement
                balance_before = stock_item.quantity_on_hand
                stock_item.quantity_on_hand = physical_count
                stock_item.save()
                
                StockMovement.objects.create(
                    stock_item=stock_item,
                    movement_type='ADJUSTMENT',
                    reference_number=f'ADJ-{timezone.now().strftime("%Y%m%d%H%M%S")}',
                    reference_type='Stock Take',
                    quantity=abs(physical_count - system_count),
                    balance_before=balance_before,
                    balance_after=physical_count,
                    performed_by=request.user,
                    remarks=request.POST.get(f'remarks_{stock_item.id}', 'Stock take adjustment')
                )
        
        messages.success(request, f'Stock take completed. {len(discrepancies)} discrepancies found.')
    
    stores = Store.objects.filter(is_active=True)
    
    context = {
        'stores': stores,
    }
    
    return render(request, 'stores/stock_takes.html', context)


@login_required
def store_stock_transfers_view(request):
    """Transfer stock between stores"""
    
    if request.method == 'POST':
        try:
            from_store = get_object_or_404(Store, id=request.POST.get('from_store'))
            to_store = get_object_or_404(Store, id=request.POST.get('to_store'))
            stock_item = get_object_or_404(StockItem, id=request.POST.get('stock_item'))
            quantity = Decimal(request.POST.get('quantity'))
            
            # Deduct from source
            if stock_item.quantity_on_hand < quantity:
                messages.error(request, 'Insufficient stock for transfer')
            else:
                balance_before = stock_item.quantity_on_hand
                stock_item.quantity_on_hand -= quantity
                stock_item.save()
                
                # Record outward movement
                transfer_ref = f'TRF-{timezone.now().strftime("%Y%m%d%H%M%S")}'
                StockMovement.objects.create(
                    stock_item=stock_item,
                    movement_type='TRANSFER',
                    reference_number=transfer_ref,
                    reference_type='Transfer Out',
                    quantity=quantity,
                    unit_cost=stock_item.average_unit_cost,
                    balance_before=balance_before,
                    balance_after=stock_item.quantity_on_hand,
                    from_store=from_store,
                    to_store=to_store,
                    performed_by=request.user
                )
                
                # Add to destination
                dest_stock, created = StockItem.objects.get_or_create(
                    store=to_store,
                    item=stock_item.item,
                    defaults={
                        'quantity_on_hand': quantity,
                        'average_unit_cost': stock_item.average_unit_cost,
                        'reorder_level': stock_item.reorder_level
                    }
                )
                
                if not created:
                    balance_before = dest_stock.quantity_on_hand
                    dest_stock.quantity_on_hand += quantity
                    dest_stock.save()
                    
                    # Record inward movement
                    StockMovement.objects.create(
                        stock_item=dest_stock,
                        movement_type='TRANSFER',
                        reference_number=transfer_ref,
                        reference_type='Transfer In',
                        quantity=quantity,
                        unit_cost=stock_item.average_unit_cost,
                        balance_before=balance_before,
                        balance_after=dest_stock.quantity_on_hand,
                        from_store=from_store,
                        to_store=to_store,
                        performed_by=request.user
                    )
                
                messages.success(request, 'Stock transferred successfully!')
                
        except Exception as e:
            messages.error(request, f'Error transferring stock: {str(e)}')
    
    stores = Store.objects.filter(is_active=True)
    
    context = {
        'stores': stores,
    }
    
    return render(request, 'stores/stock_transfers.html', context)


# ============================================================================
# ASSET MANAGEMENT VIEWS
# ============================================================================

@login_required
def store_all_assets_view(request):
    """View all assets"""
    
    assets = Asset.objects.all().select_related(
        'item', 'department', 'custodian'
    ).order_by('-created_at')
    
    # Filters
    status_filter = request.GET.get('status')
    department_filter = request.GET.get('department')
    search = request.GET.get('search')
    
    if status_filter:
        assets = assets.filter(status=status_filter)
    if department_filter:
        assets = assets.filter(department_id=department_filter)
    if search:
        assets = assets.filter(
            Q(asset_number__icontains=search) |
            Q(asset_tag__icontains=search) |
            Q(description__icontains=search)
        )
    
    departments = Department.objects.filter(is_active=True)
    
    context = {
        'assets': assets,
        'departments': departments,
    }
    
    return render(request, 'stores/all_assets.html', context)


@login_required
def store_register_asset_view(request):
    """Register new asset"""
    
    if request.method == 'POST':
        try:
            asset = Asset.objects.create(
                asset_number=request.POST.get('asset_number'),
                asset_tag=request.POST.get('asset_tag'),
                item=get_object_or_404(Item, id=request.POST.get('item')),
                description=request.POST.get('description'),
                serial_number=request.POST.get('serial_number', ''),
                model=request.POST.get('model', ''),
                brand=request.POST.get('brand', ''),
                acquisition_date=request.POST.get('acquisition_date'),
                acquisition_cost=Decimal(request.POST.get('acquisition_cost')),
                current_value=Decimal(request.POST.get('acquisition_cost')),
                depreciation_rate=Decimal(request.POST.get('depreciation_rate', '0')),
                useful_life_years=int(request.POST.get('useful_life_years', '0')),
                department=get_object_or_404(Department, id=request.POST.get('department')),
                location=request.POST.get('location'),
                custodian=get_object_or_404(User, id=request.POST.get('custodian')),
                warranty_expiry=request.POST.get('warranty_expiry') or None,
                notes=request.POST.get('notes', '')
            )
            
            messages.success(request, f'Asset {asset.asset_number} registered successfully!')
            return redirect('store_all_assets')
            
        except Exception as e:
            messages.error(request, f'Error registering asset: {str(e)}')
    
    items = Item.objects.filter(is_active=True)
    departments = Department.objects.filter(is_active=True)
    users = User.objects.filter(is_active=True)
    
    context = {
        'items': items,
        'departments': departments,
        'users': users,
    }
    
    return render(request, 'stores/register_asset.html', context)


@login_required
def store_maintenance_view(request):
    """Asset maintenance tracking"""
    
    assets = Asset.objects.filter(
        status__in=['ACTIVE', 'UNDER_MAINTENANCE']
    ).select_related('item', 'department', 'custodian')
    
    # Get assets due for maintenance (example: warranty expiring soon)
    assets_needing_attention = assets.filter(
        warranty_expiry__lte=timezone.now().date() + timedelta(days=90)
    )
    
    context = {
        'assets': assets,
        'assets_needing_attention': assets_needing_attention,
    }
    
    return render(request, 'stores/maintenance.html', context)


@login_required
def store_disposal_view(request):
    """Asset disposal management"""
    
    disposed_assets = Asset.objects.filter(
        status='DISPOSED'
    ).select_related('item', 'department').order_by('-disposal_date')
    
    if request.method == 'POST':
        try:
            asset = get_object_or_404(Asset, id=request.POST.get('asset_id'))
            asset.status = 'DISPOSED'
            asset.disposal_date = request.POST.get('disposal_date')
            asset.disposal_method = request.POST.get('disposal_method')
            asset.disposal_value = Decimal(request.POST.get('disposal_value', '0'))
            asset.notes = request.POST.get('disposal_notes', '')
            asset.save()
            
            messages.success(request, f'Asset {asset.asset_number} marked as disposed')
            return redirect('store_disposal')
            
        except Exception as e:
            messages.error(request, f'Error disposing asset: {str(e)}')
    
    # Get assets eligible for disposal
    eligible_assets = Asset.objects.filter(
        status__in=['DAMAGED', 'IDLE']
    ).select_related('item', 'department')
    
    context = {
        'disposed_assets': disposed_assets,
        'eligible_assets': eligible_assets,
    }
    
    return render(request, 'stores/disposal.html', context)


# ============================================================================
# REPORTS VIEWS
# ============================================================================

@login_required
def store_inventory_reports_view(request):
    """Generate inventory reports"""
    
    report_type = request.GET.get('report_type', 'stock_summary')
    
    if report_type == 'stock_summary':
        data = StockItem.objects.all().select_related(
            'store', 'item'
        ).values(
            'store__name', 'item__category__name'
        ).annotate(
            total_items=Count('id'),
            total_value=Sum('total_value'),
            total_quantity=Sum('quantity_on_hand')
        )
    
    elif report_type == 'low_stock':
        data = StockItem.objects.filter(
            quantity_on_hand__lte=F('reorder_level')
        ).select_related('store', 'item')
    
    elif report_type == 'stock_valuation':
        data = StockItem.objects.all().select_related(
            'store', 'item'
        ).order_by('-total_value')
    
    else:
        data = []
    
    context = {
        'report_type': report_type,
        'data': data,
    }
    
    return render(request, 'stores/inventory_reports.html', context)


@login_required
def store_movement_reports_view(request):
    """Generate stock movement reports"""
    
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).date())
    date_to = request.GET.get('date_to', timezone.now().date())
    movement_type = request.GET.get('movement_type')
    
    movements = StockMovement.objects.filter(
        movement_date__date__range=[date_from, date_to]
    ).select_related(
        'stock_item', 'stock_item__item', 'stock_item__store', 'performed_by'
    ).order_by('-movement_date')
    
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    
    # Summary statistics
    summary = movements.values('movement_type').annotate(
        total_movements=Count('id'),
        total_quantity=Sum('quantity')
    )
    
    context = {
        'movements': movements,
        'summary': summary,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'stores/movement_reports.html', context)


# ============================================================================
# HELP & SUPPORT VIEW
# ============================================================================

@login_required
def store_help_center_view(request):
    """Help and documentation"""
    
    context = {
        'page_title': 'Stores Help Center'
    }
    
    return render(request, 'stores/help_center.html', context)



# pms/views.py (Add these to your existing views.py)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from django.http import JsonResponse, HttpResponse
from django.contrib import messages

from .models import (
    User, Requisition, RequisitionApproval, Tender, Bid, PurchaseOrder,
    Invoice, Payment, Contract, Supplier, Budget, AuditLog,
    BidEvaluation, SupplierPerformance, StockMovement, Asset
)

# ============================================================================
# AUDITOR DASHBOARD & ANALYTICS
# ============================================================================
from django.db.models import Sum, Count, Q, Avg, F, Max, Min, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth, TruncDate, Coalesce
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import timedelta, datetime
from decimal import Decimal
import json

@login_required
def auditor_dashboard(request):
    """Auditor Dashboard with Analytics"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied. Auditor role required.')
        return redirect('dashboard')
    
    # Date ranges
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    six_months_ago = today - timedelta(days=180)
    this_month_start = today.replace(day=1)
    
    # ========================================================================
    # KEY METRICS
    # ========================================================================
    
    # Compliance Stats
    total_requisitions = Requisition.objects.count()
    emergency_requisitions = Requisition.objects.filter(is_emergency=True).count()
    unplanned_requisitions = Requisition.objects.filter(is_planned=False).count()
    
    total_tenders = Tender.objects.count()
    total_contracts = Contract.objects.count()
    total_purchase_orders = PurchaseOrder.objects.count()
    
    # High-value transactions
    high_value_reqs = Requisition.objects.filter(
        estimated_amount__gte=1000000
    ).count()
    
    high_value_pos = PurchaseOrder.objects.filter(
        total_amount__gte=1000000
    ).count()
    
    # Pending reviews
    pending_reviews_count = Requisition.objects.filter(
        status__in=['SUBMITTED', 'HOD_APPROVED', 'FACULTY_APPROVED']
    ).count()
    
    # Flagged items
    flagged_count = (
        emergency_requisitions + 
        unplanned_requisitions + 
        Tender.objects.annotate(bid_count=Count('bids')).filter(bid_count=1).count()
    )
    
    # Budget compliance
    budgets_with_overruns = Budget.objects.filter(
        actual_spent__gt=F('allocated_amount')
    ).count()
    
    total_budgets = Budget.objects.count()
    budget_compliance_rate = ((total_budgets - budgets_with_overruns) / total_budgets * 100) if total_budgets > 0 else 100
    
    # Payment compliance
    total_invoices = Invoice.objects.count()
    paid_on_time = Invoice.objects.filter(
        status='PAID',
        payment_date__lte=F('due_date')
    ).count()
    payment_compliance_rate = (paid_on_time / total_invoices * 100) if total_invoices > 0 else 100
    
    # Supplier compliance
    total_suppliers = Supplier.objects.filter(status='APPROVED').count()
    compliant_suppliers = Supplier.objects.filter(
        status='APPROVED',
        tax_compliance_expiry__gte=today
    ).count()
    supplier_compliance_rate = (compliant_suppliers / total_suppliers * 100) if total_suppliers > 0 else 100
    
    # ========================================================================
    # CHART 1: COMPLIANCE OVERVIEW (Donut Chart)
    # ========================================================================
    
    compliance_overview_chart = json.dumps({
        'labels': ['Budget Compliance', 'Payment Compliance', 'Supplier Compliance', 'Procurement Plan Compliance'],
        'values': [
            round(budget_compliance_rate, 1),
            round(payment_compliance_rate, 1),
            round(supplier_compliance_rate, 1),
            85.0  # Placeholder for procurement plan compliance
        ],
        'colors': ['#10B981', '#6366F1', '#F59E0B', '#2563EB']
    })
    
    # ========================================================================
    # CHART 2: RISK INDICATORS (Radar Chart)
    # ========================================================================
    
    # Calculate risk scores (0-100, higher = more risk)
    emergency_risk = (emergency_requisitions / total_requisitions * 100) if total_requisitions > 0 else 0
    unplanned_risk = (unplanned_requisitions / total_requisitions * 100) if total_requisitions > 0 else 0
    
    single_bid_tenders = Tender.objects.annotate(
        bid_count=Count('bids')
    ).filter(bid_count=1).count()
    single_bid_risk = (single_bid_tenders / total_tenders * 100) if total_tenders > 0 else 0
    
    overdue_pos = PurchaseOrder.objects.filter(
        status__in=['SENT', 'ACKNOWLEDGED'],
        delivery_date__lt=today
    ).count()
    delivery_risk = (overdue_pos / total_purchase_orders * 100) if total_purchase_orders > 0 else 0
    
    overdue_invoices = Invoice.objects.filter(
        status__in=['SUBMITTED', 'APPROVED'],
        due_date__lt=today
    ).count()
    payment_risk = (overdue_invoices / total_invoices * 100) if total_invoices > 0 else 0
    
    budget_risk = (budgets_with_overruns / total_budgets * 100) if total_budgets > 0 else 0
    
    risk_radar_chart = json.dumps({
        'labels': ['Emergency Procurements', 'Unplanned Items', 'Single Bids', 
                   'Delivery Delays', 'Payment Delays', 'Budget Overruns'],
        'values': [
            round(emergency_risk, 1),
            round(unplanned_risk, 1),
            round(single_bid_risk, 1),
            round(delivery_risk, 1),
            round(payment_risk, 1),
            round(budget_risk, 1)
        ]
    })
    
    # ========================================================================
    # CHART 3: PROCUREMENT TREND (Last 6 Months)
    # ========================================================================
    
    six_months_data = []
    for i in range(6):
        month_start = (today.replace(day=1) - timedelta(days=30*i)).replace(day=1)
        if i < 5:
            month_end = (today.replace(day=1) - timedelta(days=30*(i-1))).replace(day=1) - timedelta(days=1)
        else:
            month_end = today
        
        reqs = Requisition.objects.filter(
            created_at__range=[month_start, month_end]
        ).count()
        
        pos = PurchaseOrder.objects.filter(
            created_at__range=[month_start, month_end]
        ).count()
        
        payments = Payment.objects.filter(
            created_at__range=[month_start, month_end],
            status='COMPLETED'
        ).count()
        
        six_months_data.insert(0, {
            'month': month_start.strftime('%b %Y'),
            'requisitions': reqs,
            'pos': pos,
            'payments': payments
        })
    
    procurement_trend_chart = json.dumps({
        'labels': [d['month'] for d in six_months_data],
        'requisitions': [d['requisitions'] for d in six_months_data],
        'purchase_orders': [d['pos'] for d in six_months_data],
        'payments': [d['payments'] for d in six_months_data],
    })
    
    # ========================================================================
    # CHART 4: TRANSACTION VALUE BY DEPARTMENT (Bar Chart)
    # ========================================================================
    
    dept_spending = Budget.objects.values(
        'department__name', 'department__code'
    ).annotate(
        total_spent=Sum('actual_spent')
    ).order_by('-total_spent')[:8]
    
    dept_spending_chart = json.dumps({
        'labels': [d['department__code'] for d in dept_spending],
        'values': [float(d['total_spent']) for d in dept_spending],
        'full_names': [d['department__name'] for d in dept_spending]
    })
    
    # ========================================================================
    # FLAGGED TRANSACTIONS (FIXED)
    # ========================================================================
    
    flagged_items = []
    
    # Emergency requisitions
    emergency_reqs = Requisition.objects.filter(
        is_emergency=True,
        created_at__gte=thirty_days_ago
    ).select_related('department', 'requested_by')[:5]
    
    for req in emergency_reqs:
        flagged_items.append({
            'type': 'Emergency Requisition',
            'reference': req.requisition_number,
            'description': req.title,
            'amount': req.estimated_amount,
            'date': req.created_at.date(),  # Convert to date
            'flag': 'Emergency Procurement',
            'sort_date': req.created_at  # Keep datetime for sorting
        })
    
    # Single-bid tenders
    single_bid_tenders = Tender.objects.annotate(
        bid_count=Count('bids')
    ).filter(bid_count=1, created_at__gte=thirty_days_ago)[:5]
    
    for tender in single_bid_tenders:
        flagged_items.append({
            'type': 'Tender',
            'reference': tender.tender_number,
            'description': tender.title,
            'amount': tender.estimated_budget,
            'date': tender.created_at.date(),  # Convert to date
            'flag': 'Single Bid Received',
            'sort_date': tender.created_at  # Keep datetime for sorting
        })
    
    # Overdue purchase orders
    overdue_pos_list = PurchaseOrder.objects.filter(
        status__in=['SENT', 'ACKNOWLEDGED'],
        delivery_date__lt=today
    ).select_related('supplier')[:5]
    
    for po in overdue_pos_list:
        # Convert delivery_date (date) to datetime for consistent sorting
        delivery_datetime = datetime.combine(po.delivery_date, datetime.min.time())
        delivery_datetime = timezone.make_aware(delivery_datetime)
        
        flagged_items.append({
            'type': 'Purchase Order',
            'reference': po.po_number,
            'description': f"Overdue delivery - {po.supplier.name}",
            'amount': po.total_amount,
            'date': po.delivery_date,  # Keep as date for display
            'flag': 'Overdue Delivery',
            'sort_date': delivery_datetime  # Use datetime for sorting
        })
    
    # Sort by datetime, then remove the sort_date field
    flagged_items.sort(key=lambda x: x['sort_date'], reverse=True)
    for item in flagged_items:
        del item['sort_date']
    
    # ========================================================================
    # RECENT ACTIVITIES
    # ========================================================================
    
    recent_activities = AuditLog.objects.select_related('user').order_by('-timestamp')[:15]
    
    # Pending high-value reviews
    pending_reviews = Requisition.objects.filter(
        estimated_amount__gte=500000,
        status__in=['SUBMITTED', 'HOD_APPROVED', 'FACULTY_APPROVED']
    ).select_related('department', 'requested_by').order_by('-created_at')[:10]
    
    # High-value purchase orders
    high_value_pos_list = PurchaseOrder.objects.filter(
        total_amount__gte=1000000
    ).select_related('supplier', 'requisition').order_by('-created_at')[:10]
    
    # ========================================================================
    # COMPLIANCE ALERTS
    # ========================================================================
    
    # Budget overruns
    budget_overruns = Budget.objects.filter(
        actual_spent__gt=F('allocated_amount')
    ).select_related('department', 'category')[:5]
    
    # Expired supplier documents
    expired_supplier_docs = Supplier.objects.filter(
        Q(tax_compliance_expiry__lt=today) | 
        Q(registration_expiry__lt=today),
        status='APPROVED'
    )[:5]
    
    # Late payments 
    late_payments = Invoice.objects.filter(
        status__in=['SUBMITTED', 'APPROVED'],
        due_date__lt=today
    ).select_related('supplier').order_by('due_date')[:5]
    
    # Contracts expiring soon (within 30 days)
    expiring_contracts = Contract.objects.filter(
        status='ACTIVE',
        end_date__lte=today + timedelta(days=30),
        end_date__gte=today
    ).select_related('supplier').order_by('end_date')[:5]
    
    # ========================================================================
    # USER ACTIVITY SUMMARY
    # ========================================================================
    
    user_actions_today = AuditLog.objects.filter(
        timestamp__date=today
    ).values('action').annotate(count=Count('id'))
    
    most_active_users = AuditLog.objects.filter(
        timestamp__gte=thirty_days_ago
    ).values('user__username', 'user__first_name', 'user__last_name').annotate(
        action_count=Count('id')
    ).order_by('-action_count')[:5]
    
    # ========================================================================
    # CONTEXT
    # ========================================================================
    
    context = {
        # Key Metrics
        'total_requisitions': total_requisitions,
        'emergency_requisitions': emergency_requisitions,
        'unplanned_requisitions': unplanned_requisitions,
        'total_tenders': total_tenders,
        'total_contracts': total_contracts,
        'total_purchase_orders': total_purchase_orders,
        'high_value_reqs': high_value_reqs,
        'high_value_pos': high_value_pos,
        'pending_reviews_count': pending_reviews_count,
        'flagged_count': flagged_count,
        
        # Compliance Rates
        'budget_compliance_rate': round(budget_compliance_rate, 1),
        'payment_compliance_rate': round(payment_compliance_rate, 1),
        'supplier_compliance_rate': round(supplier_compliance_rate, 1),
        
        # Chart Data (JSON) - Only 4 essential charts
        'compliance_overview_chart': compliance_overview_chart,
        'risk_radar_chart': risk_radar_chart,
        'procurement_trend_chart': procurement_trend_chart,
        'dept_spending_chart': dept_spending_chart,
        
        # Lists
        'recent_activities': recent_activities,
        'pending_reviews': pending_reviews,
        'high_value_pos_list': high_value_pos_list,
        'flagged_items': flagged_items[:10],
        
        # Alerts
        'budget_overruns': budget_overruns,
        'expired_supplier_docs': expired_supplier_docs,
        'late_payments': late_payments,
        'expiring_contracts': expiring_contracts,
        
        # User Activity
        'user_actions_today': user_actions_today,
        'most_active_users': most_active_users,
    }
    
    return render(request, 'auditor/dashboard.html', context)

@login_required
def auditor_analytics_view(request):
    """Audit analytics and insights"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Spending patterns
    monthly_spend = PurchaseOrder.objects.filter(
        status='APPROVED',
        created_at__year=timezone.now().year
    ).values('created_at__month').annotate(
        total=Sum('total_amount')
    )
    
    # Department spending
    dept_spending = Budget.objects.values(
        'department__name'
    ).annotate(
        allocated=Sum('allocated_amount'),
        spent=Sum('actual_spent')
    )
    
    # Supplier analysis
    top_suppliers = Supplier.objects.annotate(
        total_orders=Count('purchase_orders'),
        total_value=Sum('purchase_orders__total_amount')
    ).order_by('-total_value')[:10]
    
    context = {
        'monthly_spend': monthly_spend,
        'dept_spending': dept_spending,
        'top_suppliers': top_suppliers,
    }
    
    return render(request, 'auditor/analytics.html', context)


# ============================================================================
# AUDIT MANAGEMENT
# ============================================================================

@login_required
def auditor_all_audits_view(request):
    """List all audits"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Create Audit model first
    # For now, showing audit logs as placeholder
    audit_logs = AuditLog.objects.select_related('user').all()
    
    context = {
        'audit_logs': audit_logs,
    }
    
    return render(request, 'auditor/audits/all_audits.html', context)


@login_required
def auditor_new_audit_view(request):
    """Create new audit"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Implement audit creation
    
    return render(request, 'auditor/audits/new_audit.html')


@login_required
def auditor_active_audits_view(request):
    """List active audits"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Implement when Audit model exists
    
    return render(request, 'auditor/audits/active_audits.html')


@login_required
def auditor_completed_audits_view(request):
    """List completed audits"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Implement when Audit model exists
    
    return render(request, 'auditor/audits/completed_audits.html')


# ============================================================================
# TRANSACTION REVIEWS
# ============================================================================

@login_required
def auditor_requisitions_review_view(request):
    """Review requisitions"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    status_filter = request.GET.get('status', '')
    department_filter = request.GET.get('department', '')
    
    requisitions = Requisition.objects.select_related(
        'department', 'requested_by', 'budget'
    ).all()
    
    if status_filter:
        requisitions = requisitions.filter(status=status_filter)
    
    if department_filter:
        requisitions = requisitions.filter(department_id=department_filter)
    
    # Flag suspicious requisitions
    for req in requisitions:
        req.is_suspicious = False
        # Check for red flags
        if req.is_emergency and not req.emergency_justification:
            req.is_suspicious = True
        if req.estimated_amount > 1000000:  # High value
            req.is_suspicious = True
    
    context = {
        'requisitions': requisitions,
        'status_choices': Requisition.STATUS_CHOICES,
    }
    
    return render(request, 'auditor/reviews/requisitions.html', context)


@login_required
def auditor_purchase_orders_review_view(request):
    """Review purchase orders"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    purchase_orders = PurchaseOrder.objects.select_related(
        'requisition', 'supplier', 'created_by'
    ).all()
    
    # Add compliance checks
    for po in purchase_orders:
        po.compliance_issues = []
        
        # Check if PO matches requisition
        if po.total_amount > po.requisition.estimated_amount * Decimal('1.1'):
            po.compliance_issues.append('PO amount exceeds requisition by >10%')
        
        # Check approval trail
        if not po.approved_by:
            po.compliance_issues.append('Missing approval')
        
        # Check delivery timeline
        if po.delivery_date < timezone.now().date():
            po.compliance_issues.append('Overdue delivery')
    
    context = {
        'purchase_orders': purchase_orders,
    }
    
    return render(request, 'auditor/reviews/purchase_orders.html', context)


@login_required
def auditor_payments_review_view(request):
    """Review payments"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    payments = Payment.objects.select_related(
        'invoice', 'invoice__supplier', 'processed_by'
    ).all()
    
    # Add audit checks
    for payment in payments:
        payment.audit_flags = []
        
        # Check if payment matches invoice
        if payment.payment_amount > payment.invoice.total_amount:
            payment.audit_flags.append('Payment exceeds invoice amount')
        
        # Check if invoice was approved
        if not payment.invoice.approved_by:
            payment.audit_flags.append('Invoice not approved')
        
        # Check for duplicate payments
        duplicate_count = Payment.objects.filter(
            invoice=payment.invoice,
            status='COMPLETED'
        ).count()
        if duplicate_count > 1:
            payment.audit_flags.append('Possible duplicate payment')
    
    context = {
        'payments': payments,
    }
    
    return render(request, 'auditor/reviews/payments.html', context)


@login_required
def auditor_contracts_review_view(request):
    """Review contracts"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    contracts = Contract.objects.select_related(
        'supplier', 'contract_manager'
    ).all()
    
    # Add compliance checks
    for contract in contracts:
        contract.compliance_status = 'OK'
        contract.issues = []
        
        # Check if contract is signed
        if not contract.signed_by_university or not contract.signed_by_supplier:
            contract.compliance_status = 'Warning'
            contract.issues.append('Contract not fully signed')
        
        # Check if contract is expiring soon
        if contract.end_date <= timezone.now().date() + timedelta(days=30):
            contract.issues.append('Contract expiring within 30 days')
        
        # Check performance bond if required
        if contract.performance_bond_required and contract.performance_bond_amount == 0:
            contract.compliance_status = 'Critical'
            contract.issues.append('Performance bond required but not recorded')
    
    context = {
        'contracts': contracts,
    }
    
    return render(request, 'auditor/reviews/contracts.html', context)


# ============================================================================
# COMPLIANCE
# ============================================================================

@login_required
def auditor_compliance_review_view(request):
    """Overall compliance dashboard"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Compliance metrics
    total_requisitions = Requisition.objects.count()
    approved_requisitions = Requisition.objects.filter(
        status='APPROVED'
    ).count()
    
    total_tenders = Tender.objects.count()
    properly_evaluated = Tender.objects.filter(
        status='AWARDED'
    ).count()
    
    total_payments = Payment.objects.count()
    timely_payments = Payment.objects.filter(
        status='COMPLETED',
        payment_date__lte=timezone.now().date()
    ).count()
    
    # Supplier compliance
    suppliers_compliant = Supplier.objects.filter(
        status='APPROVED',
        tax_compliance_expiry__gte=timezone.now().date()
    ).count()
    
    total_suppliers = Supplier.objects.filter(status='APPROVED').count()
    
    context = {
        'requisition_compliance': (approved_requisitions / total_requisitions * 100) if total_requisitions > 0 else 0,
        'tender_compliance': (properly_evaluated / total_tenders * 100) if total_tenders > 0 else 0,
        'payment_compliance': (timely_payments / total_payments * 100) if total_payments > 0 else 0,
        'supplier_compliance': (suppliers_compliant / total_suppliers * 100) if total_suppliers > 0 else 0,
        
        'total_requisitions': total_requisitions,
        'total_tenders': total_tenders,
        'total_payments': total_payments,
        'total_suppliers': total_suppliers,
    }
    
    return render(request, 'auditor/compliance/overview.html', context)


@login_required
def auditor_flagged_items_view(request):
    """View flagged/suspicious items"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # High-value requisitions
    high_value_reqs = Requisition.objects.filter(
        estimated_amount__gte=1000000
    ).select_related('department', 'requested_by')
    
    # Emergency procurements
    emergency_procs = Requisition.objects.filter(
        is_emergency=True
    ).select_related('department', 'requested_by')
    
    # Unplanned requisitions
    unplanned_reqs = Requisition.objects.filter(
        is_planned=False
    ).select_related('department', 'requested_by')
    
    # Direct procurements (tenders with only 1 bid)
    direct_tenders = Tender.objects.annotate(
        bid_count=Count('bids')
    ).filter(bid_count=1)
    
    # Late payments
    late_payments = Payment.objects.filter(
        status='PENDING',
        invoice__due_date__lt=timezone.now().date()
    ).select_related('invoice', 'invoice__supplier')
    
    # Overdue purchase orders
    overdue_pos = PurchaseOrder.objects.filter(
        status__in=['SENT', 'ACKNOWLEDGED'],
        delivery_date__lt=timezone.now().date()
    ).select_related('supplier', 'requisition')
    
    context = {
        'high_value_reqs': high_value_reqs,
        'emergency_procs': emergency_procs,
        'unplanned_reqs': unplanned_reqs,
        'direct_tenders': direct_tenders,
        'late_payments': late_payments,
        'overdue_pos': overdue_pos,
    }
    
    return render(request, 'auditor/flagged_items.html', context)


# ============================================================================
# SYSTEM AUDIT
# ============================================================================

@login_required
def auditor_audit_trail_view(request):
    """System audit trail"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Filters
    action_filter = request.GET.get('action', '')
    user_filter = request.GET.get('user', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    audit_logs = AuditLog.objects.select_related('user').all()
    
    if action_filter:
        audit_logs = audit_logs.filter(action=action_filter)
    
    if user_filter:
        audit_logs = audit_logs.filter(user_id=user_filter)
    
    if date_from:
        audit_logs = audit_logs.filter(timestamp__gte=date_from)
    
    if date_to:
        audit_logs = audit_logs.filter(timestamp__lte=date_to)
    
    audit_logs = audit_logs[:1000]  # Limit results
    
    context = {
        'audit_logs': audit_logs,
        'action_types': AuditLog.ACTION_TYPES,
    }
    
    return render(request, 'auditor/system_audit/audit_trail.html', context)


@login_required
def auditor_activity_logs_view(request):
    """User activity logs"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Group activities by user
    user_activities = AuditLog.objects.values(
        'user__username', 'user__get_full_name'
    ).annotate(
        total_actions=Count('id'),
        last_activity=Max('timestamp')
    ).order_by('-total_actions')
    
    # Recent activities
    recent_activities = AuditLog.objects.select_related('user').order_by(
        '-timestamp'
    )[:100]
    
    context = {
        'user_activities': user_activities,
        'recent_activities': recent_activities,
    }
    
    return render(request, 'auditor/system_audit/activity_logs.html', context)


@login_required
def auditor_access_logs_view(request):
    """Access and authentication logs"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Login activities
    login_logs = AuditLog.objects.filter(
        action__in=['LOGIN', 'LOGOUT']
    ).select_related('user').order_by('-timestamp')[:200]
    
    # Failed login attempts (would need to be tracked separately)
    # Placeholder for now
    
    context = {
        'login_logs': login_logs,
    }
    
    return render(request, 'auditor/system_audit/access_logs.html', context)


@login_required
def auditor_data_changes_view(request):
    """Track data changes and modifications"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Get all UPDATE and DELETE actions
    data_changes = AuditLog.objects.filter(
        action__in=['UPDATE', 'DELETE']
    ).select_related('user').order_by('-timestamp')[:500]
    
    # Group by model
    changes_by_model = AuditLog.objects.filter(
        action__in=['UPDATE', 'DELETE']
    ).values('model_name').annotate(
        change_count=Count('id')
    ).order_by('-change_count')
    
    context = {
        'data_changes': data_changes,
        'changes_by_model': changes_by_model,
    }
    
    return render(request, 'auditor/system_audit/data_changes.html', context)


# ============================================================================
# REPORTS & FINDINGS
# ============================================================================

@login_required
def auditor_audit_reports_view(request):
    """List of audit reports"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Create AuditReport model
    # Placeholder using ProcurementReport for now
    from .models import ProcurementReport
    
    reports = ProcurementReport.objects.filter(
        report_type='COMPLIANCE'
    ).order_by('-generated_at')
    
    context = {
        'reports': reports,
    }
    
    return render(request, 'auditor/reports/audit_reports.html', context)


@login_required
def auditor_findings_view(request):
    """Audit findings and issues"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Create Finding model
    # Placeholder - showing flagged items as findings
    
    context = {
        'findings': [],  # Placeholder
    }
    
    return render(request, 'auditor/reports/findings.html', context)


@login_required
def auditor_recommendations_view(request):
    """Audit recommendations"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # TODO: Create Recommendation model
    
    context = {
        'recommendations': [],  # Placeholder
    }
    
    return render(request, 'auditor/reports/recommendations.html', context)


@login_required
def auditor_risk_assessment_view(request):
    """Risk assessment dashboard"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    # Risk indicators
    
    # 1. Budget overruns
    budget_risks = Budget.objects.filter(
        actual_spent__gt=F('allocated_amount')
    ).select_related('department', 'category')
    
    # 2. Suppliers with poor performance
    poor_suppliers = SupplierPerformance.objects.filter(
        overall_rating__lt=3
    ).values('supplier__name').annotate(
        avg_rating=Avg('overall_rating'),
        review_count=Count('id')
    )
    
    # 3. Delayed deliveries
    delayed_pos = PurchaseOrder.objects.filter(
        status__in=['SENT', 'ACKNOWLEDGED', 'PARTIAL_DELIVERY'],
        delivery_date__lt=timezone.now().date()
    ).count()
    
    # 4. High-value single-source procurements
    single_source = Tender.objects.filter(
        procurement_method='DIRECT',
        estimated_budget__gte=500000
    ).count()
    
    context = {
        'budget_risks': budget_risks,
        'poor_suppliers': poor_suppliers,
        'delayed_deliveries': delayed_pos,
        'single_source_count': single_source,
    }
    
    return render(request, 'auditor/reports/risk_assessment.html', context)


# ============================================================================
# HELP & SUPPORT
# ============================================================================

@login_required
def auditor_help_center_view(request):
    """Auditor help center"""
    if request.user.role != 'AUDITOR':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    return render(request, 'auditor/help_center.html')